"""INFORME MENSUAL EXHAUSTIVO — reconstruye la línea de tiempo de cada cliente en un mes:
recepción de correos/documentos, envíos a mesa, respuestas (aprobación/observación/rechazo),
escrituración, detenciones con motivo, y detecta CLIENTES RECUPERABLES."""
import re
import unicodedata
import uuid
from datetime import datetime, timezone

from database import db

ETAPAS = ["recibido", "en_mesa", "observado", "rechazado", "aprobado", "escriturado"]
RX_ESCRITURA = re.compile(r"escritur|notar[ií]a|firma\b|borrador", re.I)
RX_RUIDO = re.compile(
    r"documentaci|firma|escritur|carpeta|formato|suscripci|google|coursiv|"
    r"respuestas mesa|desconocido|rechazado para|rectificatoria|con documentos|"
    r"@|propiedad|s[uú]per |pago |central mutuos|emergent|urgente|aprobaci|"
    r"t[ií]tulos|mesa clientes", re.I)


def _norm(s):
    s = unicodedata.normalize("NFD", (s or "").lower())
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return re.sub(r"\s+", " ", re.sub(r"[^a-z0-9ñ ]", " ", s)).strip()


def _fecha(v):
    return (str(v) or "")[:10]


class _Matcher:
    def __init__(self, folders):
        self.folders = [(f["nombre"], set(_norm(f["nombre"]).split()), f) for f in folders]
        self.exacto = {_norm(f["nombre"]): f["nombre"] for f in folders}

    def match(self, nombre):
        n = _norm(nombre)
        if not n or len(n) < 5:
            return None
        if n in self.exacto:
            return self.exacto[n]
        palabras = set(n.split())
        if len(palabras) < 2:
            return None
        mejor, mejor_score = None, 0
        for fname, fwords, _f in self.folders:
            if palabras <= fwords or fwords <= palabras:
                score = len(palabras & fwords)
                if score > mejor_score:
                    mejor, mejor_score = fname, score
        return mejor


async def generar_informe(mes):
    ini, fin = f"{mes}-01", f"{mes}-32"
    folders = await db.folders.find({}, {"_id": 0, "nombre": 1, "rut": 1, "created_at": 1,
                                         "credit_request": 1, "is_escrituracion": 1,
                                         "escrituracion_movida_at": 1,
                                         "faltantes_auto_lista": 1}).to_list(1000)
    m = _Matcher(folders)
    clientes = {}

    def cli(nombre):
        canon = m.match(nombre) or nombre.strip()
        if _norm(canon) == "central mutuos":
            return None
        if canon not in clientes and not m.match(nombre) and RX_RUIDO.search(nombre or ""):
            return None
        c = clientes.setdefault(canon, {"cliente": canon, "rut": "", "eventos": [],
                                        "faltantes": [], "solicito_credito": False})
        return c

    for f in folders:
        creado = _fecha(f.get("created_at"))
        c = None
        if ini <= creado < fin:
            c = cli(f["nombre"])
            if c:
                c["eventos"].append({"fecha": creado, "tipo": "carpeta_creada",
                                     "detalle": "Carpeta de cliente creada"})
        movida = _fecha(f.get("escrituracion_movida_at"))
        if ini <= movida < fin:
            c = c or cli(f["nombre"])
            if c:
                c["eventos"].append({"fecha": movida, "tipo": "escriturado",
                                     "detalle": "Movido a escrituración"})
        if c is not None:
            c["rut"] = f.get("rut") or c["rut"]
            c["solicito_credito"] = bool(f.get("credit_request")) or c["solicito_credito"]
            c["faltantes"] = f.get("faltantes_auto_lista") or c["faltantes"]

    async for e in db.mesa_enviados.find({"enviado_at": {"$gte": ini, "$lt": fin}}, {"_id": 0}):
        c = cli(e.get("cliente") or "")
        if c is None:
            continue
        c["eventos"].append({"fecha": _fecha(e.get("enviado_at")), "tipo": "enviado_a_mesa",
                             "detalle": (e.get("subject") or "")[:120]})

    async for s in db.seguimiento.find({"fecha": {"$gte": ini, "$lt": fin}}, {"_id": 0}):
        c = cli(s.get("cliente") or "")
        if c is None:
            continue
        asunto = s.get("asunto") or ""
        tipo = {"aprobacion": "aprobado", "observacion": "observado",
                "rechazo": "rechazado"}.get(s.get("estado"), "respuesta_mesa")
        if RX_ESCRITURA.search(asunto):
            tipo = "escriturado" if tipo == "aprobado" else tipo
            c["eventos"].append({"fecha": _fecha(s.get("fecha")), "tipo": "gestion_escritura",
                                 "detalle": asunto[:120]})
        c["eventos"].append({"fecha": _fecha(s.get("fecha")), "tipo": tipo,
                             "detalle": asunto[:120]})
        c["rut"] = c["rut"] or (s.get("rut") or "")

    async for cc in db.correos_clasificacion.find(
            {"fecha_recepcion": {"$gte": ini, "$lt": fin},
             "cliente_nombre": {"$nin": ["", None]}}, {"_id": 0}):
        c = cli(cc.get("cliente_nombre") or "")
        if c is None:
            continue
        c["eventos"].append({"fecha": _fecha(cc.get("fecha_recepcion")), "tipo": "correo_recibido",
                             "detalle": f"{cc.get('clasificacion')}: {(cc.get('asunto') or '')[:100]}"})
        c["rut"] = c["rut"] or (cc.get("cliente_rut") or "")

    hoy = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    salida = []
    for c in clientes.values():
        if not c["eventos"]:
            continue
        c["eventos"].sort(key=lambda e: e["fecha"])
        tipos = [e["tipo"] for e in c["eventos"]]
        ultimo = c["eventos"][-1]
        etapa = "recibido"
        for e in ["en_mesa" if "enviado_a_mesa" in tipos else "",
                  "observado" if "observado" in tipos else "",
                  "aprobado" if "aprobado" in tipos else "",
                  "rechazado" if "rechazado" in tipos else "",
                  "escriturado" if ("escriturado" in tipos or "gestion_escritura" in tipos) else ""]:
            if e:
                etapa = e
        dias_detenido = (datetime.strptime(hoy, "%Y-%m-%d")
                         - datetime.strptime(ultimo["fecha"], "%Y-%m-%d")).days
        motivo = ""
        if etapa == "recibido":
            motivo = "Documentación recibida pero NUNCA se envió a mesa"
            if c["faltantes"]:
                motivo += f" (faltan: {', '.join(c['faltantes'][:4])})"
        elif etapa == "en_mesa" and tipos[-1] == "enviado_a_mesa":
            motivo = f"Enviado a mesa el {ultimo['fecha']} SIN respuesta registrada"
        elif etapa == "observado":
            motivo = "Mesa devolvió OBSERVACIÓN y no consta reenvío posterior"
        elif etapa == "aprobado":
            motivo = "APROBADO por mesa sin avance a escrituración registrado"
        elif etapa == "rechazado":
            motivo = "RECHAZADO por mesa (terminal, revisar re-evaluación)"
        elif etapa == "escriturado":
            motivo = "En etapa de escrituración/firma"
        recuperable = etapa not in ("escriturado",) and dias_detenido >= 7
        accion = ""
        if recuperable:
            accion = {"recibido": "Completar carpeta y enviar a mesa",
                      "en_mesa": "Reclamar respuesta a la mesa",
                      "observado": "Subsanar reparos y reenviar a mesa",
                      "aprobado": "Coordinar escrituración con notaría/inmobiliaria",
                      "rechazado": "Evaluar mejoras del perfil y re-presentar"}.get(etapa, "")
        salida.append({**c, "etapa": etapa, "ultima_actividad": ultimo["fecha"],
                       "dias_sin_movimiento": dias_detenido, "motivo_detencion": motivo,
                       "recuperable": recuperable, "accion_sugerida": accion,
                       "n_eventos": len(c["eventos"])})
    salida.sort(key=lambda x: (not x["recuperable"], -x["dias_sin_movimiento"]))
    resumen = {
        "mes": mes, "total_clientes_con_actividad": len(salida),
        "enviados_a_mesa": sum(1 for c in salida if c["etapa"] not in ("recibido",)),
        "aprobados": sum(1 for c in salida if c["etapa"] == "aprobado"),
        "observados": sum(1 for c in salida if c["etapa"] == "observado"),
        "rechazados": sum(1 for c in salida if c["etapa"] == "rechazado"),
        "escriturados": sum(1 for c in salida if c["etapa"] == "escriturado"),
        "nunca_enviados_a_mesa": sum(1 for c in salida if c["etapa"] == "recibido"),
        "recuperables": sum(1 for c in salida if c["recuperable"]),
    }
    informe = {"id": str(uuid.uuid4()), "mes": mes, "resumen": resumen, "clientes": salida,
               "generado": datetime.now(timezone.utc).isoformat()}
    await db.informes_mensuales.delete_many({"mes": mes})
    await db.informes_mensuales.insert_one(dict(informe))
    informe.pop("_id", None)
    return informe


def informe_a_excel(informe):
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    wb = Workbook()
    ws = wb.active
    ws.title = "Clientes"
    head = ["Cliente", "RUT", "Etapa", "Última actividad", "Días detenido",
            "Recuperable", "Motivo detención", "Acción sugerida", "N° eventos"]
    ws.append(head)
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1F3B2C")
    verde = PatternFill("solid", fgColor="FFF3CD")
    for c in informe["clientes"]:
        ws.append([c["cliente"], c["rut"], c["etapa"], c["ultima_actividad"],
                   c["dias_sin_movimiento"], "SÍ" if c["recuperable"] else "no",
                   c["motivo_detencion"], c["accion_sugerida"], c["n_eventos"]])
        if c["recuperable"]:
            for cell in ws[ws.max_row]:
                cell.fill = verde
    for col, w in zip("ABCDEFGHI", [38, 14, 13, 15, 13, 12, 55, 45, 10]):
        ws.column_dimensions[col].width = w
    ws2 = wb.create_sheet("Timeline")
    ws2.append(["Cliente", "Fecha", "Evento", "Detalle"])
    for c in ws2[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1F3B2C")
    for c in informe["clientes"]:
        for e in c["eventos"]:
            ws2.append([c["cliente"], e["fecha"], e["tipo"], e["detalle"]])
    for col, w in zip("ABCD", [38, 12, 18, 80]):
        ws2.column_dimensions[col].width = w
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
