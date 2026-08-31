"""ALGORITMO ESPEJO HÍBRIDO (Contralor) · CONEXIÓN CONCRECES · MÓDULO POSTVENTA.

- Espejo Capa A (automática): lee el buzón IMAP de resoluciones de Concreces, extrae
  patrones (aprobación/rechazo, plazos, secuencia de documentos) y calibra el módulo.
- Espejo Capa B (manual): criterios ingresados por el Administrador. REGLA DE ORO:
  una regla manual JAMÁS se sobreescribe con una automática sin confirmación del admin.
- Postventa: seguimiento de escritura por etapas con plazos, alertas, comunicaciones
  automáticas al cliente y aprendizaje progresivo de plazos reales.
"""
import os
import re
import uuid
import asyncio
import logging
from datetime import datetime, timezone, timedelta
from fastapi import APIRouter, HTTPException, Request
from database import db

espejo = APIRouter(prefix="/contralor/espejo")
concreces = APIRouter(prefix="/config/concreces")
postventa = APIRouter(prefix="/postventa")
gpanel = APIRouter(prefix="/gerencia-panel")
brokerx = APIRouter(prefix="/broker")


def _now():
    return datetime.now(timezone.utc).isoformat()


def _rol(request):
    return (getattr(request.state, "user", {}) or {}).get("rol", "")


def _exigir(request, roles):
    if _rol(request) not in roles:
        raise HTTPException(status_code=403, detail="No está autorizado el ingreso a este módulo")


# ═══ VALIDACIÓN DE NORMATIVAS (Bloque 7): antes de aprobar/avanzar/enviar ═══
_normas_val_cache = {"t": None, "datos": []}


async def _normativas_vigentes():
    """Normativas desde la DB con caché máximo de 5 minutos (regla de inmutabilidad)."""
    ahora = datetime.now(timezone.utc)
    if _normas_val_cache["t"] and (ahora - _normas_val_cache["t"]).total_seconds() < 300:
        return _normas_val_cache["datos"]
    docs = await db.dashai_eventos.find({"motivo": "normativa"}, {"_id": 0}).to_list(200)
    _normas_val_cache.update({"t": ahora, "datos": docs})
    return docs


async def _validar_normativas_op(texto_saliente="", cc=None, rol="", contexto=""):
    """Bloquea la operación si incumple una normativa vigente, con el detalle exacto."""
    normas = {d.get("norma_clave"): d.get("patron") or "" for d in await _normativas_vigentes()}
    if texto_saliente and "concreces" in texto_saliente.lower() and "DISEÑO CORREOS" in normas:
        raise HTTPException(status_code=422, detail=(
            f"Operación bloqueada por incumplimiento de NORMATIVA FIJA 'DISEÑO CORREOS' "
            f"({contexto or 'correo saliente'}): el texto menciona 'Concreces', lo cual está "
            f"prohibido. Normativa vigente: {normas['DISEÑO CORREOS']}"))
    if cc and rol not in ("gerencia", "admin", "maestro") and "CC" in normas \
            and "nunca en salientes" in normas["CC"].lower():
        raise HTTPException(status_code=422, detail=(
            f"Operación bloqueada por incumplimiento de NORMATIVA FIJA 'CC' "
            f"({contexto or 'correo saliente'}): su rol ({rol or 'sin rol'}) no está autorizado "
            f"a agregar copias (CC) en correos salientes. Normativa vigente: {normas['CC']}"))


def _cifrar(texto):
    from cryptography.fernet import Fernet
    return Fernet(os.environ["CRED_CIPHER_KEY"].encode()).encrypt(texto.encode()).decode()


def _descifrar(token):
    from cryptography.fernet import Fernet
    return Fernet(os.environ["CRED_CIPHER_KEY"].encode()).decrypt(token.encode()).decode()


# ═══════════════ ALGORITMO ESPEJO — CAPA A (AUTOMÁTICA, IMAP) ═══════════════
DOCS_CONOCIDOS = ["carta oferta", "resolución serviu", "resolucion serviu", "promesa",
                  "tasación", "tasacion", "estudio de título", "estudio de titulo",
                  "set de crédito", "set de credito", "escritura", "pagaré", "pagare", "cbr"]


def _scan_imap_sync(correo, clave, servidor, maxn=30):
    """Lee las últimas resoluciones del buzón (solo lectura, no marca ni borra)."""
    import imaplib
    import email as emlib
    from email.header import decode_header
    M = imaplib.IMAP4_SSL(servidor, timeout=25)
    M.login(correo, clave)
    M.select("INBOX", readonly=True)
    _, data = M.search(None, "ALL")
    ids = (data[0].split() or [])[-maxn:]
    out = []
    for i in reversed(ids):
        try:
            _, msg_data = M.fetch(i, "(RFC822)")
            msg = emlib.message_from_bytes(msg_data[0][1])
            subj = ""
            for part, enc in decode_header(msg.get("Subject") or ""):
                subj += part.decode(enc or "utf-8", "ignore") if isinstance(part, bytes) else part
            body = ""
            for part in msg.walk():
                if part.get_content_type() == "text/plain":
                    body = part.get_payload(decode=True).decode(part.get_content_charset() or "utf-8", "ignore")
                    break
            out.append({"subject": subj, "body": body[:4000], "fecha": msg.get("Date") or ""})
        except Exception:
            continue
    M.logout()
    return out


def _extraer_patrones(correos):
    """Heurística de calibración: aprobaciones, rechazos+motivo, plazos y secuencia de docs."""
    patrones = []
    for c in correos:
        texto = f"{c['subject']}\n{c['body']}".lower()
        if not any(k in texto for k in ("resoluc", "aprobad", "rechaz", "curse", "mesa")):
            continue
        tipo = "aprobacion" if ("aprobad" in texto or "curse favorable" in texto) else (
            "rechazo" if "rechaz" in texto else "criterio")
        motivo = ""
        if tipo == "rechazo":
            m = re.search(r"(?:rechaz|motivo)[^.\n]{0,160}", texto)
            motivo = (m.group(0).strip() if m else "")[:200]
        plazos = re.findall(r"(\d{1,3})\s*d[ií]as?(?:\s*h[áa]biles)?", texto)[:3]
        docs = [d for d in DOCS_CONOCIDOS if d in texto][:6]
        patrones.append({
            "tipo": tipo, "asunto": c["subject"][:180], "motivo": motivo,
            "plazos_dias": [int(p) for p in plazos], "documentos": docs,
            "clave": (motivo or c["subject"][:60] or tipo).strip().lower()[:60]})
    return patrones


@espejo.post("/escanear")
async def espejo_escanear(request: Request):
    """CAPA A: escaneo bajo demanda del buzón de resoluciones + calibración automática."""
    _exigir(request, ("admin", "maestro", "contralor"))
    cfg = await db.config.find_one({"_key": "espejo_contralor"}) or {}
    if not cfg.get("email") or not cfg.get("clave_enc"):
        raise HTTPException(status_code=400, detail="Sin credenciales del buzón: complete correo y clave de aplicación primero")
    if not cfg.get("activo"):
        raise HTTPException(status_code=400, detail="La conexión está desactivada: active el buzón antes de escanear")
    try:
        clave = _descifrar(cfg["clave_enc"])
        correos = await asyncio.to_thread(_scan_imap_sync, cfg["email"], clave,
                                          cfg.get("servidor") or "imap.gmail.com")
    except Exception as e:
        await db.config.update_one({"_key": "espejo_contralor"}, {"$set": {
            "estado": "error_conexion", "ultimo_error": str(e)[:200], "ultimo_scan": _now()}})
        raise HTTPException(status_code=502, detail=f"No fue posible conectar al buzón: {str(e)[:150]}")
    patrones = _extraer_patrones(correos)
    nuevos, pendientes = 0, 0
    for p in patrones:
        if await db.espejo_bitacora.find_one({"clave": p["clave"], "origen": "capa_a"}):
            continue
        # REGLA DE ORO: si choca con una regla MANUAL activa → queda pendiente de confirmación
        manual = await db.espejo_criterios.find_one({"clave": p["clave"], "origen": "manual", "estado": "activo"})
        reg = {"id": str(uuid.uuid4()), "origen": "capa_a", "fecha": _now(), **p,
               "patron": f"[{p['tipo'].upper()}] {p['asunto']}" + (f" · Motivo: {p['motivo']}" if p['motivo'] else "")
                         + (f" · Plazos: {p['plazos_dias']} días" if p['plazos_dias'] else "")
                         + (f" · Docs: {', '.join(p['documentos'])}" if p['documentos'] else "")}
        await db.espejo_bitacora.insert_one(dict(reg))
        nuevos += 1
        if manual:
            await db.espejo_criterios.insert_one({
                "id": str(uuid.uuid4()), "clave": p["clave"], "criterio": p["asunto"][:120],
                "detalle": reg["patron"], "origen": "auto", "estado": "pendiente_confirmacion",
                "conflicto_con": manual["id"], "fecha": _now()})
            pendientes += 1
    total = await db.espejo_bitacora.count_documents({"origen": "capa_a"})
    pct = min(100, total * 10)
    await db.config.update_one({"_key": "espejo_contralor"}, {"$set": {
        "estado": "conectado", "calibracion_pct": pct, "ultimo_scan": _now(),
        "correos_leidos": len(correos), "ultimo_error": ""}})
    return {"ok": True, "correos_leidos": len(correos), "patrones_nuevos": nuevos,
            "conflictos_pendientes": pendientes, "calibracion_pct": pct, "estado": "conectado"}


async def espejo_loop():
    """CAPA A automática: escaneo cada 30 min SOLO si hay credenciales y conexión activa."""
    await asyncio.sleep(600)
    while True:
        try:
            cfg = await db.config.find_one({"_key": "espejo_contralor"}) or {}
            if cfg.get("activo") and cfg.get("email") and cfg.get("clave_enc"):
                clave = _descifrar(cfg["clave_enc"])
                correos = await asyncio.to_thread(_scan_imap_sync, cfg["email"], clave,
                                                  cfg.get("servidor") or "imap.gmail.com")
                for p in _extraer_patrones(correos):
                    if await db.espejo_bitacora.find_one({"clave": p["clave"], "origen": "capa_a"}):
                        continue
                    manual = await db.espejo_criterios.find_one(
                        {"clave": p["clave"], "origen": "manual", "estado": "activo"})
                    await db.espejo_bitacora.insert_one({
                        "id": str(uuid.uuid4()), "origen": "capa_a", "fecha": _now(), **p,
                        "patron": f"[{p['tipo'].upper()}] {p['asunto']}"})
                    if manual:
                        await db.espejo_criterios.insert_one({
                            "id": str(uuid.uuid4()), "clave": p["clave"], "criterio": p["asunto"][:120],
                            "detalle": p.get("motivo") or p["asunto"], "origen": "auto",
                            "estado": "pendiente_confirmacion", "conflicto_con": manual["id"], "fecha": _now()})
                total = await db.espejo_bitacora.count_documents({"origen": "capa_a"})
                await db.config.update_one({"_key": "espejo_contralor"}, {"$set": {
                    "estado": "conectado", "calibracion_pct": min(100, total * 10), "ultimo_scan": _now()}})
        except Exception as e:
            logging.warning(f"espejo loop: {e}")
        try:
            cfg_s = await db.config.find_one({"_key": "espejo_contralor"}) or {}
            if _creds_concreces_imap(cfg_s):
                await _sync_concreces_core()
        except Exception as e:
            logging.warning(f"sync concreces loop: {e}")
        await asyncio.sleep(1800)


# ═══════════════ SINCRONIZACIÓN CONCRECES → OPERACIONES (Algoritmo Espejo, núcleo) ═══════════════
ESTADOS_CONCRECES = [("aprobad", "Aprobada"), ("rechazad", "Rechazada"), ("cursad", "Cursada"),
                     ("escriturad", "Escriturada"), ("observa", "Con Observaciones"),
                     ("en estudio", "En Estudio"), ("pendiente", "Pendiente")]


def _creds_concreces_imap(cfg):
    """Secrets del entorno primero (norma fija); si no, credenciales guardadas en el panel."""
    user = os.environ.get("CONCRECES_IMAP_USER") or ""
    pwd = os.environ.get("CONCRECES_IMAP_PASSWORD") or ""
    if user and pwd:
        return os.environ.get("CONCRECES_IMAP_HOST") or "imap.gmail.com", user, pwd, "secrets"
    if cfg.get("email") and cfg.get("clave_enc") and cfg.get("activo"):
        return cfg.get("servidor") or "imap.gmail.com", cfg["email"], _descifrar(cfg["clave_enc"]), "panel"
    return None


def _rut_limpio(t):
    return re.sub(r"[^0-9kK]", "", t or "").lower()


def _extraer_datos_operacion(c):
    """Datos estructurados del correo Concreces: nº operación, estado, monto, observaciones, fecha."""
    texto = f"{c.get('subject') or ''}\n{c.get('body') or ''}"
    low = texto.lower()
    nro = re.search(r"operaci[oó]n\s*(?:n[°ºo]?\.?\s*)?[:#]?\s*(\d{2,10})", low)
    rut = re.search(r"(\d{1,2}\.?\d{3}\.?\d{3}-?[\dkK])", texto)
    estado = next((lb for k, lb in ESTADOS_CONCRECES if k in low), "")
    monto = re.search(r"(?:uf|monto)\s*:?\s*\$?\s*([\d.,]+)", low)
    obs_m = re.search(r"observaci[oó]n(?:es)?\s*:?\s*([^\n]{5,200})", low)
    return {"nro_operacion": nro.group(1) if nro else "", "rut": rut.group(1) if rut else "",
            "estado": estado, "monto": monto.group(1) if monto else "",
            "observaciones": obs_m.group(1).strip() if obs_m else "",
            "fecha_correo": c.get("fecha") or "", "asunto": (c.get("subject") or "")[:180]}


async def _notificar_urgencia(ia, asunto, destino):
    """Correo urgente detectado por la IA → alerta en la app (Admin + Contralor) y correo al Admin."""
    cliente = (destino or {}).get("nombre") or "operación sin clasificar"
    motivo = ia.get("motivo_urgencia") or "; ".join(ia.get("alertas") or []) or "alerta detectada por IA"
    await db.alertas.insert_one({
        "id": str(uuid.uuid4()), "tipo": "espejo_urgente", "leida": False, "creado": _now(),
        "destinatarios": ["admin", "contralor"], "cliente": cliente,
        "titulo": f"🔴 URGENTE — {cliente}",
        "mensaje": f"El análisis IA del correo de la matriz detectó: {motivo}",
        "asunto_correo": (asunto or "")[:180],
        "resumen_ia": ia.get("resumen_interpretativo") or ""})
    try:
        import email_service as mail
        admin_dest = os.environ.get("MAIL_USER") or ""
        if admin_dest:
            # Normativa DISEÑO CORREOS: jamás mencionar el nombre de la matriz en salientes
            cuerpo = (
                "<div style='font-family:Arial,Helvetica,sans-serif;background:#fff;color:#111;font-size:14px'>"
                f"<p>Estimado Administrador:</p>"
                f"<p>El Algoritmo Espejo detectó una situación <b style='color:#b91c1c'>URGENTE</b> en un "
                f"correo de la empresa matriz, asociado a: <b>{cliente}</b>.</p>"
                f"<p><b>Motivo:</b> {motivo}</p>"
                f"<p><b>Resumen interpretativo (IA):</b> {ia.get('resumen_interpretativo') or '—'}</p>"
                "<p>El detalle completo está disponible en el panel del Contralor.</p>"
                "<p style='color:#555'>Saludos cordiales,<br><b>Central Mutuos</b></p></div>")
            await asyncio.to_thread(mail.send_mail, admin_dest,
                                    f"🔴 URGENTE Algoritmo Espejo — {cliente}", cuerpo, [], "secundaria")
    except Exception as e:
        logging.warning(f"notificación urgencia espejo: {e}")


async def _sync_concreces_core():
    import hashlib
    cfg = await db.config.find_one({"_key": "espejo_contralor"}) or {}
    creds = _creds_concreces_imap(cfg)
    if not creds:
        raise ValueError("Sin credenciales del buzón Concreces: configure CONCRECES_IMAP_USER/"
                         "CONCRECES_IMAP_PASSWORD en los secrets o active el buzón en el panel")
    host, user, pwd, origen_cred = creds
    # BLOQUE 7: consultar SIEMPRE las normativas activas en cada ciclo (sin caché vieja)
    normas_activas = await db.dashai_eventos.count_documents({"motivo": "normativa"})
    correos = await asyncio.to_thread(_scan_imap_sync, user, pwd, host, 50)
    folders = await db.folders.find({}, {"id": 1, "nombre": 1, "rut": 1, "nro_operacion": 1}).to_list(1500)
    actualizadas, sin_clasificar, ahora = 0, 0, _now()
    for c in correos:
        firma = hashlib.md5(f"{c.get('subject')}|{c.get('fecha')}".encode(), usedforsecurity=False).hexdigest()
        if await db.espejo_sync_log.find_one({"firma": firma}):
            continue
        datos = _extraer_datos_operacion(c)
        # 🧠 CLAUDE (Sonnet 4.6): interpretación completa del correo de la matriz
        ia = None
        try:
            import espejo_ia
            ia = await espejo_ia.analizar_correo(c.get("subject"), c.get("body"), c.get("fecha"))
            for k in ("nro_operacion", "rut", "estado", "monto", "observaciones"):
                if ia.get(k):
                    datos[k] = ia[k]
        except Exception as e:
            logging.warning(f"espejo IA análisis: {e}")
        destino, rut_c = None, _rut_limpio(datos["rut"])
        if rut_c:
            destino = next((f for f in folders if _rut_limpio(f.get("rut")) == rut_c), None)
        if not destino and datos["nro_operacion"]:
            destino = next((f for f in folders
                            if str(f.get("nro_operacion") or "") == datos["nro_operacion"]), None)
        if not destino:
            up = f"{c.get('subject') or ''} {c.get('body') or ''}".upper()
            destino = next((f for f in folders if f.get("nombre")
                            and len(f["nombre"].split()) >= 2 and f["nombre"].upper() in up), None)
        if destino:
            await db.folders.update_one({"id": destino["id"]}, {"$set": {"concreces": {
                "nro_operacion": datos["nro_operacion"], "estado": datos["estado"] or "Informado",
                "monto": datos["monto"], "observaciones": datos["observaciones"],
                "fecha_correo": datos["fecha_correo"], "sync_at": ahora,
                "ia_analisis": ia}}})
            actualizadas += 1
        else:
            await db.espejo_no_clasificados.update_one({"firma": firma}, {"$setOnInsert": {
                "id": str(uuid.uuid4()), "firma": firma, **datos,
                "cuerpo": (c.get("body") or "")[:600], "recibido": ahora,
                "ia_analisis": ia}}, upsert=True)
            sin_clasificar += 1
        if ia:
            import espejo_ia as _eia
            await _eia.registrar_interpretacion(db, ia, c.get("subject"), (destino or {}).get("id"))
            if ia.get("urgente"):
                await _notificar_urgencia(ia, c.get("subject"), destino)
        await db.espejo_sync_log.insert_one({"firma": firma, "asignado": bool(destino),
                                             "folder_id": (destino or {}).get("id"), "fecha": ahora})
    await db.config.update_one({"_key": "espejo_contralor"}, {"$set": {
        "ultima_sync": ahora, "sync_origen_credencial": origen_cred,
        "sync_resumen": {"correos": len(correos), "actualizadas": actualizadas,
                         "sin_clasificar": sin_clasificar,
                         "normativas_consultadas": normas_activas}}}, upsert=True)
    return {"correos_leidos": len(correos), "operaciones_actualizadas": actualizadas,
            "no_clasificados_nuevos": sin_clasificar, "ultima_sync": ahora}


@espejo.post("/sincronizar")
async def espejo_sincronizar(request: Request):
    """Botón 'Sincronizar ahora' del Contralor (también corre automático cada 30 min)."""
    _exigir(request, ("admin", "maestro", "contralor"))
    try:
        r = await _sync_concreces_core()
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"No fue posible conectar al buzón Concreces: {str(e)[:150]}")
    return {"ok": True, **r}


@espejo.get("/operaciones")
async def espejo_operaciones(request: Request):
    """Estado de operaciones según Concreces (solo lectura) con marca de tiempo de última lectura."""
    _exigir(request, ("admin", "maestro", "contralor", "gerencia", "administracion", "postventa"))
    cfg = await db.config.find_one({"_key": "espejo_contralor"}) or {}
    ops = []
    async for fd in db.folders.find({"concreces": {"$exists": True}}).sort("nombre", 1):
        cz = fd.get("concreces") or {}
        ia = cz.get("ia_analisis") or {}
        ops.append({"fid": fd["id"], "cliente": fd.get("nombre"), "rut": fd.get("rut") or "",
                    "nro_operacion": cz.get("nro_operacion") or "", "estado": cz.get("estado") or "",
                    "monto": cz.get("monto") or "", "observaciones": cz.get("observaciones") or "",
                    "fecha_correo": cz.get("fecha_correo") or "", "sync_at": cz.get("sync_at") or "",
                    "ia_resumen": ia.get("resumen_interpretativo") or "",
                    "ia_urgente": bool(ia.get("urgente")), "ia_motivo_urgencia": ia.get("motivo_urgencia") or "",
                    "ia_ambiguo": bool(ia.get("ambiguo")), "ia_alertas": ia.get("alertas") or [],
                    "ia_requerimientos": ia.get("requerimientos") or [],
                    "ia_analizado_en": ia.get("analizado_en") or "",
                    "ia_correccion": ia.get("correccion") or None,
                    "simulado": bool(cz.get("simulado"))})
    return {"operaciones": ops, "total": len(ops), "ultima_sync": cfg.get("ultima_sync") or "",
            "resumen": cfg.get("sync_resumen") or {},
            "origen_credencial": cfg.get("sync_origen_credencial") or ""}


@espejo.get("/no-clasificados")
async def espejo_no_clasificados_list(request: Request):
    """Correos de Concreces sin operación asociada: SOLO Admin y Contralor."""
    _exigir(request, ("admin", "maestro", "contralor"))
    docs = await db.espejo_no_clasificados.find({}, {"_id": 0}).sort("recibido", -1).to_list(100)
    return {"correos": docs, "total": len(docs)}


@espejo.post("/probar-ia")
async def espejo_probar_ia(payload: dict, request: Request):
    """Prueba del pipeline IA con un correo simulado de la matriz (solo Admin)."""
    if _rol(request) not in ("admin", "maestro"):
        raise HTTPException(status_code=403, detail="Solo el Administrador puede ejecutar pruebas del análisis IA")
    asunto = (payload.get("asunto") or "").strip()
    cuerpo = (payload.get("cuerpo") or "").strip()
    if not asunto and not cuerpo:
        raise HTTPException(status_code=400, detail="Indique el asunto y/o el cuerpo del correo simulado")
    import espejo_ia
    try:
        ia = await espejo_ia.analizar_correo(asunto, cuerpo, _now()[:10])
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"El análisis IA no pudo completarse: {str(e)[:150]}")
    # Ruteo idéntico al sync real, marcado como simulado
    rut_c = _rut_limpio(ia.get("rut"))
    destino = None
    if rut_c:
        async for fd in db.folders.find({"rut": {"$exists": True, "$ne": ""}}, {"id": 1, "rut": 1, "nombre": 1}):
            if _rut_limpio(fd.get("rut")) == rut_c:
                destino = fd
                break
    ahora = _now()
    if destino:
        await db.folders.update_one({"id": destino["id"]}, {"$set": {"concreces": {
            "nro_operacion": ia.get("nro_operacion") or "", "estado": ia.get("estado") or "Informado",
            "monto": ia.get("monto") or "", "observaciones": ia.get("observaciones") or "",
            "fecha_correo": ia.get("fecha") or "", "sync_at": ahora,
            "ia_analisis": ia, "simulado": True}}})
    else:
        await db.espejo_no_clasificados.update_one({"firma": f"sim-{uuid.uuid4()}"}, {"$setOnInsert": {
            "id": str(uuid.uuid4()), "asunto": asunto[:180], "estado": ia.get("estado") or "",
            "nro_operacion": ia.get("nro_operacion") or "", "rut": ia.get("rut") or "",
            "monto": ia.get("monto") or "", "observaciones": ia.get("observaciones") or "",
            "cuerpo": cuerpo[:600], "recibido": ahora, "ia_analisis": ia, "simulado": True}}, upsert=True)
    await espejo_ia.registrar_interpretacion(db, ia, asunto, (destino or {}).get("id"), simulado=True)
    if ia.get("urgente"):
        await _notificar_urgencia(ia, asunto, destino)
    return {"ok": True, "simulado": True, "asignado_a": (destino or {}).get("nombre") or "",
            "analisis": ia}


@espejo.post("/operaciones/{fid}/ia-correccion")
async def espejo_ia_correccion(fid: str, payload: dict, request: Request):
    """El Admin revisa/corrige manualmente la interpretación de la IA (queda registrado)."""
    if _rol(request) not in ("admin", "maestro"):
        raise HTTPException(status_code=403, detail="Solo el Administrador puede corregir la interpretación de la IA")
    fd = await db.folders.find_one({"id": fid})
    if not fd or not fd.get("concreces"):
        raise HTTPException(status_code=404, detail="Operación sin registro del Algoritmo Espejo")
    admin = getattr(request.state, "user", {}) or {}
    cz = fd["concreces"]
    ia = cz.get("ia_analisis") or {}
    cambios, anterior = {}, {}
    for campo in ("nro_operacion", "estado", "monto", "observaciones", "resumen_interpretativo",
                  "motivo_urgencia"):
        if campo in (payload or {}):
            nuevo = str(payload.get(campo) or "").strip()
            anterior[campo] = ia.get(campo) if campo in ("resumen_interpretativo", "motivo_urgencia") else cz.get(campo)
            cambios[campo] = nuevo
    if "urgente" in (payload or {}):
        anterior["urgente"], cambios["urgente"] = bool(ia.get("urgente")), bool(payload.get("urgente"))
    if not cambios:
        raise HTTPException(status_code=400, detail="Indique al menos un campo a corregir")
    for campo, nuevo in cambios.items():
        if campo in ("resumen_interpretativo", "motivo_urgencia", "urgente"):
            ia[campo] = nuevo
        else:
            cz[campo] = nuevo
            if campo in ia:
                ia[campo] = nuevo
    ia["correccion"] = {"por": admin.get("nombre") or admin.get("sub") or "", "fecha": _now(),
                        "campos": list(cambios.keys())}
    cz["ia_analisis"] = ia
    await db.folders.update_one({"id": fid}, {"$set": {"concreces": cz}})
    await db.espejo_ia_log.insert_one({
        "id": str(uuid.uuid4()), "fecha": _now(), "accion": "correccion_manual", "folder_id": fid,
        "por": ia["correccion"]["por"], "anterior": anterior, "nuevo": cambios})
    return {"ok": True, "correccion": ia["correccion"], "campos": list(cambios.keys())}


# ═══════════════ ALGORITMO ESPEJO — CAPA B (MANUAL) ═══════════════
@espejo.get("/hallazgos")
async def espejo_hallazgos(request: Request):
    """🛡️ Panel de Hallazgos del Contralor: alertas de auditoría agrupadas por carpeta,
    con la regla incumplida y la fuente exacta citada."""
    _exigir(request, ("admin", "maestro", "contralor", "gerencia", "administracion"))
    alertas = await db.alertas.find(
        {"tipo": {"$regex": "^(auditoria71|mesa_verdad)"}},
        {"_id": 0}).sort("fecha", -1).limit(400).to_list(400)
    grupos = {}
    for a in alertas:
        k = a.get("folder_id") or a.get("cliente") or "general"
        g = grupos.setdefault(k, {"folder_id": a.get("folder_id") or "",
                                  "cliente": a.get("cliente") or "(sistema)",
                                  "ultima_fecha": a.get("fecha", ""), "hallazgos": []})
        g["hallazgos"].append({
            "id": a.get("id"), "tipo": a.get("tipo"), "nivel": a.get("nivel") or "media",
            "regla": a.get("regla") or ("Fuente de Verdad de Mesa" if str(a.get("tipo", "")).startswith("mesa_verdad") else ""),
            "detalle": a.get("mensaje", ""), "recomendacion": a.get("recomendacion", ""),
            "fuente": a.get("fuente", ""), "bloqueante": bool(a.get("bloqueante")),
            "fecha": a.get("fecha", ""), "leida": bool(a.get("leida"))})
    carpetas = sorted(grupos.values(), key=lambda g: g["ultima_fecha"], reverse=True)
    return {"carpetas": carpetas, "total_alertas": len(alertas),
            "criticas": sum(1 for a in alertas if (a.get("nivel") == "critica" or a.get("bloqueante")) and not a.get("leida")),
            "sin_leer": sum(1 for a in alertas if not a.get("leida"))}


@espejo.get("/criterios")
async def espejo_criterios_list(request: Request):
    _exigir(request, ("admin", "maestro", "contralor", "gerencia", "administracion", "postventa"))
    regs = await db.espejo_criterios.find({}, {"_id": 0}).sort("fecha", -1).to_list(100)
    return {"criterios": regs, "total": len(regs),
            "pendientes": sum(1 for r in regs if r.get("estado") == "pendiente_confirmacion")}


@espejo.post("/criterios")
async def espejo_criterios_add(payload: dict, request: Request):
    """CAPA B: criterio humano manual — solo el Administrador."""
    _exigir(request, ("admin", "maestro"))
    criterio = (payload.get("criterio") or "").strip()
    if not criterio:
        raise HTTPException(status_code=400, detail="Falta el texto del criterio")
    reg = {"id": str(uuid.uuid4()), "clave": criterio.lower()[:60], "criterio": criterio[:200],
           "detalle": (payload.get("detalle") or "").strip()[:500], "origen": "manual",
           "estado": "activo", "fecha": _now(),
           "por": (getattr(request.state, "user", {}) or {}).get("sub") or "admin"}
    await db.espejo_criterios.insert_one(dict(reg))
    await db.espejo_bitacora.insert_one({"id": str(uuid.uuid4()), "origen": "capa_b",
                                         "fecha": _now(), "clave": reg["clave"], "tipo": "criterio_manual",
                                         "patron": f"[MANUAL] {criterio[:160]}"})
    return {"ok": True, "criterio": reg}


@espejo.post("/criterios/{cid}/resolver")
async def espejo_criterio_resolver(cid: str, payload: dict, request: Request):
    """Conflicto auto vs manual: SOLO el admin confirma (aplicar) o rechaza la regla automática."""
    _exigir(request, ("admin", "maestro"))
    accion = (payload.get("accion") or "").strip()
    reg = await db.espejo_criterios.find_one({"id": cid})
    if not reg or reg.get("estado") != "pendiente_confirmacion":
        raise HTTPException(status_code=404, detail="Criterio pendiente no encontrado")
    if accion == "confirmar":
        await db.espejo_criterios.update_one({"id": cid}, {"$set": {"estado": "activo", "confirmado_en": _now()}})
        if reg.get("conflicto_con"):
            await db.espejo_criterios.update_one({"id": reg["conflicto_con"]},
                                                 {"$set": {"estado": "reemplazado", "reemplazado_en": _now()}})
        return {"ok": True, "accion": "confirmado", "nota": "Regla automática aplicada con autorización del administrador"}
    await db.espejo_criterios.update_one({"id": cid}, {"$set": {"estado": "rechazado", "rechazado_en": _now()}})
    return {"ok": True, "accion": "rechazado", "nota": "La regla manual se mantiene intacta"}


# ═══════════════ CONEXIÓN CONCRECES (empresas vinculadas legalmente) ═══════════════
@concreces.get("")
async def concreces_get(request: Request):
    _exigir(request, ("admin", "maestro", "gerencia", "contralor"))
    cfg = await db.config.find_one({"_key": "conexion_concreces"}, {"_id": 0, "clave_enc": 0}) or {}
    tiene = bool((await db.config.find_one({"_key": "conexion_concreces"}) or {}).get("clave_enc"))
    return {"usuario": cfg.get("usuario") or "", "url": cfg.get("url") or "",
            "tiene_clave": tiene, "activo": bool(cfg.get("activo")),
            "estado": ("Credenciales guardadas — sin conexión activa" if tiene else "Pendiente de credenciales"),
            "nota": "Conexión dentro de la normativa de empresas vinculadas legalmente. No se conecta hasta que el administrador lo autorice."}


async def _reconfirmar_identidad_local(request, payload):
    """Configuración avanzada (Bloque 7): reingreso de contraseña obligatorio."""
    import bcrypt as _b
    clave = ((payload or {}).get("confirmacion_clave") or "").strip()
    sub = (getattr(request.state, "user", {}) or {}).get("sub") or ""
    user = await db.users.find_one({"codigo": sub})
    ok = False
    if user and clave:
        if user.get("clave_hash"):
            ok = _b.checkpw(clave.encode(), user["clave_hash"].encode())
        else:
            ok = user.get("password") == clave
    if not ok:
        raise HTTPException(status_code=403, detail=(
            "Confirmación de identidad requerida: reingrese su contraseña para "
            "modificar la configuración avanzada."))


@concreces.post("")
async def concreces_save(payload: dict, request: Request):
    _exigir(request, ("admin", "maestro"))
    await _reconfirmar_identidad_local(request, payload)
    cambios = {"actualizado": _now(), "activo": False}
    if "usuario" in payload:
        cambios["usuario"] = (payload.get("usuario") or "").strip()
    if "url" in payload:
        cambios["url"] = (payload.get("url") or "").strip()
    if payload.get("clave"):
        cambios["clave_enc"] = _cifrar(str(payload["clave"]).strip())
    await db.config.update_one({"_key": "conexion_concreces"}, {"$set": cambios}, upsert=True)
    cfg = await db.config.find_one({"_key": "conexion_concreces"}) or {}
    return {"ok": True, "tiene_clave": bool(cfg.get("clave_enc")),
            "estado": "Credenciales guardadas — sin conexión activa"}


# ═══════════════ MÓDULO POSTVENTA REFORZADO ═══════════════
ETAPAS_PV = [("firma", "Firma"), ("escritura", "Escritura"),
             ("entrega_pagare", "Entrega de Pagaré"), ("doc_posterior", "Documentación Posterior")]
PLAZOS_DEFAULT = {"firma": 7, "escritura": 15, "entrega_pagare": 7, "doc_posterior": 10}
RESPONSABLE_PV = "Javier Urrutia"


async def _plazos_pv():
    cfg = await db.config.find_one({"_key": "postventa_plazos"}) or {}
    return {k: int(cfg.get(k) or v) for k, v in PLAZOS_DEFAULT.items()}


async def _aprendizaje_pv():
    """Aprendizaje progresivo: promedio de días reales por etapa (escrituras completadas)."""
    agg = {}
    async for r in db.postventa_aprendizaje.find({}):
        for k, dias in (r.get("duraciones") or {}).items():
            agg.setdefault(k, []).append(dias)
    return {k: round(sum(v) / len(v), 1) for k, v in agg.items() if v}


def _dias_desde(iso):
    try:
        d = datetime.fromisoformat(iso)
        if d.tzinfo is None:
            d = d.replace(tzinfo=timezone.utc)
        return round((datetime.now(timezone.utc) - d).total_seconds() / 86400, 1)
    except Exception:
        return 0


@postventa.get("/panel")
async def postventa_panel(request: Request):
    """Vista de control: etapa de cada cliente, plazos, alertas y cumplimiento del responsable."""
    plazos = await _plazos_pv()
    aprendido = await _aprendizaje_pv()
    casos, alertas = [], 0
    async for c in db.postventa_casos.find({}).sort("creado", -1):
        etapa = c.get("etapa_actual") or ""
        dias = _dias_desde(c.get("inicio_etapa") or c.get("creado") or "")
        atrasada = bool(etapa and etapa != "completado" and dias > plazos.get(etapa, 99))
        if atrasada:
            alertas += 1
        detalle = []
        for k, lb in ETAPAS_PV:
            e = (c.get("etapas") or {}).get(k) or {}
            detalle.append({"clave": k, "etapa": lb, "plazo_dias": plazos[k],
                            "plazo_estimado_aprendido": aprendido.get(k),
                            "completada": bool(e.get("completada")),
                            "fecha": e.get("fecha") or "", "dias_reales": e.get("dias_reales"),
                            "en_tiempo_y_forma": e.get("en_plazo"),
                            "en_curso": k == etapa,
                            "dias_en_curso": dias if k == etapa else None,
                            "atrasada": atrasada if k == etapa else False})
        casos.append({"id": c["id"], "cliente": c.get("cliente"), "email": c.get("email") or "",
                      "responsable": RESPONSABLE_PV, "etapa_actual": etapa,
                      "etapa_label": dict(ETAPAS_PV).get(etapa, "✅ Completado"),
                      "dias_en_etapa": dias, "atrasada": atrasada, "etapas": detalle,
                      "comunicaciones": (c.get("comunicaciones") or [])[-4:], "creado": c.get("creado")})
    return {"casos": casos, "total": len(casos), "alertas_atraso": alertas,
            "plazos": plazos, "plazos_aprendidos": aprendido,
            "escrituras_completadas": await db.postventa_aprendizaje.count_documents({}),
            "responsable": RESPONSABLE_PV, "etapas": [{"clave": k, "label": lb} for k, lb in ETAPAS_PV]}


@postventa.post("/casos")
async def postventa_crear(payload: dict, request: Request):
    _exigir(request, ("admin", "maestro", "postventa", "gerencia"))
    cliente = (payload.get("cliente") or "").strip()
    if not cliente:
        raise HTTPException(status_code=400, detail="Falta el nombre del cliente")
    reg = {"id": str(uuid.uuid4()), "cliente": cliente, "email": (payload.get("email") or "").strip(),
           "etapa_actual": "firma", "inicio_etapa": _now(), "etapas": {}, "comunicaciones": [],
           "responsable": RESPONSABLE_PV, "creado": _now()}
    await db.postventa_casos.insert_one(dict(reg))
    return {"ok": True, "caso": {k: v for k, v in reg.items()}}


@postventa.post("/casos/{cid}/avanzar")
async def postventa_avanzar(cid: str, request: Request):
    """Completa la etapa actual, genera la comunicación al cliente y aprende plazos reales."""
    _exigir(request, ("admin", "maestro", "postventa", "gerencia"))
    c = await db.postventa_casos.find_one({"id": cid})
    if not c:
        raise HTTPException(status_code=404, detail="Caso no encontrado")
    etapa = c.get("etapa_actual")
    claves = [k for k, _ in ETAPAS_PV]
    if etapa not in claves:
        raise HTTPException(status_code=400, detail="El caso ya está completado")
    plazos = await _plazos_pv()
    dias = _dias_desde(c.get("inicio_etapa") or c.get("creado"))
    idx = claves.index(etapa)
    siguiente = claves[idx + 1] if idx + 1 < len(claves) else "completado"
    lb = dict(ETAPAS_PV)
    # Comunicación automática al cliente (texto generado al avanzar la etapa)
    if siguiente != "completado":
        msg = (f"Estimado(a) {c.get('cliente')}: le informamos que la etapa '{lb[etapa]}' de su proceso "
               f"de escrituración fue completada con éxito. La siguiente etapa es '{lb[siguiente]}' "
               f"(plazo estimado: {plazos[siguiente]} días). Le mantendremos informado(a). — Central Mutuos")
    else:
        msg = (f"Estimado(a) {c.get('cliente')}: ¡felicitaciones! Su proceso de escrituración fue "
               f"completado en su totalidad. Gracias por confiar en Central Mutuos.")
    await _validar_normativas_op(texto_saliente=msg, rol=_rol(request),
                                 contexto=f"avance de etapa '{lb[etapa]}'")
    comunicacion = {"etapa": etapa, "texto": msg, "generada": _now(), "estado": "generada"}
    etapas = c.get("etapas") or {}
    etapas[etapa] = {"completada": True, "fecha": _now(), "dias_reales": dias,
                     "en_plazo": dias <= plazos.get(etapa, 99), "por": RESPONSABLE_PV}
    await db.postventa_casos.update_one({"id": cid}, {"$set": {
        "etapas": etapas, "etapa_actual": siguiente, "inicio_etapa": _now()},
        "$push": {"comunicaciones": comunicacion}})
    # APRENDIZAJE PROGRESIVO: la escritura completada alimenta los plazos futuros
    if siguiente == "completado":
        await db.postventa_aprendizaje.insert_one({
            "id": str(uuid.uuid4()), "caso_id": cid, "cliente": c.get("cliente"),
            "duraciones": {k: (etapas.get(k) or {}).get("dias_reales", 0) for k in claves},
            "fecha": _now()})
    return {"ok": True, "etapa_completada": lb[etapa], "dias_reales": dias,
            "en_plazo": dias <= plazos.get(etapa, 99), "siguiente": siguiente,
            "comunicacion_cliente": msg}


@postventa.post("/plazos")
async def postventa_plazos_set(payload: dict, request: Request):
    """Plazos por etapa: configurables SOLO por el Administrador."""
    _exigir(request, ("admin", "maestro"))
    cambios = {}
    for k, _ in ETAPAS_PV:
        if k in payload:
            try:
                v = int(payload[k])
            except (TypeError, ValueError):
                raise HTTPException(status_code=400, detail=f"Plazo inválido para {k}")
            if not 1 <= v <= 365:
                raise HTTPException(status_code=400, detail=f"El plazo de {k} debe estar entre 1 y 365 días")
            cambios[k] = v
    if not cambios:
        raise HTTPException(status_code=400, detail="Sin plazos para actualizar")
    cambios["actualizado"] = _now()
    await db.config.update_one({"_key": "postventa_plazos"}, {"$set": cambios}, upsert=True)
    return {"ok": True, "plazos": await _plazos_pv()}


# ═══════════════ DASHBOARD GERENCIA COMERCIAL (indicadores) ═══════════════
@gpanel.get("/rol")
async def gerencia_panel_rol(request: Request):
    _exigir(request, ("admin", "maestro", "gerencia", "contralor"))
    from datetime import timedelta
    hoy = datetime.now(timezone.utc)
    h7 = (hoy - timedelta(days=7)).isoformat()
    h14 = (hoy - timedelta(days=14)).isoformat()
    por_inmo, por_proy, por_broker, carpetas = {}, {}, {}, []
    async for fd in db.folders.find({"oculto_supercarpeta": {"$ne": True}}):
        inmo = (fd.get("inmobiliaria") or ("Casa Usada" if "usad" in (fd.get("tipo_operacion") or "") else "Directa")).strip()
        proy = (fd.get("proyecto") or "—").strip() or "—"
        brk = (fd.get("broker_origen") or fd.get("broker_codigo") or fd.get("broker_nombre") or "Directo").strip() or "Directo"
        for d, k in ((por_inmo, inmo), (por_proy, proy), (por_broker, brk)):
            d[k] = d.get(k, 0) + 1
        if len(carpetas) < 40:
            carpetas.append({"cliente": fd.get("nombre"), "inmobiliaria": inmo, "proyecto": proy,
                             "broker_origen": brk,
                             "estado": ("Escriturada" if fd.get("escritura_confirmada_at")
                                        else "Estudio aprobado" if fd.get("estudio_recibido_at")
                                        else "Tasación recibida" if fd.get("tasacion_informe_recibido_at")
                                        else "En proceso")})
    # Actividad del equipo administrativo (volumen operativo, no cierres)
    equipo = {"Victoria Vilchez": {"hoy": 0, "semana": 0}, "Daniela Galindo": {"hoy": 0, "semana": 0}, "Otros": {"hoy": 0, "semana": 0}}
    _ops_eq = {k: set() for k in equipo}
    dia_hoy = hoy.strftime("%Y-%m-%d")
    async for r in db.estado_manual_log.find({"fecha": {"$gte": h7}}):
        por = (r.get("por") or "").lower()
        key = "Victoria Vilchez" if "victoria" in por or "vilche" in por else (
            "Daniela Galindo" if "daniela" in por or "galindo" in por else "Otros")
        equipo[key]["semana"] += 1
        if r.get("folder_id"):
            _ops_eq[key].add(r["folder_id"])
        if str(r.get("fecha") or "").startswith(dia_hoy):
            equipo[key]["hoy"] += 1
    for k in equipo:
        equipo[k]["operaciones"] = len(_ops_eq[k])
    # Volumen diario (7 días): gestiones + correos detectados + envíos
    volumen = {}
    for col, campo in ((db.estado_manual_log, "fecha"), (db.hitos_externos, "creado"), (db.correos_smtp_log, "fecha")):
        async for r in col.find({campo: {"$gte": h7}}, {campo: 1}):
            d = str(r.get(campo) or "")[:10]
            if d:
                volumen[d] = volumen.get(d, 0) + 1
    # Comparativa semanal del equipo
    sem_actual = await db.estado_manual_log.count_documents({"fecha": {"$gte": h7}}) + \
        await db.hitos_externos.count_documents({"creado": {"$gte": h7}})
    sem_previa = (await db.estado_manual_log.count_documents({"fecha": {"$gte": h14, "$lt": h7}}) +
                  await db.hitos_externos.count_documents({"creado": {"$gte": h14, "$lt": h7}}))
    return {"por_inmobiliaria": sorted(por_inmo.items(), key=lambda x: -x[1]),
            "por_proyecto": sorted(por_proy.items(), key=lambda x: -x[1])[:10],
            "por_broker": sorted(por_broker.items(), key=lambda x: -x[1]),
            "carpetas": carpetas, "equipo": equipo,
            "volumen_diario": sorted(volumen.items()),
            "comparativa": {"semana_actual": sem_actual, "semana_anterior": sem_previa,
                            "variacion_pct": round((sem_actual - sem_previa) * 100 / sem_previa, 1) if sem_previa else None}}


def _dias_habiles_desde(iso):
    try:
        d0 = datetime.fromisoformat(str(iso)[:19]).date()
    except Exception:
        return None
    hoy, n, d = datetime.now(timezone.utc).date(), 0, None
    d = d0
    while d < hoy:
        d += timedelta(days=1)
        if d.weekday() < 5:
            n += 1
    return n


@gpanel.get("/command-center")
async def gerencia_command_center(request: Request):
    """DASHBOARD UNIFICADO DE GERENCIA: Command Center, Brokers, Carga Administrativa y Bandeja."""
    _exigir(request, ("admin", "maestro", "gerencia", "contralor"))
    ahora = datetime.now(timezone.utc)
    mes = ahora.strftime("%Y-%m")
    mes_prev = (ahora.replace(day=1) - timedelta(days=1)).strftime("%Y-%m")
    uf_cfg = await db.config.find_one({"_key": "uf"}) or {}
    valor_uf = float(uf_cfg.get("valor_uf") or 0)
    activas, cerradas_mes, cerradas_prev, dias_cierre = 0, 0, 0, []
    bloqueadas, monto_uf = 0, 0.0
    nuevas_mes, nuevas_prev = 0, 0
    brokers, bandeja = {}, []
    async for fd in db.folders.find({"oculto_supercarpeta": {"$ne": True}}):
        brk = (fd.get("broker_origen") or fd.get("broker_codigo") or "Directo").strip() or "Directo"
        b = brokers.setdefault(brk, {"broker": brk, "clientes": 0, "tramitacion": 0,
                                     "cerradas_mes": 0, "cerradas_total": 0, "monto_uf": 0.0,
                                     "dias_resp": []})
        b["clientes"] += 1
        creado = str(fd.get("created_at") or fd.get("created") or "")
        if creado[:7] == mes:
            nuevas_mes += 1
        elif creado[:7] == mes_prev:
            nuevas_prev += 1
        cerrada_at = str(fd.get("escritura_confirmada_at") or "")
        if cerrada_at:
            b["cerradas_total"] += 1
            if cerrada_at[:7] == mes:
                cerradas_mes += 1
                b["cerradas_mes"] += 1
            elif cerrada_at[:7] == mes_prev:
                cerradas_prev += 1
            try:
                dias_cierre.append(max((datetime.fromisoformat(cerrada_at[:19])
                                        - datetime.fromisoformat(creado[:19])).days, 0))
            except Exception:
                pass
            continue
        activas += 1
        b["tramitacion"] += 1
        p_uf = float(fd.get("proyeccion_uf") or 0)
        b["monto_uf"] += p_uf
        monto_uf += p_uf
        faltan = fd.get("faltantes_auto_lista") or []
        if faltan:
            bloqueadas += 1
        ult = str(fd.get("updated_at") or creado or "")
        dh = _dias_habiles_desde(ult)
        if dh is not None:
            b["dias_resp"].append(dh)
        s = (fd.get("subsidio_proyeccion") or "").upper()
        tipo = ("DS49" if "49" in s else "DS1" if "DS1" in s or "DS 1" in s
                else "SERVIU" if "SERVIU" in s else "")
        if not tipo:
            tipo = "Vivienda usada" if "usad" in (fd.get("tipo_operacion") or "").lower() else "Mutuo hipotecario"
        estado = ("Estudio aprobado" if fd.get("estudio_recibido_at")
                  else "Tasación recibida" if fd.get("tasacion_informe_recibido_at")
                  else "SET emitido" if fd.get("set_credito_at") else "En tramitación")
        if faltan:
            estado = f"Bloqueada por normativa · {len(faltan)} doc. faltante(s)"
        bandeja.append({"fid": fd["id"], "cliente": fd.get("nombre"), "rut": fd.get("rut") or "",
                        "broker": brk, "tipo": tipo, "estado": estado,
                        "ultimo_movimiento": ult[:10] or "No disponible",
                        "dias_sin_movimiento": dh,
                        "alerta": bool(dh is not None and dh > 5),
                        "urgente": bool(fd.get("urgente"))})
    bandeja.sort(key=lambda x: (not x["urgente"], -(x["dias_sin_movimiento"] or 0)))
    lista_brokers = []
    for b in brokers.values():
        b["monto_uf"] = round(b["monto_uf"], 1)
        b["tasa_cierre"] = round(b["cerradas_total"] * 100 / b["clientes"]) if b["clientes"] else 0
        dr = b.pop("dias_resp")
        b["dias_respuesta"] = round(sum(dr) / len(dr), 1) if dr else None
        b["semaforo"] = ("verde" if (b["cerradas_mes"] > 0 or b["tasa_cierre"] >= 40
                                     or (b["dias_respuesta"] is not None and b["dias_respuesta"] <= 5))
                         else "amarillo" if (b["dias_respuesta"] is not None and b["dias_respuesta"] <= 10)
                         else "rojo")
        lista_brokers.append(b)
    lista_brokers.sort(key=lambda x: (-x["cerradas_mes"], -x["tasa_cierre"], -x["clientes"]))
    if lista_brokers:
        lista_brokers[0]["mejor_mes"] = True
    # ZONA 3: carga administrativa real del mes
    ini_mes = f"{mes}-01"
    equipo = {"Daniela Galindo": {"docs": 0, "correos": 0, "ops": set()},
              "Victoria Vilchez": {"docs": 0, "correos": 0, "ops": set()}}
    emails_eq = {"Daniela Galindo": "danielagalindo@centralmutuos.cl",
                 "Victoria Vilchez": "victoriavilches@centralmutuos.cl"}
    tiempos = {k: {} for k in equipo}
    async for r in db.estado_manual_log.find({"fecha": {"$gte": ini_mes}}):
        por = (r.get("por") or "").lower()
        key = ("Victoria Vilchez" if "victoria" in por or "vilche" in por else
               "Daniela Galindo" if "daniela" in por or "galindo" in por else None)
        if not key:
            continue
        equipo[key]["docs"] += 1
        if r.get("folder_id"):
            equipo[key]["ops"].add(r["folder_id"])
            tiempos[key].setdefault(r["folder_id"], []).append(str(r.get("fecha") or ""))
    async for r in db.correos_smtp_log.find({"fecha": {"$gte": ini_mes}}):
        blob = str(r).lower()
        for key, em in emails_eq.items():
            if em in blob:
                equipo[key]["correos"] += 1
    carga = {}
    for key, e in equipo.items():
        horas_ops = []
        for fechas in tiempos[key].values():
            fs = sorted(f for f in fechas if f)
            if len(fs) >= 2:
                try:
                    horas_ops.append((datetime.fromisoformat(fs[-1][:19])
                                      - datetime.fromisoformat(fs[0][:19])).total_seconds() / 3600)
                except Exception:
                    pass
        carga[key] = {"documentos_procesados": e["docs"], "correos_gestionados": e["correos"],
                      "operaciones_tramitadas": len(e["ops"]),
                      "horas_promedio_resolucion": round(sum(horas_ops) / len(horas_ops), 1) if horas_ops else None}
    total_docs = sum(c["documentos_procesados"] for c in carga.values())
    for c in carga.values():
        share = c["documentos_procesados"] / total_docs if total_docs else 0
        vol = c["documentos_procesados"] + c["correos_gestionados"]
        c["indicador_carga"] = "Alta" if (vol >= 40 or share >= 0.65) else ("Media" if vol >= 15 else "Normal")
    # ZONA 1
    base_mes = activas + cerradas_mes
    base_prev = activas + cerradas_prev
    tasa_mes = round(cerradas_mes * 100 / base_mes, 1) if base_mes else 0
    tasa_prev = round(cerradas_prev * 100 / base_prev, 1) if base_prev else 0
    zona1 = {
        "operaciones_activas": {"valor": activas, "tendencia": nuevas_mes - nuevas_prev,
                                "nuevas_mes": nuevas_mes, "nuevas_mes_anterior": nuevas_prev},
        "monto_tramitacion": {"uf": round(monto_uf, 1), "clp": round(monto_uf * valor_uf),
                              "valor_uf_dia": valor_uf},
        "tasa_cierre": {"mes_actual": tasa_mes, "mes_anterior": tasa_prev,
                        "tendencia": round(tasa_mes - tasa_prev, 1),
                        "cierres_mes": cerradas_mes, "cierres_mes_anterior": cerradas_prev},
        "tiempo_promedio_cierre_dias": round(sum(dias_cierre) / len(dias_cierre)) if dias_cierre else None,
        "bloqueadas_normativa": {"n": bloqueadas,
                                 "pct": round(bloqueadas * 100 / activas, 1) if activas else 0},
        "docs_sin_clasificar": await db.docs_sin_clasificar.count_documents({}),
    }
    tendencia = {}
    async for fd in db.folders.find({}, {"created_at": 1}):
        m = str(fd.get("created_at") or "")[:7]
        if m:
            tendencia[m] = tendencia.get(m, 0) + 1
    # ── CUMPLEAÑOS DE LA SEMANA (campo fecha_nacimiento de la carpeta) ──
    cumpleanos = []
    hoy_d = ahora.date()
    async for fd in db.folders.find({"fecha_nacimiento": {"$exists": True, "$ne": ""},
                                     "oculto_supercarpeta": {"$ne": True}},
                                    {"id": 1, "nombre": 1, "fecha_nacimiento": 1, "broker_origen": 1}):
        try:
            fn = datetime.fromisoformat(fd["fecha_nacimiento"]).date()
        except (ValueError, TypeError):
            continue
        for anio in (hoy_d.year, hoy_d.year + 1):
            try:
                prox = fn.replace(year=anio)
            except ValueError:
                prox = fn.replace(year=anio, day=28)
            delta = (prox - hoy_d).days
            if 0 <= delta <= 7:
                cumpleanos.append({"fid": fd["id"], "cliente": fd.get("nombre"),
                                   "fecha": prox.strftime("%d/%m/%Y"), "dias": delta,
                                   "broker": fd.get("broker_origen") or ""})
                break
    cumpleanos.sort(key=lambda x: x["dias"])
    return {"mes": mes, "zona1": zona1, "brokers": lista_brokers,
            "carga_administrativa": carga, "bandeja": bandeja[:80],
            "cumpleanos_semana": cumpleanos,
            "serie_mensual": sorted(tendencia.items())[-6:], "generado": _now()}


@gpanel.post("/fecha-nacimiento")
async def gerencia_fecha_nacimiento(payload: dict, request: Request):
    """Registra la fecha de nacimiento del cliente (para alertas de cumpleaños)."""
    _exigir(request, ("admin", "maestro", "gerencia", "administracion"))
    fid = (payload or {}).get("fid") or ""
    fecha = ((payload or {}).get("fecha") or "").strip()
    fd = await db.folders.find_one({"id": fid})
    if not fd:
        raise HTTPException(status_code=404, detail="Operación no encontrada")
    m = re.match(r"^(\d{2})/(\d{2})/(\d{4})$", fecha)
    if m:
        fecha = f"{m.group(3)}-{m.group(2)}-{m.group(1)}"
    if not re.match(r"^\d{4}-\d{2}-\d{2}$", fecha):
        raise HTTPException(status_code=400, detail="Formato de fecha inválido — use DD/MM/AAAA")
    await db.folders.update_one({"id": fid}, {"$set": {"fecha_nacimiento": fecha}})
    return {"ok": True, "fid": fid, "fecha_nacimiento": fecha}


@gpanel.post("/urgente")
async def gerencia_marcar_urgente(payload: dict, request: Request):
    _exigir(request, ("admin", "maestro", "gerencia"))
    fid = (payload or {}).get("fid") or ""
    fd = await db.folders.find_one({"id": fid})
    if not fd:
        raise HTTPException(status_code=404, detail="Operación no encontrada")
    nuevo = not fd.get("urgente")
    await db.folders.update_one({"id": fid}, {"$set": {"urgente": nuevo}})
    return {"ok": True, "fid": fid, "urgente": nuevo}


@gpanel.get("/inteligencia")
async def gerencia_inteligencia(request: Request, broker: str = ""):
    """CENTRO DE INTELIGENCIA COMERCIAL: panel por cliente (docs + fechas + preview),
    estadísticas por broker/inmobiliaria, subsidios y real vs proyectado."""
    _exigir(request, ("admin", "maestro", "gerencia", "contralor"))
    mes = datetime.now(timezone.utc).strftime("%Y-%m")
    clientes, stats_broker, subsidios, real_uf = [], {}, {}, 0.0
    q = {"oculto_supercarpeta": {"$ne": True}}
    async for fd in db.folders.find(q).sort("nombre", 1):
        brk = (fd.get("broker_origen") or fd.get("broker_codigo") or "Directo").strip() or "Directo"
        if broker and broker.lower() not in brk.lower():
            continue
        arch = [a if isinstance(a, str) else (a.get("ruta") or a.get("nombre") or "")
                for a in (fd.get("archivos") or [])]

        def _ruta(prefs):
            return next((a for a in arch if any(a.startswith(p) for p in prefs)), "")
        docs = [
            {"doc": "Tasación",
             "estado": "Recibida" if fd.get("tasacion_informe_recibido_at") else ("Solicitada" if fd.get("tasacion_solicitada_at") else "Pendiente"),
             "fecha": str(fd.get("tasacion_informe_recibido_at") or fd.get("tasacion_solicitada_at") or "")[:10],
             "ruta": _ruta(["99_otros/TASACION_"]), "accion": "tasacion"},
            {"doc": "Estudio de Título",
             "estado": "Con Observaciones" if (fd.get("reparos_alertas") or []) else ("Recibido" if fd.get("estudio_recibido_at") else "Pendiente"),
             "fecha": str(fd.get("estudio_recibido_at") or "")[:10],
             "ruta": _ruta(["07_estudio_titulo/"]), "accion": "estudio"},
            {"doc": "Cédula de Crédito (SET)",
             "estado": fd.get("set_credito_estado") or "Pendiente",
             "fecha": str(fd.get("set_credito_at") or "")[:10], "ruta": "", "accion": ""},
            {"doc": "DPS", "estado": "Recibido" if fd.get("dps_recibido_at") else "Pendiente",
             "fecha": str(fd.get("dps_recibido_at") or "")[:10], "ruta": "", "accion": ""},
            {"doc": "Actualización de Documentos",
             "estado": "Actualizados" if fd.get("datos_financieros_ocr_fecha") else "Pendiente",
             "fecha": str(fd.get("datos_financieros_ocr_fecha") or "")[:10], "ruta": "", "accion": "docs"},
        ]
        clientes.append({"fid": fd["id"], "cliente": fd.get("nombre"), "rut": fd.get("rut") or "",
                         "rut_propiedad": (fd.get("rut_propiedad") or fd.get("rol_propiedad") or ""),
                         "broker": brk, "inmobiliaria": fd.get("inmobiliaria") or "Directa",
                         "proyecto": fd.get("proyecto") or "", "docs": docs})
        # Estadísticas
        sb = stats_broker.setdefault(brk, {"clientes": 0, "aprobadas": 0, "cerradas": 0, "dias_cierre": []})
        sb["clientes"] += 1
        if fd.get("set_credito_at") or fd.get("escritura_confirmada_at"):
            sb["aprobadas"] += 1
        if fd.get("escritura_confirmada_at"):
            sb["cerradas"] += 1
            try:
                ini = datetime.fromisoformat(str(fd.get("created") or fd.get("created_at"))[:19])
                fin = datetime.fromisoformat(str(fd.get("escritura_confirmada_at"))[:19])
                sb["dias_cierre"].append((fin - ini).days)
            except Exception:
                pass
        s = (fd.get("subsidio_proyeccion") or "").upper()
        tipo_s = ("DS49" if "49" in s else "DS1" if "DS1" in s or "DS 1" in s
                  else "SERVIU" if "SERVIU" in s else "Con Subsidio (otro)" if "CON" in s else "Sin Subsidio")
        subsidios[tipo_s] = subsidios.get(tipo_s, 0) + 1
        if (fd.get("mes_proyeccion") or mes) == mes:
            real_uf += float(fd.get("proyeccion_uf") or 0)
    brokers_out = []
    for b, s in sorted(stats_broker.items(), key=lambda x: -x[1]["clientes"]):
        brokers_out.append({"broker": b, "clientes": s["clientes"],
                            "avance_pct": round(s["aprobadas"] * 100 / s["clientes"]) if s["clientes"] else 0,
                            "tasa_aprobacion": round(s["aprobadas"] * 100 / s["clientes"]) if s["clientes"] else 0,
                            "cierres": s["cerradas"],
                            "dias_promedio_cierre": round(sum(s["dias_cierre"]) / len(s["dias_cierre"])) if s["dias_cierre"] else None})
    meta_cfg = await db.config.find_one({"_key": f"proyeccion_{mes}"}) or await db.config.find_one({"_key": "proyeccion_agosto"}) or {}
    ratios_cfg = await db.config.find_one({"_key": "gerencia_ratios"}, {"_id": 0}) or {}
    return {"clientes": clientes, "total": len(clientes), "brokers": brokers_out,
            "subsidios": sorted(subsidios.items(), key=lambda x: -x[1]),
            "proyeccion": {"mes": mes, "meta_uf": meta_cfg.get("meta_uf") or 0,
                           "real_uf": round(real_uf, 1),
                           "cumplimiento_pct": round(real_uf * 100 / meta_cfg["meta_uf"], 1) if meta_cfg.get("meta_uf") else 0},
            "ratios_configurables": ratios_cfg.get("ratios") or []}


@gpanel.post("/accion")
async def gerencia_accion(payload: dict, request: Request):
    """Botones de acción por cliente: correo al destinatario con CC de libre elección (Gerencia)."""
    _exigir(request, ("admin", "maestro", "gerencia"))
    fid, tipo = (payload.get("fid") or ""), (payload.get("tipo") or "")
    ACC = {"tasacion": ("Solicitar tasación al broker", "broker"),
           "estudio": ("Consultar estudio de título", "administrativo"),
           "docs": ("Pedir actualización de documentos", "administrativo"),
           "seguimiento": ("Solicitar estado y seguimiento de la operación", "administrativo")}
    if tipo not in ACC:
        raise HTTPException(status_code=400, detail="Acción inválida")
    fd = await db.folders.find_one({"id": fid})
    if not fd:
        raise HTTPException(status_code=404, detail="Carpeta no existe")
    label, destino_tipo = ACC[tipo]
    destino = ""
    if destino_tipo == "broker":
        bk = await db.brokers_fuentes.find_one({"$or": [{"codigo": fd.get("broker_codigo")},
                                                        {"nombre": fd.get("broker_origen")}]}) or {}
        destino = bk.get("email") or bk.get("correo") or ""
    else:
        async for e in db.ejecutivos_correo.find({"email": {"$ne": ""}}):
            destino = e["email"]
            break
    destino = destino or os.environ.get("MAIL2_USER", "")
    if not destino:
        raise HTTPException(status_code=400, detail="Sin destinatario configurado para esta acción")
    # CC LIBRE (Gerencia Comercial): Daniela, Victoria, ambas u otros — sin restricción bloqueante
    cc = payload.get("cc") or []
    if isinstance(cc, str):
        cc = [cc]
    cc = [c.strip() for c in cc if isinstance(c, str) and "@" in c
          and c.strip().lower() != destino.lower()]
    cliente = fd.get("nombre") or ""
    rut_txt = f" (RUT {fd.get('rut')})" if fd.get("rut") else ""
    cuerpo = (f"<div style='font-family:Arial,Helvetica,sans-serif;background:#fff;color:#111;font-size:14px'>"
              f"<p>Estimados, junto con saludar:</p>"
              f"<p>Gerencia Comercial solicita: <b>{label}</b> para el cliente <b>{cliente}</b>{rut_txt}.</p>"
              f"<p>Agradeceremos gestionar a la brevedad e informar el estado.</p>"
              f"<p style='color:#555'>Saludos cordiales,<br><b>Central Mutuos</b></p></div>")
    await _validar_normativas_op(texto_saliente=cuerpo, cc=cc, rol=_rol(request),
                                 contexto=f"acción de Gerencia '{label}'")
    import email_service as mail
    asyncio.create_task(asyncio.to_thread(
        mail.send_mail, destino, f"{label} — {cliente}", cuerpo, [], "secundaria", cc or None))
    await db.folders.update_one({"id": fid}, {"$push": {"bitacora_solicitudes": {
        "tipo": f"gerencia_{tipo}", "asunto": f"{label} — {cliente}", "para": destino, "cc": cc,
        "en": _now(), "estado": "en_envio", "documentos": [label]}}})
    return {"ok": True, "accion": label, "para": destino, "cc": cc,
            "nota": ("Correo generado y despachado en segundo plano"
                     + (f" con copia a: {', '.join(cc)}" if cc else " (sin copias)"))}


@gpanel.get("/cc-opciones")
async def gerencia_cc_opciones(request: Request):
    """Destinatarios disponibles para copiar (CC) en los correos de Gerencia."""
    _exigir(request, ("admin", "maestro", "gerencia"))
    vistos, opciones = set(), []
    for nombre, email in [("Daniela Galindo", "danielagalindo@centralmutuos.cl"),
                          ("Victoria Vilchez", "victoriavilches@centralmutuos.cl")]:
        opciones.append({"nombre": nombre, "email": email})
        vistos.add(email)
    async for e in db.ejecutivos_correo.find({"email": {"$ne": ""}}):
        if e.get("email") and e["email"] not in vistos:
            opciones.append({"nombre": e.get("nombre") or e["email"], "email": e["email"]})
            vistos.add(e["email"])
    async for u in db.users.find({"email": {"$exists": True, "$nin": ["", None]}}):
        if u.get("email") and u["email"] not in vistos:
            opciones.append({"nombre": u.get("nombre") or u["email"], "email": u["email"]})
            vistos.add(u["email"])
    return {"opciones": opciones}


# ═══════════════ MÓDULO BROKER — EXCEL OFICIAL DE PROYECCIÓN ═══════════════
COLS_EXCEL = ["Nombre Cliente", "RUT", "Inmobiliaria", "Proyecto", "Ciudad", "Notaría",
              "Monto UF", "Subsidio (Con/Sin)", "Estudio de Títulos", "Tasación",
              "Actualización de Documentos", "Fecha Firma Estimada (AAAA-MM-DD)"]


@brokerx.get("/ventana-proyeccion")
async def broker_ventana_proyeccion(request: Request):
    """Estado de la ventana de carga (día 1 al 5° hábil del mes)."""
    from malla_inteligencia import _ventana_proyeccion
    abierta, limite = _ventana_proyeccion()
    return {"abierta": abierta, "limite": limite,
            "mensaje": "" if abierta else ("La ventana de carga de proyecciones está cerrada. "
                                           "Disponible entre el día 1 y 5 hábil de cada mes.")}


@brokerx.get("/formato-excel")
async def broker_formato_excel(request: Request):
    """Formato oficial generado desde las columnas de la Supercarpeta (pre-llenado con sus clientes)."""
    from fastapi.responses import Response as _Resp
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    sub = (getattr(request.state, "user", {}) or {}).get("sub") or ""
    wb = Workbook()
    ws = wb.active
    ws.title = "Proyección"
    ws.append(COLS_EXCEL)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="14263F")
    async for fd in db.folders.find({"broker_codigo": sub, "oculto_supercarpeta": {"$ne": True}}).sort("nombre", 1):
        em = fd.get("estados_manuales") or {}
        ws.append([fd.get("nombre") or "", fd.get("rut") or "", fd.get("inmobiliaria") or "",
                   fd.get("proyecto") or "", fd.get("ciudad") or "", fd.get("notaria") or "",
                   fd.get("proyeccion_uf") or "", fd.get("subsidio_proyeccion") or "",
                   (em.get("estudio") or {}).get("estado") or "", (em.get("tasacion") or {}).get("estado") or "",
                   "", str(fd.get("fecha_firma") or "")[:10]])
    for i, w in enumerate([28, 14, 18, 22, 14, 16, 10, 16, 18, 16, 24, 26], start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    buf = io.BytesIO()
    wb.save(buf)
    return _Resp(content=buf.getvalue(),
                 media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                 headers={"Content-Disposition": 'attachment; filename="FORMATO_PROYECCION_SUPERCARPETA.xlsx"'})


@brokerx.post("/cargar-excel")
async def broker_cargar_excel(request: Request):
    """Sube el Excel oficial → alimenta directamente las carpetas de la Supercarpeta.
    Ventana: solo del día 1 al 5° día hábil del mes."""
    from malla_inteligencia import _ventana_proyeccion
    claims = getattr(request.state, "user", {}) or {}
    sub = claims.get("sub") or ""
    abierta, limite = _ventana_proyeccion()
    if not abierta and claims.get("rol") not in ("admin", "maestro"):
        raise HTTPException(status_code=423, detail=(
            f"⛔ Ventana de carga cerrada: la proyección solo se puede subir entre el día 1 y el "
            f"5° día hábil de cada mes (última fecha de este mes: {limite})."))
    form = await request.form()
    archivo = form.get("archivo")
    if archivo is None:
        raise HTTPException(status_code=400, detail="Adjunte el archivo Excel del formato oficial")
    import io
    from openpyxl import load_workbook
    try:
        wb = load_workbook(io.BytesIO(await archivo.read()), data_only=True)
        ws = wb.active
    except Exception:
        raise HTTPException(status_code=400, detail="El archivo no es un Excel válido (.xlsx del formato oficial)")
    mes_actual = datetime.now(timezone.utc).strftime("%Y-%m")
    creados, actualizados, errores = 0, 0, []
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        vals = [str(v).strip() if v is not None else "" for v in (list(row) + [""] * 12)[:12]]
        nombre, rut, inmo, proy, ciudad, notaria, monto, subsidio, estudio, tasacion, _docs, ffirma = vals
        if not nombre:
            continue
        try:
            monto_v = float(str(monto).replace(".", "").replace(",", ".")) if monto else None
        except ValueError:
            monto_v = None
        q = {"$or": [{"rut": rut}, {"nombre": {"$regex": f"^{re.escape(nombre)}$", "$options": "i"}}]} if rut \
            else {"nombre": {"$regex": f"^{re.escape(nombre)}$", "$options": "i"}}
        fd = await db.folders.find_one(q)
        # REGLA RUT ÚNICO: el primer broker que registró el RUT lo retiene permanentemente
        if fd and rut and (fd.get("broker_codigo") or "") not in ("", sub) \
                and claims.get("rol") not in ("admin", "maestro"):
            errores.append(f"Fila {idx} — {nombre}: Este RUT ya está registrado en el sistema por otro ejecutivo.")
            continue
        setd = {"broker_codigo": sub, "broker_origen": claims.get("nombre") or sub, "updated_at": _now()}
        for k, v in (("rut", rut), ("inmobiliaria", inmo), ("proyecto", proy), ("ciudad", ciudad),
                     ("notaria", notaria), ("subsidio_proyeccion", subsidio), ("fecha_firma", ffirma)):
            if v:
                setd[k] = v
        if monto_v:
            setd["proyeccion_uf"] = monto_v
        est_man = {}
        for hito, val in (("estudio", estudio), ("tasacion", tasacion)):
            if val:
                est_man[hito] = {"estado": val, "en": _now(), "por": f"broker:{sub}"}
        if fd:
            upd = dict(setd)
            for h, v in est_man.items():
                upd[f"estados_manuales.{h}"] = v
            await db.folders.update_one({"id": fd["id"]}, {"$set": upd})
            actualizados += 1
        else:
            nuevo = {"id": str(uuid.uuid4()), "nombre": nombre.upper(),
                     "mes_proyeccion": mes_actual, "created": _now(), **setd}
            if est_man:
                nuevo["estados_manuales"] = est_man
            await db.folders.insert_one(nuevo)
            creados += 1
    await db.broker_activity_log.insert_one({
        "id": str(uuid.uuid4()), "broker_codigo": sub, "accion": "carga_excel_proyeccion",
        "detalle": {"creados": creados, "actualizados": actualizados}, "fecha": _now()})
    return {"ok": True, "creados": creados, "actualizados": actualizados,
            "errores": errores, "nota": "Las carpetas de la Supercarpeta fueron alimentadas sin ingreso manual"}


# ═══ 🛡️ REGLAS DE ORO #71/#72 — AUDITORÍA PRE-MESA DEL CONTRALOR ═══
# #71: Bóveda de Criterios verificada ANTES de todo envío a mesa (bloqueo + alerta admin).
# #72: edad calculada automáticamente desde la cédula (OCR) para validar plazo máximo.
import io
from datetime import date

import folders_service as fsvc
import ocr_service
from criterios_data import DEFAULT_CRITERIOS, DEFAULT_UF

MIN_CREDITO_SIN_SUBSIDIO_UF = 2000  # Regla de Oro #71 — SOLO viviendas SIN subsidio (con subsidio NO hay mínimo)

MESES = {"ene": 1, "feb": 2, "mar": 3, "abr": 4, "may": 5, "jun": 6, "jul": 7,
         "ago": 8, "sep": 9, "sept": 9, "set": 9, "oct": 10, "nov": 11, "dic": 12,
         "enero": 1, "febrero": 2, "marzo": 3, "abril": 4, "mayo": 5, "junio": 6,
         "julio": 7, "agosto": 8, "septiembre": 9, "setiembre": 9, "octubre": 10,
         "noviembre": 11, "diciembre": 12}
RX_NAC_TXT = re.compile(r"nacimi?ento\D{0,50}?(\d{1,2})\s+([a-záéíóú]{3,12})\.?\s+(\d{4})", re.I)
RX_NAC_NUM = re.compile(r"nacimi?ento\D{0,50}?(\d{1,2})[./\- ](\d{1,2})[./\- ](\d{4})", re.I)
RX_FECHA_SUELTA = re.compile(r"\b(\d{1,2})\s+(ene|feb|mar|abr|may|jun|jul|ago|sept?|set|oct|nov|dic)[a-z]*\.?\s+(\d{4})\b", re.I)


def _edad_de_fecha(f):
    hoy = date.today()
    return hoy.year - f.year - ((hoy.month, hoy.day) < (f.month, f.day))


def _parse_nacimiento(texto):
    m = RX_NAC_TXT.search(texto)
    if m:
        mes = MESES.get(m.group(2).lower().rstrip("."))
        if mes:
            try:
                return date(int(m.group(3)), mes, int(m.group(1)))
            except ValueError:
                pass
    m = RX_NAC_NUM.search(texto)
    if m:
        try:
            return date(int(m.group(3)), int(m.group(2)), int(m.group(1)))
        except ValueError:
            pass
    # Respaldo: la fecha con año más antiguo de la cédula (emisión/vencimiento son recientes)
    hoy = date.today()
    candidatas = []
    for m in RX_FECHA_SUELTA.finditer(texto):
        mes = MESES.get(m.group(2).lower())
        if not mes:
            continue
        try:
            f = date(int(m.group(3)), mes, int(m.group(1)))
        except ValueError:
            continue
        if 1930 <= f.year <= hoy.year - 17:
            candidatas.append(f)
    return min(candidatas) if candidatas else None


def extraer_nacimiento_cedula_sync(nombre_folder):
    """Regla #72: OCR de la cédula → fecha de nacimiento. Devuelve date o None."""
    base = fsvc.folder_dir(nombre_folder) / "01_cedula"
    if not base.exists():
        return None
    for p in sorted(base.iterdir()):
        try:
            if p.suffix.lower() == ".pdf":
                texto, _ = ocr_service.extraer_texto(p.read_bytes(), p.name, force_ocr=True)
            elif p.suffix.lower() in (".jpg", ".jpeg", ".png", ".webp"):
                from PIL import Image
                texto, _ = ocr_service.ocr_imagen(Image.open(io.BytesIO(p.read_bytes())))
            else:
                continue
        except Exception as e:
            logging.warning(f"OCR cédula {p.name}: {e}")
            continue
        f = _parse_nacimiento(texto or "")
        if f and 18 <= _edad_de_fecha(f) <= 99:
            return f
    return None


async def obtener_edad(doc):
    """Edad del titular: 1° la guardada en la carpeta; 2° cédula por OCR (y se persiste)."""
    if doc.get("edad_titular"):
        return int(doc["edad_titular"])
    df = doc.get("datos_financieros") or {}
    if df.get("edad"):
        return int(df["edad"])
    import asyncio
    if doc.get("edad_ocr_fallido"):
        return None
    f = await asyncio.to_thread(extraer_nacimiento_cedula_sync, doc.get("nombre", ""))
    if not f:
        await db.folders.update_one({"id": doc["id"]}, {"$set": {"edad_ocr_fallido": True}})
        return None
    edad = _edad_de_fecha(f)
    await db.folders.update_one({"id": doc["id"]}, {
        "$set": {"fecha_nacimiento": f.isoformat(), "edad_titular": edad},
        "$push": {"historial": {"fecha": _now(),
                                "accion": f"🪪 Regla #72: fecha de nacimiento leída de la cédula "
                                          f"({f.strftime('%d-%m-%Y')}) — edad calculada: {edad} años"}}})
    doc["fecha_nacimiento"] = f.isoformat()
    doc["edad_titular"] = edad
    return edad


async def _criterios():
    cfg = await db.config.find_one({"_key": "criterios"}, {"_id": 0})
    return cfg or DEFAULT_CRITERIOS


def _num(v):
    try:
        x = float(str(v).replace(",", "."))
        return x if x > 0 else None
    except (TypeError, ValueError):
        return None


RX_SIM_PLAZO = re.compile(r"plazo(?:\s+cr[eé]dito)?\s*:?\s*(\d{1,2})\s*a[ñn]os|·\s*(\d{1,2})\s*a[ñn]os", re.I)
RX_SIM_PLAZO_TASA = re.compile(r"\b([1-4]\d)\s+\d{1,2},\d{2}%")
RX_SIM_MONTO = re.compile(r"(?:monto\s+)?cr[eé]dito\s*:?\s*([\d.]+,\d{2})", re.I)
RX_SIM_DIV = re.compile(r"total\s+dividendo\s+mensual\s+([\d.,]+)\s*uf", re.I)
RX_SIM_TASA = re.compile(r"tasa\s+(?:anual|del\s+cr[eé]dito)\s*:?\s*(\d{1,2}[.,]\d{1,2})\s*%", re.I)
RX_CMF_TOTAL = re.compile(r"total\s+\$\s*([\d.]+)\s+\$\s*([\d.]+)\s+\$\s*([\d.]+)\s+\$\s*([\d.]+)\s+\$\s*([\d.]+)", re.I)


def _clp_miles(s):
    """Montos CMF vienen en MILES de pesos → CLP reales."""
    return int(re.sub(r"[^\d]", "", str(s)) or 0) * 1000


RX_PAGO_KW = re.compile(r"comprobante|transferencia|pago|abono|dep[oó]sito|cancelaci[oó]n|recibo|voucher|pagad[oa]", re.I)
RX_MONTOS = re.compile(r"\$\s?([\d.]{4,15})|(?:monto|total|valor)\D{0,12}([\d.]{4,15})", re.I)


def validar_comprobante_mora(raw, filename, morosidad_clp):
    """🧾 Validación automática del comprobante de pago de mora (sin intervención del admin).
    Reglas: legible + parece comprobante de pago + monto detectado ≥ 95% de la mora."""
    try:
        texto, _m = ocr_service.extraer_texto(raw, filename, force_ocr=False)
        if not texto or len(texto.strip()) < 30:
            texto, _m = ocr_service.extraer_texto(raw, filename, force_ocr=True)
    except Exception as e:
        return {"ok": False, "motivo": f"No se pudo leer el archivo ({str(e)[:80]}). Suba un PDF o imagen legible."}
    t = " ".join((texto or "").split())
    if len(t) < 30:
        return {"ok": False, "motivo": "El archivo es ilegible para el sistema. Suba un comprobante nítido (PDF o foto clara)."}
    if not RX_PAGO_KW.search(t):
        return {"ok": False, "motivo": "El documento no parece un comprobante de pago (no contiene palabras como "
                                       "'transferencia', 'pago', 'abono' o 'comprobante'). Verifique el archivo subido."}
    montos = []
    for m in RX_MONTOS.finditer(t):
        v = int(re.sub(r"[^\d]", "", m.group(1) or m.group(2) or "") or 0)
        if 1000 <= v <= 10_000_000_000:
            montos.append(v)
    if not montos:
        return {"ok": False, "motivo": "No se detectó ningún monto en el comprobante. Verifique que el documento muestre el valor pagado."}
    monto_max = max(montos)
    if morosidad_clp and monto_max < morosidad_clp * 0.95:
        return {"ok": False, "monto_detectado": monto_max,
                "motivo": (f"El comprobante indica un pago de ${monto_max:,.0f}, pero la mora registrada en el CMF "
                           f"es de ${morosidad_clp:,.0f}. El pago debe cubrir el total de la deuda morosa "
                           "(o suba el comprobante del saldo restante).")}
    return {"ok": True, "monto_detectado": monto_max, "motivo": ""}


async def cerrar_alertas_mora(fid):
    r = await db.alertas.update_many(
        {"folder_id": fid, "tipo": "auditoria71:morosidad", "leida": {"$ne": True}},
        {"$set": {"leida": True, "cerrada_por": "comprobante_pago_validado", "cerrada_at": _now()}})
    return r.modified_count


RX_REGULARIZACION = re.compile(r"regularizaci[oó]n|repactaci[oó]n|convenio\s+de\s+pago|compromiso\s+de\s+pago|"
                               r"plan\s+de\s+pago|aclaraci[oó]n\s+de\s+(deuda|mora)|renegociaci[oó]n", re.I)


def validar_formulario_regularizacion(raw, filename, nombre_cliente="", rut=""):
    """📋 Validación automática del formulario manual de regularización de mora."""
    try:
        texto, _m = ocr_service.extraer_texto(raw, filename, force_ocr=False)
        if not texto or len(texto.strip()) < 40:
            texto, _m = ocr_service.extraer_texto(raw, filename, force_ocr=True)
    except Exception as e:
        return {"ok": False, "motivo": f"No se pudo leer el archivo ({str(e)[:80]}). Suba un PDF o imagen legible."}
    t = " ".join((texto or "").split())
    if len(t) < 40:
        return {"ok": False, "motivo": "El formulario es ilegible para el sistema. Suba un documento nítido (PDF o foto clara)."}
    if not RX_REGULARIZACION.search(t):
        return {"ok": False, "motivo": "El documento no parece un formulario de regularización (no contiene términos como "
                                       "'regularización', 'convenio de pago', 'compromiso de pago' o 'repactación')."}
    tl = t.lower()
    rut_dig = re.sub(r"[^\dkK]", "", rut or "")
    nombre_ok = any(p.lower() in tl for p in (nombre_cliente or "").split() if len(p) > 3)
    rut_ok = bool(rut_dig) and rut_dig[:-1] in re.sub(r"[^\dkK]", "", t)
    if not (nombre_ok or rut_ok):
        return {"ok": False, "motivo": f"El formulario no menciona al cliente ({nombre_cliente or rut}). "
                                       "Verifique que el documento corresponda a este cliente."}
    return {"ok": True, "motivo": ""}


def leer_cmf_sync(nombre_folder):
    """🧾 Lector CMF: morosidad real desde el Informe de Deudas (filas 'Total' de
    Deuda Directa + Indirecta, columnas 30-59 / 60-89 / 90+ días de atraso)."""
    base = fsvc.folder_dir(nombre_folder) / "04_cmf"
    if not base.exists():
        return None
    pdfs = sorted(base.glob("*.pdf"), key=lambda p: p.stat().st_mtime, reverse=True)
    for p in pdfs[:4]:
        try:
            texto, _m = ocr_service.extraer_texto(p.read_bytes(), p.name)
        except Exception as e:
            logging.warning(f"lector CMF {p.name}: {e}")
            continue
        t = " ".join((texto or "").split())
        filas = RX_CMF_TOTAL.findall(t)
        if not filas:
            continue
        a30 = sum(_clp_miles(f[2]) for f in filas)
        a60 = sum(_clp_miles(f[3]) for f in filas)
        a90 = sum(_clp_miles(f[4]) for f in filas)
        return {"morosidad_clp": a30 + a60 + a90, "atraso_30_59_clp": a30,
                "atraso_60_89_clp": a60, "atraso_90_mas_clp": a90,
                "deuda_total_clp": sum(_clp_miles(f[0]) for f in filas),
                "vigente_clp": sum(_clp_miles(f[1]) for f in filas), "archivo": p.name}
    return None


def _num_cl(s):
    """Número chileno: '2.000,00'→2000.0 · '5.50'→5.5."""
    s = str(s or "").strip()
    if "," in s:
        s = s.replace(".", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def leer_simulador_sync(nombre_folder):
    """📄 Lector de Simulador (ORO-71/72): extrae plazo, monto, dividendo y tasa reales
    desde el PDF Simulador más reciente de la carpeta."""
    base = fsvc.folder_dir(nombre_folder)
    if not base.exists():
        return None
    pdfs = sorted([p for p in base.rglob("*.pdf") if re.search(r"simulad", p.name, re.I)],
                  key=lambda p: p.stat().st_mtime, reverse=True)
    for p in pdfs[:4]:
        try:
            texto, _m = ocr_service.extraer_texto(p.read_bytes(), p.name)
        except Exception as e:
            logging.warning(f"lector simulador {p.name}: {e}")
            continue
        t = " ".join((texto or "").split())
        if not t:
            continue
        m = RX_SIM_PLAZO.search(t)
        plazo = int(m.group(1) or m.group(2)) if m else None
        if not plazo:
            m2 = RX_SIM_PLAZO_TASA.search(t)
            plazo = int(m2.group(1)) if m2 else None
        monto = None
        mm = RX_SIM_MONTO.search(t)
        if mm:
            monto = _num_cl(mm.group(1))
        div = None
        md = RX_SIM_DIV.search(t)
        if md:
            div = _num_cl(md.group(1))
        tasa = None
        mt = RX_SIM_TASA.search(t)
        if mt:
            tasa = _num_cl(mt.group(1))
        if plazo or monto or div:
            return {"plazo_anos": plazo, "monto_uf": monto, "dividendo_uf": div,
                    "tasa_pct": tasa, "archivo": p.name}
    return None


async def _avisar_mora_cliente(doc, cm):
    """📧 SOLICITUD DE CRÉDITO — aviso institucional automático al CLIENTE (usted) al
    detectar deuda CMF, con portal para subir el comprobante de pago a SU acreedor.
    REGLA DURA: jamás se incluyen cuentas ni instrucciones de pago hacia Central Mutuos."""
    try:
        email_cliente = (doc.get("email") or doc.get("email_cliente")
                         or (doc.get("credit_request") or {}).get("email_cliente") or "").strip()
        if not email_cliente or "@" not in email_cliente:
            return
        token = str(uuid.uuid4())
        claim = await db.folders.update_one(
            {"id": doc["id"], "cmf_morosidad.aviso_cliente_at": {"$exists": False}},
            {"$set": {"cmf_morosidad.aviso_cliente_at": _now(),
                      "cmf_morosidad.portal_token": token}})
        if not claim.modified_count:
            return
        import email_service as mail
        app_url = ""
        try:
            for line in open("/app/frontend/.env"):
                if line.startswith("REACT_APP_BACKEND_URL="):
                    app_url = line.split("=", 1)[1].strip().strip('"').rstrip("/")
        except Exception:
            pass
        link = f"{app_url}/api/mora/portal/{token}"
        nombre = (doc.get("nombre") or "").title()
        entidad = cm.get("entidad") or "la institución financiera informada en su Informe CMF"
        monto = f"${float(cm.get('morosidad_clp') or 0):,.0f}"
        cuerpo = (
            "<div style='background:#0a0a0a;padding:30px 18px;font-family:Georgia,serif'>"
            "<table role='presentation' width='100%' cellpadding='0' cellspacing='0' style='max-width:620px;margin:0 auto'>"
            "<tr><td style='text-align:center;padding-bottom:18px'>"
            "<div style='display:inline-block;border:2px solid #d4af37;width:48px;height:48px;line-height:48px;color:#d4af37;font-size:18px;font-weight:bold'>CM</div>"
            "<div style='color:#f4f2ec;letter-spacing:4px;font-size:13px;margin-top:8px;font-weight:bold'>CENTRAL MUTUOS</div>"
            "<div style='color:#d4af37;letter-spacing:5px;font-size:9px;margin-top:2px'>CON CRECES</div></td></tr>"
            "<tr><td style='border-top:2px solid #d4af37;padding-top:22px'>"
            f"<p style='color:#f4f2ec;font-size:15px;line-height:1.8'>Estimado/a {nombre}, hemos detectado una deuda "
            f"registrada en la CMF con <b style='color:#d4af37'>{entidad}</b> por <b style='color:#d4af37'>{monto}</b>. "
            "Para continuar con su solicitud de crédito, le solicitamos adjuntar el comprobante de pago de dicha deuda.</p>"
            "<p style='color:#f4f2ec;font-size:15px;line-height:1.8'>Puede subir su comprobante de forma segura en el "
            "siguiente portal; el sistema lo validará automáticamente contra el registro CMF:</p>"
            f"<p style='text-align:center;margin:22px 0'><a href='{link}' style='display:inline-block;background:#d4af37;"
            "color:#0a0a0a;text-decoration:none;font-weight:bold;font-size:14px;letter-spacing:1px;padding:13px 32px'>"
            "Adjuntar comprobante de pago</a></p>"
            f"<p style='color:#9c9a92;font-size:12px;word-break:break-all'>{link}</p>"
            "<p style='color:#9c9a92;font-size:12px;line-height:1.7'>Importante: el pago debe realizarse directamente a la "
            "entidad acreedora de su deuda. Central Mutuos <b>nunca</b> le solicitará transferencias a cuentas propias por este concepto.</p>"
            "<p style='color:#f4f2ec;font-size:15px;line-height:1.7'>Atentamente,<br><b style='color:#d4af37'>Central Mutuos</b><br>"
            "<span style='color:#9c9a92;font-size:13px'>Mutuaria regulada por la CMF</span></p></td></tr></table></div>")
        r = await asyncio.to_thread(mail.send_mail, email_cliente,
                                    "Solicitud de Crédito — Regularización de deuda CMF",
                                    cuerpo, [], "secundaria")
        await db.folders.update_one({"id": doc["id"]}, {"$push": {"historial": {"fecha": _now(), "accion": (
            f"📧 Aviso automático de deuda CMF enviado al CLIENTE ({email_cliente}) con portal de comprobante"
            f"{'' if r.get('success') else ' — ⚠ envío falló: ' + str(r.get('error'))[:80]}")}}})
    except Exception as e:
        logging.warning(f"aviso mora cliente: {e}")


async def _avisar_mora_ejecutivo(doc, cm):
    """📧 Aviso de Mora (ORO-73): un solo correo al ejecutivo apenas la auditoría
    detecta mora, con el link directo a la ficha para subir el comprobante."""
    try:
        claim = await db.folders.update_one(
            {"id": doc["id"], "cmf_morosidad.aviso_ejecutivo_at": {"$exists": False}},
            {"$set": {"cmf_morosidad.aviso_ejecutivo_at": _now()}})
        if not claim.modified_count:
            return
        import email_service as mail
        cfg = await db.config.find_one({"_key": "reglas_auto"}) or {}
        modo_prueba = bool(cfg.get("modo_prueba_clasificacion"))
        m = re.search(r"[\w.+-]+@[\w-]+\.[\w.-]+", doc.get("source_email") or "")
        dest_real = m.group(0) if m else ""
        admin = os.environ.get("MAIL_USER") or "gerardo.ext@centralmutuos.cl"
        destino = "gerardo.ext@centralmutuos.cl" if modo_prueba else (dest_real or admin)
        app_url = ""
        try:
            for line in open("/app/frontend/.env"):
                if line.startswith("REACT_APP_BACKEND_URL="):
                    app_url = line.split("=", 1)[1].strip().strip('"').rstrip("/")
        except Exception:
            pass
        link = f"{app_url}/#cliente-{doc['id']}"
        cliente = doc.get("nombre", "")
        cuerpo = (
            "<div style='font-family:Arial,Helvetica,sans-serif;background:#fff;color:#111;font-size:14px'>"
            "<p>Estimado ejecutivo:</p>"
            f"<p>La auditoría del Contralor detectó <b style='color:#b91c1c'>morosidad vigente</b> en el informe "
            f"CMF de su cliente <b>{cliente}</b>:</p>"
            f"<p style='background:#fef2f2;border-left:4px solid #b91c1c;padding:10px 14px'>"
            f"<b>Mora total: ${cm.get('morosidad_clp', 0):,.0f}</b><br>"
            f"30-59 días: ${cm.get('atraso_30_59_clp', 0):,.0f} · "
            f"60-89 días: ${cm.get('atraso_60_89_clp', 0):,.0f} · "
            f"90+ días: ${cm.get('atraso_90_mas_clp', 0):,.0f}<br>"
            f"Fuente: {cm.get('archivo', 'informe CMF')}</p>"
            "<p>La operación <b>no puede avanzar a mesa</b> mientras la mora no sea aclarada. "
            "Desde la ficha del cliente puede:</p>"
            "<ol><li>Enviar el link/instrucciones de pago al cliente</li>"
            "<li>Subir el comprobante de pago (validación automática)</li>"
            "<li>Subir el formulario de regularización</li></ol>"
            f"<p><a href='{link}' style='background:#101012;color:#d4af37;padding:10px 22px;"
            "text-decoration:none;font-weight:bold;display:inline-block'>Abrir ficha y subir comprobante</a></p>"
            f"<p style='color:#555;font-size:12px'>{link}</p>"
            "<p style='color:#555'>Saludos cordiales,<br><b>Central Mutuos</b></p></div>")
        await asyncio.to_thread(mail.send_mail, destino,
                                f"⚠️ Aviso de mora — {cliente}", cuerpo, [], "secundaria")
        await db.folders.update_one({"id": doc["id"]}, {
            "$set": {"cmf_morosidad.aviso_ejecutivo_a": destino},
            "$push": {"historial": {"fecha": _now(), "accion": (
                f"📧 Aviso de mora enviado al ejecutivo ({destino}"
                f"{' · MODO PRUEBA: interceptado al administrador' if modo_prueba else ''}) "
                "con link directo a la ficha para subir el comprobante")}}})
        logging.info(f"📧 Aviso de mora enviado — {cliente} → {destino}")
    except Exception as e:
        logging.warning(f"aviso mora ejecutivo: {e}")


async def auditar_folder(doc):
    """Auditoría completa de la carpeta contra la Bóveda de Criterios.
    Devuelve {violaciones:[{clave, regla, detalle, recomendacion, bloqueante}], edad, ...}."""
    crit = await _criterios()
    df = doc.get("datos_financieros") or {}
    cr = doc.get("credit_request") or {}
    con_sub = bool(df.get("con_subsidio"))
    btg = (crit.get("btg_pactual") or {}).get("con_subsidio" if con_sub else "sin_subsidio") or {}
    monto = _num(df.get("monto_credito"))
    valor = _num(df.get("valor_propiedad"))
    plazo = _num(df.get("plazo_anos") or df.get("plazo") or cr.get("plazo_anos"))
    antig = _num(df.get("antiguedad_laboral_meses"))

    # 📄 Lector de Simulador: plazo/monto/dividendo reales desde el PDF (una sola lectura)
    se = doc.get("simulador_extraido") or {}
    if not se and not doc.get("simulador_scan_at"):
        try:
            sim = await asyncio.to_thread(leer_simulador_sync, doc.get("nombre", ""))
            upd = {"simulador_scan_at": _now()}
            op = {"$set": upd}
            if sim:
                se = {**sim, "extraido_at": _now()}
                upd["simulador_extraido"] = se
                if sim.get("plazo_anos"):
                    upd["datos_financieros.plazo_anos"] = sim["plazo_anos"]
                op["$push"] = {"historial": {"fecha": _now(), "accion": (
                    f"📄 Lector de Simulador ({sim['archivo']}): plazo {sim.get('plazo_anos') or '—'} años · "
                    f"crédito UF {sim.get('monto_uf') or '—'} · dividendo {sim.get('dividendo_uf') or '—'} UF · "
                    f"tasa {sim.get('tasa_pct') or '—'}%")}}
            await db.folders.update_one({"id": doc["id"]}, op)
        except Exception as e:
            logging.warning(f"lector simulador #71: {e}")
    plazo = plazo or _num(se.get("plazo_anos"))
    if monto is None:
        monto = _num(se.get("monto_uf"))
    v = []

    def add(clave, regla, detalle, recomendacion, bloqueante=False, nivel=None, fuente=""):
        v.append({"clave": clave, "regla": regla, "detalle": detalle,
                  "recomendacion": recomendacion, "bloqueante": bloqueante,
                  "nivel": nivel or ("critica" if bloqueante else "media"), "fuente": fuente})

    if monto is None:
        add("sin_monto", "Monto de crédito no registrado",
            "La carpeta no tiene monto de crédito en UF — la auditoría queda incompleta",
            "Registrar el monto del crédito antes de enviar a mesa",
            fuente="ORO-65 Certeza Absoluta")
    else:
        if not con_sub and monto < MIN_CREDITO_SIN_SUBSIDIO_UF:
            add("min_sin_subsidio", "INV-3 — Crédito mínimo 2.000 UF sin subsidio (REGLA INVIOLABLE DEL DUEÑO)",
                f"Crédito de UF {monto:,.0f} SIN subsidio, bajo el tope duro de UF {MIN_CREDITO_SIN_SUBSIDIO_UF:,}. "
                "«Ninguna evaluación puede aprobarse bajo ese monto»",
                "No enviar a mesa: subir el monto, cambiar a operación con subsidio o descartar",
                bloqueante=True, fuente="dashai_eventos INV-3 · bloqueo 422 según OP-7")
        minimo = _num(btg.get("monto_credito_min_uf"))
        if minimo and monto < minimo and (con_sub or minimo > MIN_CREDITO_SIN_SUBSIDIO_UF):
            add("monto_min_boveda", "Bóveda — monto mínimo",
                f"Crédito de UF {monto:,.0f} bajo el mínimo de la Bóveda (UF {minimo:,.0f})",
                "Ajustar el monto o revisar la operación con el ejecutivo",
                fuente="Bóveda btg_pactual.monto_credito_min_uf")
        maximo = _num(btg.get("monto_credito_max_uf"))
        if maximo and monto > maximo:
            add("monto_max_boveda", "Bóveda — monto máximo",
                f"Crédito de UF {monto:,.0f} sobre el máximo de la Bóveda (UF {maximo:,.0f})",
                "Reducir el monto solicitado o aumentar pie/ahorro",
                fuente="Bóveda btg_pactual.monto_credito_max_uf")
        ltv_max = _num(btg.get("ltv_max"))
        if valor and ltv_max and monto / valor > ltv_max + 0.0001:
            add("ltv", "Bóveda — LTV máximo",
                f"LTV {monto / valor * 100:.1f}% supera el máximo permitido ({ltv_max * 100:.0f}%)",
                f"Aumentar el pie: crédito máximo para esta propiedad UF {valor * ltv_max:,.0f}",
                fuente="Bóveda btg_pactual.ltv_max · ORO-11 LTV sin redondeo hacia arriba")

    antig_min = _num(btg.get("antiguedad_laboral_min_meses"))
    if antig is not None and antig_min and antig < antig_min:
        add("antiguedad", "Bóveda — antigüedad laboral mínima",
            f"Antigüedad de {antig:.0f} meses, bajo el mínimo de {antig_min:.0f} meses",
            "Esperar a cumplir la antigüedad o incorporar un codeudor con antigüedad suficiente",
            fuente="Bóveda btg_pactual.antiguedad_laboral_min_meses · Mesa real exigió ≥6 meses plazo fijo (caso Yan Carmona)")

    # Carga financiera — fórmula de endeudamiento de la Bóveda (2% mensual, motor credit_engine)
    renta = _num(df.get("renta_liquida"))
    uf_val = None
    if renta:
        try:
            import credit_engine as ce
            uf_doc = await db.config.find_one({"_key": "uf"})
            uf_val = float(uf_doc["valor_uf"]) if uf_doc else float(DEFAULT_UF)
            fe = crit.get("formula_endeudamiento") or {}
            endeud = ce.endeudamiento_mensual(df, uf_val, _num(fe.get("cuota_pct_mensual")) or 0.02)
            cuota_clp = float(endeud.get("endeudamiento_mensual_clp") or 0)
            tope = (_num(btg.get("carga_financiera_max") or btg.get("carga_financiera_max_sin_codeudor"))
                    or _num(fe.get("tope_carga_financiera")) or 0.4)
            if cuota_clp:
                carga = cuota_clp / renta
                if carga > tope + 0.0001:
                    add("carga_financiera", "ORO-9 / Bóveda — carga financiera máxima 40% (RIESGO CRÍTICO)",
                        f"Carga financiera {carga * 100:.1f}% (cuota teórica ${cuota_clp:,.0f} / renta ${renta:,.0f}), "
                        f"sobre el tope de {tope * 100:.0f}%. «Es RIESGO CRÍTICO, mande lo que mande la MESA» (ORO-9)",
                        f"Depurar deudas CMF, acreditar más renta o incorporar codeudor: "
                        f"deuda compatible máxima ≈ ${renta * tope / (_num(fe.get('cuota_pct_mensual')) or 0.02):,.0f}",
                        nivel="critica",
                        fuente="dashai_eventos ORO-9 · Bóveda formula_endeudamiento (2% mensual CMF+PAV)")
        except Exception as e:
            logging.warning(f"carga financiera #71: {e}")

    # Algoritmo Espejo de Mesa — tope empírico por vecindad de veredictos reales
    if renta and monto:
        try:
            modelo = await db.config.find_one({"_key": "espejo_mesa_modelo"}) or {}
            casos = modelo.get("casos") or []
            if modelo.get("listo") and casos:
                cercanos = [c for c in casos if 0.75 * renta <= c["renta_liquida_clp"] <= 1.25 * renta]
                if not cercanos:
                    cercanos = sorted(casos, key=lambda c: abs(c["renta_liquida_clp"] - renta))[:3]
                tope_real = sum(c["tope_uf"] for c in cercanos) / len(cercanos)
                if monto > tope_real * 1.15:
                    n_cod = sum(1 for c in cercanos if c["con_codeudor"])
                    sug = " Los casos comparables aprobados incluían codeudor." if cercanos and n_cod / len(cercanos) >= 0.5 else ""
                    add("tope_espejo_mesa", "Algoritmo Espejo — tope empírico de la Mesa",
                        f"Crédito de UF {monto:,.0f} supera en más de 15% el tope que la Mesa aprobó a rentas "
                        f"similares: ~UF {tope_real:,.0f} ({len(cercanos)} veredicto(s) real(es), "
                        f"precisión modelo {modelo.get('precision_pct', 0)}%)",
                        f"Anticipar aprobación con tope o ajustar la operación a ~UF {tope_real:,.0f}.{sug}",
                        fuente="config espejo_mesa_modelo · limites_reales_mesa (veredictos aprobaciones@)")
        except Exception as e:
            logging.warning(f"tope espejo mesa #71: {e}")

    # Dividendo/renta — con el dividendo REAL del PDF Simulador (Lector ORO-71)
    div_uf = _num(se.get("dividendo_uf"))
    if renta and div_uf:
        try:
            if uf_val is None:
                uf_doc = await db.config.find_one({"_key": "uf"})
                uf_val = float(uf_doc["valor_uf"]) if uf_doc else float(DEFAULT_UF)
            con_cod = bool(doc.get("codeudor_nombre"))
            dr_max = _num(btg.get("div_renta_max")
                          or (btg.get("div_renta_max_con_codeudor_conjunto") if con_cod
                              else btg.get("div_renta_max_sin_codeudor"))) or 0.3
            dr = div_uf * uf_val / renta
            if dr > dr_max + 0.0001:
                add("div_renta", "Bóveda — dividendo/renta máximo",
                    f"Dividendo real del Simulador {div_uf:.2f} UF (${div_uf * uf_val:,.0f}) equivale al "
                    f"{dr * 100:.1f}% de la renta, sobre el máximo de {dr_max * 100:.0f}%",
                    f"Bajar el dividendo (más plazo o menor monto) o complementar renta: "
                    f"dividendo máximo compatible ≈ ${renta * dr_max:,.0f}",
                    fuente="Bóveda btg_pactual.div_renta_max · Lector de Simulador (PDF real)")
        except Exception as e:
            logging.warning(f"div/renta #71: {e}")

    # 🧾 Morosidad — informe CMF real (Bóveda: morosidad_permitida = No)
    cm = doc.get("cmf_morosidad") or {}
    if not cm and not doc.get("cmf_scan_at"):
        try:
            r_cmf = await asyncio.to_thread(leer_cmf_sync, doc.get("nombre", ""))
            upd = {"cmf_scan_at": _now()}
            op = {"$set": upd}
            if r_cmf:
                cm = {**r_cmf, "extraido_at": _now()}
                upd["cmf_morosidad"] = cm
                op["$push"] = {"historial": {"fecha": _now(), "accion": (
                    f"🧾 Lector CMF ({r_cmf['archivo']}): deuda total ${r_cmf['deuda_total_clp']:,.0f} · "
                    f"morosidad ${r_cmf['morosidad_clp']:,.0f} "
                    f"(30-59d ${r_cmf['atraso_30_59_clp']:,.0f} · 60-89d ${r_cmf['atraso_60_89_clp']:,.0f} · "
                    f"90+d ${r_cmf['atraso_90_mas_clp']:,.0f})")}}
            await db.folders.update_one({"id": doc["id"]}, op)
        except Exception as e:
            logging.warning(f"lector CMF #71: {e}")
    if cm.get("morosidad_clp") and not cm.get("aclarada"):
        morosidad_permitida = str(btg.get("morosidad_permitida") or "No").strip().lower()
        if morosidad_permitida in ("no", "false", "0"):
            add("morosidad", "Bóveda — morosidad NO permitida",
                (f"El informe CMF ({cm.get('archivo', '')}) registra ${cm['morosidad_clp']:,.0f} en atraso "
                 f"(30-59d ${cm.get('atraso_30_59_clp', 0):,.0f} · 60-89d ${cm.get('atraso_60_89_clp', 0):,.0f} · "
                 f"90+d ${cm.get('atraso_90_mas_clp', 0):,.0f}) y la Bóveda no permite morosidad"),
                "Regularizar la deuda morosa y adjuntar comprobante de pago o aclaración antes de enviar a mesa",
                nivel="critica",
                fuente="Bóveda btg_pactual.morosidad_permitida · Informe CMF real (04_cmf)")
            if not cm.get("aviso_ejecutivo_at"):
                asyncio.create_task(_avisar_mora_ejecutivo(doc, cm))
            if not cm.get("aviso_cliente_at"):
                asyncio.create_task(_avisar_mora_cliente(doc, cm))

    edad = None
    try:
        edad = await obtener_edad(doc)
    except Exception as e:
        logging.warning(f"edad titular: {e}")
    edad_plazo_max = _num(btg.get("edad_plazo_max") or btg.get("edad_termino_max")) or 80
    plazo_min = _num(btg.get("plazo_min_anos")) or 5
    plazo_max_permitido = None
    if edad:
        edad_min = _num(btg.get("edad_min"))
        edad_max = _num(btg.get("edad_max"))
        if edad_min and edad < edad_min:
            add("edad_min", "Bóveda — edad mínima",
                f"Titular de {edad} años, bajo la edad mínima de {edad_min:.0f}",
                "El titular no califica por edad: evaluar otro titular")
        if edad_max and edad > edad_max:
            add("edad_max", "Bóveda — edad máxima del titular",
                f"Titular de {edad} años supera la edad máxima de {edad_max:.0f}",
                "Evaluar cambio de titular o codeudor joven")
        plazo_max_permitido = max(0, int(edad_plazo_max - edad))
        if plazo and edad + plazo > edad_plazo_max:
            add("plazo_edad", "Bóveda — edad al término del crédito",
                f"Edad {edad} + plazo {plazo:.0f} años = {edad + plazo:.0f} al término, "
                f"supera el límite de {edad_plazo_max:.0f} años",
                f"Acortar el plazo a máximo {plazo_max_permitido} años (subirá el dividendo)")
        if plazo_max_permitido < plazo_min:
            add("plazo_inviable", "Bóveda — plazo inviable por edad",
                f"Por edad ({edad}) el plazo máximo sería {plazo_max_permitido} años, "
                f"bajo el mínimo de {plazo_min:.0f} años",
                "Operación inviable como titular único: requiere codeudor o descartar")
        elif not plazo and plazo_max_permitido < 25:
            add("plazo_acotado", "Aviso — plazo acotado por edad",
                f"Titular de {edad} años: plazo máximo permitido {plazo_max_permitido} años "
                f"(edad término {edad_plazo_max:.0f}) — el crédito posible baja",
                "Simular con ese plazo antes de enviar; anticipar tope de monto de la Mesa",
                bloqueante=False)

    return {"folder_id": doc.get("id"), "cliente": doc.get("nombre"),
            "con_subsidio": con_sub, "monto_credito_uf": monto, "edad": edad,
            "plazo_max_permitido": plazo_max_permitido,
            "violaciones": v, "bloqueantes": sum(1 for x in v if x["bloqueante"]),
            "auditado_at": _now()}


async def registrar_alertas(doc, violaciones):
    """Alerta inmediata al administrador — una por carpeta+regla (sin duplicar)."""
    nuevas = 0
    for vi in violaciones:
        clave = f"auditoria71:{vi['clave']}"
        if await db.alertas.find_one({"tipo": clave, "folder_id": doc.get("id"), "leida": {"$ne": True}}):
            continue
        await db.alertas.insert_one({
            "id": str(uuid.uuid4()), "tipo": clave,
            "nivel": vi.get("nivel") or ("critica" if vi["bloqueante"] else "media"),
            "cliente": doc.get("nombre", ""), "folder_id": doc.get("id"),
            "mensaje": (f"🛡️ AUDITORÍA PRE-MESA — {doc.get('nombre', '')}: {vi['regla']}. "
                        f"{vi['detalle']}. Acción recomendada: {vi['recomendacion']}"),
            "regla": vi["regla"], "recomendacion": vi["recomendacion"],
            "fuente": vi.get("fuente", ""),
            "bloqueante": vi["bloqueante"], "fecha": _now(), "leida": False})
        nuevas += 1
    return nuevas


async def auditoria_proactiva(nombre_cliente):
    """Se ejecuta al ingresar documentos: detecta ANTES de que la carpeta viaje a mesa."""
    doc = await db.folders.find_one({"nombre": nombre_cliente}, {"_id": 0})
    if not doc or doc.get("mesa_enviado_at"):
        return None
    res = await auditar_folder(doc)
    if res["violaciones"]:
        await registrar_alertas(doc, res["violaciones"])
    return res


async def auditar_carpetas_activas(dias=30, limite=100):
    """Barrido: audita todas las carpetas activas aún no enviadas a mesa."""
    from datetime import timedelta, timezone
    corte = (datetime.now(timezone.utc) - timedelta(days=dias)).isoformat()
    cur = db.folders.find({"mesa_enviado_at": {"$in": [None, ""]},
                           "created_at": {"$gte": corte}}, {"_id": 0}).limit(limite)
    n = 0
    async for doc in cur:
        try:
            res = await auditar_folder(doc)
            if res["violaciones"]:
                await registrar_alertas(doc, res["violaciones"])
                n += 1
        except Exception as e:
            logging.warning(f"auditoría {doc.get('nombre')}: {e}")
    return n


REGLAS_ORO_AUDITORIA = [
    ("ORO-71", "Auditoría Pre-Mesa del Contralor",
     "REGLA DE ORO #71 — AUDITORÍA PRE-MESA DEL CONTRALOR: antes de CUALQUIER envío a mesa el sistema "
     "verifica la carpeta ÚNICAMENTE contra lo escrito en: 1) la Bóveda de Criterios (montos mín/máx, LTV, "
     "edad, plazo por edad, antigüedad, carga financiera con fórmula de endeudamiento 2% mensual vía "
     "credit_engine, dividendo/renta con el dividendo real del PDF Simulador, y morosidad extraída del "
     "informe CMF real — columnas de atraso 30-59/60-89/90+ días), 2) las reglas de dashai_eventos, y "
     "3) el Algoritmo Espejo de Mesa (topes empíricos "
     "de veredictos reales de aprobaciones@, config espejo_mesa_modelo). JERARQUÍA ESCRITA: "
     "a) INV-3 (crédito mínimo UF 2.000 SIN subsidio — 'ninguna evaluación puede aprobarse bajo ese monto') "
     "es la ÚNICA violación que BLOQUEA el envío, con error 422 y el detalle exacto según OP-7; "
     "solo la clave maestra (MASTER_PIN) permite forzar. Para viviendas CON subsidio NO existe mínimo. "
     "b) Carga financiera >40% es RIESGO CRÍTICO según ORO-9 ('mande lo que mande la MESA') → alerta crítica. "
     "c) Todo el resto de la Bóveda y del Espejo genera alerta informativa inmediata al administrador "
     "(cliente, regla, fuente y acción recomendada) SIN bloquear, en cumplimiento de ORO-35 "
     "(el Contralor audita e informa). La auditoría corre proactivamente al ingresar documentos. "
     "PERMANENTE E INAMOVIBLE."),
    ("ORO-72", "Edad Automática desde la Cédula",
     "REGLA DE ORO #72 — EDAD AUTOMÁTICA DESDE LA CÉDULA: si la carpeta tiene cédula de identidad "
     "cargada (01_cedula), el sistema extrae por OCR la fecha de nacimiento, calcula la edad actual y "
     "la persiste en la carpeta (fecha_nacimiento, edad_titular) con evento en el historial. Esa edad "
     "se usa para verificar edad mínima/máxima y el plazo máximo permitido (edad al término ≤ 80 años "
     "según Bóveda); si el plazo solicitado lo supera, se alerta al administrador antes de mesa. "
     "PERMANENTE E INAMOVIBLE."),
    ("ORO-73", "Gestión de Pago de Mora (autovalidada)",
     "REGLA DE ORO #73 — GESTIÓN DE PAGO DE MORA: en la ficha del cliente moroso el ejecutivo dispone de "
     "tres acciones: a) enviar link/instrucciones de pago directamente al cliente (correo con el monto de la "
     "mora y la cuenta oficial MUTUARIAS Y LEASING LIMITADA), b) subir comprobante de pago, c) subir "
     "formulario manual de regularización. Al subir comprobante o formulario el sistema los VALIDA "
     "AUTOMÁTICAMENTE (legibilidad OCR + contenido + monto ≥95% de la mora en comprobantes + identidad del "
     "cliente en formularios) y CIERRA la alerta de mora SIN intervención del administrador, dejando registro "
     "en el historial y archivo en 04_cmf. Si la validación falla, el ejecutivo recibe el motivo exacto. "
     "PERMANENTE E INAMOVIBLE — modificable solo con PIN maestro."),
    ("ORO-74", "Gestor de Credenciales Crece",
     "REGLA DE ORO #74 — GESTOR DE CREDENCIALES CRECE: las credenciales de la plataforma Crece se administran "
     "en un gestor central (colección credenciales_crece). Los ejecutivos acceden EXCLUSIVAMENTE en modo "
     "lectura; crear, editar o eliminar credenciales es potestad EXCLUSIVA del Administrador (roles "
     "admin/maestro), con bloqueo 403 en el backend para cualquier otro rol. "
     "PERMANENTE E INAMOVIBLE — modificable solo con PIN maestro."),
]


async def seed_reglas_oro_auditoria():
    """Registra las reglas #71/#72 en dashai_eventos y en el módulo controlador (espejo)."""
    for clave, titulo, patron in REGLAS_ORO_AUDITORIA:
        await db.dashai_eventos.update_one({"motivo": "regla_oro", "norma_clave": clave}, {"$set": {
            "titulo": titulo, "patron": patron, "categoria": "auditoria_trazabilidad",
            "inamovible": True, "nivel_calibracion": 100, "fecha": _now(),
        }, "$setOnInsert": {"id": str(uuid.uuid4()), "motivo": "regla_oro", "norma_clave": clave}},
            upsert=True)
        clave_esp = f"regla-oro-{clave.lower()}"
        if not await db.espejo_criterios.find_one({"clave": clave_esp}):
            await db.espejo_criterios.insert_one({
                "id": str(uuid.uuid4()), "clave": clave_esp,
                "criterio": f"{clave} — {titulo}", "detalle": patron[:500],
                "origen": "manual", "estado": "activo", "fecha": _now(), "por": "DashAI"})
            await db.espejo_bitacora.insert_one({
                "id": str(uuid.uuid4()), "origen": "capa_b", "fecha": _now(),
                "clave": clave_esp, "tipo": "criterio_manual",
                "patron": f"[REGLA DE ORO {clave}] {titulo}"})
        else:
            # Paridad: mantener el texto del criterio sincronizado con el código
            await db.espejo_criterios.update_one({"clave": clave_esp}, {"$set": {
                "criterio": f"{clave} — {titulo}", "detalle": patron[:500]}})
    logging.info("🛡️ Reglas de Oro #71/#72 (Auditoría Pre-Mesa + Edad Cédula) registradas")


async def seed_paridad_produccion():
    """🔁 PARIDAD PREVIEW↔PRODUCCIÓN: aplica al arrancar (idempotente) todo lo que en
    preview se ajustó por base de datos, para que un redeploy deje ambos entornos iguales.
    1) Bóveda: mínimos de crédito según INV-3 (sin subsidio 2.000 UF, con subsidio sin mínimo).
    2) proc_rules: reglas de clasificación aprendidas (seeds/proc_rules_seed.json)."""
    import json
    from pathlib import Path
    # 1) Bóveda — corrección INV-3 si el entorno trae valores antiguos
    c = await db.config.find_one({"_key": "criterios"})
    if c:
        btg = c.get("btg_pactual") or {}
        min_sin = (btg.get("sin_subsidio") or {}).get("monto_credito_min_uf")
        min_con = (btg.get("con_subsidio") or {}).get("monto_credito_min_uf")
        if min_sin != 2000 or min_con != 0:
            version = int(float(str(c.get("version", 0)).split(".")[-1] or 0)) + 1
            await db.config.update_one({"_key": "criterios"}, {"$set": {
                "btg_pactual.sin_subsidio.monto_credito_min_uf": 2000,
                "btg_pactual.sin_subsidio.nota_minimo":
                    "Mínimo UF 2.000 aplica ÚNICAMENTE a viviendas sin subsidio (Regla de Oro #71 / INV-3)",
                "btg_pactual.con_subsidio.monto_credito_min_uf": 0,
                "btg_pactual.con_subsidio.nota_minimo":
                    "SIN mínimo de crédito para viviendas con subsidio (Regla de Oro #71 / INV-3)",
                "updated_at": _now(), "version": version}})
            await db.criterios_auditoria.insert_one({
                "id": str(uuid.uuid4()), "fecha": _now(), "version": version,
                "usuario": "DashAI — migración de paridad (INV-3)",
                "detalle": "Paridad preview↔producción: mínimo UF 2.000 SOLO sin subsidio; con subsidio sin mínimo",
                "cambios": [f"sin_subsidio.monto_credito_min_uf: {min_sin} → 2000",
                            f"con_subsidio.monto_credito_min_uf: {min_con} → 0"]})
            logging.info(f"🔁 Paridad: Bóveda corregida según INV-3 (v{version})")
    # 2) proc_rules — reglas aprendidas (upsert por nombre, sin duplicar)
    ruta = Path(__file__).parent / "seeds" / "proc_rules_seed.json"
    if ruta.exists():
        n = 0
        for r in json.loads(ruta.read_text()):
            res = await db.proc_rules.update_one({"name": r["name"]}, {"$set": r}, upsert=True)
            if res.upserted_id or res.modified_count:
                n += 1
        if n:
            logging.info(f"🔁 Paridad: {n} proc_rules sembradas/actualizadas desde seed")
    # 3) Cuenta bancaria oficial única (mandato del Administrador 2026-08-22):
    #    reemplaza cualquier cuenta anterior en config y plantillas de gastos.
    cuenta_oficial = {"nombre": "MUTUARIAS Y LEASING LIMITADA", "rut": "77.771.552-6",
                      "banco": "Mercado Pago", "tipo_cuenta": "Cuenta Vista",
                      "numero_cuenta": "1030937838", "email": "gerardo.ext@centralmutuos.cl"}
    g = await db.config.find_one({"_key": "gastos_op"}) or {}
    if (g.get("datos_pago") or {}) != cuenta_oficial:
        await db.config.update_one({"_key": "gastos_op"},
                                   {"$set": {"datos_pago": cuenta_oficial}}, upsert=True)
        logging.info("🔁 Paridad: datos bancarios oficiales aplicados en config gastos_op")
    r = await db.plantillas.update_many(
        {"tipo": "gastos", "data.datos_pago.numero_cuenta": {"$ne": cuenta_oficial["numero_cuenta"]}},
        {"$set": {"data.datos_pago": cuenta_oficial}})
    if r.modified_count:
        logging.info(f"🔁 Paridad: {r.modified_count} plantilla(s) de gastos con cuenta antigua corregida(s)")
    # 4) Perfiles de acceso del menú (mandato del Administrador 2026-08-22):
    #    ventas = Yerile/Deisy · gerencia_comercial = Daniela/Victoria/Javier
    perfiles = {
        "yerile.barrera@centralmutuos.cl": "ventas",
        "deysi.salazar@centralmutuos.cl": "ventas",
        "daniela.galindo@centralmutuos.cl": "gerencia_comercial",
        "daniela": "gerencia_comercial",
        "victoria.vilches@centralmutuos.cl": "gerencia_comercial",
        "victoria": "gerencia_comercial",
        "javier.urrutia@centralmutuos.cl": "gerencia_comercial",
        "javier": "gerencia_comercial",
        "javierurrutia@centralmutuos.cl": "gerencia_comercial",
    }
    for cod, perfil in perfiles.items():
        await db.users.update_one({"codigo": cod, "perfil": {"$ne": perfil}},
                                  {"$set": {"perfil": perfil}})
    # El usuario 'javier' antiguo era rol postventa → administracion para que las APIs del perfil funcionen
    await db.users.update_many({"codigo": {"$in": ["javier", "javierurrutia@centralmutuos.cl"]},
                                "rol": "postventa"}, {"$set": {"rol": "administracion"}})
    if not await db.users.find_one({"codigo": "javier.urrutia@centralmutuos.cl"}):
        import bcrypt as _bc
        await db.users.insert_one({
            "codigo": "javier.urrutia@centralmutuos.cl",
            "email": "javier.urrutia@centralmutuos.cl", "nombre": "Javier Urrutia",
            "rol": "administracion", "perfil": "gerencia_comercial",
            "clave_hash": _bc.hashpw("Urrutia2026!".encode(), _bc.gensalt()).decode(),
            "clave_temporal": True, "activo": True, "created": _now()})
        logging.info("🔁 Paridad: usuario Javier Urrutia creado (perfil gerencia_comercial)")
