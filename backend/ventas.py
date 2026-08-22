"""🧲 MÓDULO VENTAS — Yerile Barrera & Deisy Salazar (independiente, Regla de Oro de arquitectura)"""
from datetime import datetime, timezone
from fastapi import APIRouter, HTTPException, Request
from database import db
from victoria_independiente import (_exigir, _now, _docs_validos, _buscar_o_crear_cliente,
                                    DOCS_REQUERIDOS, ETIQUETAS, EJECUTIVOS_VENTAS,
                                    asignar_a_ventas_si_corresponde)

vtas = APIRouter(prefix="/ventas")

ESTADOS_VENTAS = {"en_gestion": "En gestión", "contactado": "Contactado",
                  "esperando_documentos": "Esperando documentos",
                  "documentacion_completa": "Documentación completa",
                  "sin_respuesta": "Sin respuesta",
                  "enviado_mesa": "Enviado a mesa",
                  "aprobado": "Aprobado", "rechazado": "Rechazado"}
CERRADOS = ("aprobado", "rechazado")

EMBUDO_ETAPAS = [("ingresado", "Ingresado"), ("documentacion_proceso", "Documentación en proceso"),
                 ("completo", "Completo"), ("enviado_mesa", "Enviado a mesa"),
                 ("aprobado", "Aprobado"), ("rechazado", "Rechazado")]


def _fecha(v):
    try:
        return datetime.fromisoformat(str(v).replace("Z", "+00:00"))
    except Exception:
        return None


async def _evento(cid, texto, por=""):
    """Línea de tiempo: cada acción queda registrada con fecha, hora y ejecutivo."""
    reg = {"fecha": _now(), "accion": texto, "por": por}
    await db.victoria_clientes.update_one({"id": cid}, {
        "$push": {"ventas.timeline": reg}, "$set": {"ventas.ultimo_evento": _now()}})
    return reg


def _semaforo(v, resumen):
    """Verde: activo · Amarillo: 3+ días sin movimiento · Rojo: 5+ días paralizado."""
    if resumen["estado"] in CERRADOS:
        return {"color": "cerrado", "dias_sin_movimiento": 0}
    fechas = [f for f in (_fecha(v.get("asignado_en")), _fecha(v.get("ultimo_contacto")),
                          _fecha(v.get("ultimo_evento"))) if f]
    if not fechas:
        return {"color": "verde", "dias_sin_movimiento": 0}
    dias = max(0, (datetime.now(timezone.utc) - max(fechas)).days)
    color = "rojo" if dias >= 5 else ("amarillo" if dias >= 3 else "verde")
    return {"color": color, "dias_sin_movimiento": dias}


def _etapa_embudo(x):
    if x["estado"] in ("aprobado", "rechazado", "enviado_mesa"):
        return x["estado"]
    if x["docs_completos"]:
        return "completo"
    if x["estado"] in ("contactado", "esperando_documentos") or x["n_contactos"] > 0:
        return "documentacion_proceso"
    return "ingresado"


async def _resumen_cliente(c):
    docs = await _docs_validos(c["id"])
    presentes = {d["tipo"] for d in docs}
    faltantes = [ETIQUETAS[t] for t in DOCS_REQUERIDOS if t not in presentes]
    v = c.get("ventas") or {}
    contactos = v.get("contactos") or []
    dias = 0
    try:
        asig = datetime.fromisoformat(str(v.get("asignado_en", "")).replace("Z", "+00:00"))
        dias = max(0, (datetime.now(timezone.utc) - asig).days)
    except Exception:
        pass
    return {"id": c["id"], "nombre": c["nombre"], "rut": c.get("rut", ""),
            "email": c.get("email", ""), "telefono": c.get("telefono", ""),
            "estado": v.get("estado", "en_gestion"),
            "estado_etiqueta": ESTADOS_VENTAS.get(v.get("estado", "en_gestion"), v.get("estado", "")),
            "faltantes": faltantes, "n_docs": len(docs), "docs_completos": not faltantes,
            "asignado_en": v.get("asignado_en", ""), "dias_gestion": dias,
            "ejecutivo": v.get("ejecutivo", ""), "ejecutivo_nombre": v.get("ejecutivo_nombre", ""),
            "ultimo_contacto": (contactos[-1] if contactos else None),
            "n_contactos": len(contactos),
            "cerrado_en": v.get("cerrado_en", ""),
            **(lambda r: {"semaforo": _semaforo(v, r), "etapa_embudo": _etapa_embudo(r)})(
                {"estado": v.get("estado", "en_gestion"), "docs_completos": not faltantes,
                 "n_contactos": len(contactos)})}


@vtas.get("/panel/{ejecutivo}")
async def panel_ventas(ejecutivo: str, request: Request):
    _exigir(request)
    if ejecutivo not in EJECUTIVOS_VENTAS:
        raise HTTPException(status_code=404, detail="Ejecutivo no registrado en el Módulo Ventas")
    clientes = await db.victoria_clientes.find({"ventas.ejecutivo": ejecutivo}, {"_id": 0}).sort("ventas.asignado_en", -1).to_list(200)
    out = [await _resumen_cliente(c) for c in clientes]
    kpis = {"asignados": len(out),
            "incompletos": sum(1 for x in out if not x["docs_completos"]),
            "completos": sum(1 for x in out if x["docs_completos"]),
            "faltantes_total": sum(len(x["faltantes"]) for x in out)}
    return {"ejecutivo": ejecutivo, "nombre": EJECUTIVOS_VENTAS[ejecutivo],
            "kpis": kpis, "clientes": out, "estados": ESTADOS_VENTAS}


@vtas.post("/solicitudes")
async def nueva_solicitud(payload: dict, request: Request):
    _exigir(request)
    nombre = (payload.get("nombre") or "").strip()
    rut = (payload.get("rut") or "").strip()
    if not nombre or not rut:
        raise HTTPException(status_code=400, detail="Indique el nombre y el RUT del solicitante")
    import re as _re
    if not _re.fullmatch(r"\d{1,2}\.?\d{3}\.?\d{3}-[0-9kK]", rut):
        raise HTTPException(status_code=400, detail="RUT inválido: use el formato 12.345.678-9")
    cliente = await _buscar_o_crear_cliente(rut, nombre, "ventas_manual")
    if not cliente:
        raise HTTPException(status_code=400, detail="RUT inválido: no se pudo crear la solicitud")
    sets = {}
    for k in ("email", "telefono"):
        val = (payload.get(k) or "").strip()
        if val:
            sets[k] = val
    if payload.get("entrega_inmediata"):
        sets["entrega_inmediata"] = True
    if sets:
        await db.victoria_clientes.update_one({"id": cliente["id"]}, {"$set": sets})
    ej = await asignar_a_ventas_si_corresponde(cliente["id"])
    if ej:
        return {"ok": True, "asignado": True, "cliente_id": cliente["id"],
                "ejecutivo": EJECUTIVOS_VENTAS[ej],
                "mensaje": f"Solicitud asignada automáticamente a {EJECUTIVOS_VENTAS[ej]} (turno alternado)"}
    c2 = await db.victoria_clientes.find_one({"id": cliente["id"]}, {"_id": 0})
    if (c2 or {}).get("ventas"):
        n = c2["ventas"].get("ejecutivo_nombre", "")
        return {"ok": True, "asignado": True, "cliente_id": cliente["id"], "ejecutivo": n,
                "mensaje": f"Esta solicitud ya estaba asignada a {n}"}
    return {"ok": True, "asignado": False, "cliente_id": cliente["id"],
            "mensaje": "NO cumple las condiciones del Módulo Ventas (requiere documentación "
                       "incompleta Y entrega inmediata): quedó como cliente normal de la bóveda"}


@vtas.post("/clientes/{cid}/contacto-registro")
async def registrar_contacto(cid: str, payload: dict, request: Request):
    u = _exigir(request)
    canal = payload.get("canal") or "llamada"
    nota = (payload.get("nota") or "").strip()
    if not nota:
        raise HTTPException(status_code=400, detail="Describa el contacto realizado con el cliente")
    c = await db.victoria_clientes.find_one({"id": cid, "ventas": {"$exists": True}})
    if not c:
        raise HTTPException(status_code=404, detail="Cliente no encontrado en el Módulo Ventas")
    reg = {"canal": canal, "nota": nota, "por": u.get("sub", ""), "fecha": _now()}
    await db.victoria_clientes.update_one({"id": cid}, {
        "$push": {"ventas.contactos": reg}, "$set": {"ventas.ultimo_contacto": _now()}})
    await _evento(cid, f"Contacto registrado vía {canal}: {nota[:140]}", u.get("sub", ""))
    return {"ok": True, "contacto": reg}


@vtas.get("/clientes/{cid}/timeline")
async def timeline_cliente(cid: str, request: Request):
    """Línea de tiempo: trazabilidad completa de cada acción con fecha, hora y ejecutivo."""
    _exigir(request)
    c = await db.victoria_clientes.find_one({"id": cid, "ventas": {"$exists": True}}, {"_id": 0})
    if not c:
        raise HTTPException(status_code=404, detail="Cliente no encontrado en el Módulo Ventas")
    v = c.get("ventas") or {}
    eventos = list(v.get("timeline") or [])
    if not any("asignaci" in str(e.get("accion", "")).lower() for e in eventos):
        eventos.insert(0, {"fecha": v.get("asignado_en", ""), "por": "sistema",
                           "accion": f"Cliente asignado a {v.get('ejecutivo_nombre', '')} en el Módulo Ventas."})
    eventos.sort(key=lambda e: str(e.get("fecha", "")))
    return {"cliente": c["nombre"], "ejecutivo": v.get("ejecutivo_nombre", ""), "eventos": eventos}


@vtas.put("/clientes/{cid}/estado")
async def estado_ventas(cid: str, payload: dict, request: Request):
    u = _exigir(request)
    estado = payload.get("estado")
    if estado not in ESTADOS_VENTAS:
        raise HTTPException(status_code=400, detail="Estado inválido")
    sets = {"ventas.estado": estado}
    if estado in CERRADOS:
        sets["ventas.cerrado_en"] = _now()
    r = await db.victoria_clientes.update_one({"id": cid, "ventas": {"$exists": True}}, {"$set": sets})
    if not r.matched_count:
        raise HTTPException(status_code=404, detail="Cliente no encontrado en el Módulo Ventas")
    await _evento(cid, f"Estado actualizado a «{ESTADOS_VENTAS[estado]}»"
                       + (" — gestión cerrada." if estado in CERRADOS else "."), u.get("sub", ""))
    return {"ok": True, "estado": estado, "etiqueta": ESTADOS_VENTAS[estado]}


@vtas.get("/rendimiento")
async def rendimiento_ventas(request: Request):
    """Panel de rendimiento lado a lado: activos, cerrados del mes, conversión y tiempo a cierre."""
    _exigir(request)
    ahora = datetime.now(timezone.utc)
    out = {}
    for ej, nombre in EJECUTIVOS_VENTAS.items():
        clientes = await db.victoria_clientes.find({"ventas.ejecutivo": ej}, {"_id": 0}).to_list(300)
        res = [await _resumen_cliente(c) for c in clientes]
        activos = [x for x in res if x["estado"] not in CERRADOS]
        cerrados = [x for x in res if x["estado"] in CERRADOS]
        cerrados_mes = [x for x in cerrados if (_fecha(x["cerrado_en"]) or ahora.replace(year=2000)).month == ahora.month
                        and (_fecha(x["cerrado_en"]) or ahora.replace(year=2000)).year == ahora.year]
        aprobados = [x for x in cerrados if x["estado"] == "aprobado"]
        tiempos = []
        for x in cerrados:
            fa, fc = _fecha(x["asignado_en"]), _fecha(x["cerrado_en"])
            if fa and fc:
                tiempos.append(max(0, (fc - fa).days))
        out[ej] = {"nombre": nombre, "activos": len(activos), "cerrados_mes": len(cerrados_mes),
                   "cerrados_total": len(cerrados), "aprobados": len(aprobados),
                   "tasa_conversion": round(100 * len(aprobados) / len(res), 1) if res else 0,
                   "dias_promedio_cierre": round(sum(tiempos) / len(tiempos), 1) if tiempos else 0,
                   "total_asignados": len(res),
                   "semaforos": {"verde": sum(1 for x in activos if x["semaforo"]["color"] == "verde"),
                                 "amarillo": sum(1 for x in activos if x["semaforo"]["color"] == "amarillo"),
                                 "rojo": sum(1 for x in activos if x["semaforo"]["color"] == "rojo")}}
    return {"ejecutivos": out, "generado": _now()}


@vtas.get("/embudo")
async def embudo_ventas(request: Request):
    """Embudo de ventas por etapas: ingresado → documentación → completo → mesa → aprobado/rechazado."""
    _exigir(request)
    clientes = await db.victoria_clientes.find({"ventas": {"$exists": True}}, {"_id": 0}).to_list(500)
    res = [await _resumen_cliente(c) for c in clientes]
    embudo = []
    for clave, etiqueta in EMBUDO_ETAPAS:
        grupo = [x for x in res if x["etapa_embudo"] == clave]
        embudo.append({"etapa": clave, "etiqueta": etiqueta, "total": len(grupo),
                       "por_ejecutivo": {ej: sum(1 for x in grupo if x["ejecutivo"] == ej)
                                         for ej in EJECUTIVOS_VENTAS}})
    return {"embudo": embudo, "total": len(res), "ejecutivos": EJECUTIVOS_VENTAS, "generado": _now()}


@vtas.get("/export")
async def export_ventas(request: Request, ejecutivo: str = "", desde: str = "", hasta: str = "",
                        estado: str = "", resultado: str = ""):
    """Exportación a Excel con filtros por ejecutivo, fecha, estado y resultado."""
    _exigir(request)
    q = {"ventas": {"$exists": True}}
    if ejecutivo in EJECUTIVOS_VENTAS:
        q["ventas.ejecutivo"] = ejecutivo
    clientes = await db.victoria_clientes.find(q, {"_id": 0}).sort("ventas.asignado_en", -1).to_list(500)
    res = [await _resumen_cliente(c) for c in clientes]
    if estado and estado in ESTADOS_VENTAS:
        res = [x for x in res if x["estado"] == estado]
    if resultado == "aprobado":
        res = [x for x in res if x["estado"] == "aprobado"]
    elif resultado == "rechazado":
        res = [x for x in res if x["estado"] == "rechazado"]
    elif resultado == "abierto":
        res = [x for x in res if x["estado"] not in CERRADOS]
    if desde:
        res = [x for x in res if str(x["asignado_en"])[:10] >= desde]
    if hasta:
        res = [x for x in res if str(x["asignado_en"])[:10] <= hasta]
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    wb = Workbook()
    ws = wb.active
    ws.title = "Módulo Ventas"
    cab = ["Cliente", "RUT", "Ejecutiva", "Estado", "Etapa embudo", "Semáforo", "Docs completos",
           "Docs faltantes", "Contactos", "Asignado", "Cerrado", "Días en gestión"]
    ws.append(cab)
    dorado = PatternFill(start_color="1A1A1A", end_color="1A1A1A", fill_type="solid")
    for celda in ws[1]:
        celda.font = Font(bold=True, color="C9A227")
        celda.fill = dorado
    etq_embudo = dict(EMBUDO_ETAPAS)
    for x in res:
        ws.append([x["nombre"], x["rut"], x["ejecutivo_nombre"], x["estado_etiqueta"],
                   etq_embudo.get(x["etapa_embudo"], x["etapa_embudo"]),
                   x["semaforo"]["color"].upper(), "Sí" if x["docs_completos"] else "No",
                   ", ".join(x["faltantes"]) or "—", x["n_contactos"],
                   str(x["asignado_en"])[:16].replace("T", " "),
                   str(x["cerrado_en"])[:16].replace("T", " ") or "—", x["dias_gestion"]])
    for col in "ABCDEFGHIJKL":
        ws.column_dimensions[col].width = 22
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    from fastapi.responses import Response as _Resp
    nombre_arch = f"Ventas_CentralMutuos_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M')}.xlsx"
    return _Resp(content=buf.read(),
                 media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                 headers={"Content-Disposition": f"attachment; filename={nombre_arch}"})


@vtas.get("/reporte")
async def reporte_ventas(request: Request):
    """Reporte transversal en tiempo real para el Administrador."""
    _exigir(request)
    ejecutivos = {}
    for ej, nombre in EJECUTIVOS_VENTAS.items():
        clientes = await db.victoria_clientes.find({"ventas.ejecutivo": ej}, {"_id": 0}).sort("ventas.asignado_en", -1).to_list(200)
        res = [await _resumen_cliente(c) for c in clientes]
        inc = [x for x in res if not x["docs_completos"]]
        ejecutivos[ej] = {"nombre": nombre, "clientes": res, "total": len(res),
                          "incompletos": len(inc),
                          "faltantes_total": sum(len(x["faltantes"]) for x in res),
                          "dias_promedio_sin_completar":
                              round(sum(x["dias_gestion"] for x in inc) / len(inc), 1) if inc else 0}
    return {"ejecutivos": ejecutivos, "estados": ESTADOS_VENTAS, "generado": _now()}


@vtas.put("/ejecutivos/{ej}/avisos-email")
async def set_avisos_email(ej: str, payload: dict, request: Request):
    """Espacios de correo por ejecutiva para recibir avisos del sistema de gestión."""
    _exigir(request)
    if ej not in EJECUTIVOS_VENTAS:
        raise HTTPException(status_code=404, detail="Ejecutivo no registrado")
    emails = [e.strip().lower() for e in (payload.get("emails") or []) if "@" in e][:5]
    await db.config.update_one({"_key": "ventas_emails"}, {"$set": {ej: emails}}, upsert=True)
    return {"ok": True, "ejecutivo": ej, "emails": emails}


@vtas.get("/avisos-email")
async def get_avisos_email(request: Request):
    _exigir(request)
    cfg = await db.config.find_one({"_key": "ventas_emails"}, {"_id": 0}) or {}
    return {ej: cfg.get(ej, []) for ej in EJECUTIVOS_VENTAS}
