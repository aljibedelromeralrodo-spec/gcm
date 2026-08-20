"""📋 AUDITORÍA DE CRÉDITOS → MESA
Solicitudes recibidas en los últimos N días, cruzadas contra:
  1) envíos por el sistema (folders.mesa_enviado_at)
  2) envíos DIRECTOS por correo detectados en el espejo de la casilla de mesa
     (colección mesa_enviados — match por RUT o nombre normalizado).
"""
import io
import re
import unicodedata
from datetime import datetime, timezone, timedelta

from fastapi import APIRouter, HTTPException, Request
from fastapi.responses import Response
from database import db
import folders_service as fsvc

audimesa = APIRouter(prefix="/autocorreo")

ROLES = ("admin", "maestro", "administracion", "gerencia", "contralor")


def _exigir(request):
    c = getattr(request.state, "user", {}) or {}
    if c.get("rol") not in ROLES:
        raise HTTPException(status_code=403, detail="Su rol no tiene acceso a la Auditoría de Créditos")
    return c


def _norm_tokens(s):
    s = unicodedata.normalize("NFD", (s or "").lower())
    s = "".join(ch for ch in s if unicodedata.category(ch) != "Mn")
    return [t for t in re.findall(r"[a-z0-9]+", s) if len(t) > 1]


def _match_nombre(ftoks, mtoks):
    if not ftoks or not mtoks:
        return False
    fs, ms = set(ftoks), set(mtoks)
    inter = fs & ms
    if len(inter) >= 2 and (inter == ms or inter == fs):
        return True
    return len(ms) == 1 and ms == fs


def _rut_limpio(r):
    return re.sub(r"[^0-9kK]", "", r or "").lower()


def _monto_uf(f):
    df = f.get("datos_financieros") or {}
    try:
        v = float(df.get("monto_credito") or 0)
        if v:
            return v
    except (TypeError, ValueError):
        pass
    cr = f.get("credit_request") or {}
    for k in ("monto_credito_uf", "monto_uf", "monto_credito", "monto"):
        try:
            x = float(str(cr.get(k) or 0).replace(",", "."))
            if x:
                return x
        except (TypeError, ValueError):
            continue
    return None


def _motivos_retencion(f, contraste_map, ahora):
    motivos = []
    cr = f.get("credit_request") or {}
    df = f.get("datos_financieros") or {}
    ct = cr.get("client_type") or "dependiente"
    try:
        cats = {fsvc.cat_de_archivo(a["nombre"], a["subfolder"]) for a in fsvc.scan_archivos(f.get("nombre", ""))}
        faltan = [fsvc.MISSING_LABELS.get(c, c) for c in fsvc.required_cats(ct) if c not in cats]
    except Exception:
        faltan = []
    if faltan:
        motivos.append("📄 Documentación incompleta — faltan: " + ", ".join(faltan))
    if not (df.get("fecha_entrega") or "").strip():
        motivos.append("📅 Falta fecha de entrega (inmediata/futura)")
    if not (f.get("ejecutivo_interno") or "").strip():
        motivos.append("👤 Sin ejecutivo interno asignado")
    if _monto_uf(f) is None:
        motivos.append("💰 Sin monto de crédito registrado")
    reg = contraste_map.get(f.get("id")) or {}
    if (reg.get("estado") or "pendiente") == "pendiente":
        motivos.append("🔍 Contraste Bodega/OCR pendiente (Regla #24)")
    ts = f.get("updated_at") or f.get("created_at") or ""
    try:
        dt = datetime.fromisoformat(str(ts).replace("Z", "+00:00"))
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        if (ahora - dt).total_seconds() >= 48 * 3600:
            motivos.append("⏸ Sin actividad reciente (48h+)")
    except (TypeError, ValueError):
        pass
    if not motivos:
        motivos.append("⚠️ Sin motivo registrado — revisar")
    return motivos


async def _auditoria(dias):
    ahora = datetime.now(timezone.utc)
    dias = max(1, min(int(dias or 3), 30))
    corte = (ahora - timedelta(days=dias)).isoformat()
    folders = await db.folders.find({"created_at": {"$gte": corte}}, {
        "_id": 0, "id": 1, "nombre": 1, "rut": 1, "created_at": 1, "updated_at": 1,
        "datos_financieros": 1, "credit_request": 1, "ejecutivo_interno": 1,
        "broker_nombre": 1, "mesa_enviado_at": 1, "mesa_message_id": 1}).sort("created_at", -1).to_list(1000)
    mesa_docs = await db.mesa_enviados.find({}, {"_id": 0, "cliente": 1, "nombre": 1,
                                                 "rut": 1, "enviado_at": 1}).to_list(5000)
    mesa_idx = [(_norm_tokens(m.get("cliente") or m.get("nombre")), _rut_limpio(m.get("rut")), m)
                for m in mesa_docs]
    contraste_map = {d.get("folder_id"): d async for d in
                     db.bodega_contraste.find({}, {"_id": 0, "folder_id": 1, "estado": 1})}
    enviados, pendientes = [], []
    for f in folders:
        base = {
            "folder_id": f.get("id"), "cliente": f.get("nombre") or "—",
            "rut": f.get("rut") or "", "monto_uf": _monto_uf(f),
            "ejecutivo": (f.get("ejecutivo_interno") or "").strip() or (f.get("broker_nombre") or "").strip() or "—",
            "fecha_recepcion": str(f.get("created_at") or "")[:16].replace("T", " ")}
        via, fecha_envio = None, ""
        if f.get("mesa_enviado_at"):
            via = "Sistema"
            fecha_envio = str(f["mesa_enviado_at"])[:16].replace("T", " ")
        else:
            ftoks = _norm_tokens(f.get("nombre"))
            frut = _rut_limpio(f.get("rut"))
            for mtoks, mrut, m in mesa_idx:
                if (frut and mrut and frut == mrut) or _match_nombre(ftoks, mtoks):
                    via = "Correo directo"
                    fecha_envio = str(m.get("enviado_at") or "")[:16].replace("T", " ")
                    break
        if via:
            enviados.append({**base, "via": via, "fecha_envio_mesa": fecha_envio})
        else:
            base["motivos_retencion"] = _motivos_retencion(f, contraste_map, ahora)
            pendientes.append(base)
    return {"dias": dias, "generado": ahora.isoformat(),
            "resumen": {"recibidas": len(folders), "enviadas_mesa": len(enviados),
                        "pendientes": len(pendientes),
                        "enviadas_sistema": sum(1 for e in enviados if e["via"] == "Sistema"),
                        "enviadas_correo_directo": sum(1 for e in enviados if e["via"] == "Correo directo")},
            "enviados": enviados, "pendientes": pendientes}


@audimesa.get("/auditoria-mesa")
async def auditoria_mesa(request: Request, dias: int = 3):
    _exigir(request)
    return await _auditoria(dias)


@audimesa.get("/auditoria-mesa/export-xlsx")
async def auditoria_mesa_xlsx(request: Request, dias: int = 3):
    _exigir(request)
    d = await _auditoria(dias)
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = Workbook()
    negro = PatternFill("solid", fgColor="0A0A0A")
    oro = Font(color="C9A227", bold=True, size=11)

    def _hoja(ws, titulo, cols, filas):
        ws.title = titulo
        ws.append([f"CENTRAL MUTUOS — Auditoría de Créditos ({d['dias']} días) — {titulo}"])
        ws["A1"].font = Font(color="C9A227", bold=True, size=13)
        ws["A1"].fill = negro
        ws.append([c[0] for c in cols])
        for cell in ws[2]:
            cell.font = oro
            cell.fill = negro
            cell.alignment = Alignment(horizontal="left")
        for fila in filas:
            ws.append([fila.get(c[1], "") for c in cols])
        for i, c in enumerate(cols, 1):
            ws.column_dimensions[ws.cell(row=2, column=i).column_letter].width = c[2]

    _hoja(wb.active, "Enviados a Mesa",
          [("Cliente", "cliente", 30), ("RUT", "rut", 15), ("Monto UF", "monto_uf", 12),
           ("Ejecutivo responsable", "ejecutivo", 22), ("Fecha recepción", "fecha_recepcion", 18),
           ("Fecha envío a mesa", "fecha_envio_mesa", 18), ("Vía", "via", 14)],
          d["enviados"])
    ws2 = wb.create_sheet()
    _hoja(ws2, "No Enviados",
          [("Cliente", "cliente", 30), ("RUT", "rut", 15), ("Monto UF", "monto_uf", 12),
           ("Ejecutivo responsable", "ejecutivo", 22), ("Fecha recepción", "fecha_recepcion", 18),
           ("Motivo de retención", "motivos_txt", 80)],
          [{**p, "motivos_txt": " | ".join(p["motivos_retencion"])} for p in d["pendientes"]])
    buf = io.BytesIO()
    wb.save(buf)
    hoy = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    return Response(content=buf.getvalue(),
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": f'attachment; filename="Auditoria_Creditos_Mesa_{hoy}.xlsx"'})
