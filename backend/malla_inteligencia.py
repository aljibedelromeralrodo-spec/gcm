"""MALLA DE INTELIGENCIA (Regla de Oro #34) + MÓDULO BROKERS (Usuarios D).

- Brokers (perfil D): cargan Set de Crédito, Proyección Mensual y ven SOLO sus carpetas
  con subcarpetas automáticas: Solicitud, Set Crédito y Estudio Título.
- Fuentes IMAP operativas por panel (victoria, daniela, postventa): correo principal
  + hasta 3 correos de aliados externos (Value Property, Guillermo Mardones, Notarías).
- Motor de seguimiento en tiempo real: DashAI rastrea envíos/recepciones con aliados y
  marca hitos externos. REGLA DE HIERRO: sin RUT en asunto/cuerpo, el hito NO se marca.
"""
import re
import os
import uuid
import base64
import bcrypt
import asyncio
import hashlib
import logging
from datetime import datetime, timezone
from fastapi import APIRouter, HTTPException, Request, UploadFile, File, Form
from fastapi.responses import Response
from database import db
import email_service as mail
import folders_service as fsvc

broker = APIRouter(prefix="/broker")
fuentes = APIRouter(prefix="/fuentes")
hitos = APIRouter(prefix="/hitos")
flujos = APIRouter(prefix="/flujos")
micorreo = APIRouter(prefix="/mi-correo")
buzon = APIRouter(prefix="/buzon-aprendizaje")
supercarpeta = APIRouter(prefix="/supercarpeta")

# MOTOR DE REPAROS: remitentes de los abogados de estudio de título / escrituración
REPARO_REMITENTES = ("mardluf", "majluf", "gmardones", "olave", "ibarra",
                     "victoriavilches", "ecerda", "amvabogados")

# RADAR DE ESCRITURACIÓN — Documentación 2.0 y Log de Firmas
DOC20_REQ = {"03_afp": "AFP", "02_liquidaciones": "Liquidación actual", "04_cmf": "CMF actualizado"}
_MESES = {"enero": 1, "febrero": 2, "marzo": 3, "abril": 4, "mayo": 5, "junio": 6, "julio": 7,
          "agosto": 8, "septiembre": 9, "octubre": 10, "noviembre": 11, "diciembre": 12}
FECHA_RE = re.compile(r"\b(\d{1,2})[-/](\d{1,2})[-/](\d{2,4})\b")
FECHA_TXT_RE = re.compile(r"\b(\d{1,2})\s+de\s+(enero|febrero|marzo|abril|mayo|junio|julio|agosto|septiembre|octubre|noviembre|diciembre)(?:\s+de\s+(\d{4}))?", re.I)


def doc20_folder(fd):
    """Hito 'Documentación 2.0': AFP + Liquidación actual + CMF actualizado.
    Si falta uno → 'Pendiente de Información' (jamás se inventa)."""
    base = fsvc.folder_dir(fd.get("nombre") or "")
    faltantes = []
    for sub, label in DOC20_REQ.items():
        d = base / sub
        if not (d.exists() and any(p.is_file() for p in d.iterdir())):
            faltantes.append(label)
    return {"estado": "ok" if not faltantes else "pendiente_informacion", "faltantes": faltantes}


def firmas_folder(fd):
    """LOG DE FIRMAS según la simulación. REGLA DE HIERRO: con codeudor, el hito solo
    es VERDE cuando AMBOS firmaron — no se aceptan firmas parciales."""
    df = fd.get("datos_financieros") or {}
    aplica_codeudor = bool(fd.get("codeudor_rut") or fd.get("codeudor_nombre") or df.get("renta_codeudor"))
    log = fd.get("firmas_log") or {}
    roles = [("titular", "Titular", True), ("codeudor", "Codeudor", aplica_codeudor),
             ("mandatario", "Mandatario Judicial", True), ("anexos", "Anexos Notaría", True)]
    firmas = [{"rol": k, "label": lb, "estado": log.get(k, "pendiente")}
              for k, lb, aplica in roles if aplica]
    if firmas and all(f["estado"] == "firmado" for f in firmas):
        hito = "ok"
    elif any(f["estado"] == "firmado" for f in firmas):
        hito = "proceso"  # parcial = ⏳ amarillo, JAMÁS verde
    else:
        hito = "pendiente"
    return firmas, hito


def _capturar_fecha(texto):
    m = FECHA_TXT_RE.search(texto or "")
    if m:
        anio = int(m.group(3) or datetime.now(timezone.utc).year)
        try:
            return f"{anio:04d}-{_MESES[m.group(2).lower()]:02d}-{int(m.group(1)):02d}"
        except Exception:
            pass
    m = FECHA_RE.search(texto or "")
    if m:
        d, mth, a = int(m.group(1)), int(m.group(2)), int(m.group(3))
        a = a + 2000 if a < 100 else a
        if 1 <= mth <= 12 and 1 <= d <= 31:
            return f"{a:04d}-{mth:02d}-{d:02d}"
    return ""

_now = lambda: datetime.now(timezone.utc).isoformat()
_rut_limpio = lambda r: re.sub(r"[^0-9kK]", "", str(r or "")).lower()
RUT_RE = re.compile(r"\b(\d{1,2}\.?\d{3}\.?\d{3}\s*-\s*[0-9kK])\b")
# REGLA DE HIERRO: la firma SOLO se confirma con evidencia REAL en pasado (jamás agendamientos)
FIRMA_REAL_RE = re.compile(
    r"(escritura\s+(?:fue\s+)?firmad[ao]|se\s+firm[óo]\b|firmad[ao]\s+(?:la\s+)?escritura"
    r"|firma\s+(?:realizada|efectuada|concretada|exitosa)|firmaron\s+la\s+escritura"
    r"|cesi[óo]n\s+firmad[ao]|firmad[ao]\s+(?:la\s+)?cesi[óo]n)", re.I)
EMAIL_RE = re.compile(r"^[\w.+-]+@[\w-]+\.[\w.-]+$")

PANELES = ("victoria", "daniela", "postventa", "brokers")
SUBCARPETAS_BROKER = [("06_solicitud", "Solicitud"),
                      ("08_set_credito", "Set Crédito"),
                      ("07_estudio_titulo", "Estudio Título"),
                      ("09_cartas_enmienda", "Cartas Enmienda"),
                      ("10_actualizaciones", "Doc. Actualización"),
                      ("99_otros", "Otros Documentos")]
# CENTRO DE CARGA MASIVA: categorías → subcarpeta física
CATEGORIAS_CARGA = {
    "set_credito": "08_set_credito", "carta_enmienda": "09_cartas_enmienda",
    "actualizacion": "10_actualizaciones", "otros": "99_otros",
}


def _panel_valido(panel):
    """Paneles fijos + fuentes personales por broker (broker_<codigo>)."""
    return panel in PANELES or panel.startswith("broker_")


async def _log_broker(claims, accion, detalle):
    """HUELLA DIGITAL DASHAI: cada interacción del broker queda registrada."""
    await db.broker_activity_log.insert_one({
        "id": str(uuid.uuid4()), "broker_codigo": claims.get("sub") or "",
        "broker_nombre": claims.get("nombre") or "", "accion": accion,
        "detalle": detalle, "fecha": _now()})

# Hitos conocidos por dominio aliado según dirección del correo
HITOS_BASE = {
    "valueproperty": {"enviado": "Tasación Solicitada", "recibido": "Respuesta Tasador"},
    "gmardones": {"enviado": "Estudio Solicitado", "recibido": "Estudio Recibido"},
}


def _claims(request):
    return getattr(request.state, "user", None) or {}


# ───────────────────────── MÓDULO BROKERS (Usuarios D) ─────────────────────
def _archivos_broker(nombre):
    base = fsvc.folder_dir(nombre)
    out = {}
    for sub, label in SUBCARPETAS_BROKER:
        d = base / sub
        out[sub] = sorted(p.name for p in d.iterdir() if p.is_file()) if d.exists() else []
    return out


@broker.get("/carpetas")
async def broker_carpetas(request: Request):
    c = _claims(request)
    codigo = c.get("sub") or ""
    q = {"broker_codigo": {"$exists": True}} if c.get("rol") in ("admin", "maestro") else {"broker_codigo": codigo}
    docs = await db.folders.find(q).sort("created_at", -1).to_list(300)
    carpetas = []
    for d in docs:
        carpetas.append({"id": d["id"], "nombre": d.get("nombre"), "rut": d.get("rut") or "",
                         "created_at": d.get("created_at"), "broker_codigo": d.get("broker_codigo"),
                         "subcarpetas": _archivos_broker(d.get("nombre") or "")})
    return {"carpetas": carpetas, "total": len(carpetas),
            "subcarpetas": [{"key": k, "label": l} for k, l in SUBCARPETAS_BROKER]}


@broker.post("/carpetas")
async def broker_crear_carpeta(payload: dict, request: Request):
    c = _claims(request)
    nombre = (payload.get("nombre") or "").strip().upper()
    rut = (payload.get("rut") or "").strip()
    if not nombre or not rut:
        raise HTTPException(status_code=400, detail="Nombre y RUT del cliente son obligatorios")
    if not _rut_limpio(rut) or len(_rut_limpio(rut)) < 8:
        raise HTTPException(status_code=400, detail="RUT inválido — el RUT es el pegamento del sistema (Regla #34)")
    # REGLA RUT ÚNICO: el primer broker que registró el RUT retiene al cliente permanentemente
    rutn = _rut_limpio(rut)
    async for fd0 in db.folders.find({"rut": {"$exists": True, "$ne": ""}},
                                     {"rut": 1, "broker_codigo": 1}):
        if _rut_limpio(fd0.get("rut") or "") == rutn:
            if (fd0.get("broker_codigo") or "") and fd0.get("broker_codigo") != (c.get("sub") or ""):
                raise HTTPException(status_code=409,
                                    detail="Este RUT ya está registrado en el sistema por otro ejecutivo.")
            raise HTTPException(status_code=409, detail="Este RUT ya está registrado en el sistema.")
    doc = {"id": str(uuid.uuid4()), "nombre": nombre, "rut": rut, "archivos": [],
           "broker_codigo": c.get("sub") or "", "broker_nombre": c.get("nombre") or "",
           "broker_origen": c.get("nombre") or c.get("sub") or "",  # SELLO DE ORIGEN (Regla #38)
           "origen": "broker", "created_at": _now()}
    await db.folders.insert_one(dict(doc))
    base = fsvc.folder_dir(nombre)
    for sub, _ in SUBCARPETAS_BROKER:
        (base / sub).mkdir(parents=True, exist_ok=True)
    await _log_broker(c, "carpeta_creada", {"cliente": nombre, "rut": rut})
    return {"ok": True, "id": doc["id"], "nombre": nombre,
            "subcarpetas": [l for _, l in SUBCARPETAS_BROKER]}


@broker.post("/carpetas/{fid}/upload")
async def broker_upload(fid: str, request: Request, subcarpeta: str = Form(""),
                        categoria: str = Form(""), descripcion: str = Form(""),
                        archivo: UploadFile = File(...)):
    c = _claims(request)
    fd = await db.folders.find_one({"id": fid})
    if not fd:
        raise HTTPException(status_code=404, detail="Carpeta no existe")
    if c.get("rol") not in ("admin", "maestro") and fd.get("broker_codigo") != (c.get("sub") or ""):
        raise HTTPException(status_code=403, detail="Solo puede subir archivos a sus propias carpetas")
    if categoria:
        subcarpeta = CATEGORIAS_CARGA.get(categoria) or ""
    if subcarpeta not in {k for k, _ in SUBCARPETAS_BROKER}:
        raise HTTPException(status_code=400, detail="Subcarpeta o categoría inválida")
    nombre_arch = re.sub(r"[^\w.\- ]", "_", archivo.filename or "documento.pdf")
    contenido = await archivo.read()
    # REGLA DE HIERRO: DashAI audita el RUT del documento antes de aceptarlo
    if nombre_arch.lower().endswith(".pdf"):
        try:
            import ocr_service
            res_ocr = await asyncio.to_thread(ocr_service.extraer_texto, contenido, nombre_arch)
            texto = res_ocr[0] if isinstance(res_ocr, tuple) else (res_ocr or "")
            ruts_doc = {_rut_limpio(m) for m in RUT_RE.findall(texto or "")}
            ruts_ok = {_rut_limpio(fd.get("rut")), _rut_limpio(fd.get("codeudor_rut"))} - {""}
            if ruts_doc and ruts_ok and not (ruts_doc & ruts_ok):
                await _log_broker(c, "archivo_rechazado", {
                    "archivo": nombre_arch, "carpeta": fd.get("nombre"),
                    "motivo": "RUT del documento no coincide con el cliente"})
                raise HTTPException(status_code=422,
                    detail=f"⛔ DashAI rechazó el archivo: el RUT del documento no coincide con {fd.get('nombre')} ({fd.get('rut')})")
        except HTTPException:
            raise
        except Exception as e:
            logging.warning(f"broker auditoria rut {nombre_arch}: {e}")
    destino = fsvc.folder_dir(fd["nombre"]) / subcarpeta
    destino.mkdir(parents=True, exist_ok=True)
    (destino / nombre_arch).write_bytes(contenido)
    # ☁️ DUAL WRITE: copia persistente en el storage integrado (operación/RUT)
    try:
        import media_storage as _ms
        asyncio.create_task(_ms.registrar_documento(
            contenido, nombre_arch, fd, origen="broker", subido_por=c.get("sub") or "",
            rol=c.get("rol") or "", rel=f"{subcarpeta}/{nombre_arch}"))
    except Exception as _e:
        logging.warning(f"storage dual broker: {_e}")
    await db.folders.update_one({"id": fid}, {"$addToSet": {"archivos": f"{subcarpeta}/{nombre_arch}"},
                                              "$set": {"updated_at": _now()}})
    await _log_broker(c, "archivo_subido", {"archivo": nombre_arch, "subcarpeta": subcarpeta,
                                            "categoria": categoria or subcarpeta,
                                            "descripcion": descripcion[:200],
                                            "carpeta": fd.get("nombre"), "rut": fd.get("rut")})
    return {"ok": True, "archivo": nombre_arch, "subcarpeta": subcarpeta, "auditado_dashai": True}


INMO_KW = {"boetsch": "Boetsch", "boetch": "Boetsch", "word": "Word", "urbanizate": "Urbanizate",
           "maestra": "Maestra", "ecomac": "Ecomac"}


def _parsear_proyeccion(ruta):
    """P11 — extrae filas (nombre, inmobiliaria, tipo, subsidio, monto UF) desde Excel/PDF/CSV."""
    filas, lineas = [], []
    suf = str(ruta).lower()
    if suf.endswith((".xlsx", ".xls")):
        import openpyxl
        wb = openpyxl.load_workbook(ruta, data_only=True)
        for ws in wb.worksheets:
            for row in ws.iter_rows(values_only=True):
                lineas.append(" | ".join(str(c) for c in row if c is not None))
    elif suf.endswith(".pdf"):
        from pypdf import PdfReader
        r = PdfReader(str(ruta))
        for p in r.pages:
            lineas += (p.extract_text() or "").splitlines()
    else:
        lineas = ruta.read_text(errors="ignore").splitlines()
    nombre_re = re.compile(r"([A-ZÁÉÍÓÚÑ][a-zA-ZÁÉÍÓÚÑáéíóúñ]+(?:\s+[A-ZÁÉÍÓÚÑa-zA-Záéíóúñ]+){1,3})")
    monto_re = re.compile(r"\$?\s*(\d{1,2}[.,]\d{3}(?:[.,]\d+)?|\d{3,5})")
    for ln in lineas:
        t = ln.strip()
        if len(t) < 8:
            continue
        tl = t.lower()
        inmo = next((v for k, v in INMO_KW.items() if k in tl), "")
        usada = "usad" in tl
        monto = None
        for m in monto_re.finditer(t):
            try:
                v = float(m.group(1).replace(".", "").replace(",", "."))
                if 100 <= v <= 99999:
                    monto = v
                    break
            except Exception:
                continue
        m_n = nombre_re.search(t)
        nombre = (m_n.group(1).strip().upper() if m_n else "")
        nombre = re.sub(r"\b(CASA USADA|SIN SUBSIDIO|CON SUBSIDIO|UF)\b.*", "", nombre).strip()
        if not nombre or len(nombre.split()) < 2 or not (monto or inmo or usada):
            continue
        filas.append({"nombre": nombre, "inmobiliaria": inmo,
                      "tipo": "usada" if usada else "nueva",
                      "subsidio": ("Con Subsidio" if ("con subsidio" in tl or "ds1" in tl)
                                   else "Sin Subsidio" if "sin subsidio" in tl else ""),
                      "monto_uf": monto})
    vistos, out = set(), []
    for f in filas:
        if f["nombre"] not in vistos:
            vistos.add(f["nombre"])
            out.append(f)
    return out


async def _aplicar_proyeccion(broker_codigo, broker_nombre, mes, clientes_p):
    """P11 — CARGA AUTOMÁTICA: upsert de los clientes de la proyección en carpetas + Bóveda ADN.
    Reemplaza la lista anterior de ese broker para ese mes (sin duplicados) y alerta a Gerencia."""
    creados, actualizados, faltantes = 0, 0, []
    ahora = _now()
    nombres_nuevos = []
    for c in clientes_p:
        nombre = c["nombre"]
        nombres_nuevos.append(nombre)
        fd = await db.folders.find_one({"nombre": {"$regex": f"^{re.escape(nombre)}$", "$options": "i"}})
        setd = {"inmobiliaria": c.get("inmobiliaria") or "", "tipo_operacion": c.get("tipo") or "nueva",
                "subsidio_proyeccion": c.get("subsidio") or "", "proyeccion_uf": c.get("monto_uf"),
                "broker_origen": broker_nombre, "proyeccion_broker": broker_codigo,
                "proyeccion_mes": mes, "updated_at": ahora}
        setd = {k: v for k, v in setd.items() if v not in (None, "")}
        if fd:
            await db.folders.update_one({"id": fd["id"]},
                                        {"$set": setd, "$unset": {"oculto_supercarpeta": ""}})
            actualizados += 1
            fid = fd["id"]
        else:
            fid = str(uuid.uuid4())
            await db.folders.insert_one({"id": fid, "nombre": nombre, "rut": "", "archivos": [],
                                         "created": ahora, **setd})
            creados += 1
        if not c.get("monto_uf"):
            faltantes.append(f"{nombre}: monto UF")
        await _sync_adn(fid)
    async for f in db.folders.find({"proyeccion_broker": broker_codigo, "proyeccion_mes": mes}):
        if (f.get("nombre") or "").strip().upper() not in nombres_nuevos:
            await db.folders.update_one({"id": f["id"]}, {"$set": {"oculto_supercarpeta": {
                "en": ahora, "por": f"proyeccion_{broker_codigo}"}}})
    cfg = await db.config.find_one({"_key": "flota_agosto"}) or {}
    if cfg.get("activo") and len(nombres_nuevos) >= 5:
        await db.config.update_one({"_key": "flota_agosto"}, {"$set": {"nombres": nombres_nuevos}})
    total_uf = round(sum(c.get("monto_uf") or 0 for c in clientes_p), 1)
    meta_cfg = await db.config.find_one({"_key": "proyeccion_agosto"}) or {}
    meta_uf = meta_cfg.get("meta_uf") or 41717
    resumen = (f"📊 Proyección {mes} de {broker_nombre}: {len(clientes_p)} clientes cargados, "
               f"{total_uf} UF proyectadas (meta {meta_uf} UF). "
               + (f"⚠️ Faltan datos: {'; '.join(faltantes[:6])}. " if faltantes else "")
               + (f"⚠️ Diferencia vs meta: {round(meta_uf - total_uf, 1)} UF."
                  if abs(meta_uf - total_uf) > 0.5 else "✅ Cuadra con la meta."))
    await db.alertas.insert_one({"id": str(uuid.uuid4()), "tipo": "proyeccion_broker",
                                 "destino": "Gerencia Comercial", "mensaje": resumen,
                                 "fecha": ahora, "leida": False})
    return {"clientes": len(clientes_p), "creados": creados, "actualizados": actualizados,
            "total_uf": total_uf, "faltantes": faltantes, "resumen": resumen}


# VENTANA DE PROYECCIÓN: solo del día 1 al 5 HÁBIL de cada mes
def _ventana_proyeccion():
    from datetime import timedelta
    hoy = datetime.now(timezone.utc).date()
    d, habiles = hoy.replace(day=1), []
    while len(habiles) < 5:
        if d.weekday() < 5:
            habiles.append(d)
        d += timedelta(days=1)
    return hoy in habiles, habiles[-1].strftime("%d/%m/%Y")


@broker.post("/proyeccion")
async def broker_proyeccion(request: Request, mes: str = Form(...), archivo: UploadFile = File(...)):
    c = _claims(request)
    codigo = c.get("sub") or ""
    abierta, limite = _ventana_proyeccion()
    if not abierta and c.get("rol") not in ("admin", "maestro"):
        raise HTTPException(status_code=423, detail=(
            f"⛔ Ventana de carga cerrada: la proyección solo se puede subir entre el día 1 y el "
            f"5° día hábil de cada mes (última fecha de este mes: {limite})."))
    if not re.match(r"^\d{4}-\d{2}$", mes or ""):
        raise HTTPException(status_code=400, detail="Mes inválido (formato AAAA-MM)")
    nombre_arch = re.sub(r"[^\w.\- ]", "_", archivo.filename or "proyeccion.pdf")
    base = fsvc.CLIENTES_DIR.parent / "brokers" / re.sub(r"[^\w-]", "_", codigo)
    base.mkdir(parents=True, exist_ok=True)
    destino = base / f"PROYECCION_{mes}_{nombre_arch}"
    destino.write_bytes(await archivo.read())
    reg = {"id": str(uuid.uuid4()), "broker_codigo": codigo, "broker_nombre": c.get("nombre") or codigo,
           "mes": mes, "archivo": destino.name, "ruta": str(destino), "subido_en": _now()}
    await db.broker_proyecciones.insert_one(dict(reg))
    resultado = {}
    try:
        filas = await asyncio.to_thread(_parsear_proyeccion, destino)
        if filas:
            resultado = await _aplicar_proyeccion(codigo, c.get("nombre") or codigo, mes, filas)
            await db.broker_proyecciones.update_one({"id": reg["id"]}, {"$set": {"parseo": resultado}})
    except Exception as e:
        logging.warning(f"proyeccion parse: {e}")
    await _log_broker(c, "proyeccion_subida", {"archivo": destino.name, "mes": mes})
    return {"ok": True, "archivo": destino.name, "mes": mes, "carga_automatica": resultado}


@broker.get("/estado-situacion")
async def broker_estado_situacion(request: Request):
    """ESTADO_DE_SITUACION: el broker solo ve la situación de SUS clientes asociados."""
    c = _claims(request)
    # AISLAMIENTO TOTAL: cada broker (y el admin en modo broker) ve SOLO sus propios clientes
    q = {"broker_codigo": c.get("sub") or ""}
    out = []
    async for fd in db.folders.find(q).sort("nombre", 1):
        hitos_r = await db.hitos_externos.find({"folder_id": fd["id"]}, {"_id": 0, "hito": 1, "fecha": 1, "fuente": 1}).sort("creado", -1).to_list(3)
        out.append({
            "id": fd["id"], "cliente": fd.get("nombre"), "rut": fd.get("rut") or "",
            "tipo_operacion": (fd.get("tipo_operacion") or "").upper() or "—",
            "documentos": len(fd.get("archivos") or []),
            "tasacion": ("Recibida" if fd.get("tasacion_informe_recibido_at")
                         else ("Solicitada" if fd.get("tasacion_solicitada_at") else "Pendiente de Información")),
            "estudio": ("Con Reparos" if (fd.get("reparos_alertas") or [])
                        else ("Recibido" if fd.get("estudio_recibido_at") else "Pendiente de Información")),
            "reparos": len(fd.get("reparos_alertas") or []),
            "escrituracion": bool(fd.get("is_escrituracion")),
            "hitos_recientes": hitos_r})
    return {"situacion": out, "total": len(out)}


@broker.get("/actividad")
async def broker_actividad(request: Request):
    """HUELLA DIGITAL: registro DashAI de cada interacción del broker."""
    c = _claims(request)
    q = {} if c.get("rol") in ("admin", "maestro") else {"broker_codigo": c.get("sub") or ""}
    regs = await db.broker_activity_log.find(q, {"_id": 0}).sort("fecha", -1).to_list(50)
    return {"actividad": regs, "total": len(regs)}


@broker.get("/proyecciones")
async def broker_proyecciones(request: Request):
    c = _claims(request)
    q = {} if c.get("rol") in ("admin", "maestro") else {"broker_codigo": c.get("sub") or ""}
    regs = await db.broker_proyecciones.find(q, {"_id": 0, "ruta": 0}).sort("subido_en", -1).to_list(100)
    return {"proyecciones": regs, "total": len(regs)}


# ──── PANEL DE FUENTES OPERATIVAS + GESTOR DINÁMICO (Reglas #34 y #36) ─────
async def _validar_clave_usuario(request, clave):
    """Regla #36: todo cambio en la red de escucha exige la firma digital (clave)."""
    c = _claims(request)
    codigo = c.get("sub") or ""
    user = await db.users.find_one({"codigo": {"$regex": f"^{re.escape(codigo)}$", "$options": "i"}})
    if not user:
        raise HTTPException(status_code=401, detail="Usuario no identificado")
    ok = False
    if user.get("clave_hash"):
        ok = bool(clave) and bcrypt.checkpw(clave.encode(), user["clave_hash"].encode())
    else:
        ok = bool(clave) and user.get("password") == clave
    if not ok and clave == os.environ.get("MASTER_PIN", "!") and user.get("rol") == "admin":
        ok = True
    if not ok:
        raise HTTPException(status_code=403,
                            detail="Clave incorrecta — el cambio de fuentes exige su firma digital (Regla de Oro #36)")
    return user


@fuentes.get("")
async def fuentes_todas():
    out = []
    for p in PANELES:
        cfg = await db.config.find_one({"_key": f"fuentes_imap_{p}"}, {"_id": 0}) or {}
        out.append({"panel": p, "correo_principal": cfg.get("correo_principal", ""),
                    "aliados": cfg.get("aliados", [])})
    return {"fuentes": out}


@fuentes.get("/auditoria")
async def fuentes_auditoria():
    """LOG DE AUDITORÍA DE RED (DashAI): registro inmutable de altas/bajas de fuentes."""
    regs = await db.fuentes_auditoria_log.find({}, {"_id": 0}).sort("fecha", -1).to_list(100)
    return {"auditoria": regs, "total": len(regs)}


@fuentes.get("/{panel}")
async def fuentes_get(panel: str):
    if not _panel_valido(panel):
        raise HTTPException(status_code=404, detail="Panel desconocido")
    cfg = await db.config.find_one({"_key": f"fuentes_imap_{panel}"}, {"_id": 0}) or {}
    return {"panel": panel, "correo_principal": cfg.get("correo_principal", ""),
            "aliados": cfg.get("aliados", [])}


async def _auditar_cambios(user, panel, antes, despues):
    """Regla #36: cada alta, baja o edición queda firmada en fuentes_auditoria_log."""
    prev = {a.get("email"): a for a in (antes or [])}
    nuev = {a.get("email"): a for a in (despues or [])}
    cambios = []
    for em in nuev:
        if em not in prev:
            cambios.append(("alta", f"Alta de fuente {em} ({nuev[em].get('etiqueta','')})"))
        elif prev[em] != nuev[em]:
            cambios.append(("edicion", f"Edición de fuente {em} → etiqueta '{nuev[em].get('etiqueta','')}', nombre '{nuev[em].get('nombre','')}'"))
    for em in prev:
        if em not in nuev:
            cambios.append(("baja", f"Baja de fuente {em} — DashAI deja de escucharla de inmediato"))
    for tipo, detalle in cambios:
        await db.fuentes_auditoria_log.insert_one({
            "id": str(uuid.uuid4()), "usuario": user.get("nombre", ""), "codigo": user.get("codigo", ""),
            "modulo": panel, "cambio": tipo, "detalle": detalle,
            "firmado_digitalmente": True, "fecha": _now()})
    return len(cambios)


@fuentes.post("/{panel}")
async def fuentes_set(panel: str, payload: dict, request: Request):
    if not _panel_valido(panel):
        raise HTTPException(status_code=404, detail="Panel desconocido")
    # VALIDACIÓN DE CLAVE (firma digital del responsable del módulo)
    user = await _validar_clave_usuario(request, (payload.get("clave") or "").strip())
    correo = (payload.get("correo_principal") or "").strip().lower()
    if correo and not EMAIL_RE.match(correo):
        raise HTTPException(status_code=400, detail="Correo principal inválido")
    aliados = []
    for a in (payload.get("aliados") or [])[:3]:
        em = (a.get("email") or "").strip().lower()
        if not em:
            continue
        if not EMAIL_RE.match(em):
            raise HTTPException(status_code=400, detail=f"Correo de aliado inválido: {em}")
        etiqueta = (a.get("etiqueta") or "").strip()
        if not etiqueta:
            raise HTTPException(status_code=400, detail=f"La etiqueta es obligatoria para {em} (ej. Tasador, Abogado, Notaría)")
        aliados.append({"nombre": (a.get("nombre") or "").strip() or em.split("@")[-1],
                        "email": em, "etiqueta": etiqueta})
    cfg_prev = await db.config.find_one({"_key": f"fuentes_imap_{panel}"}) or {}
    await db.config.update_one({"_key": f"fuentes_imap_{panel}"}, {"$set": {
        "correo_principal": correo, "aliados": aliados,
        "actualizado_por": user.get("nombre", ""), "actualizado": _now()}}, upsert=True)
    cambios = await _auditar_cambios(user, panel, cfg_prev.get("aliados"), aliados)
    return {"ok": True, "panel": panel, "aliados": len(aliados), "cambios_auditados": cambios}


# ─────────── MOTOR DE SEGUIMIENTO EN TIEMPO REAL (Regla de Oro #34) ────────
async def _aliados_config():
    dominios = {}
    for p in PANELES:
        cfg = await db.config.find_one({"_key": f"fuentes_imap_{p}"}) or {}
        for a in (cfg.get("aliados") or []):
            dom = (a.get("email") or "").split("@")[-1].split(".")[0].lower()
            if dom:
                dominios[dom] = {"nombre": a.get("nombre") or dom, "panel": p}
    for d in HITOS_BASE:
        dominios.setdefault(d, {"nombre": "Value Property" if d == "valueproperty" else "Guillermo Mardones", "panel": ""})
    return dominios


def _dtx(s):
    try:
        return datetime.fromisoformat(str(s)).astimezone(timezone.utc).replace(tzinfo=None)
    except Exception:
        return None


async def _archivar_adjuntos(fd, email_id, prefijo, subdir="07_estudio_titulo"):
    """Archiva los PDF adjuntos de un correo dentro de la carpeta del cliente."""
    try:
        atts = await asyncio.to_thread(mail.fetch_attachments_by_id, email_id)
        destino = fsvc.folder_dir(fd.get("nombre") or "") / subdir
        destino.mkdir(parents=True, exist_ok=True)
        rutas = []
        for a in atts or []:
            fn = a.get("filename") or ""
            if fn.lower().endswith(".pdf") and a.get("content_bytes"):
                nombre_f = f"{prefijo}_" + re.sub(r"[^\w.\- ]", "_", fn)
                (destino / nombre_f).write_bytes(a["content_bytes"])
                await db.folders.update_one({"id": fd["id"]},
                    {"$addToSet": {"archivos": f"{subdir}/{nombre_f}"}})
                rutas.append(f"{subdir}/{nombre_f}")
        return rutas
    except Exception as e:
        logging.warning(f"malla archivar {fd.get('nombre')}: {e}")
        return []


async def _procesar_hito(correo, dom, info, direccion, por_rut, texto, email_id=None):
    """REGLA DE HIERRO #34: el RUT es el pegamento. Sin RUT válido → NO se marca hito."""
    asunto = correo.get("subject") or ""
    _raw = f"{dom}|{direccion}|{asunto}|{correo.get('date','')}".encode()
    clave = hashlib.sha256(_raw).hexdigest()
    # _legacy: clave de deduplicación histórica (MD5 no criptográfico, solo lectura de compatibilidad)
    _legacy = hashlib.md5(_raw, usedforsecurity=False).hexdigest()
    if (await db.hitos_externos.find_one({"clave": {"$in": [clave, _legacy]}})
            or await db.hitos_descartados.find_one({"clave": {"$in": [clave, _legacy]}})):
        return "duplicado"
    m = RUT_RE.search(texto or "")
    rut_n = _rut_limpio(m.group(1)) if m else ""
    fd = por_rut.get(rut_n)
    if not fd:
        motivo = "sin RUT en asunto/cuerpo" if not rut_n else f"RUT {m.group(1)} sin carpeta asociada"
        await db.hitos_descartados.insert_one({"clave": clave, "motivo": motivo, "asunto": asunto[:180],
                                               "fuente": info["nombre"], "dominio": dom,
                                               "direccion": direccion, "fecha": correo.get("date", ""),
                                               "creado": _now()})
        logging.info(f"Malla #34: hito descartado ({motivo}) — {asunto[:60]}")
        return "descartado"
    hito_nombre = HITOS_BASE.get(dom, {}).get(direccion) or (
        f"Enviado a {info['nombre']}" if direccion == "enviado" else f"Recibido de {info['nombre']}")
    reg = {"id": str(uuid.uuid4()), "clave": clave, "folder_id": fd["id"], "cliente": fd.get("nombre"),
           "rut": fd.get("rut") or "", "hito": hito_nombre, "fuente": info["nombre"], "dominio": dom,
           "panel": info.get("panel", ""), "direccion": direccion, "asunto": asunto[:180],
           "fecha": correo.get("date", ""), "validado_rut": True, "creado": _now()}
    await db.hitos_externos.insert_one(dict(reg))
    marcas = {"updated_at": _now()}
    if dom == "valueproperty" and direccion == "enviado" and not fd.get("tasacion_solicitada_at"):
        marcas["tasacion_solicitada_at"] = _now()
    if dom == "gmardones" and direccion == "recibido":
        marcas["estudio_recibido_at"] = _now()
        if email_id:
            adjs_g = await _archivar_adjuntos(fd, email_id, "ESTUDIO")
            if adjs_g:
                await db.hitos_externos.update_one({"id": reg["id"]}, {"$set": {"adjuntos": adjs_g}})
    await db.folders.update_one({"id": fd["id"]}, {"$set": marcas})
    return "marcado"


async def malla_scan():
    est = await db.config.find_one({"_key": "malla_estado"}) or {}
    since = est.get("since")
    if not since:  # Regla #15: filtro temporal — nada retroactivo
        since = _now()
        await db.config.update_one({"_key": "malla_estado"}, {"$set": {"since": since}}, upsert=True)
    since_dt = _dtx(since)
    dominios = await _aliados_config()
    folders = await db.folders.find({}, {"id": 1, "nombre": 1, "rut": 1,
                                         "tasacion_solicitada_at": 1}).to_list(2000)
    por_rut = {_rut_limpio(f.get("rut")): f for f in folders if len(_rut_limpio(f.get("rut"))) >= 8}
    res = {"marcados": 0, "descartados": 0, "duplicados": 0}
    enviados = await asyncio.to_thread(mail.fetch_sent_headers, 80)
    for e in enviados or []:
        dom = next((d for d in dominios if d in (e.get("to") or "").lower()), None)
        fch = _dtx(e.get("date"))
        if not dom or (since_dt and fch and fch < since_dt):
            continue
        r = await _procesar_hito(e, dom, dominios[dom], "enviado", por_rut, e.get("subject") or "")
        res[{"marcado": "marcados", "descartado": "descartados"}.get(r, "duplicados")] += 1
    recibidos = await asyncio.to_thread(mail.fetch_recent_full, 30)
    for e in recibidos or []:
        dom = next((d for d in dominios if d in (e.get("from") or "").lower()), None)
        fch = _dtx(e.get("date"))
        if not dom or (since_dt and fch and fch < since_dt):
            continue
        texto = f"{e.get('subject') or ''} {e.get('body') or ''}"
        r = await _procesar_hito(e, dom, dominios[dom], "recibido", por_rut, texto, email_id=e.get("id"))
        res[{"marcado": "marcados", "descartado": "descartados"}.get(r, "duplicados")] += 1
    # ── FLUJOS DIFERENCIADOS (Regla #37): vendedores transitorios + reparos ──
    vendedores = {}
    async for f in db.folders.find({"vendedor_usada.email": {"$exists": True, "$ne": ""}},
                                   {"id": 1, "nombre": 1, "rut": 1, "vendedor_usada": 1}):
        em = ((f.get("vendedor_usada") or {}).get("email") or "").lower()
        if em:
            vendedores[em] = f
    for e in recibidos or []:
        fch = _dtx(e.get("date"))
        if since_dt and fch and fch < since_dt:
            continue
        de = (e.get("from") or "").lower()
        asunto = e.get("subject") or ""
        texto = f"{asunto} {e.get('body') or ''}"
        # 1) Fuente TRANSITORIA: documentos del vendedor de una usada → carpeta vinculada
        fd_v = next((vendedores[em] for em in vendedores if em and em in de), None)
        if fd_v:
            _raw_v = f"{de}|{asunto}|{e.get('date','')}".encode()
            clave = "vend-" + hashlib.sha256(_raw_v).hexdigest()
            _legacy_v = "vend-" + hashlib.md5(_raw_v, usedforsecurity=False).hexdigest()
            if not await db.hitos_externos.find_one({"clave": {"$in": [clave, _legacy_v]}}):
                archivados = (await _archivar_adjuntos(fd_v, e.get("id"), "VENDEDOR")) if e.get("id") else []
                await db.hitos_externos.insert_one({
                    "id": str(uuid.uuid4()), "clave": clave, "folder_id": fd_v["id"],
                    "cliente": fd_v.get("nombre"), "rut": fd_v.get("rut") or "",
                    "hito": "Documento de Vendedor Recibido", "fuente": (fd_v.get("vendedor_usada") or {}).get("nombre") or "Vendedor",
                    "dominio": "vendedor_usada", "panel": "victoria", "direccion": "recibido",
                    "asunto": asunto[:180], "fecha": e.get("date", ""), "archivados": len(archivados),
                    "adjuntos": archivados,
                    "validado_rut": True, "tipo_operacion": "usada", "creado": _now()})
                res["marcados"] += 1
        # 2) MOTOR DE REPAROS (DashAI): "reparo" en correos de los abogados
        if "reparo" in texto.lower() and any(r in de for r in REPARO_REMITENTES):
            _raw_r = f"{de}|{asunto}|{e.get('date','')}".encode()
            clave = "rep-" + hashlib.sha256(_raw_r).hexdigest()
            _legacy_r = "rep-" + hashlib.md5(_raw_r, usedforsecurity=False).hexdigest()
            if (await db.hitos_externos.find_one({"clave": {"$in": [clave, _legacy_r]}})
                    or await db.hitos_descartados.find_one({"clave": {"$in": [clave, _legacy_r]}})):
                continue
            m = RUT_RE.search(texto)
            fd_r = por_rut.get(_rut_limpio(m.group(1))) if m else None
            if not fd_r:  # REGLA DE HIERRO: el RUT es el único eje — sin RUT no hay alerta
                await db.hitos_descartados.insert_one({"clave": clave, "motivo": "reparo sin RUT del cliente",
                    "asunto": asunto[:180], "fuente": de[:80], "dominio": "reparos",
                    "direccion": "recibido", "fecha": e.get("date", ""), "creado": _now()})
                res["descartados"] += 1
                continue
            reparo = {"asunto": asunto[:180], "texto": (e.get("body") or "")[:600],
                      "de": de[:120], "fecha": e.get("date", ""), "detectado": _now()}
            await db.folders.update_one({"id": fd_r["id"]}, {"$push": {"reparos_alertas": reparo}})
            await db.hitos_externos.insert_one({
                "id": str(uuid.uuid4()), "clave": clave, "folder_id": fd_r["id"],
                "cliente": fd_r.get("nombre"), "rut": fd_r.get("rut") or "",
                "hito": "⚠️ Reparo Detectado", "fuente": "Abogados (Estudio de Título)",
                "dominio": "reparos", "panel": "victoria", "direccion": "recibido",
                "asunto": asunto[:180], "fecha": e.get("date", ""),
                "validado_rut": True, "creado": _now()})
            await db.alertas.insert_one({"id": str(uuid.uuid4()), "tipo": "reparo",
                "mensaje": f"⚠️ REPARO detectado por DashAI en carpeta {fd_r.get('nombre')}: {asunto[:120]}",
                "folder_id": fd_r["id"], "fecha": _now(), "leida": False})
            res["marcados"] += 1
    await db.config.update_one({"_key": "malla_estado"}, {"$set": {"ultima_revision": _now(),
                                                                   "ultimo_resultado": res}}, upsert=True)
    return res


async def auditoria_loop():
    """REGLA DE ORO (Supercarpeta V2): actualización en tiempo real — cada 10 minutos se
    revisan los correos recientes de brokers y tasadores autorizados."""
    await asyncio.sleep(240)
    while True:
        try:
            await flujos_auditoria_real(limit=40, dias=2)
        except Exception as e:
            if "after close" in str(e):
                break
            logging.warning(f"auditoria loop: {e}")
        await asyncio.sleep(600)


async def malla_loop():
    """DashAI rastrea envíos y recepciones con los aliados externos cada 5 minutos."""
    await asyncio.sleep(90)
    while True:
        try:
            await malla_scan()
        except Exception as e:
            logging.warning(f"malla_inteligencia: {e}")
        await asyncio.sleep(300)


@hitos.get("/feed")
async def hitos_feed():
    regs = await db.hitos_externos.find({}, {"_id": 0}).sort("creado", -1).to_list(40)
    est = await db.config.find_one({"_key": "malla_estado"}) or {}
    descartados = await db.hitos_descartados.count_documents({})
    return {"hitos": regs, "total": len(regs), "descartados": descartados,
            "ultima_revision": est.get("ultima_revision", ""), "activo_desde": est.get("since", "")}


@hitos.post("/scan")
async def hitos_scan_manual():
    return {"ok": True, "resultado": await malla_scan()}


# ─── FLUJOS DIFERENCIADOS: USADA (transitoria) vs INMOBILIARIA (permanente) ─
INMOBILIARIAS_SEED = ["Maestra", "Comac", "Bestal"]


@flujos.get("/carpetas")
async def flujos_carpetas():
    out = []
    async for fd in db.folders.find({}).sort("nombre", 1):
        out.append({"id": fd["id"], "nombre": fd.get("nombre"), "rut": fd.get("rut") or "",
                    "tipo_operacion": fd.get("tipo_operacion", ""),
                    "broker_origen": fd.get("broker_origen") or fd.get("broker_nombre") or "DIRECTO",
                    "vendedor_usada": fd.get("vendedor_usada") or {},
                    "contacto_inmobiliario": fd.get("contacto_inmobiliario") or {},
                    "reparos": fd.get("reparos_alertas") or []})
    return {"carpetas": out, "total": len(out)}


@flujos.post("/vendedor/{fid}")
async def flujos_set_vendedor(fid: str, payload: dict):
    """[🏠 Vivienda Usada] Fuente TRANSITORIA: vendedor por carpeta (Regla #37).
    DashAI rastrea y archiva los documentos de este mail bajo el RUT del cliente."""
    fd = await db.folders.find_one({"id": fid})
    if not fd:
        raise HTTPException(status_code=404, detail="Carpeta no existe")
    em = (payload.get("email") or "").strip().lower()
    if not em or not EMAIL_RE.match(em):
        raise HTTPException(status_code=400, detail="Correo del vendedor inválido")
    vendedor = {"nombre": (payload.get("nombre") or "").strip(),
                "email": em, "telefono": (payload.get("telefono") or "").strip()}
    await db.folders.update_one({"id": fid}, {"$set": {
        "vendedor_usada": vendedor, "tipo_operacion": "usada", "updated_at": _now()}})
    return {"ok": True, "tipo_operacion": "usada", "vendedor": vendedor,
            "nota": "DashAI escucha este correo y archiva sus documentos bajo el RUT del cliente"}


@flujos.get("/inmobiliarias")
async def flujos_inmobiliarias():
    if await db.contactos_inmobiliarios.count_documents({}) == 0:
        for n in INMOBILIARIAS_SEED:
            await db.contactos_inmobiliarios.insert_one({"id": str(uuid.uuid4()), "inmobiliaria": n,
                "encargado": "", "email": "", "telefono": "", "creado": _now()})
    regs = await db.contactos_inmobiliarios.find({}, {"_id": 0}).sort("inmobiliaria", 1).to_list(100)
    return {"contactos": regs, "total": len(regs)}


@flujos.post("/inmobiliarias")
async def flujos_inmobiliaria_guardar(payload: dict):
    """Fuente PERMANENTE: contactos inmobiliarios reutilizables (Regla #37)."""
    nombre = (payload.get("inmobiliaria") or "").strip()
    if not nombre:
        raise HTTPException(status_code=400, detail="Nombre de la inmobiliaria obligatorio")
    em = (payload.get("email") or "").strip().lower()
    if em and not EMAIL_RE.match(em):
        raise HTTPException(status_code=400, detail="Correo inválido")
    doc = {"inmobiliaria": nombre, "encargado": (payload.get("encargado") or "").strip(),
           "email": em, "telefono": (payload.get("telefono") or "").strip()}
    cid = (payload.get("id") or "").strip()
    if cid:
        await db.contactos_inmobiliarios.update_one({"id": cid}, {"$set": doc})
    else:
        cid = str(uuid.uuid4())
        await db.contactos_inmobiliarios.insert_one({"id": cid, **doc, "creado": _now()})
    return {"ok": True, "id": cid}


@flujos.post("/inmobiliaria/{fid}")
async def flujos_asignar_inmobiliaria(fid: str, payload: dict):
    """[🏢 Vivienda Inmobiliaria] asigna el contacto permanente a la carpeta."""
    fd = await db.folders.find_one({"id": fid})
    if not fd:
        raise HTTPException(status_code=404, detail="Carpeta no existe")
    ct = await db.contactos_inmobiliarios.find_one({"id": (payload.get("contacto_id") or "").strip()}, {"_id": 0})
    if not ct:
        raise HTTPException(status_code=404, detail="Contacto inmobiliario no existe")
    await db.folders.update_one({"id": fid}, {"$set": {
        "contacto_inmobiliario": ct, "tipo_operacion": "inmobiliaria", "updated_at": _now()}})
    return {"ok": True, "tipo_operacion": "inmobiliaria", "contacto": ct}


@flujos.get("/contactos-visita")
async def flujos_contactos_visita():
    """GESTIÓN DE TASACIÓN: Contacto de Visita por carpeta.
    Usada → mail/teléfono del Vendedor · Inmobiliaria → encargado del proyecto."""
    out = {}
    async for fd in db.folders.find({"tipo_operacion": {"$in": ["usada", "inmobiliaria"]}}):
        if fd.get("tipo_operacion") == "usada":
            v = fd.get("vendedor_usada") or {}
            out[fd["id"]] = {"tipo": "USADA", "origen": "Vendedor (auto)", "nombre": v.get("nombre", ""),
                             "email": v.get("email", ""), "telefono": v.get("telefono", "")}
        else:
            ci = fd.get("contacto_inmobiliario") or {}
            out[fd["id"]] = {"tipo": "INMOBILIARIA", "origen": f"Encargado {ci.get('inmobiliaria', '')}",
                             "nombre": ci.get("encargado", ""), "email": ci.get("email", ""),
                             "telefono": ci.get("telefono", "")}
    return {"contactos": out, "total": len(out)}


# ── AUDITORÍA DE CORREO DASHAI (Regla de Oro #43) — flujo real, sin inventar ─
NOTARIA_KW = ("notar", "escritura", "repertorio", "cesión", "cesion", "serie de crédito", "serie de credito", "firma confirmada")


OBS_RE = re.compile(r"(observaci[oó]n(?:es)?|reparos?)\s*[:\-\n]?\s*(.{40,900}?)(?:\n\s*\n|$)",
                    re.I | re.S)


async def _minar_reparos_pdf(fd):
    """REGLA DE MINADO: extrae el texto EXACTO de 'Observaciones'/'Reparos' desde los PDFs
    de Estudio de Títulos archivados en la Bóveda Local (celda NARANJA en la Supercarpeta)."""
    import ocr_service as _ocr
    base = fsvc.folder_dir(fd.get("nombre") or "")
    if not base.exists():
        return 0
    pdfs = sorted([p for p in base.rglob("*.pdf")
                   if "estudio" in p.name.lower() or p.parent.name == "07_estudio_titulo"],
                  key=lambda x: -x.stat().st_mtime)[:2]
    existentes = {(r.get("texto") or "")[:80] for r in (fd.get("reparos_alertas") or [])}
    nuevos = 0
    for p in pdfs:
        try:
            texto = await asyncio.to_thread(_ocr.ocr_texto, p.read_bytes(), 8) or ""
        except Exception:
            continue
        for m in OBS_RE.finditer(texto):
            exacto = re.sub(r"\s+", " ", m.group(2)).strip()[:800]
            if len(exacto) < 40 or exacto[:80] in existentes:
                continue
            reparo = {"asunto": f"Minado PDF: {p.name[:120]}", "texto": exacto,
                      "de": "minado_pdf", "fecha": _now(), "detectado": _now(),
                      "origen": "minado_pdf", "celda": "naranja", "archivo_fuente": p.name}
            await db.folders.update_one({"id": fd["id"]}, {"$push": {"reparos_alertas": reparo}})
            existentes.add(exacto[:80])
            nuevos += 1
            break  # una sección por PDF: el texto íntegro ya quedó copiado
    return nuevos


def _match_carpeta(texto, por_rut, folders):
    """Eje 1: RUT (Regla #34). Eje 2: nombre completo (≥2 tokens) como respaldo."""
    m = RUT_RE.search(texto or "")
    if m:
        fd = por_rut.get(_rut_limpio(m.group(1)))
        if fd:
            return fd, "rut"
    t = (texto or "").lower()
    for fd in folders:
        toks = [x for x in (fd.get("nombre") or "").lower().split() if len(x) > 2]
        if len(toks) >= 2 and sum(1 for x in toks if x in t) >= 2:
            return fd, "nombre"
    return None, ""


FUENTES_HITOS = ("tasacion", "estudio", "cesion", "set_credito", "notaria",
                 "serviu", "promesa", "carta_oferta", "carpeta_notaria", "escritura", "fecha_firma")


def _norm_fuente(x):
    if isinstance(x, dict):
        return {"correo": (x.get("correo") or "").strip().lower(),
                "nombre": (x.get("nombre") or "").strip()}
    return {"correo": str(x).strip().lower(), "nombre": ""}


def _correos_de(lst):
    return [f["correo"] for f in (_norm_fuente(x) for x in (lst or [])) if f["correo"]]


async def _fuentes_documentos_meta():
    cfg = await db.config.find_one({"_key": "fuentes_documentos"}) or {}
    fu = cfg.get("fuentes") or {}
    base = {"tasacion": [{"correo": "contacto@valueproperty.cl", "nombre": "Value Property - Tasaciones"},
                         {"correo": "contacto@valuedproperty.cl", "nombre": "Value Property - Tasaciones"}],
            "estudio": [{"correo": "victoriavilches@centralmutuos.cl", "nombre": "Victoria Vilches - Estudios"},
                        {"correo": "gmajluf@amvabogados.cl", "nombre": "Guillermo Majluf - Abogado"}]}
    out = {}
    for h in FUENTES_HITOS:
        lst = fu.get(h) if h in fu else base.get(h, [])
        out[h] = [f for f in (_norm_fuente(x) for x in (lst or [])) if f["correo"]]
    return out


async def _fuentes_documentos():
    """Regla #67 (P8): correos fuente configurables por tipo de documento (solo direcciones)."""
    meta = await _fuentes_documentos_meta()
    return {h: [f["correo"] for f in lst] for h, lst in meta.items()}


async def _sync_adn(fid):
    """Regla #67: todo cambio termina ESCRITO en la Bóveda ADN_CLIENTES_360."""
    import adn_clientes as _adn
    fd = await db.folders.find_one({"id": fid})
    if not fd:
        return False
    reg = _adn._registro_desde_folder(fd)
    reg["expediente_360"] = await _adn._expediente_360(fd)
    return await _adn._upsert_adn(reg)


def _verificar_firmas_pdf(pdf_bytes):
    """P9 — REGLA DE HIERRO: 'Set Firmado' en el asunto NO basta; debe haber evidencia
    documental de firma dentro del PDF (firma digital, texto de firma o firma manuscrita)."""
    try:
        import io
        from pypdf import PdfReader
        r = PdfReader(io.BytesIO(pdf_bytes))
        try:
            af = r.trailer["/Root"].get("/AcroForm")
            if af and int(af.get("/SigFlags", 0) or 0) >= 1:
                return True, "firma digital detectada (AcroForm/SigFlags)"
        except Exception:
            pass
        texto = " ".join((p.extract_text() or "") for p in r.pages[-3:]).lower()
        kws = ("firmado electronicamente", "firmado electrónicamente", "firma electrónica",
               "firma electronica", "e-cert", "ecert", "firmado ante notario", "firman en señal")
        if any(k in texto for k in kws):
            return True, "texto de firma electrónica dentro del documento"
        tiene_img = False
        for p in r.pages[-2:]:
            try:
                xo = (p.get("/Resources") or {}).get("/XObject")
                if xo:
                    for k in xo:
                        if xo[k].get_object().get("/Subtype") == "/Image":
                            tiene_img = True
            except Exception:
                continue
        if tiene_img and "firma" in texto:
            return True, "imagen de firma manuscrita en las últimas páginas"
        if not texto.strip():
            try:
                import ocr_service as _ocr
                texto_ocr = (_ocr.ocr_texto(pdf_bytes, 3) or "").lower()
                if any(k in texto_ocr for k in kws) or ("firma" in texto_ocr and tiene_img):
                    return True, "firma detectada vía OCR del documento escaneado"
            except Exception:
                pass
        return False, "sin evidencia verificable de firmas dentro del PDF"
    except Exception as e:
        return False, f"no se pudo abrir el PDF para verificar firmas ({str(e)[:80]})"


_PEND_VERIF_PROMESA = "Pendiente verificación manual"


async def _verificar_compromiso_ia(fd, e):
    """VERIFICACIÓN DE FIRMA (IA): el compromiso/promesa de compraventa adjunto solo se
    marca 'Firmada (verificada IA)' con evidencia de firma de alta confianza; en cualquier
    duda queda azul 'Pendiente verificación manual' (jamás verde sin respaldo)."""
    import ai_extract as _ai
    estado, evidencia, archivo, firmado_ok = _PEND_VERIF_PROMESA, "correo sin adjunto PDF verificable", "", False
    try:
        atts = (await asyncio.to_thread(mail.fetch_attachments_by_id, e["id"])) if e.get("id") else []
        pdfs = [a for a in (atts or [])
                if (a.get("filename") or "").lower().endswith(".pdf") and a.get("content_bytes")]
        pref = [a for a in pdfs if re.search(r"promes|compromis|compravent", (a.get("filename") or "").lower())]
        for a in (pref or pdfs)[:3]:
            archivo = a.get("filename") or ""
            heur_ok, heur_ev = await asyncio.to_thread(_verificar_firmas_pdf, a["content_bytes"])
            texto = ""
            try:
                import io
                from pypdf import PdfReader
                r = PdfReader(io.BytesIO(a["content_bytes"]))
                texto = " ".join((p.extract_text() or "") for p in r.pages)
            except Exception:
                pass
            if len(texto.strip()) < 120:
                try:
                    import ocr_service as _ocr
                    texto = (await asyncio.to_thread(_ocr.ocr_texto, a["content_bytes"], 8)) or texto
                except Exception:
                    pass
            ia = await _ai.verificar_firma_compromiso(texto, fd.get("nombre") or "",
                                                      heur_ev if heur_ok else "")
            if ia:
                firmado_ok = bool(ia.get("firmado")) and (ia.get("confianza") == "alta" or heur_ok)
                evidencia = (ia.get("evidencia") or "").strip()[:400] or heur_ev
            else:
                firmado_ok = heur_ok
                evidencia = f"IA no disponible — verificación técnica del PDF: {heur_ev}"
            if firmado_ok:
                break
        if firmado_ok:
            estado = "Firmada (verificada IA)"
    except Exception as ex:
        evidencia = f"error al verificar el adjunto ({str(ex)[:80]})"
    await db.folders.update_one({"id": fd["id"]}, {"$set": {
        "promesa_verificacion": {"estado": estado, "firmado": firmado_ok,
                                 "evidencia": evidencia, "archivo": archivo,
                                 "fecha": e.get("date") or _now(), "origen": "ia"},
        "promesa_verificada_at": e.get("date") or _now()}})
    return estado, evidencia


async def _capturar_remitente(fd, hito, e):
    """CAPTURA AUTOMÁTICA DE REMITENTES: aprende quién envía qué. Queda 'Pendiente de
    Confirmación' hasta que Gerencia lo valide (Regla de Hierro) + registro por broker."""
    try:
        de_raw = e.get("from") or ""
        m = re.search(r"[\w.+-]+@[\w.-]+", de_raw)
        if not m:
            return
        correo = m.group(0).lower()
        bloqueados = (await db.config.find_one({"_key": "remitentes_bloqueados"}) or {}).get("correos") or []
        if correo in bloqueados or correo in {a["user"].lower() for a in mail.ACCOUNTS}:
            return
        nm = re.match(r'^\s*"?([^"<]+?)"?\s*<', de_raw)
        nombre_rem = (nm.group(1).strip() if nm else correo.split("@")[0])
        apr = await db.remitentes_registro.find_one({"correo": correo}) or {}
        hito_final = apr.get("hito_forzado") or hito
        ya_fuente = correo in _correos_de((fd.get("fuentes_doc") or {}).get(hito_final))
        detectadas = (fd.get("fuentes_detectadas") or {}).get(hito_final) or []
        if not ya_fuente and not any(d.get("correo") == correo for d in detectadas):
            hoy = _now()[:10]
            await db.folders.update_one({"id": fd["id"]}, {"$push": {
                f"fuentes_detectadas.{hito_final}": {
                    "correo": correo, "nombre": nombre_rem, "primera_vez": _now(),
                    "estado": "pendiente_confirmacion",
                    "etiqueta": f"Detectado automáticamente el {hoy}"}}})
            await db.alertas.insert_one({"id": str(uuid.uuid4()), "tipo": "remitente_detectado",
                "mensaje": f"📡 Nuevo remitente detectado: {nombre_rem} ({correo}) agregado como fuente de "
                           f"{hito_final} para {fd.get('nombre')} — Pendiente de Confirmación.",
                "fecha": _now(), "leida": False})
        await db.remitentes_registro.update_one({"correo": correo}, {
            "$set": {"nombre": nombre_rem, "ultima": _now()},
            "$inc": {f"hitos.{hito_final}": 1,
                     f"brokers.{(fd.get('inmobiliaria') or fd.get('broker_origen') or 'sin_broker').replace('.', '·').replace('$', '')}": 1}},
            upsert=True)
    except Exception as ex:
        logging.warning(f"captura remitente: {ex}")


async def _auditar_lote(correos):
    """Procesa un lote de correos (Regla #67): detecta tasaciones (Value Property),
    estudios de títulos (Victoria Vilches / Guillermo Majluf) y notaría, y ESCRIBE
    cada hallazgo en la carpeta y en el EXPEDIENTE_360 de la Bóveda ADN.
    REGLA DE HIERRO: leer correos sin persistir el hallazgo es violación constitucional."""
    folders = await db.folders.find({}, {"id": 1, "nombre": 1, "rut": 1}).to_list(2000)
    por_rut = {_rut_limpio(f.get("rut")): f for f in folders if len(_rut_limpio(f.get("rut"))) >= 8}
    res = {"correos_revisados": len(correos or []), "tasaciones_detectadas": 0,
           "estudios_detectados": 0, "sets_detectados": 0, "promesas_detectadas": 0,
           "docs_detectados": 0,
           "reparos_transcritos": 0, "sin_respaldo": 0, "detalle": []}
    fu = await _fuentes_documentos()
    tas_src = [s.lower() for s in fu.get("tasacion", []) if s]
    est_src = [s.lower() for s in fu.get("estudio", []) if s]
    ces_src = [s.lower() for s in fu.get("cesion", []) if s]
    set_src = [s.lower() for s in fu.get("set_credito", []) if s]
    async for f in db.folders.find({"fuentes_doc": {"$exists": True}}, {"fuentes_doc": 1}):
        fdoc = f.get("fuentes_doc") or {}
        tas_src += _correos_de(fdoc.get("tasacion"))
        est_src += _correos_de(fdoc.get("estudio"))
        ces_src += _correos_de(fdoc.get("cesion"))
        set_src += _correos_de(fdoc.get("set_credito"))
    notaria_srcs = []
    async for f in db.folders.find({"fuentes_doc.notaria": {"$exists": True}},
                                   {"id": 1, "nombre": 1, "ciudad": 1, "fuentes_doc": 1}):
        for s in _correos_de((f.get("fuentes_doc") or {}).get("notaria")):
            notaria_srcs.append((s, f))
    for e in correos or []:
        de = (e.get("from") or "").lower()
        asunto = e.get("subject") or ""
        cuerpo = e.get("body") or ""
        texto = f"{asunto} {cuerpo}"
        # ── NOTARÍA (fuente por cliente): confirma firma de escritura + coherencia de ciudad ──
        fd_n = next((f for s, f in notaria_srcs if s in de), None)
        if fd_n:
            tx = texto.lower()
            partes = [p for p in (fd_n.get("nombre") or "").lower().split() if len(p) > 3]
            if partes and not any(p in tx for p in partes):
                continue
            setn = {"escritura_notaria_detectada_at": e.get("date") or _now()}
            hito_txt = "Actividad de notaría detectada"
            if FIRMA_REAL_RE.search(tx):
                setn["escritura_confirmada_at"] = e.get("date") or _now()
                setn["fecha_firma_detectada"] = str(e.get("date") or _now())[:10]
                hito_txt = "FIRMA DE ESCRITURA confirmada por correo de la notaría"
            ciudad = (fd_n.get("ciudad") or "").strip().lower()
            if ciudad and ciudad not in tx:
                setn["alerta_notaria_ciudad"] = True
            await db.folders.update_one({"id": fd_n["id"]}, {"$set": setn})
            await db.hitos_externos.insert_one({
                "id": str(uuid.uuid4()), "folder_id": fd_n["id"], "hito": "notaria",
                "asunto": asunto[:160], "fecha": e.get("date") or "", "fuente": "correo notaría",
                "direccion": de[:80], "creado": _now()})
            await _sync_adn(fd_n["id"])
            await _capturar_remitente(fd_n, "notaria", e)
            res["detalle"].append({"asunto": asunto[:100], "cliente": fd_n.get("nombre"),
                                   "hito": hito_txt, "match": "fuente notaría"})
            continue
        es_tasacion = ("valueproperty" in de or "valuedproperty" in de
                       or any(s in de for s in tas_src)
                       or ("tasaci" in asunto.lower()
                           and ("valueproperty" in texto.lower() or "valuedproperty" in texto.lower())))
        es_estudio = (any(r in de for r in REPARO_REMITENTES) or any(s in de for s in est_src)
                      or "estudio de titulo" in asunto.lower().replace("í", "i"))
        es_notaria = (any(s in de for s in ces_src)
                      or (any(k in asunto.lower() for k in NOTARIA_KW)
                          and (any(r in de for r in REPARO_REMITENTES) or "notar" in de)))
        asunto_l = asunto.lower()
        frase_set = "set firmado" in asunto_l or "set para la firma" in asunto_l
        es_set = frase_set and (not set_src or any(s in de for s in set_src))
        es_serviu = any(k in asunto_l for k in ("serviu", "subsidio", "resolucion", "resolución"))
        es_promesa = bool(re.search(r"(promesa|compromiso)\s+(de\s+)?compraventa|promesa\s+firmada|compromiso\s+firmado",
                                    f"{asunto_l} {cuerpo[:800].lower()}"))
        es_carta_doc = "carta oferta" in asunto_l
        es_cert_sub = bool(re.search(r"certificado.{0,14}subsidio", asunto_l))
        es_carta_pie = bool(re.search(r"carta\s+(de\s+)?pie", asunto_l))
        if not (es_tasacion or es_estudio or es_notaria or es_set or es_serviu or es_promesa
                or es_carta_doc or es_cert_sub or es_carta_pie):
            continue
        fd, metodo = _match_carpeta(texto, por_rut, folders)
        if not fd:
            res["sin_respaldo"] += 1
            res["detalle"].append({"asunto": asunto[:100], "estado": "Pendiente de Información",
                                   "motivo": "sin RUT ni nombre de cliente identificable"})
            continue
        hito_cap = ("tasacion" if es_tasacion else "estudio" if es_estudio else
                    "cesion" if es_notaria else "set_credito" if es_set else
                    "promesa" if (es_promesa or es_cert_sub or es_carta_pie) else
                    "carta_oferta" if es_carta_doc else "serviu")
        await _capturar_remitente(fd, hito_cap, e)
        if es_serviu and not (es_tasacion or es_estudio or es_notaria or es_set or es_promesa
                              or es_carta_doc or es_cert_sub or es_carta_pie):
            # CUENTA DE BARRIDO: resolución serviu que llega → marcado azul (verificación manual)
            if re.search(r"resoluci", asunto_l):
                await _marcar_doc_llegada(fd, "serviu", e)
            continue
        _raw_a = f"{de}|{asunto}|{e.get('date','')}".encode()
        clave = "aud-" + hashlib.sha256(_raw_a).hexdigest()
        _legacy_a = "aud-" + hashlib.md5(_raw_a, usedforsecurity=False).hexdigest()
        if await db.hitos_externos.find_one({"clave": {"$in": [clave, _legacy_a]}}):
            continue
        adjs = []  # rutas de PDF archivados de este correo (para el Hilo del Cliente)
        if es_set:
            # P9 — SET DE CRÉDITO: 'Set Para la Firma' = pendiente; 'Set Firmado' exige
            # verificación de firmas reales dentro del PDF adjunto
            if "set firmado" in asunto_l:
                firmado, evidencia = False, "sin adjunto PDF para verificar firmas"
                if e.get("id"):
                    try:
                        atts = await asyncio.to_thread(mail.fetch_attachments_by_id, e["id"])
                        for a in atts or []:
                            if (a.get("filename") or "").lower().endswith(".pdf") and a.get("content_bytes"):
                                firmado, evidencia = await asyncio.to_thread(
                                    _verificar_firmas_pdf, a["content_bytes"])
                                if firmado:
                                    break
                    except Exception as ex:
                        evidencia = f"error al abrir adjuntos ({str(ex)[:60]})"
                estado_set = "firmado" if firmado else "verificacion_pendiente"
            else:
                estado_set = "esperando_firma"
                evidencia = "asunto 'Set Para la Firma': pendiente de firma del cliente (NO firmado)"
            await db.folders.update_one({"id": fd["id"]}, {"$set": {
                "set_credito_estado": estado_set, "set_credito_asunto": asunto[:180],
                "set_credito_evidencia": evidencia, "set_credito_at": e.get("date") or _now()}})
            if e.get("id"):
                adjs = await _archivar_adjuntos(fd, e["id"], "SETCRED", subdir="99_otros")
            hito_n, res_k = f"Set de Crédito: {estado_set} ({evidencia[:80]})", "sets_detectados"
        elif es_promesa:
            # VERIFICACIÓN DE FIRMA (IA): solo verde con evidencia de firma de alta confianza
            estado_p, evidencia_p = await _verificar_compromiso_ia(fd, e)
            if e.get("id"):
                adjs = await _archivar_adjuntos(fd, e["id"], "PROMESA", subdir="99_otros")
            hito_n = f"Promesa de Compraventa: {estado_p} ({evidencia_p[:80]})"
            res_k = "promesas_detectadas"
        elif es_carta_doc or es_cert_sub or es_carta_pie:
            # CUENTA DE BARRIDO: documento solicitado que llega → 🔵 azul, jamás confirma sola
            hito_d = "carta_oferta" if es_carta_doc else "cert_subsidio" if es_cert_sub else "carta_pie"
            await _marcar_doc_llegada(fd, hito_d, e)
            if e.get("id"):
                adjs = await _archivar_adjuntos(fd, e["id"], hito_d.upper(), subdir="99_otros")
            hito_n = f"{_DOC_LABEL[hito_d]}: recibido — {_PEND_VERIF}"
            res_k = "docs_detectados"
        elif es_tasacion:
            await db.folders.update_one({"id": fd["id"]}, {"$set": {
                "tasacion_informe_recibido_at": e.get("date") or _now(),
                "tasacion_informe_asunto": asunto[:180]}})
            # MOTOR DE COSECHA (Regla #64): el dato queda en DashAI, sin re-consultar IMAP
            try:
                import perfil_consolidado as _pc
                await _pc.cosechar(fd["id"], {"fecha_informe_tasacion": e.get("date") or _now()},
                                   "informe_tasacion")
            except Exception:
                pass
            if e.get("id"):
                adjs = await _archivar_adjuntos(fd, e["id"], "TASACION", subdir="99_otros")
            hito_n, res_k = "Informe de Tasación Recibido", "tasaciones_detectadas"
        elif es_notaria:
            # RADAR ESCRITURACIÓN: envío a notaría / cesión / firma serie de créditos
            marcas = {"escritura_notaria_detectada_at": e.get("date") or _now()}
            if FIRMA_REAL_RE.search(texto):
                marcas["firma_cesion_confirmada_at"] = e.get("date") or _now()
            fecha_f = _capturar_fecha(texto)
            if fecha_f and not (await db.folders.find_one({"id": fd["id"]}, {"fecha_firma": 1}) or {}).get("fecha_firma"):
                marcas["fecha_firma_detectada"] = fecha_f
            await db.folders.update_one({"id": fd["id"]}, {"$set": marcas})
            if e.get("id"):
                adjs = await _archivar_adjuntos(fd, e["id"], "NOTARIA", subdir="99_otros")
            hito_n, res_k = "Escritura enviada a Notaría", "estudios_detectados"
        else:
            marcas = {"estudio_recibido_at": e.get("date") or _now()}
            if "reparo" in texto.lower():
                reparo = {"asunto": asunto[:180], "texto": cuerpo[:600], "de": de[:120],
                          "fecha": e.get("date", ""), "detectado": _now(), "origen": "auditoria_real"}
                await db.folders.update_one({"id": fd["id"]}, {"$push": {"reparos_alertas": reparo}})
                res["reparos_transcritos"] += 1
            await db.folders.update_one({"id": fd["id"]}, {"$set": marcas})
            if e.get("id"):
                adjs = await _archivar_adjuntos(fd, e["id"], "ESTUDIO")
                # REGLA DE MINADO: si el PDF del estudio trae 'Observaciones'/'Reparos',
                # el texto EXACTO se copia a la columna Detalle de Reparos (celda naranja)
                try:
                    n_min = await _minar_reparos_pdf(fd)
                    res["reparos_transcritos"] += n_min
                except Exception as ex:
                    logging.warning(f"minado reparos pdf: {ex}")
            hito_n, res_k = "Informe Estudio de Títulos Recibido", "estudios_detectados"
        # ALERTAS DE ACCIÓN: notificación automática a la Malla (Ejecutivos A y B) + hito ✅
        try:
            await db.alertas.insert_one({
                "id": str(uuid.uuid4()), "tipo": "malla_inteligencia",
                "destino": "Ejecutivos A y B", "cliente": fd.get("nombre") or "",
                "mensaje": f"✅ {hito_n}: {fd.get('nombre')} ({fd.get('rut') or 'sin RUT'}) — "
                           f"detectado en el correo y guardado en el EXPEDIENTE_360",
                "fecha": _now(), "leida": False})
        except Exception:
            pass
        # VINCULACIÓN EXPEDIENTE_360: el hallazgo queda para siempre en la Bóveda de ADN
        try:
            import adn_clientes as _adn
            fd_full = await db.folders.find_one({"id": fd["id"]})
            if fd_full:
                reg = _adn._registro_desde_folder(fd_full)
                reg["expediente_360"] = await _adn._expediente_360(fd_full)
                await _adn._upsert_adn(reg)
        except Exception:
            pass
        await db.hitos_externos.insert_one({
            "id": str(uuid.uuid4()), "clave": clave, "folder_id": fd["id"],
            "cliente": fd.get("nombre"), "rut": fd.get("rut") or "", "hito": hito_n,
            "fuente": "Value Property" if es_tasacion else "Abogados (Estudio de Título)",
            "dominio": "auditoria_real", "direccion": "recibido", "asunto": asunto[:180],
            "fecha": e.get("date", ""), "validado_rut": metodo == "rut", "match": metodo,
            "adjuntos": adjs, "creado": _now()})
        res[res_k] += 1
        res["detalle"].append({"asunto": asunto[:100], "cliente": fd.get("nombre"),
                               "hito": hito_n, "match": metodo})
    return res


@flujos.post("/auditoria-real")
async def flujos_auditoria_real(limit: int = 60, dias: int = 0):
    """Escaneo inmediato del correo (Regla #43) sobre los correos recientes de TODAS
    las cuentas. REGLA DE HIERRO: sin correo de respaldo → el hito queda 'Pendiente
    de Información' (jamás se inventan datos)."""
    correos = await asyncio.to_thread(mail.fetch_recent_full, min(max(limit, 10), 400))
    if dias > 0:
        from datetime import timedelta
        corte = (datetime.now(timezone.utc) - timedelta(days=dias)).isoformat()
        correos = [e for e in (correos or []) if (e.get("date") or "9999") >= corte[:10] or not e.get("date")]
    res = await _auditar_lote(correos)
    await db.config.update_one({"_key": "auditoria_real"}, {"$set": {
        "ultima": _now(), "resultado": {k: v for k, v in res.items() if k != "detalle"}}}, upsert=True)
    return res


BARRIDO_REMITENTES = ("valueproperty", "valuedproperty", "victoriavilches", "gmajluf",
                      "amvabogados", "mardluf", "majluf", "gmardones")


@flujos.post("/barrido-forzado")
async def flujos_barrido_forzado(dias: int = 60):
    """BARRIDO FORZADO: rastrea los últimos N días en TODAS las cuentas IMAP priorizando
    a Value Property, Victoria Vilches y Guillermo Majluf, y persiste cada hallazgo en
    la carpeta y el EXPEDIENTE_360 de la Bóveda ADN (Regla #67)."""
    fu = await _fuentes_documentos()
    senders = set(BARRIDO_REMITENTES)
    for lst in fu.values():
        senders.update(s.strip().lower() for s in (lst or []) if s)
    async for f in db.folders.find({"fuentes_doc": {"$exists": True}}, {"fuentes_doc": 1}):
        for lst in (f.get("fuentes_doc") or {}).values():
            senders.update(_correos_de(lst))
    async def _run():
        try:
            await db.config.update_one({"_key": "barrido_forzado"}, {"$set": {
                "estado": "en_proceso", "inicio": _now(), "dias": dias}}, upsert=True)
            correos = await asyncio.to_thread(mail.fetch_since_by_senders, dias, sorted(senders))
            res = await _auditar_lote(correos)
            await db.config.update_one({"_key": "barrido_forzado"}, {"$set": {
                "estado": "completado", "ultima": _now(), "dias": dias,
                "resultado": {k: v for k, v in res.items() if k != "detalle"}}}, upsert=True)
        except Exception as ex:
            logging.warning(f"barrido forzado: {ex}")
            await db.config.update_one({"_key": "barrido_forzado"}, {"$set": {
                "estado": f"error: {str(ex)[:120]}", "ultima": _now()}}, upsert=True)
    asyncio.create_task(_run())
    return {"ok": True, "lanzado": True, "dias": dias,
            "seguimiento": "GET /api/flujos/barrido-estado"}


@flujos.get("/barrido-estado")
async def flujos_barrido_estado():
    return await db.config.find_one({"_key": "barrido_forzado"}, {"_id": 0}) or {"estado": "nunca_ejecutado"}


# ── RADAR DE ESCRITURACIÓN Y CONTROL DE FIRMAS ──────────────────────────────
@flujos.get("/radar")
async def flujos_radar():
    out = []
    async for fd in db.folders.find({}).sort("nombre", 1):
        firmas, hito = firmas_folder(fd)
        out.append({"id": fd["id"], "cliente": fd.get("nombre"), "rut": fd.get("rut") or "",
                    "doc20": doc20_folder(fd), "firmas": firmas, "hito_firmas": hito,
                    "fecha_firma": fd.get("fecha_firma") or fd.get("fecha_firma_detectada") or "",
                    "fecha_firma_origen": "manual" if fd.get("fecha_firma") else ("dashai" if fd.get("fecha_firma_detectada") else "")})
    return {"radar": out, "total": len(out)}


@flujos.post("/firmas/{fid}")
async def flujos_marcar_firma(fid: str, payload: dict):
    rol = (payload.get("rol") or "").strip()
    estado = (payload.get("estado") or "").strip()
    if rol not in ("titular", "codeudor", "mandatario", "anexos") or estado not in ("pendiente", "firmado"):
        raise HTTPException(status_code=400, detail="Rol o estado de firma inválido")
    fd = await db.folders.find_one({"id": fid})
    if not fd:
        raise HTTPException(status_code=404, detail="Carpeta no existe")
    if estado == "firmado":
        # REGLA DE HIERRO #62: sin salida SMTP confirmada, el hito no se completa
        import monitor_envios as _me
        await _me.exigir_correo_ok(fd.get("nombre") or "")
    await db.folders.update_one({"id": fid}, {"$set": {f"firmas_log.{rol}": estado, "updated_at": _now()}})
    fd = await db.folders.find_one({"id": fid})
    firmas, hito = firmas_folder(fd)
    return {"ok": True, "firmas": firmas, "hito_firmas": hito,
            "regla_hierro": "Con codeudor, el hito solo es VERDE cuando AMBOS firman"}


@flujos.post("/fecha-firma/{fid}")
async def flujos_fecha_firma(fid: str, payload: dict):
    fecha = (payload.get("fecha") or "").strip()
    if fecha and not re.match(r"^\d{4}-\d{2}-\d{2}$", fecha):
        raise HTTPException(status_code=400, detail="Fecha inválida (AAAA-MM-DD)")
    r = await db.folders.update_one({"id": fid}, {"$set": {"fecha_firma": fecha, "updated_at": _now()}})
    if not r.matched_count:
        raise HTTPException(status_code=404, detail="Carpeta no existe")
    return {"ok": True, "fecha_firma": fecha, "origen": "manual"}


# ─── AUTOGESTIÓN DE CLAVES + SELLO DE ORIGEN (Regla de Oro #38) ─────────────
def _aes_key():
    """Clave AES-256 derivada del secreto del sistema (nunca sale del entorno)."""
    return hashlib.sha256(os.environ["JWT_SECRET"].encode()).digest()


def _enc_aes(texto):
    from cryptography.hazmat.primitives.ciphers.aead import AESGCM
    nonce = os.urandom(12)
    return base64.b64encode(nonce + AESGCM(_aes_key()).encrypt(nonce, texto.encode(), None)).decode()


def _dec_aes(b64):
    from cryptography.hazmat.primitives.ciphers.aead import AESGCM
    raw = base64.b64decode(b64)
    return AESGCM(_aes_key()).decrypt(raw[:12], raw[12:], None).decode()


def _probar_imap(host, email_u, clave):
    import imaplib
    import socket
    socket.setdefaulttimeout(20)
    m = imaplib.IMAP4_SSL(host)
    m.login(email_u, clave)
    m.logout()


@micorreo.get("")
async def micorreo_estado(request: Request):
    c = _claims(request)
    cfg = await db.correos_ejecutivos.find_one({"codigo": c.get("sub") or ""},
                                               {"_id": 0, "cred_enc": 0}) or {}
    return {"configurado": bool(cfg), **cfg}


@micorreo.post("/configurar")
async def micorreo_configurar(payload: dict, request: Request):
    """ONBOARDING: valida IMAP y clave de aplicación SIN intervención del administrador.
    La credencial se guarda cifrada con AES-256 (Regla de Oro #38)."""
    c = _claims(request)
    em = (payload.get("email") or "").strip().lower()
    clave = (payload.get("app_password") or "").strip()
    if not em or not EMAIL_RE.match(em) or not clave:
        raise HTTPException(status_code=400, detail="Correo y clave de aplicación son obligatorios")
    imap_host = (payload.get("imap_host") or "imap.gmail.com").strip()
    smtp_host = (payload.get("smtp_host") or "smtp.gmail.com").strip()
    try:
        await asyncio.to_thread(_probar_imap, imap_host, em, clave)
    except Exception:
        raise HTTPException(status_code=400,
            detail="⚠️ Su conexión de correo necesita actualización — verifique el correo y la clave de aplicación")
    await db.correos_ejecutivos.update_one({"codigo": c.get("sub") or ""}, {"$set": {
        "codigo": c.get("sub") or "", "nombre": c.get("nombre") or "",
        "email": em, "imap_host": imap_host, "smtp_host": smtp_host,
        "cred_enc": _enc_aes(clave), "cifrado": "AES-256-GCM",
        "estado": "ok", "alerta_enviada": False, "configurado_en": _now()}}, upsert=True)
    return {"ok": True, "email": em, "estado": "ok",
            "nota": "Credencial validada y guardada con cifrado AES-256"}


@micorreo.post("/revelar")
async def micorreo_revelar(payload: dict, request: Request):
    """Acceso a la credencial: SOLO el dueño del módulo o Gerardo vía PIN 0586."""
    c = _claims(request)
    propio = c.get("sub") or ""
    objetivo = (payload.get("codigo") or propio).strip()
    if objetivo != propio:
        pin = (payload.get("pin") or "").strip()
        if c.get("rol") not in ("admin", "maestro") or pin != os.environ.get("MASTER_PIN", "!"):
            raise HTTPException(status_code=403,
                detail="Solo el dueño del módulo o el administrador con PIN maestro pueden acceder (Regla #38)")
    cfg = await db.correos_ejecutivos.find_one({"codigo": objetivo})
    if not cfg:
        raise HTTPException(status_code=404, detail="Sin configuración de correo")
    return {"email": cfg.get("email"), "imap_host": cfg.get("imap_host"),
            "smtp_host": cfg.get("smtp_host"), "app_password": _dec_aes(cfg["cred_enc"])}


async def lector_ejecutivos_loop():
    """MOTOR DE LECTURA SOSTENIBLE (Regla #38): procesa el buzón de cada ejecutivo de
    forma independiente, con pausas de seguridad. REGLA DE HIERRO: si una clave falla,
    se notifica al ejecutivo — jamás se detiene el Maserati."""
    await asyncio.sleep(120)
    while True:
        try:
            async for cfg in db.correos_ejecutivos.find({}):
                try:
                    clave = _dec_aes(cfg["cred_enc"])
                    await asyncio.to_thread(_probar_imap, cfg.get("imap_host", "imap.gmail.com"),
                                            cfg.get("email", ""), clave)
                    await db.correos_ejecutivos.update_one({"codigo": cfg["codigo"]}, {"$set": {
                        "estado": "ok", "alerta_enviada": False, "ultima_lectura": _now()}})
                except Exception:
                    if not cfg.get("alerta_enviada"):
                        await db.alertas.insert_one({"id": str(uuid.uuid4()), "tipo": "correo_ejecutivo",
                            "mensaje": f"⚠️ Su conexión de correo necesita actualización — {cfg.get('nombre','')} ({cfg.get('email','')})",
                            "fecha": _now(), "leida": False})
                    await db.correos_ejecutivos.update_one({"codigo": cfg["codigo"]}, {"$set": {
                        "estado": "requiere_actualizacion", "alerta_enviada": True}})
                await asyncio.sleep(5)  # pausa de seguridad anti-bloqueo de servidores externos
        except Exception as e:
            logging.warning(f"lector ejecutivos: {e}")
        await asyncio.sleep(600)


# ── BUZÓN DE APRENDIZAJE (2º IMAP, SOLO LECTURA) — entrena DashAI ───────────
_CLASIF = [("tasac", "tasacion"), ("reparo", "reparo"), ("estudio", "estudio_titulo"),
           ("notar", "notaria"), ("escritura", "escrituracion"), ("subsidio", "subsidio")]


def _leer_buzon_ro(host, email_u, clave, limit=25):
    """Lectura SOLO LECTURA (readonly + BODY.PEEK): jamás marca ni altera correos."""
    import imaplib
    import socket
    import email as _em
    from email.header import decode_header
    socket.setdefaulttimeout(25)
    m = imaplib.IMAP4_SSL(host)
    m.login(email_u, clave)
    m.select("INBOX", readonly=True)
    _, data = m.search(None, "ALL")
    ids = (data[0].split() or [])[-limit:]
    out = []
    for i in reversed(ids):
        try:
            _, md = m.fetch(i, "(BODY.PEEK[HEADER.FIELDS (FROM SUBJECT DATE)])")
            msg = _em.message_from_bytes(md[0][1])
            def _dec(v):
                try:
                    return " ".join(t.decode(e or "utf-8", "ignore") if isinstance(t, bytes) else t
                                    for t, e in decode_header(v or ""))
                except Exception:
                    return v or ""
            out.append({"de": _dec(msg.get("From")), "asunto": _dec(msg.get("Subject")),
                        "fecha": msg.get("Date", "")})
        except Exception:
            continue
    m.logout()
    return out


@buzon.get("")
async def buzon_estado():
    cfg = await db.config.find_one({"_key": "buzon_aprendizaje"}, {"_id": 0, "cred_enc": 0}) or {}
    total = await db.buzon_aprendizaje.count_documents({})
    return {"configurado": bool(cfg.get("email")), "email": cfg.get("email", ""),
            "ingeridos": total, "ultima_lectura": cfg.get("ultima_lectura", ""),
            "estado": cfg.get("estado", ""), "modo": "solo_lectura"}


@buzon.post("/configurar")
async def buzon_configurar(payload: dict):
    em = (payload.get("email") or "").strip().lower()
    clave = (payload.get("app_password") or "").strip()
    host = (payload.get("imap_host") or "imap.gmail.com").strip()
    if not em or not EMAIL_RE.match(em) or not clave:
        raise HTTPException(status_code=400, detail="Correo y clave de aplicación son obligatorios")
    try:
        await asyncio.to_thread(_probar_imap, host, em, clave)
    except Exception:
        raise HTTPException(status_code=400, detail="⚠️ No fue posible conectar: verifique el correo y la clave de aplicación")
    await db.config.update_one({"_key": "buzon_aprendizaje"}, {"$set": {
        "email": em, "imap_host": host, "cred_enc": _enc_aes(clave),
        "estado": "ok", "configurado_en": _now()}}, upsert=True)
    return {"ok": True, "email": em, "modo": "solo_lectura",
            "nota": "DashAI leerá este buzón sin marcar ni mover correos (AES-256)"}


async def buzon_aprendizaje_loop():
    """Cada 15 min ingiere encabezados del buzón de aprendizaje para entrenar DashAI."""
    await asyncio.sleep(150)
    while True:
        try:
            cfg = await db.config.find_one({"_key": "buzon_aprendizaje"}) or {}
            if cfg.get("email") and cfg.get("cred_enc"):
                try:
                    correos = await asyncio.to_thread(_leer_buzon_ro, cfg.get("imap_host", "imap.gmail.com"),
                                                      cfg["email"], _dec_aes(cfg["cred_enc"]))
                    nuevos = 0
                    for c in correos or []:
                        _raw_b = f"{c['de']}|{c['asunto']}|{c['fecha']}".encode()
                        clave_h = hashlib.sha256(_raw_b).hexdigest()
                        _legacy_b = hashlib.md5(_raw_b, usedforsecurity=False).hexdigest()
                        if await db.buzon_aprendizaje.find_one({"clave": {"$in": [clave_h, _legacy_b]}}):
                            continue
                        t = c["asunto"].lower()
                        clasif = next((v for k, v in _CLASIF if k in t), "otro")
                        m = RUT_RE.search(c["asunto"])
                        await db.buzon_aprendizaje.insert_one({
                            "id": str(uuid.uuid4()), "clave": clave_h, **c,
                            "clasificacion": clasif, "rut_detectado": m.group(1) if m else "",
                            "ingerido": _now()})
                        nuevos += 1
                    await db.config.update_one({"_key": "buzon_aprendizaje"}, {"$set": {
                        "estado": "ok", "ultima_lectura": _now(), "ultimos_nuevos": nuevos}})
                except Exception as e:
                    await db.config.update_one({"_key": "buzon_aprendizaje"}, {"$set": {"estado": "requiere_actualizacion"}})
                    logging.warning(f"buzón aprendizaje: {e}")
        except Exception as e:
            logging.warning(f"buzón loop: {e}")
        await asyncio.sleep(900)


# ── SUPERCARPETA DE MANAGEMENT (Regla de Oro #55) — vista de control primario ─
def _informes_folder(fd):
    base = fsvc.folder_dir(fd.get("nombre") or "")
    inf = {"tasacion": None, "estudio": None, "borrador": None}
    if base.exists():
        for p in sorted(base.rglob("*.pdf"), key=lambda x: -x.stat().st_mtime):
            n = p.name.lower()
            rel = f"{p.parent.name}/{p.name}" if p.parent != base else p.name
            mt = datetime.fromtimestamp(p.stat().st_mtime, tz=timezone.utc).isoformat()
            if not inf["tasacion"] and "tasac" in n:
                inf["tasacion"] = {"disponible": True, "archivo": rel, "fecha": mt}
            elif not inf["estudio"] and (p.parent.name == "07_estudio_titulo" or "estudio" in n):
                inf["estudio"] = {"disponible": True, "archivo": rel, "fecha": mt}
            elif not inf["borrador"] and ("escritura" in n or "borrador" in n):
                inf["borrador"] = {"disponible": True, "archivo": rel, "fecha": mt}
    for k in inf:
        if not inf[k]:
            inf[k] = {"disponible": False, "archivo": "", "fecha": ""}
    return inf


def _estado_tasacion(fd):
    if fd.get("tasacion_informe_recibido_at"):
        return "Informe Recibido"
    if fd.get("tasacion_fecha"):
        return "Visita"
    if (fd.get("reclamos_gerencia") or {}).get("tasacion") or fd.get("tasacion_solicitada_at"):
        return "Solicitada"
    return "Pendiente"


def _estado_estudio(fd):
    reparos = [r for r in (fd.get("reparos_alertas") or []) if not r.get("resuelto")]
    if reparos:
        return "Con Reparos", " | ".join((r.get("texto") or r.get("asunto") or "")[:300] for r in reparos[:4])
    if fd.get("estudio_titulo_terminado_at"):
        return "Aprobado", ""
    return "En Proceso", ""


def _estado_cesion(fd):
    """REGLA DE HIERRO: 'Confirmada' SOLO con evidencia real de firma detectada en un
    correo de la fuente configurada. Actividad de notaría o inferencias NO confirman."""
    if fd.get("firma_cesion_confirmada_at") or fd.get("escritura_confirmada_at"):
        return "Confirmada"
    return "Pendiente"


def _parse_fecha(s):
    from email.utils import parsedate_to_datetime
    s = str(s or "")
    try:
        return datetime.fromisoformat(s.replace("Z", "+00:00"))
    except Exception:
        try:
            return parsedate_to_datetime(s)
        except Exception:
            return None


async def _bitacora_hito(fd, hito):
    """BITÁCORA DE TIEMPOS: respaldo de fecha/hora de cada solicitud 'En Proceso'.
    REGLA DE HIERRO: sin registro de solicitud → 'ERROR DE SEGUIMIENTO'."""
    nombre = (fd.get("nombre") or "").strip()
    kw = "TASACION" if hito == "tasacion" else "ESTUDIO"
    rx_n = {"$regex": re.escape(nombre[:18]), "$options": "i"}
    envio = await db.correos_smtp_log.find_one(
        {"success": True, "$and": [{"subject": rx_n},
                                   {"subject": {"$regex": kw, "$options": "i"}}]},
        sort=[("fecha", -1)])
    seg = None
    hx = None
    if not envio:
        seg = await db.seguimiento.find_one(
            {"$and": [{"$or": [{"cliente": rx_n}, {"asunto": rx_n}]},
                      {"asunto": {"$regex": kw, "$options": "i"}}]}, sort=[("fecha", -1)])
    if not envio and not seg:
        hx = await db.hitos_externos.find_one(
            {"$and": [{"$or": [{"cliente": rx_n}, {"asunto": rx_n}]},
                      {"asunto": {"$regex": kw, "$options": "i"}}]}, sort=[("fecha", -1)])
    rec = (fd.get("reclamos_gerencia") or {}).get(hito) or {}
    fecha_raw = ((envio or {}).get("fecha") or (seg or {}).get("fecha") or (hx or {}).get("fecha")
                 or rec.get("fecha") or rec.get("en") or "")
    if not fecha_raw:
        return {"hito": hito, "error_seguimiento": True, "estado": "ERROR DE SEGUIMIENTO",
                "detalle": "Regla de Hierro: no hay registro de cuándo se solicitó este hito"}
    destinatario = ((envio or {}).get("to")
                    or ((seg or {}).get("de") or "registro de seguimiento" if seg else "")
                    or ((hx or {}).get("fuente") or "radar de correos" if hx else "")
                    or ("Gerencia Comercial (reclamo manual)" if rec else ""))
    resumen = ((envio or {}).get("subject") or (seg or {}).get("asunto")
               or (hx or {}).get("asunto") or rec.get("hito") or "")
    respondido_at = (fd.get("tasacion_informe_recibido_at") if hito == "tasacion"
                     else fd.get("estudio_recibido_at") or fd.get("estudio_titulo_terminado_at"))
    dt = _parse_fecha(fecha_raw)
    horas = dias = None
    if dt:
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        delta = datetime.now(timezone.utc) - dt
        horas = round(delta.total_seconds() / 3600, 1)
        dias = delta.days
    return {"hito": hito, "error_seguimiento": False,
            "fecha_solicitud": str(fecha_raw)[:19],
            "destinatario": destinatario, "resumen": resumen[:220],
            "fuente": "correo SMTP" if envio else ("seguimiento" if seg else "huella de gestión"),
            "dias_transcurridos": dias, "horas_transcurridas": horas,
            "respondido": bool(respondido_at), "respondido_at": str(respondido_at or "")[:19],
            "demora_48h": bool(horas is not None and horas > 48 and not respondido_at)}


@supercarpeta.get("/bitacora/{fid}")
async def supercarpeta_bitacora(fid: str, hito: str = "tasacion"):
    fd = await db.folders.find_one({"id": fid})
    if not fd:
        raise HTTPException(status_code=404, detail="Carpeta no existe")
    b = await _bitacora_hito(fd, "tasacion" if hito == "tasacion" else "estudio")
    b["cliente"] = fd.get("nombre") or ""
    return b


@supercarpeta.get("/flota")
async def flota_ver():
    cfg = await db.config.find_one({"_key": "flota_agosto"}, {"_id": 0}) or {}
    return {"activo": bool(cfg.get("activo")), "nombres": cfg.get("nombres") or [],
            "definida": _now() if cfg else None}


@supercarpeta.post("/flota")
async def flota_definir(request: Request, payload: dict):
    """PURGA TOTAL (Flota Agosto): solo el administrador define el universo de trabajo.
    Bloquea el ingreso de nuevos prospectos a las vistas sin autorización expresa."""
    user = getattr(request.state, "user", {}) or {}
    if (user.get("rol") or "") not in ("admin", "maestro"):
        raise HTTPException(status_code=403, detail="Solo el administrador define la Flota Agosto")
    nombres = [str(n).strip().upper() for n in (payload.get("nombres") or []) if str(n).strip()]
    activo = bool(payload.get("activo", True)) and len(nombres) > 0
    await db.config.update_one({"_key": "flota_agosto"}, {"$set": {
        "nombres": nombres, "activo": activo, "definida": _now(),
        "por": user.get("sub") or "admin"}}, upsert=True)
    return {"ok": True, "activo": activo, "total": len(nombres),
            "regla_hierro": "Solo existen estos clientes en las vistas activas hasta nueva autorización"}


_AVANCE_OK_RE = re.compile(
    r"(aprobad|firmad[ao]|verificado\b|informe recibido|recibid[ao]|confirmad[ao]|limpio|inscrita)", re.I)


async def _avance_notificar(mes_sel, clientes, pct_global):
    """Notifica a Rodrigo Ibáñez (100% por cliente) y a Gerencia General (hitos 50/75/100%)."""
    try:
        cfg = await db.config.find_one({"_key": f"avance_hitos_{mes_sel}"}) or {}
        upd, notif100 = {}, list(cfg.get("clientes_100") or [])
        for c in clientes:
            if (c.get("avance") or {}).get("pct", 0) >= 100 and c["id"] not in notif100:
                await db.alertas.insert_one({"id": str(uuid.uuid4()), "tipo": "avance_100",
                    "mensaje": f"🏆 Cliente {c['cliente']} ha completado todas las etapas. "
                               f"Crédito listo para escriturar. UF: {c.get('monto_uf') or 0}",
                    "fecha": _now(), "leida": False, "destino": "gerencia"})
                notif100.append(c["id"])
                upd["clientes_100"] = notif100
        for hito in (50, 75, 100):
            if pct_global >= hito and not cfg.get(f"hito_{hito}"):
                await db.alertas.insert_one({"id": str(uuid.uuid4()), "tipo": "avance_global",
                    "mensaje": f"📈 Proyección {mes_sel}: el cumplimiento global superó el {hito}% "
                               f"(actual: {pct_global}%). Hito informado a Gerencia General.",
                    "fecha": _now(), "leida": False, "destino": "gerencia_general"})
                upd[f"hito_{hito}"] = _now()
        if upd:
            await db.config.update_one({"_key": f"avance_hitos_{mes_sel}"}, {"$set": upd}, upsert=True)
    except Exception as e:
        logging.warning(f"avance notificar: {e}")


async def avance_snapshot_loop():
    """HISTÓRICO DE AVANCE: foto diaria del % de cada cliente, guardada en ADN_CLIENTES_360."""
    import adn_clientes as _adn
    await asyncio.sleep(300)
    while True:
        try:
            hoy = _now()[:10]
            cfg = await db.config.find_one({"_key": "avance_snapshot"}) or {}
            if cfg.get("dia") != hoy:
                meses = {m for m in await db.folders.distinct("mes_proyeccion") if m} | {"2026-08"}
                for mes_s in meses:
                    data = await supercarpeta_vista(mes_s)
                    for c in data.get("clientes", []):
                        if c.get("rut"):
                            await db.adn_clientes_360.update_one(
                                {"rut_norm": _adn._norm_rut(c["rut"])},
                                {"$push": {"avance_historico": {"fecha": hoy, "mes": mes_s,
                                           "pct": (c.get("avance") or {}).get("pct", 0)}}})
                await db.config.update_one({"_key": "avance_snapshot"},
                                           {"$set": {"dia": hoy, "en": _now()}}, upsert=True)
        except Exception as e:
            logging.warning(f"avance snapshot: {e}")
        await asyncio.sleep(3600)


@supercarpeta.get("")
async def supercarpeta_vista(mes: str = ""):
    """Regla #55 (V3): consulta directa a la Bóveda ADN_CLIENTES_360 — sin escanear PDFs físicos.
    NAVEGACIÓN MENSUAL: ?mes=YYYY-MM muestra la proyección de ese mes (bóveda compartida)."""
    from datetime import timedelta
    import adn_clientes as _adn
    from base_historica import validar_rut_chileno as _vrut, _fmt_rut
    mes_cal = datetime.now(timezone.utc).strftime("%Y-%m")
    mes_sel = (mes or "").strip() or "2026-08"
    limite24 = (datetime.now(timezone.utc) - timedelta(hours=24)).isoformat()
    flota_cfg = await db.config.find_one({"_key": "flota_agosto"}) or {}
    flota = ({n.strip().upper() for n in (flota_cfg.get("nombres") or [])}
             if flota_cfg.get("activo") and mes_sel == "2026-08" else set())
    adn_map = {}
    async for r in db.adn_clientes_360.find({}, {"rut_norm": 1, "rut": 1, "expediente_360": 1,
                                                 "actualizado": 1, "identidad": 1, "propiedad": 1,
                                                 "financiero": 1, "origen": 1}):
        adn_map[r.get("rut_norm")] = r
    # ── ALERTA HILO FRÍO: último evento del hilo (recibidos) por carpeta ──
    hilo_recibidos = {}
    async for h in db.hitos_externos.aggregate([{"$group": {"_id": "$folder_id", "ult": {"$max": "$creado"}}}]):
        if h.get("_id"):
            hilo_recibidos[h["_id"]] = str(h.get("ult") or "")
    ahora_dt = datetime.now(timezone.utc)
    clientes = []
    async for fd in db.folders.find({}).sort("nombre", 1):
        nombre_u = (fd.get("nombre") or "").strip().upper()
        if fd.get("oculto_supercarpeta"):
            continue  # P10: eliminado de la vista por Gerencia (ficha ADN intacta)
        if (fd.get("mes_proyeccion") or "2026-08") != mes_sel:
            continue  # NAVEGACIÓN MENSUAL: cada mes muestra solo su proyección
        if flota:
            # REGLA DE HIERRO (Flota Agosto): solo existe la flota autorizada
            if not any(nf in nombre_u or nombre_u in nf for nf in flota):
                continue
        elif mes_sel == "2026-08":
            act = (fd.get("updated_at") or fd.get("created") or fd.get("created_at") or "")[:7]
            if act != mes_cal and not (fd.get("datos_financieros") or {}).get("monto_credito"):
                continue
        reg = adn_map.get(_adn._norm_rut(fd.get("rut"))) or {}
        exp = reg.get("expediente_360") or {}
        docs = exp.get("documentos") or []

        def _doc(kws, preferir="informe"):
            cands = [d for d in docs if any(k in (d.get("archivo") or "").lower() for k in kws)]
            cands.sort(key=lambda d: 0 if preferir in (d.get("archivo") or "").lower() else 1)
            return cands[0] if cands else None

        inf = {}
        for k, sel in (("tasacion", _doc(["tasac"])), ("estudio", _doc(["estudio"])),
                       ("borrador", _doc(["escritura", "borrador"], preferir="borrador"))):
            inf[k] = ({"disponible": True, "archivo": sel.get("archivo") or "",
                       "fecha": reg.get("actualizado") or ""} if sel
                      else {"disponible": False, "archivo": "", "fecha": ""})
        est_estudio, detalle_reparos = _estado_estudio(fd)
        hl = exp.get("hitos_legales") or {}
        if not detalle_reparos and hl.get("reparos"):
            detalle_reparos = hl["reparos"]
            est_estudio = "Con Reparos"
        # ── P7: ESTADOS MANUALES (Gerencia) con trazabilidad y detección de conflicto ──
        em = fd.get("estados_manuales") or {}
        est_tas = _estado_tasacion(fd)
        est_ces = _estado_cesion(fd)
        man = {"tasacion": False, "estudio": False, "cesion": False, "set_credito": False}
        conf = {"tasacion": False, "estudio": False, "cesion": False, "set_credito": False}
        auto_marks = {"tasacion": str(fd.get("tasacion_informe_recibido_at") or ""),
                      "estudio": str(fd.get("estudio_titulo_terminado_at")
                                     or fd.get("estudio_recibido_at") or ""),
                      "cesion": str(fd.get("firma_cesion_confirmada_at")
                                    or fd.get("escritura_confirmada_at") or ""),
                      "set_credito": str(fd.get("set_credito_at") or ""),
                      "serviu": "", "promesa": str(fd.get("promesa_verificada_at") or ""), "carta_oferta": "",
                      "carpeta_notaria": str(fd.get("escritura_notaria_detectada_at") or ""),
                      "escritura": str(fd.get("escritura_confirmada_at") or "")}
        set_est = fd.get("set_credito_estado") or ""
        pv = fd.get("promesa_verificacion") or {}
        est_serviu = "Pendiente"
        est_promesa = pv.get("estado") or "Pendiente"
        est_carta = "Pendiente"
        est_carpeta = "Enviada" if fd.get("escritura_notaria_detectada_at") else "Pendiente"
        est_escritura = ("Firmada" if fd.get("escritura_confirmada_at")
                         else "Agendada" if (fd.get("fecha_firma") or fd.get("fecha_firma_detectada"))
                         else "Pendiente")
        for h in ("tasacion", "estudio", "cesion", "set_credito",
                  "serviu", "promesa", "carta_oferta", "carpeta_notaria", "escritura"):
            mh = em.get(h) or {}
            if mh.get("estado"):
                man[h] = True
                conf[h] = bool(auto_marks[h] and auto_marks[h] > str(mh.get("en") or "")
                               and not mh.get("resuelto"))
                if h == "tasacion":
                    est_tas = mh["estado"]
                elif h == "estudio":
                    est_estudio = mh["estado"]
                elif h == "cesion":
                    est_ces = mh["estado"]
                elif h == "set_credito":
                    set_est = mh["estado"]
                elif h == "serviu":
                    est_serviu = mh["estado"]
                elif h == "carta_oferta":
                    est_carta = mh["estado"]
                elif h == "promesa":
                    est_promesa = mh["estado"]
                elif h == "carpeta_notaria":
                    est_carpeta = mh["estado"]
                else:
                    est_escritura = mh["estado"]
        estado_legal = ("⚠️ Con Reparos" if est_estudio == "Con Reparos"
                        else "✅ Limpio" if est_estudio in ("Aprobado", "Aprobada") else "⏳ En Proceso")
        # ── P1/P5: la Bóveda ADN manda — cada campo se lee primero de ADN_CLIENTES_360 ──
        prop_adn = reg.get("propiedad") or {}
        exp_prop = exp.get("propiedad") or {}
        ident = reg.get("identidad") or {}
        fin_adn = reg.get("financiero") or {}
        monto_v = fd.get("proyeccion_uf") or fin_adn.get("monto_credito_uf")
        sub_adn = fin_adn.get("con_subsidio")
        subsidio_v = (fd.get("subsidio_proyeccion")
                      or ("Con Subsidio" if sub_adn is True else "Sin Subsidio" if sub_adn is False else ""))
        rut_v = fd.get("rut") or reg.get("rut") or ""
        tipo_op = (fd.get("tipo_operacion") or prop_adn.get("tipo_operacion") or "").lower()
        inmo_dato = (fd.get("inmobiliaria") or prop_adn.get("inmobiliaria")
                     or exp_prop.get("inmobiliaria") or "").strip()
        if "usad" in tipo_op:
            inmobiliaria = "Casa Usada"
        elif inmo_dato:
            inmobiliaria = inmo_dato
        else:
            inmobiliaria = "Directa"
        if flota:
            broker_v = "Mutuaria y Leasing Limitada"
        else:
            broker_v = (fd.get("broker_origen") or fd.get("broker_nombre")
                        or (reg.get("origen") or {}).get("broker_origen") or "").strip()
        proyecto_v = (fd.get("proyecto") or prop_adn.get("proyecto") or exp_prop.get("proyecto")
                      or _limpiar_proyecto(reg.get("nombre_proyecto"))
                      or (fd.get("perfil_consolidado") or {}).get("proyecto") or "").strip()
        if not proyecto_v and "usad" in tipo_op:
            proyecto_v = (exp_prop.get("direccion") or prop_adn.get("direccion")
                          or (fd.get("perfil_consolidado") or {}).get("direccion") or "").strip()
        ciudad_v = (fd.get("ciudad") or ident.get("ciudad")
                    or (fd.get("perfil_consolidado") or {}).get("ciudad")
                    or exp_prop.get("comuna") or prop_adn.get("comuna") or "").strip()
        notaria_v = (fd.get("notaria") or exp_prop.get("notaria") or "").strip()
        # ── P6: FALTANTES = alerta de fallo de cosecha → botón de ingreso manual ──
        faltantes = []
        if not _vrut(rut_v):
            faltantes.append("rut")
        if not inmo_dato and "usad" not in tipo_op:
            faltantes.append("inmobiliaria")
        if not broker_v:
            faltantes.append("broker")
        if not monto_v:
            faltantes.append("monto")
        if not proyecto_v:
            faltantes.append("proyecto")
        if not ciudad_v:
            faltantes.append("ciudad")
        # BITÁCORA rápida por hito pendiente (respaldo de fecha o ERROR DE SEGUIMIENTO)
        bit = {}
        try:
            if _estado_tasacion(fd) != "Informe Recibido":
                bit["tasacion"] = await _bitacora_hito(fd, "tasacion")
            if est_estudio not in ("Aprobado", "Aprobada"):
                bit["estudio"] = await _bitacora_hito(fd, "estudio")
        except Exception:
            bit = {}
        fechas = [v["fecha"] for v in inf.values() if v["disponible"]]
        # ── AVANCE POR CLIENTE (Regla de Hierro): solo etapas realmente completadas ──
        con_sub_av = "con" in subsidio_v.lower()
        etapas_av = [
            ("docs", "Documentos Comerciales recibidos y actualizados", 15,
             bool(fd.get("datos_financieros_ocr_fecha"))),
            ("tasacion", "Tasación aprobada", 15, bool(_AVANCE_OK_RE.search(est_tas or ""))),
            ("estudio", "Estudio de Títulos aprobado", 20, est_estudio in ("Aprobado", "Aprobada")),
            ("serviu", "Resolución Serviu aprobada", 15, bool(_AVANCE_OK_RE.search(est_serviu or ""))),
            ("set", "Cédula de Crédito (SET) firmada y verificada", 15,
             bool(re.search(r"firmad|verificado\b|aprobad", set_est or "", re.I))
             and "esperando" not in (set_est or "").lower()),
            ("borrador", "Borrador de Escritura listo en Notaría", 10,
             est_carpeta == "Enviada" or inf["borrador"]["disponible"]),
            ("escritura", "Firma de Escritura en Notaría", 10, est_escritura == "Firmada"),
        ]
        if not con_sub_av:
            etapas_av = [e for e in etapas_av if e[0] != "serviu"]
        peso_tot = sum(e[2] for e in etapas_av) or 1
        avance_pct = round(sum(e[2] for e in etapas_av if e[3]) * 100 / peso_tot)
        avance = {"pct": avance_pct,
                  "etapas": [{"clave": k, "etapa": lb, "peso": round(p * 100 / peso_tot, 1),
                              "completada": ok_} for k, lb, p, ok_ in etapas_av]}
        # ── ALERTA HILO FRÍO: más de 7 días sin movimiento en el hilo de correos ──
        _evs_hilo = ([hilo_recibidos.get(fd["id"], "")] +
                     [str(r.get("en") or "") for r in (fd.get("bitacora_solicitudes") or [])])
        _ult_hilo = max([e for e in _evs_hilo if e] or [""])
        _ref_frio = _ult_hilo or str(fd.get("created_at") or fd.get("created") or "")
        try:
            hilo_frio = (ahora_dt - datetime.fromisoformat(_ref_frio[:19])
                         .replace(tzinfo=timezone.utc)).days > 7
        except (ValueError, TypeError):
            hilo_frio = False
        clientes.append({"id": fd["id"], "cliente": fd.get("nombre"),
                         "avance": avance,
                         "hilo_frio": hilo_frio,
                         "hilo_ultimo": _ult_hilo[:10],
                         "fecha_nacimiento": fd.get("fecha_nacimiento") or "",
                         "rut": _fmt_rut(rut_v) if _vrut(rut_v) else (rut_v or ""),
                         "manual_identidad": list((fd.get("ingreso_manual") or {}).keys()),
                         "arrastre": fd.get("arrastre_desde") or "",
                         "inmobiliaria": inmobiliaria,
                         "proyecto": proyecto_v,
                         "ciudad": ciudad_v or "Por Confirmar",
                         "notaria": notaria_v,
                         "alerta_notaria_ciudad": bool(fd.get("alerta_notaria_ciudad")),
                         "alerta_reparos": bool(fd.get("alerta_reparos_sin_procesar")),
                         "broker": broker_v or "",
                         "broker_origen": broker_v or inmobiliaria,
                         "monto_uf": monto_v,
                         "subsidio": subsidio_v,
                         "reactivacion": bool(fd.get("reactivacion")),
                         "faltantes": faltantes,
                         "contacto": {"email": ident.get("email") or "",
                                      "telefono": ident.get("telefono") or ""},
                         "informes": inf,
                         "estado_tasacion": est_tas,
                         "estudio_titulos": est_estudio,
                         "estado_legal": estado_legal,
                         "detalle_reparos": detalle_reparos,
                         "cesion": est_ces,
                         "serviu": est_serviu, "promesa": est_promesa,
                         "promesa_ia": ({"evidencia": pv.get("evidencia") or "",
                                         "archivo": pv.get("archivo") or "",
                                         "firmado": bool(pv.get("firmado")),
                                         "fecha": str(pv.get("fecha") or "")[:10]} if pv else None),
                         "carta_oferta": est_carta,
                         "docs_co_rs": _marcado_documentos(fd, "con" in subsidio_v.lower(),
                                                           bool(fd.get("co_rs_reenviado_at"))),
                         "carpeta_notaria": est_carpeta, "escritura": est_escritura,
                         "fecha_firma": fd.get("fecha_firma") or fd.get("fecha_firma_detectada") or "",
                         "con_subsidio": "con" in subsidio_v.lower(),
                         "set_credito": {"estado": set_est,
                                         "evidencia": fd.get("set_credito_evidencia") or "",
                                         "asunto": fd.get("set_credito_asunto") or "",
                                         "fecha": str(fd.get("set_credito_at") or "")[:19]},
                         "manual": man, "conflicto": conf,
                         "bitacora": bit,
                         "notas": sorted([{**n, "hito": h}
                                          for h, lst in (fd.get("notas_estados") or {}).items()
                                          for n in (lst or [])],
                                         key=lambda n: n.get("en") or "", reverse=True),
                         "en_adn": bool(reg),
                         "resumen_hilo": {"texto": (fd.get("resumen_hilo") or {}).get("texto") or "",
                                          "en": (fd.get("resumen_hilo") or {}).get("en") or ""},
                         "recien_24h": any(f >= limite24 for f in fechas)})
    if mes_sel == "2026-08":
        meta_cfg = await db.config.find_one({"_key": "proyeccion_agosto"}) or {}
        meta_uf = meta_cfg.get("meta_uf") or 41717
    else:
        meta_cfg = await db.config.find_one({"_key": f"proyeccion_{mes_sel}"})
        if meta_cfg is None:
            # PROYECCIÓN DEL MES: estructura creada automáticamente, lista para recibir clientes
            await db.config.update_one({"_key": f"proyeccion_{mes_sel}"},
                                       {"$setOnInsert": {"creada": _now(), "meta_uf": 0}}, upsert=True)
            meta_cfg = {}
        meta_uf = meta_cfg.get("meta_uf") or 0
    suma_uf = sum(c.get("monto_uf") or 0 for c in clientes)
    pendientes_monto = [c["cliente"] for c in clientes if not c.get("monto_uf")]
    proyeccion = {"meta_uf": meta_uf, "suma_uf": round(suma_uf, 1),
                  "avance_pct": round(suma_uf * 100 / meta_uf, 1) if meta_uf else 0,
                  "diferencia_uf": round(meta_uf - suma_uf, 1),
                  "alerta_diferencia": bool(meta_uf and abs(meta_uf - suma_uf) > 0.5),
                  "pendientes_monto": pendientes_monto,
                  "broker": "Mutuaria y Leasing Limitada"}
    meses = sorted({m for m in await db.folders.distinct("mes_proyeccion") if m} | {"2026-08", "2026-09"})
    # ── META DE PROYECCIÓN: consolidado hacia Gerencia General y Rodrigo Ibáñez ──
    avance_prom = round(sum((c.get("avance") or {}).get("pct", 0) for c in clientes) / len(clientes), 1) if clientes else 0
    uf_en_avance = round(sum(c.get("monto_uf") or 0 for c in clientes if (c.get("avance") or {}).get("pct", 0) > 50), 1)
    uf_cerradas = round(sum(c.get("monto_uf") or 0 for c in clientes if (c.get("avance") or {}).get("pct", 0) >= 100), 1)
    pct_global = round(uf_cerradas * 100 / meta_uf, 1) if meta_uf else 0
    proyeccion.update({"avance_promedio": avance_prom, "uf_en_avance": uf_en_avance,
                       "uf_cerradas": uf_cerradas, "pct_global": pct_global})
    await _avance_notificar(mes_sel, clientes, pct_global)
    await db.config.update_one({"_key": f"avance_global_{mes_sel}"}, {"$set": {
        "pct_global": pct_global, "avance_promedio": avance_prom, "uf_cerradas": uf_cerradas,
        "uf_en_avance": uf_en_avance, "meta_uf": meta_uf, "actualizado": _now()}}, upsert=True)
    return {"mes": mes_cal, "mes_proyeccion": mes_sel, "meses": meses,
            "clientes": clientes, "total": len(clientes),
            "lista_maestra": {k: (await db.config.find_one({"_key": "lista_maestra_origenes"}) or {}).get(k) or []
                              for k in ("inmobiliarias", "brokers")},
            "proyeccion": proyeccion,
            "flota_activa": bool(flota), "flota_total": len(flota),
            "fuente": "ADN_CLIENTES_360 (Regla #66) — sin escaneo de PDFs físicos",
            "recien_llegados": sum(1 for c in clientes if c["recien_24h"])}


HITOS_VALIDOS = ("tasacion", "estudio", "cesion", "set_credito",
                 "serviu", "promesa", "carta_oferta", "carpeta_notaria", "escritura", "notaria",
                 "cert_subsidio", "carta_pie")

_PEND_VERIF = "Pendiente verificación manual"


# ── MATRIZ DEFINITIVA: DOCUMENTOS POR TIPO DE CLIENTE (detectado desde ADN) ──
_DOCS_CONFIRMADOS = ("Recibida", "Aprobada", "Recibido", "Aprobado",
                     "Firmada", "Firmada (verificada IA)")
_DOC_LABEL = {"carta_oferta": "Carta Oferta", "serviu": "Resolución SERVIU",
              "promesa": "Compromiso de Compraventa", "cert_subsidio": "Certificado de Subsidio",
              "carta_pie": "Carta Pie"}
_DOC_ADJ_RE = {"carta_oferta": r"carta.?oferta|cartaoferta|oferta", "serviu": r"serviu|resoluc",
               "promesa": r"promes|compromis|compravent", "cert_subsidio": r"cert.*subsidio|subsidio",
               "carta_pie": r"carta.?pie|cartapie"}


def _tipo_cliente(fd, con_subsidio):
    usada = _inmo_de_folder(fd) == "Casa Usada"
    return (("usada_con_subsidio" if con_subsidio else "usada_sin_subsidio") if usada
            else ("nueva_con_subsidio" if con_subsidio else "nueva_sin_subsidio"))


def _docs_de_tipo(tipo):
    """Cada requisito es una lista de alternativas (la 'O' del ejecutivo en usada sin subsidio)."""
    return {"nueva_con_subsidio": [["carta_oferta"], ["serviu"]],
            "nueva_sin_subsidio": [["carta_oferta"]],
            "usada_con_subsidio": [["promesa"], ["cert_subsidio"]],
            "usada_sin_subsidio": [["promesa", "carta_pie"]]}[tipo]


async def _con_subsidio_fd(fd):
    sub = (fd.get("subsidio_proyeccion") or "").lower()
    if sub:
        return sub.startswith("con")
    try:
        import adn_clientes as _adn
        reg = await db.adn_clientes_360.find_one(
            {"rut_norm": _adn._norm_rut(fd.get("rut") or "")}, {"financiero": 1}) if fd.get("rut") else None
        return bool(((reg or {}).get("financiero") or {}).get("con_subsidio"))
    except Exception:
        return False


def _estado_doc(fd, hito):
    em = fd.get("estados_manuales") or {}
    est = (em.get(hito) or {}).get("estado") or ""
    if hito == "promesa" and not est:
        est = (fd.get("promesa_verificacion") or {}).get("estado") or ""
    return est


def _color_doc(est):
    """✅ recibido y confirmado · 🔵 recibido pendiente verificación · 🟡 solicitado sin respuesta · 🔴 no solicitado."""
    if est in _DOCS_CONFIRMADOS:
        return "verde", "✅"
    if est == _PEND_VERIF:
        return "azul", "🔵"
    if not est or est == "Pendiente":
        return "rojo", "🔴"
    return "amarillo", "🟡"


def _marcado_documentos(fd, con_subsidio, reenviado=False):
    """Marcado por documento + color global del cliente según la matriz de su tipo."""
    tipo = _tipo_cliente(fd, con_subsidio)
    rango = {"rojo": 0, "amarillo": 1, "azul": 2, "verde": 3}
    docs, faltan, hay_azul, llegaron_todos, alguno_llego = [], [], False, True, False
    for alts in _docs_de_tipo(tipo):
        estados = [(h, _estado_doc(fd, h)) for h in alts]
        h, est = max(estados, key=lambda x: rango[_color_doc(x[1])[0]])
        color, icono = _color_doc(est)
        label = " / ".join(_DOC_LABEL[x] for x in alts)
        docs.append({"hito": h, "alternativas": alts, "label": label,
                     "estado": est or "No solicitado", "color": color, "icono": icono})
        if color == "azul":
            hay_azul = True
        if color in ("verde", "azul"):
            alguno_llego = True
        else:
            llegaron_todos = False
            faltan.append(label)
    if llegaron_todos and not hay_azul:
        det = ("Documentos completos — reenviados ✓ a los destinatarios globales" if reenviado
               else "Documentos completos — reenvío automático en curso")
        return {"tipo": tipo, "color": "verde", "icono": "✅", "detalle": det,
                "documentos": docs, "reenviado": reenviado}
    if llegaron_todos and hay_azul:
        return {"tipo": tipo, "color": "azul", "icono": "🔵", "documentos": docs,
                "detalle": "Llegaron todos — hay verificación manual pendiente (no se reenvía hasta confirmar)"}
    if alguno_llego:
        return {"tipo": tipo, "color": "amarillo", "icono": "🟡", "documentos": docs,
                "detalle": "Falta: " + ", ".join(faltan)}
    return {"tipo": tipo, "color": "rojo", "icono": "🔴", "documentos": docs,
            "detalle": "No ha llegado ningún documento (" + ", ".join(faltan) + ")"}


ESTADOS_MANUALES_BASE = ["Tasación Piloto", "Solicitada", "En Proceso",
                         "Con Observaciones", "Aprobada", "Rechazada"]


def _exigir_gerencia(request):
    user = getattr(request.state, "user", {}) or {}
    if (user.get("rol") or "") not in ("admin", "maestro", "gerencia"):
        raise HTTPException(status_code=403, detail="Solo el perfil Gerencia/Administrador puede realizar esta acción")
    return user


@supercarpeta.post("/manual/{fid}")
async def supercarpeta_manual(fid: str, payload: dict, request: Request):
    """P6 — INGRESO MANUAL DE RESPALDO: alerta de fallo de cosecha. El dato se guarda
    de inmediato en la carpeta y en la Bóveda ADN_CLIENTES_360 (Regla #67).
    BITÁCORA: valor anterior, valor nuevo, fecha/hora y usuario (inmutable)."""
    from base_historica import validar_rut_chileno, _fmt_rut
    user = getattr(request.state, "user", {}) or {}
    campo = (payload.get("campo") or "").strip().lower()
    valor = (payload.get("valor") or "").strip()
    if campo not in ("rut", "nombre", "inmobiliaria", "broker", "tipo_operacion", "monto",
                     "proyecto", "ciudad", "notaria", "subsidio") or not valor:
        raise HTTPException(status_code=400, detail="Campo o valor inválido")
    fd = await db.folders.find_one({"id": fid})
    if not fd:
        raise HTTPException(status_code=404, detail="Carpeta no existe")
    if campo == "rut":
        if not validar_rut_chileno(valor):
            raise HTTPException(status_code=400,
                                detail="RUT inválido: no pasa el Dígito Verificador (Regla #66)")
        valor = _fmt_rut(valor)
    if campo == "nombre":
        valor = valor.upper()
        if len(valor) < 5 or len(valor.split()) < 2:
            raise HTTPException(status_code=400, detail="Nombre inválido (nombre y apellido)")
    if campo == "monto":
        try:
            valor = float(str(valor).replace("$", "").replace(" ", "").replace(".", "").replace(",", "."))
        except Exception:
            raise HTTPException(status_code=400, detail="Monto UF inválido")
    # REGLA: ningún cliente sin origen configurado — si la inmobiliaria/broker no existe
    # en el panel de fuentes, se bloquea el guardado hasta registrarla (HTTP 409)
    if campo in ("inmobiliaria", "broker") and str(valor).strip():
        chk = await fuentes_verificar(request,
                                      inmobiliaria=valor if campo == "inmobiliaria" else "",
                                      broker=valor if campo == "broker" else "")
        ok_key = "inmobiliaria_ok" if campo == "inmobiliaria" else "broker_ok"
        if not chk.get(ok_key):
            raise HTTPException(status_code=409, detail={
                "code": "ORIGEN_NO_CONFIGURADO", "campo": campo, "valor": valor,
                "mensaje": f"'{valor}' no existe en el Panel de Fuentes. Registre sus contactos "
                           f"(tasación, estudio de títulos, carta oferta/RS) antes de guardar."})
    destino = {"rut": "rut", "nombre": "nombre", "inmobiliaria": "inmobiliaria",
               "broker": "broker_origen", "tipo_operacion": "tipo_operacion",
               "monto": "proyeccion_uf", "proyecto": "proyecto", "ciudad": "ciudad",
               "notaria": "notaria", "subsidio": "subsidio_proyeccion"}[campo]
    anterior = fd.get(destino)
    ahora = _now()
    await db.estado_manual_log.insert_one({
        "id": str(uuid.uuid4()), "folder_id": fid, "cliente": fd.get("nombre") or "",
        "hito": f"identidad_{campo}", "estado_anterior": str(anterior or ""),
        "estado_nuevo": str(valor), "por": user.get("sub") or "usuario",
        "fecha": ahora, "inmutable": True})
    await db.folders.update_one({"id": fid}, {"$set": {
        destino: valor, "updated_at": ahora,
        f"ingreso_manual.{campo}": {"valor": valor, "anterior": anterior,
                                    "por": user.get("sub") or "usuario", "en": ahora}}})
    if campo == "nombre" and (fd.get("nombre") or "") != valor:
        try:
            viejo, nuevo = fsvc.folder_dir(fd.get("nombre") or ""), fsvc.folder_dir(valor)
            if viejo.exists() and not nuevo.exists():
                viejo.rename(nuevo)
        except Exception:
            pass
    en_adn = await _sync_adn(fid)
    return {"ok": True, "campo": campo, "valor": valor, "anterior": anterior,
            "en_adn": bool(en_adn), "marca": "✏️ manual",
            "alerta": "Ingreso manual registrado en bitácora inmutable (Regla #67)"}


@supercarpeta.post("/estado/{fid}")
async def supercarpeta_estado_manual(fid: str, payload: dict, request: Request):
    """P7 — EDICIÓN MANUAL DE ESTADOS (solo Gerencia): bitácora inmutable con quién,
    fecha/hora, estado anterior y nuevo. Se marca ✏️ y se guarda en la Bóveda ADN."""
    user = _exigir_gerencia(request)
    hito = (payload.get("hito") or "").strip().lower()
    estado = (payload.get("estado") or "").strip()
    if hito not in HITOS_VALIDOS or not estado or len(estado) > 60:
        raise HTTPException(status_code=400, detail="Hito o estado inválido")
    fd = await db.folders.find_one({"id": fid})
    if not fd:
        raise HTTPException(status_code=404, detail="Carpeta no existe")
    anterior = (((fd.get("estados_manuales") or {}).get(hito) or {}).get("estado")
                or (_estado_tasacion(fd) if hito == "tasacion"
                    else _estado_estudio(fd)[0] if hito == "estudio"
                    else fd.get("set_credito_estado") or "Pendiente" if hito == "set_credito"
                    else _estado_cesion(fd)))
    ahora = _now()
    await db.estado_manual_log.insert_one({
        "id": str(uuid.uuid4()), "folder_id": fid, "cliente": fd.get("nombre") or "",
        "hito": hito, "estado_anterior": anterior, "estado_nuevo": estado,
        "por": user.get("sub") or "gerencia", "fecha": ahora, "inmutable": True})
    await db.folders.update_one({"id": fid}, {"$set": {
        f"estados_manuales.{hito}": {"estado": estado, "por": user.get("sub") or "gerencia",
                                     "en": ahora, "anterior": anterior},
        "updated_at": ahora}})
    await _sync_adn(fid)
    # REENVÍO AUTOMÁTICO: si con este estado quedan TODOS los documentos del tipo confirmados
    if hito in ("carta_oferta", "serviu", "promesa", "cert_subsidio", "carta_pie"):
        try:
            fd2 = await db.folders.find_one({"id": fid})
            if fd2:
                await _reenvio_co_rs(fd2)
        except Exception as ex:
            logging.warning(f"reenvio co+rs tras estado manual: {ex}")
    return {"ok": True, "hito": hito, "estado": estado, "anterior": anterior,
            "marca": "✏️ manual", "bitacora": "estado_manual_log (inmutable)"}


@supercarpeta.post("/estado/{fid}/resolver")
async def supercarpeta_estado_resolver(fid: str, payload: dict, request: Request):
    """P7 — CONFLICTO: el correo detectó un dato posterior al estado manual.
    Gerencia decide: mantener el manual o sobreescribir con lo detectado."""
    user = _exigir_gerencia(request)
    hito = (payload.get("hito") or "").strip().lower()
    accion = (payload.get("accion") or "").strip().lower()
    if hito not in HITOS_VALIDOS or accion not in ("mantener", "sobreescribir"):
        raise HTTPException(status_code=400, detail="Hito o acción inválida")
    fd = await db.folders.find_one({"id": fid})
    if not fd:
        raise HTTPException(status_code=404, detail="Carpeta no existe")
    ahora = _now()
    if accion == "sobreescribir":
        await db.folders.update_one({"id": fid}, {"$unset": {f"estados_manuales.{hito}": ""},
                                                  "$set": {"updated_at": ahora}})
        detalle = "Estado manual eliminado: manda lo detectado automáticamente por correo"
    else:
        await db.folders.update_one({"id": fid}, {"$set": {
            f"estados_manuales.{hito}.resuelto": ahora, "updated_at": ahora}})
        detalle = "Estado manual confirmado por Gerencia sobre el dato detectado"
    await db.estado_manual_log.insert_one({
        "id": str(uuid.uuid4()), "folder_id": fid, "cliente": fd.get("nombre") or "",
        "hito": hito, "accion_conflicto": accion, "detalle": detalle,
        "por": user.get("sub") or "gerencia", "fecha": ahora, "inmutable": True})
    await _sync_adn(fid)
    return {"ok": True, "accion": accion, "detalle": detalle}


@supercarpeta.get("/estado-log/{fid}")
async def supercarpeta_estado_log(fid: str):
    regs = await db.estado_manual_log.find({"folder_id": fid}, {"_id": 0}).sort("fecha", -1).to_list(50)
    return {"log": regs, "total": len(regs)}


@supercarpeta.get("/fuentes-doc")
async def fuentes_doc_get():
    """P8 — Fuentes por columna: BLOQUE GLOBAL (todos los clientes) + BLOQUE INDIVIDUAL (por cliente)."""
    meta = await _fuentes_documentos_meta()
    for h, lst in meta.items():
        for f in lst:
            hx = await db.hitos_externos.find_one(
                {"direccion": {"$regex": re.escape(f["correo"]), "$options": "i"}}, sort=[("creado", -1)])
            f["ultima_deteccion"] = (hx or {}).get("creado") or ""
            f["tipo"] = "Global"
    alternativas = []
    async for f in db.folders.find({"fuentes_doc": {"$exists": True}},
                                   {"_id": 0, "id": 1, "nombre": 1, "fuentes_doc": 1}):
        fdoc = {}
        for h, lst in (f.get("fuentes_doc") or {}).items():
            fdoc[h] = [{**_norm_fuente(x), "tipo": "Individual"}
                       for x in (lst or []) if _norm_fuente(x)["correo"]]
        alternativas.append({"id": f["id"], "nombre": f.get("nombre"), "fuentes_doc": fdoc})
    return {"fuentes": meta, "alternativas_cliente": alternativas,
            "estados_disponibles": ESTADOS_MANUALES_BASE}


EMAIL_VALID_RE = re.compile(r"^[\w.+-]+@[\w-]+\.[\w.-]+$")


async def _log_fuente(user, ambito, hito, accion, correo, nombre, cliente=""):
    await db.fuentes_log.insert_one({
        "id": str(uuid.uuid4()), "ambito": ambito, "hito": hito, "accion": accion,
        "correo": correo, "nombre": nombre, "cliente": cliente,
        "por": user.get("sub") or "gerencia", "fecha": _now(), "inmutable": True})


@supercarpeta.post("/fuentes-doc")
async def fuentes_doc_set(payload: dict, request: Request):
    """P8 — FUENTES GLOBALES: agregar/quitar casillas sin límite; efecto inmediato + bitácora."""
    user = _exigir_gerencia(request)
    hito = (payload.get("hito") or "").strip().lower()
    accion = (payload.get("accion") or "").strip().lower()
    if hito in FUENTES_HITOS and accion in ("agregar", "quitar"):
        correo = (payload.get("correo") or "").strip().lower()
        nombre = (payload.get("nombre") or "").strip()
        if not EMAIL_VALID_RE.match(correo):
            raise HTTPException(status_code=400, detail="Correo fuente inválido")
        meta = await _fuentes_documentos_meta()
        lst = [f for f in meta.get(hito, []) if f["correo"] != correo]
        if accion == "agregar":
            lst.append({"correo": correo, "nombre": nombre})
        meta[hito] = lst
        await db.config.update_one({"_key": "fuentes_documentos"}, {"$set": {
            "fuentes": meta, "actualizado": _now(), "por": user.get("sub") or "gerencia"}}, upsert=True)
        await _log_fuente(user, "global", hito, accion, correo, nombre)
        return {"ok": True, "hito": hito, "accion": accion, "fuentes": meta[hito],
                "nota": "Monitoreo actualizado de inmediato (sin reiniciar)"}
    meta = await _fuentes_documentos_meta()
    cambiado = {}
    for h in FUENTES_HITOS:
        if h in payload:
            meta[h] = [{"correo": c, "nombre": ""} for c in _correos_de(payload.get(h))
                       if EMAIL_VALID_RE.match(c)]
            cambiado[h] = meta[h]
    if not cambiado:
        raise HTTPException(status_code=400, detail="Nada que actualizar")
    await db.config.update_one({"_key": "fuentes_documentos"}, {"$set": {
        "fuentes": meta, "actualizado": _now(), "por": user.get("sub") or "gerencia"}}, upsert=True)
    return {"ok": True, "fuentes": cambiado}


@supercarpeta.post("/fuentes-doc/{fid}")
async def fuentes_doc_cliente(fid: str, payload: dict, request: Request):
    """P8 — FUENTES INDIVIDUALES: solo para ese cliente (ej. vivienda usada). Sin límite."""
    user = _exigir_gerencia(request)
    fd = await db.folders.find_one({"id": fid})
    if not fd:
        raise HTTPException(status_code=404, detail="Carpeta no existe")
    fdoc = fd.get("fuentes_doc") or {}
    hito = (payload.get("hito") or "").strip().lower()
    accion = (payload.get("accion") or "").strip().lower()
    if hito in FUENTES_HITOS and accion in ("agregar", "quitar"):
        correo = (payload.get("correo") or "").strip().lower()
        nombre = (payload.get("nombre") or "").strip()
        if not EMAIL_VALID_RE.match(correo):
            raise HTTPException(status_code=400, detail="Correo fuente inválido")
        lst = [f for f in (_norm_fuente(x) for x in (fdoc.get(hito) or []))
               if f["correo"] and f["correo"] != correo]
        if accion == "agregar":
            lst.append({"correo": correo, "nombre": nombre})
        fdoc[hito] = lst
        await _log_fuente(user, "individual", hito, accion, correo, nombre, fd.get("nombre") or "")
    else:
        for h in FUENTES_HITOS:
            if h in payload:
                fdoc[h] = [{"correo": c, "nombre": ""} for c in _correos_de(payload.get(h))
                           if EMAIL_VALID_RE.match(c)]
    await db.folders.update_one({"id": fid}, {"$set": {"fuentes_doc": fdoc, "updated_at": _now()}})
    await _sync_adn(fid)
    return {"ok": True, "cliente": fd.get("nombre"), "fuentes_doc": fdoc}


# ── INMOBILIARIAS (encargado + correo) y SOLICITUD CARTA OFERTA + RES. SERVIU ──
def _norm_inmo(s):
    # unifica variantes de escritura (Boetsch/Boetch → boetch)
    return mail._sin_acentos((s or "").strip().lower()).replace("boetsch", "boetch")


def _inmo_de_folder(fd):
    if "usad" in (fd.get("tipo_operacion") or "").lower():
        return "Casa Usada"
    return (fd.get("inmobiliaria") or "").strip()


@supercarpeta.get("/inmobiliarias")
async def inmobiliarias_get(request: Request):
    _exigir_gerencia(request)
    regs = {r["nombre_norm"]: r async for r in db.inmobiliarias.find({}, {"_id": 0})}
    detectadas = set()
    async for fd in db.folders.find({}, {"inmobiliaria": 1, "tipo_operacion": 1, "oculto_supercarpeta": 1}):
        if fd.get("oculto_supercarpeta"):
            continue
        n = _inmo_de_folder(fd)
        if n:
            detectadas.add(n)
    out = []
    for n in sorted(detectadas):
        r = regs.pop(_norm_inmo(n), None) or {}
        out.append({"nombre": n, "encargado": r.get("encargado") or "", "email": r.get("email") or "",
                    "detectada": True, "configurada": bool(r.get("email"))})
    for r in regs.values():
        out.append({"nombre": r.get("nombre") or "", "encargado": r.get("encargado") or "",
                    "email": r.get("email") or "", "detectada": False, "configurada": bool(r.get("email"))})
    return {"inmobiliarias": out}


@supercarpeta.post("/inmobiliarias")
async def inmobiliarias_set(payload: dict, request: Request):
    user = _exigir_gerencia(request)
    nombre = (payload.get("nombre") or "").strip()
    email_enc = (payload.get("email") or "").strip()
    if not nombre:
        raise HTTPException(status_code=400, detail="El nombre de la inmobiliaria es obligatorio")
    if email_enc and "@" not in email_enc:
        raise HTTPException(status_code=400, detail="Correo del encargado inválido")
    await db.inmobiliarias.update_one({"nombre_norm": _norm_inmo(nombre)}, {"$set": {
        "nombre": nombre, "nombre_norm": _norm_inmo(nombre),
        "encargado": (payload.get("encargado") or "").strip(),
        "email": email_enc, "actualizado": _now(), "por": user.get("sub") or ""}}, upsert=True)
    return {"ok": True}


# ── DESTINATARIOS GLOBALES (CC obligatoria en todos los envíos) ─────────────
_CC_SEMILLA = [
    {"nombre": "Victoria Vilche", "email": "victoriavilches@centralmutuos.cl", "activo": True},
    {"nombre": "Daniela Galindo", "email": "daniela.galindo@centralmutuos.cl", "activo": True},
]


async def _cc_globales():
    cfg = await db.config.find_one({"_key": "cc_globales"}) or {}
    lista = cfg.get("lista")
    if not lista:
        lista = [dict(x) for x in _CC_SEMILLA]
        await db.config.update_one({"_key": "cc_globales"}, {"$set": {
            "lista": lista, "actualizado": _now()}}, upsert=True)
    return lista


@supercarpeta.get("/cc-globales")
async def cc_globales_get(request: Request):
    _exigir_gerencia(request)
    return {"lista": await _cc_globales()}


@supercarpeta.post("/cc-globales")
async def cc_globales_set(payload: dict, request: Request):
    """Editable solo por el Admin General. REGLA: no se eliminan — solo se desactivan."""
    _exigir_admin_general(request)
    nombre = (payload.get("nombre") or "").strip()
    email_d = (payload.get("email") or "").strip()
    if not nombre or "@" not in email_d:
        raise HTTPException(status_code=400, detail="Nombre y correo válido son obligatorios")
    lista = await _cc_globales()
    n_norm = mail._sin_acentos(nombre.lower())
    for x in lista:
        if mail._sin_acentos((x.get("nombre") or "").lower()) == n_norm:
            x.update({"nombre": nombre, "email": email_d,
                      "activo": bool(payload.get("activo", x.get("activo", True)))})
            break
    else:
        lista.append({"nombre": nombre, "email": email_d, "activo": bool(payload.get("activo", True))})
    await db.config.update_one({"_key": "cc_globales"}, {"$set": {
        "lista": lista, "actualizado": _now()}}, upsert=True)
    return {"ok": True, "lista": lista}


# ── CONTACTOS CARTA OFERTA: por INMOBILIARIA + PROYECTO específico ──────────
_CONTACTOS_SEMILLA = [
    {"inmobiliaria": "BOETCH", "proyecto": "", "contacto": "Celinda Soria", "email": ""},
    {"inmobiliaria": "BOETCH", "proyecto": "Uvas y el Viento", "contacto": "Rodrigo Quintero", "email": ""},
    {"inmobiliaria": "BOETCH", "proyecto": "Fuchslocker", "contacto": "Rodrigo Salazar", "email": ""},
    {"inmobiliaria": "MAESTRA", "proyecto": "", "contacto": "", "email": ""},
    {"inmobiliaria": "ECOMAC", "proyecto": "", "contacto": "", "email": ""},
]


async def _seed_contactos_conocidos():
    """Semilla de contactos conocidos (correo vacío = a configurar → alerta antes de enviar)."""
    for s in _CONTACTOS_SEMILLA:
        existe = await db.contactos_carta.find_one({
            "inmobiliaria_norm": {"$in": [_norm_inmo(s["inmobiliaria"]),
                                          _norm_inmo(s["inmobiliaria"]).replace("boetsch", "boetch")]},
            "proyecto_norm": _norm_inmo(s["proyecto"])})
        if not existe:
            await db.contactos_carta.insert_one({
                "id": str(uuid.uuid4()), "inmobiliaria": s["inmobiliaria"],
                "inmobiliaria_norm": _norm_inmo(s["inmobiliaria"]),
                "proyecto": s["proyecto"], "proyecto_norm": _norm_inmo(s["proyecto"]),
                "contacto": s["contacto"], "email": s["email"],
                "activo": True, "actualizado": _now(), "origen": "semilla"})


@supercarpeta.get("/contactos-carta")
async def contactos_carta_get(request: Request):
    _exigir_gerencia(request)
    await _seed_contactos_conocidos()
    contactos = [c async for c in db.contactos_carta.find({}, {"_id": 0}).sort(
        [("inmobiliaria", 1), ("proyecto", 1)])]
    if not contactos:
        # migración: los contactos generales legados de db.inmobiliarias
        async for r in db.inmobiliarias.find({}):
            if r.get("email"):
                doc = {"id": str(uuid.uuid4()), "inmobiliaria": r.get("nombre") or "",
                       "inmobiliaria_norm": r.get("nombre_norm") or "",
                       "proyecto": "", "proyecto_norm": "",
                       "contacto": r.get("encargado") or "", "email": r["email"],
                       "activo": True, "actualizado": _now()}
                await db.contactos_carta.insert_one(dict(doc))
                contactos.append(doc)
    detectadas = set()
    async for fd in db.folders.find({}, {"inmobiliaria": 1, "tipo_operacion": 1, "oculto_supercarpeta": 1}):
        if not fd.get("oculto_supercarpeta"):
            n = _inmo_de_folder(fd)
            if n:
                detectadas.add(n)
    return {"contactos": contactos, "inmobiliarias_detectadas": sorted(detectadas)}


@supercarpeta.post("/contactos-carta")
async def contactos_carta_set(payload: dict, request: Request):
    """Alta/edición/desactivación (nunca eliminación) — Admin General."""
    _exigir_admin_general(request)
    inmob = (payload.get("inmobiliaria") or "").strip()
    email_c = (payload.get("email") or "").strip()
    if not inmob:
        raise HTTPException(status_code=400, detail="La inmobiliaria/corredor es obligatoria")
    if email_c and "@" not in email_c:
        raise HTTPException(status_code=400, detail="Correo del contacto inválido")
    proyecto = (payload.get("proyecto") or "").strip()
    doc = {"inmobiliaria": inmob, "inmobiliaria_norm": _norm_inmo(inmob),
           "proyecto": proyecto, "proyecto_norm": _norm_inmo(proyecto),
           "contacto": (payload.get("contacto") or "").strip(), "email": email_c,
           "tasacion_nombre": (payload.get("tasacion_nombre") or "").strip(),
           "tasacion_email": (payload.get("tasacion_email") or "").strip(),
           "estudio_nombre": (payload.get("estudio_nombre") or "").strip(),
           "estudio_email": (payload.get("estudio_email") or "").strip(),
           "activo": bool(payload.get("activo", True)), "actualizado": _now()}
    if payload.get("id"):
        await db.contactos_carta.update_one({"id": payload["id"]}, {"$set": doc})
    else:
        doc["id"] = str(uuid.uuid4())
        await db.contactos_carta.insert_one(dict(doc))
    return {"ok": True}


async def _contacto_para(fd, proyecto):
    """LÓGICA DE SELECCIÓN: inmobiliaria+proyecto exacto → contacto general → legado → nada."""
    inmo = _inmo_de_folder(fd) or ""
    inmo_n, proy_n = _norm_inmo(inmo), _norm_inmo(proyecto)
    # tolerancia de escritura (Boetch/Boetsch): match exacto o por contención
    cands = [c async for c in db.contactos_carta.find({"activo": {"$ne": False}})
             if inmo_n and (c.get("inmobiliaria_norm") == inmo_n
                            or inmo_n in (c.get("inmobiliaria_norm") or "zzz")
                            or (c.get("inmobiliaria_norm") or "zzz") in inmo_n)]
    if proy_n:
        for c in cands:
            pn = c.get("proyecto_norm") or ""
            if pn and (pn == proy_n or pn in proy_n or proy_n in pn):
                return inmo, c
    general = next((c for c in cands if not c.get("proyecto_norm")), None)
    if general:
        return inmo, general
    legado = await db.inmobiliarias.find_one({"nombre_norm": inmo_n}) or {}
    if legado.get("email"):
        return inmo, {"contacto": legado.get("encargado") or "", "email": legado["email"]}
    # ORDEN 4: broker configurado en fuentes
    broker_n = _norm_inmo(fd.get("broker_origen") or fd.get("broker_nombre") or "")
    if broker_n:
        async for b in db.brokers_fuentes.find({"activo": {"$ne": False}}):
            bn = b.get("nombre_norm") or ""
            if bn and (bn in broker_n or broker_n in bn):
                em_b = b.get("carta_email") or b.get("email")
                if em_b:
                    return inmo, {"contacto": b.get("carta_nombre") or b.get("nombre") or "", "email": em_b}
    return inmo, None


# ── VENDEDORES (VIVIENDA USADA): cada cliente usada tiene su vendedor directo ──
@supercarpeta.get("/vendedores-usada")
async def vendedores_usada_get(request: Request):
    _exigir_gerencia(request)
    out = []
    async for fd in db.folders.find({}, {"id": 1, "nombre": 1, "rut": 1, "tipo_operacion": 1,
                                         "inmobiliaria": 1, "vendedor_usada": 1, "oculto_supercarpeta": 1}):
        if fd.get("oculto_supercarpeta") or _inmo_de_folder(fd) != "Casa Usada":
            continue
        v = fd.get("vendedor_usada") or {}
        con_sub_v = (fd.get("subsidio_proyeccion") or "").lower().startswith("con")
        out.append({"fid": fd["id"], "cliente": fd.get("nombre") or "", "rut": fd.get("rut") or "",
                    "vendedor": v.get("nombre") or "", "email": v.get("email") or "",
                    "telefono": v.get("telefono") or "",
                    "activo": v.get("activo", True) is not False,
                    "tipo_propiedad": "usada_con_subsidio" if con_sub_v else "usada_sin_subsidio",
                    "docs": (["Compromiso de Compraventa", "Certificado de Subsidio"] if con_sub_v
                             else ["Compromiso de Compraventa O Carta Pie (elige el ejecutivo)"]),
                    "configurado": bool(v.get("email"))})
    return {"vendedores": sorted(out, key=lambda x: x["cliente"])}


@supercarpeta.post("/vendedores-usada")
async def vendedores_usada_set(payload: dict, request: Request):
    """Edición del vendedor directo por cliente usada (nunca se elimina — solo se desactiva)."""
    user = _exigir_gerencia(request)
    fid = (payload.get("fid") or "").strip()
    fd = await db.folders.find_one({"id": fid})
    if not fd:
        raise HTTPException(status_code=404, detail="Carpeta no existe")
    email_v = (payload.get("email") or "").strip()
    if email_v and "@" not in email_v:
        raise HTTPException(status_code=400, detail="Correo del vendedor inválido")
    v = fd.get("vendedor_usada") or {}
    v.update({"nombre": (payload.get("vendedor") or payload.get("nombre") or "").strip(),
              "email": email_v, "activo": bool(payload.get("activo", v.get("activo", True))),
              "actualizado": _now(), "por": user.get("sub") or ""})
    if payload.get("telefono") is not None:
        v["telefono"] = (payload.get("telefono") or "").strip()
    await db.folders.update_one({"id": fid}, {"$set": {"vendedor_usada": v, "updated_at": _now()}})
    await _sync_adn(fid)
    return {"ok": True, "vendedor_usada": v}


# ── PANEL DE FUENTES: Sección 1 inmobiliarias/proyectos · 2 brokers · 3 individuales ──
@supercarpeta.get("/fuentes-panel")
async def fuentes_panel_get(request: Request):
    _exigir_gerencia(request)
    await _seed_contactos_conocidos()
    if not await db.brokers_fuentes.find_one({}):
        await db.brokers_fuentes.insert_one({
            "id": str(uuid.uuid4()), "nombre": "MUTUARIA Y LEASING LIMITADA",
            "nombre_norm": _norm_inmo("Mutuaria y Leasing Limitada"), "tipo": "word_consultor",
            "email": "", "tasacion_nombre": "", "tasacion_email": "",
            "estudio_nombre": "", "estudio_email": "", "carta_nombre": "", "carta_email": "",
            "activo": True, "origen": "semilla", "actualizado": _now()})
    contactos = [c async for c in db.contactos_carta.find({}, {"_id": 0}).sort(
        [("inmobiliaria", 1), ("proyecto", 1)])]
    generales = {g["nombre_norm"]: g async for g in db.inmobiliarias.find({}, {"_id": 0})}
    arbol, orden = {}, []
    for c in contactos:
        k = c.get("inmobiliaria_norm") or ""
        if k not in arbol:
            g = next((v for gk, v in generales.items() if gk and (gk in k or k in gk)), {})
            arbol[k] = {"inmobiliaria": (c.get("inmobiliaria") or "").upper(),
                        "correo_general": g.get("email") or "", "proyectos": []}
            orden.append(k)
        arbol[k]["proyectos"].append(c)
    brokers = [b async for b in db.brokers_fuentes.find({}, {"_id": 0}).sort("nombre", 1)]
    vend = await vendedores_usada_get(request)
    lm = await db.config.find_one({"_key": "lista_maestra_origenes"}, {"_id": 0}) or {}
    # ── SEMÁFORO FUENTES: proyectos/orígenes detectados en carpetas SIN contacto configurado ──
    semaforo = []
    detectadas = {}
    async for fd in db.folders.find({}, {"inmobiliaria": 1, "tipo_operacion": 1,
                                         "proyecto": 1, "oculto_supercarpeta": 1}):
        if fd.get("oculto_supercarpeta"):
            continue
        n = _inmo_de_folder(fd)
        if n:
            detectadas.setdefault(n, set())
            if (fd.get("proyecto") or "").strip():
                detectadas[n].add(fd["proyecto"].strip())
    for inmo, proys in sorted(detectadas.items()):
        inmo_n = _norm_inmo(inmo)
        cts = [c for c in contactos
               if c.get("activo") is not False and (c.get("inmobiliaria_norm") == inmo_n
               or inmo_n in (c.get("inmobiliaria_norm") or "") or (c.get("inmobiliaria_norm") or "") in inmo_n)]
        con_email = [c for c in cts if (c.get("email") or "").strip()]
        gen = generales.get(inmo_n) or {}
        if con_email or (gen.get("email") or "").strip():
            estado = "verde"
        elif cts:
            estado = "amarillo"
        else:
            estado = "rojo"
        semaforo.append({"origen": inmo, "proyectos": sorted(proys), "estado": estado,
                         "detalle": ("Contacto con correo configurado" if estado == "verde"
                                     else "Contacto registrado SIN correo — complételo antes de enviar"
                                     if estado == "amarillo"
                                     else "Sin contacto configurado — bloqueará las cartas oferta")})
    return {"inmobiliarias": [arbol[k] for k in orden], "brokers": brokers,
            "individuales": vend["vendedores"],
            "semaforo": semaforo,
            "lista_maestra": {"inmobiliarias": lm.get("inmobiliarias") or [],
                              "brokers": lm.get("brokers") or []}}


@supercarpeta.post("/inmobiliaria-general")
async def inmobiliaria_general_set(payload: dict, request: Request):
    _exigir_admin_general(request)
    nombre = (payload.get("inmobiliaria") or "").strip()
    email_g = (payload.get("email") or "").strip()
    if not nombre:
        raise HTTPException(status_code=400, detail="Nombre de inmobiliaria obligatorio")
    if email_g and "@" not in email_g:
        raise HTTPException(status_code=400, detail="Correo general inválido")
    await db.inmobiliarias.update_one({"nombre_norm": _norm_inmo(nombre)},
        {"$set": {"nombre": nombre.upper(), "nombre_norm": _norm_inmo(nombre),
                  "email": email_g, "actualizado": _now()}}, upsert=True)
    await db.config.update_one({"_key": "lista_maestra_origenes"},
        {"$addToSet": {"inmobiliarias": nombre.upper()}}, upsert=True)
    return {"ok": True}


@supercarpeta.post("/brokers-fuentes")
async def brokers_fuentes_set(payload: dict, request: Request):
    """Registro completo de broker: correo + contactos de tasación, estudio y carta oferta/RS."""
    _exigir_admin_general(request)
    nombre = (payload.get("nombre") or "").strip()
    if not nombre:
        raise HTTPException(status_code=400, detail="Nombre del broker obligatorio")
    for k in ("email", "tasacion_email", "estudio_email", "carta_email"):
        if (payload.get(k) or "").strip() and "@" not in payload[k]:
            raise HTTPException(status_code=400, detail=f"Correo inválido en {k}")
    doc = {"nombre": nombre.upper(), "nombre_norm": _norm_inmo(nombre),
           "tipo": payload.get("tipo") if payload.get("tipo") in ("word_consultor", "autocorredor") else "word_consultor",
           "email": (payload.get("email") or "").strip(),
           "tasacion_nombre": (payload.get("tasacion_nombre") or "").strip(),
           "tasacion_email": (payload.get("tasacion_email") or "").strip(),
           "estudio_nombre": (payload.get("estudio_nombre") or "").strip(),
           "estudio_email": (payload.get("estudio_email") or "").strip(),
           "carta_nombre": (payload.get("carta_nombre") or "").strip(),
           "carta_email": (payload.get("carta_email") or "").strip(),
           "activo": bool(payload.get("activo", True)), "actualizado": _now()}
    if payload.get("id"):
        await db.brokers_fuentes.update_one({"id": payload["id"]}, {"$set": doc})
    else:
        doc["id"] = str(uuid.uuid4())
        await db.brokers_fuentes.insert_one(dict(doc))
    await db.config.update_one({"_key": "lista_maestra_origenes"},
        {"$addToSet": {"brokers": nombre.upper()}}, upsert=True)
    return {"ok": True}


@supercarpeta.get("/fuentes/verificar")
async def fuentes_verificar(request: Request, inmobiliaria: str = "", broker: str = ""):
    """REGISTRO INTELIGENTE: ¿existe ya este origen en el panel de fuentes?"""
    _exigir_gerencia(request)
    out = {}
    if inmobiliaria:
        n = _norm_inmo(inmobiliaria)
        existe = (n in ("casa usada", "directa")
                  or await db.contactos_carta.find_one({"inmobiliaria_norm": {"$regex": f"^{re.escape(n[:6])}"}}) is not None
                  or await db.inmobiliarias.find_one({"nombre_norm": n}) is not None)
        if not existe:
            existe = any([c async for c in db.contactos_carta.find({"activo": {"$ne": False}})
                          if (c.get("inmobiliaria_norm") or "zzz") in n or n in (c.get("inmobiliaria_norm") or "zzz")])
        out["inmobiliaria_ok"] = bool(existe)
    if broker:
        n = _norm_inmo(broker)
        existe = any([b async for b in db.brokers_fuentes.find({"activo": {"$ne": False}})
                      if (b.get("nombre_norm") or "zzz") in n or n in (b.get("nombre_norm") or "zzz")])
        out["broker_ok"] = bool(existe)
    return out


def _saludo_genero(nombre):
    primer = (nombre or "").strip().split(" ")[0]
    return "Estimada" if primer.lower().endswith("a") else "Estimado"


async def _firma_html():
    """Firma formal de los correos salientes: logo + firma personal (solo 'Central Mutuos')."""
    html = ""
    firma = await db.config.find_one({"_key": "firma_correo"}) or {}
    if firma.get("url"):
        html += (f'<br><img src="{firma["url"]}" alt="Central Mutuos" '
                 f'width="340" style="max-width:340px;border-radius:6px;display:block;">')
    html += ('<div style="margin-top:10px;font-family:Arial,Helvetica,sans-serif;'
             'color:#1e293b;font-size:14px;line-height:1.5;">'
             '<b style="color:#0f2557;">Gerardo Barrera P.</b><br>'
             '<span style="color:#8a6d1d;">Asesor Jefe Externo</span><br>'
             'Canal Inmobiliarias y Brokers<br>'
             '<span style="color:#0f2557;font-weight:bold;">Central Mutuos</span>'
             '</div>')
    return html


# ── PARTE 2: REENVÍO AUTOMÁTICO CO+RS a los destinatarios globales ──────────
_PRES_CO_RS = ("Recibida", "Aprobada")


async def _reenvio_co_rs(fd):
    """ÚNICA automatización de envío permitida (REGLA ABSOLUTA): solo cuando TODOS los
    documentos que corresponden al TIPO del cliente están confirmados, se reenvían JUNTOS
    a los destinatarios globales activos (Victoria/Daniela). Jamás documentos parciales."""
    if fd.get("co_rs_reenviado_at"):
        return {"ok": False, "motivo": "ya reenviado"}
    con_sub = await _con_subsidio_fd(fd)
    tipo = _tipo_cliente(fd, con_sub)
    elegidos = []
    for alts in _docs_de_tipo(tipo):
        conf = next((h for h in alts if _estado_doc(fd, h) in _DOCS_CONFIRMADOS), None)
        if not conf:
            return {"ok": False, "motivo": "espera: falta " + " / ".join(_DOC_LABEL[a] for a in alts)}
        elegidos.append(conf)
    destinos = [x["email"] for x in await _cc_globales() if x.get("activo") and x.get("email")]
    if not destinos:
        return {"ok": False, "motivo": "sin destinatarios globales activos"}
    adjuntos, tiene = [], {h: False for h in elegidos}
    try:
        base = fsvc.folder_dir(fd.get("nombre") or "")
        for p in sorted(base.rglob("*.pdf")):
            n = mail._sin_acentos(p.name.lower())
            for h in elegidos:
                if not tiene[h] and re.search(_DOC_ADJ_RE[h], n):
                    adjuntos.append({"filename": p.name,
                                     "content_b64": base64.b64encode(p.read_bytes()).decode()})
                    tiene[h] = True
                    break
    except Exception as ex:
        logging.warning(f"reenvio docs adjuntos {fd.get('nombre')}: {ex}")
    nombre, rut = fd.get("nombre") or "", fd.get("rut") or ""
    filas = "".join(
        f"<li><b>{_DOC_LABEL[h]}</b> ({_estado_doc(fd, h)}): "
        f"{'adjunto ✓' if tiene[h] else 'confirmado en sistema (archivo no localizado en la carpeta digital)'}</li>"
        for h in elegidos)
    tipo_txt = {"nueva_con_subsidio": "Vivienda Nueva con Subsidio",
                "nueva_sin_subsidio": "Vivienda Nueva sin Subsidio",
                "usada_con_subsidio": "Vivienda Usada con Subsidio",
                "usada_sin_subsidio": "Vivienda Usada sin Subsidio"}[tipo]
    html = (f"<p>Estimadas,</p>"
            f"<p>Se reenvían de forma automática los documentos completos del cliente "
            f"<b>{nombre}</b> (RUT {rut or 'por confirmar'}) — {tipo_txt}:</p>"
            f"<ul>{filas}</ul>"
            f"<p>Todos los documentos que corresponden a este tipo de cliente fueron "
            f"recibidos y confirmados en el sistema.</p>"
            f"<p>Atentamente,<br>Central Mutuos — reenvío automático</p>")
    html += await _firma_html()
    asunto = " + ".join(_DOC_LABEL[h] for h in elegidos) + f" — {nombre}" + (f" ({rut})" if rut else "")
    res = await asyncio.to_thread(
        lambda: mail.send_mail(destinos, asunto, html, attachments=adjuntos, desde="secundaria"))
    if not res.get("success"):
        logging.warning(f"reenvio docs {nombre}: {res.get('error')}")
        return {"ok": False, "motivo": res.get("error") or "error SMTP"}
    ahora = _now()
    registro = {"tipo": "reenvio_automatico_docs", "tipo_cliente": tipo, "para": destinos,
                "documentos": [_DOC_LABEL[h] for h in elegidos],
                "adjuntos": [a["filename"] for a in adjuntos], "asunto": asunto,
                "en": ahora, "por": "sistema (automático)", "estado": "enviado"}
    await db.folders.update_one({"id": fd["id"]}, {
        "$set": {"co_rs_reenviado_at": ahora, "co_rs_reenvio": registro, "updated_at": ahora},
        "$push": {"bitacora_solicitudes": registro}})
    await db.alertas.insert_one({"id": str(uuid.uuid4()), "tipo": "reenvio_co_rs",
        "mensaje": f"📤 Documentos completos de {nombre} ({tipo_txt}) reenviados automáticamente a "
                   f"{', '.join(destinos)} ({len(adjuntos)} adjunto(s))",
        "fecha": ahora, "leida": False})
    await _sync_adn(fd["id"])
    return {"ok": True, "para": destinos, "documentos": elegidos,
            "adjuntos": [a["filename"] for a in adjuntos]}


async def _marcar_doc_llegada(fd, hito, e):
    """CUENTA DE BARRIDO: al detectar la llegada de un documento solicitado lo marca
    🔵 azul (recibido — pendiente verificación manual). Nunca confirma sola."""
    actual = _estado_doc(fd, hito)
    if actual in _DOCS_CONFIRMADOS or actual == _PEND_VERIF:
        return False
    ahora = _now()
    await db.folders.update_one({"id": fd["id"]}, {"$set": {
        f"estados_manuales.{hito}": {"estado": _PEND_VERIF, "por": "cuenta de barrido",
                                     "en": ahora, "via": "barrido", "anterior": actual},
        "updated_at": ahora}})
    await db.alertas.insert_one({"id": str(uuid.uuid4()), "tipo": "doc_recibido",
        "mensaje": f"🔵 {_DOC_LABEL.get(hito, hito)} de {fd.get('nombre')} recibido — "
                   f"pendiente verificación manual (asunto: {(e.get('subject') or '')[:70]})",
        "fecha": ahora, "leida": False})
    await _sync_adn(fd["id"])
    return True


async def reenvio_co_rs_loop():
    """Red de seguridad cada 30 min: reenvía los casos que quedaron completos."""
    while True:
        await asyncio.sleep(1800)
        try:
            async for fd in db.folders.find({
                    "co_rs_reenviado_at": {"$exists": False},
                    "estados_manuales": {"$exists": True}}):
                if not fd.get("oculto_supercarpeta"):
                    await _reenvio_co_rs(fd)
        except Exception as e:
            logging.warning(f"reenvio docs loop: {e}")


# ── RESUMEN SEMANAL A GERENCIA (lunes): avances y cuellos de botella de la Flota ──
async def _resumen_gerencia_destinatarios():
    cfg = await db.config.find_one({"_key": "resumen_gerencia"}) or {}
    dest = cfg.get("destinatarios")
    if not dest:
        dest = sorted({"rodrigoibanez@centralmutuos.cl"}
                      | {x["email"] for x in await _cc_globales() if x.get("activo") and x.get("email")})
        await db.config.update_one({"_key": "resumen_gerencia"}, {"$set": {
            "destinatarios": dest, "actualizado": _now()}}, upsert=True)
    return dest


def _cuellos_cliente(c):
    cb = []
    for f in (c.get("faltantes") or []):
        cb.append(f"Dato faltante: {f}")
    if ((c.get("bitacora") or {}).get("tasacion") or {}).get("demora_48h"):
        cb.append("Tasación +48h sin respuesta")
    if c.get("estudio_titulos") == "Con Reparos" or c.get("alerta_reparos"):
        cb.append("Estudio de Títulos con reparos pendientes")
    if c.get("promesa") == _PEND_VERIF:
        cb.append("Promesa/Compromiso pendiente de verificación manual")
    docs = c.get("docs_co_rs") or {}
    if docs.get("color") in ("rojo", "amarillo", "azul"):
        cb.append(f"CO+RS: {docs.get('detalle')}")
    if not c.get("fecha_firma") and (c.get("avance") or {}).get("pct", 0) >= 70:
        cb.append("Sin fecha de firma agendada pese al avance")
    if (c.get("set_credito") or {}).get("estado") == "verificacion_pendiente":
        cb.append("Set de Crédito con verificación de firmas pendiente")
    return cb


async def _resumen_gerencia_html(mes: str = ""):
    data = await supercarpeta_vista(mes)
    p = data.get("proyeccion") or {}
    filas = []
    for i, c in enumerate(sorted(data["clientes"], key=lambda x: (x.get("cliente") or "")), 1):
        cb = _cuellos_cliente(c)
        pct = (c.get("avance") or {}).get("pct", 0)
        color = "#15803d" if pct >= 90 else "#b45309" if pct >= 40 else "#b91c1c"
        filas.append(
            f"<tr style='border-bottom:1px solid #e2e8f0'>"
            f"<td style='padding:6px 8px;color:#64748b;font-weight:bold'>{i}</td>"
            f"<td style='padding:6px 8px;font-weight:bold'>{c.get('cliente') or ''}"
            f"<div style='color:#64748b;font-weight:normal;font-size:11px'>{c.get('rut') or ''}</div></td>"
            f"<td style='padding:6px 8px;text-align:right'>{c.get('monto_uf') or '—'}</td>"
            f"<td style='padding:6px 8px;text-align:center;color:{color};font-weight:bold'>{pct}%</td>"
            f"<td style='padding:6px 8px;font-size:12px;color:#7c2d12'>"
            + ("<br>".join("⚠ " + x for x in cb) if cb else "<span style='color:#15803d'>Sin cuellos de botella</span>")
            + "</td></tr>")
    hoy = _now()[:10]
    html = (
        f"<div style='font-family:Arial,Helvetica,sans-serif;max-width:760px;margin:auto;color:#1e293b'>"
        f"<div style='background:#1a1f2e;color:#d4af37;padding:16px 22px;border-radius:8px 8px 0 0'>"
        f"<h2 style='margin:0;font-size:19px'>Resumen Semanal — Gerencia</h2>"
        f"<div style='color:#94a3b8;font-size:12px'>Central Mutuos · Flota {data.get('mes_proyeccion')} · {hoy}</div></div>"
        f"<div style='border:1px solid #e2e8f0;border-top:none;padding:16px 22px'>"
        f"<p style='font-size:13px'><b>Proyección:</b> {p.get('suma_uf')} / {p.get('meta_uf')} UF "
        f"({p.get('avance_pct')}%) · Avance promedio de etapas: <b>{p.get('avance_promedio')}%</b> · "
        f"UF en avance (&gt;50%): <b>{p.get('uf_en_avance')}</b> · UF cerradas: <b>{p.get('uf_cerradas')}</b></p>"
        f"<table style='width:100%;border-collapse:collapse;font-size:13px'>"
        f"<tr style='background:#f1f5f9;color:#0f2557;text-transform:uppercase;font-size:11px'>"
        f"<th style='padding:6px 8px;text-align:left'>N°</th><th style='padding:6px 8px;text-align:left'>Cliente</th>"
        f"<th style='padding:6px 8px;text-align:right'>Monto UF</th><th style='padding:6px 8px'>Avance</th>"
        f"<th style='padding:6px 8px;text-align:left'>Cuellos de botella</th></tr>"
        + "".join(filas) + "</table>"
        f"<p style='color:#64748b;font-size:11px;margin-top:14px'>Generado automáticamente desde la "
        f"Supercarpeta (Bóveda ADN_CLIENTES_360). Los estados manuales de Gerencia prevalecen.</p></div></div>")
    return html + await _firma_html()


async def _enviar_resumen_gerencia():
    dest = await _resumen_gerencia_destinatarios()
    html = await _resumen_gerencia_html()
    hoy = datetime.now(timezone.utc).strftime("%d/%m/%Y")
    res = await asyncio.to_thread(
        lambda: mail.send_mail(dest, f"📊 Resumen Semanal Gerencia — Flota Supercarpeta — {hoy}",
                               html, desde="secundaria"))
    return res, dest


async def resumen_gerencia_loop():
    """Cada lunes ≥ 08:00 (hora Chile): resumen de avances y cuellos de botella a Gerencia."""
    from zoneinfo import ZoneInfo
    while True:
        await asyncio.sleep(1800)
        try:
            ahora = datetime.now(ZoneInfo("America/Santiago"))
            if ahora.weekday() != 0 or ahora.hour < 8:
                continue
            semana_key = ahora.strftime("%G-W%V")
            cfg = await db.config.find_one({"_key": "resumen_gerencia"}) or {}
            if cfg.get("last_sent_week") == semana_key:
                continue
            res, dest = await _enviar_resumen_gerencia()
            if res.get("success"):
                await db.config.update_one({"_key": "resumen_gerencia"}, {"$set": {
                    "last_sent_week": semana_key, "last_sent_at": _now(),
                    "ultimo_envio_a": dest}}, upsert=True)
        except Exception as e:
            logging.warning(f"resumen gerencia: {e}")


@supercarpeta.get("/resumen-gerencia")
async def resumen_gerencia_get(request: Request):
    _exigir_gerencia(request)
    cfg = await db.config.find_one({"_key": "resumen_gerencia"}, {"_id": 0}) or {}
    return {"destinatarios": await _resumen_gerencia_destinatarios(),
            "last_sent_at": cfg.get("last_sent_at") or "", "programado": "Lunes 08:00 (hora Chile)"}


@supercarpeta.post("/resumen-gerencia/config")
async def resumen_gerencia_config(payload: dict, request: Request):
    _exigir_gerencia(request)
    dest = [d.strip().lower() for d in (payload.get("destinatarios") or []) if "@" in str(d)]
    if not dest:
        raise HTTPException(status_code=400, detail="Debe indicar al menos un correo válido")
    await db.config.update_one({"_key": "resumen_gerencia"}, {"$set": {
        "destinatarios": dest, "actualizado": _now()}}, upsert=True)
    return {"ok": True, "destinatarios": dest}


@supercarpeta.post("/resumen-gerencia/enviar")
async def resumen_gerencia_enviar(payload: dict, request: Request):
    """Envío manual con vista previa (confirm:false = solo preview)."""
    _exigir_gerencia(request)
    if not (payload or {}).get("confirm"):
        return {"body": await _resumen_gerencia_html(),
                "destinatarios": await _resumen_gerencia_destinatarios()}
    res, dest = await _enviar_resumen_gerencia()
    if not res.get("success"):
        raise HTTPException(status_code=502, detail=res.get("error") or "Error de envío SMTP")
    await db.config.update_one({"_key": "resumen_gerencia"}, {"$set": {
        "last_sent_at": _now(), "ultimo_envio_a": dest}}, upsert=True)
    return {"ok": True, "destinatarios": dest}


@supercarpeta.get("/solicitud-doc/{fid}")
async def solicitud_doc_preview(fid: str, request: Request, doc: str = ""):
    """VISTA PREVIA OBLIGATORIA. MATRIZ POR TIPO: nueva c/sub → CO+RS (inmobiliaria/proyecto);
    nueva s/sub → CO; usada c/sub → Compromiso+Cert Subsidio (vendedor);
    usada s/sub → Compromiso O Carta Pie (elige el ejecutivo, vendedor)."""
    _exigir_gerencia(request)
    import adn_clientes as _adn
    fd = await db.folders.find_one({"id": fid})
    if not fd:
        raise HTTPException(status_code=404, detail="Carpeta no encontrada")
    reg = (await db.adn_clientes_360.find_one({"rut_norm": _adn._norm_rut(fd["rut"])})
           if fd.get("rut") else None) or {}
    proyecto = (fd.get("proyecto") or (reg.get("propiedad") or {}).get("proyecto")
                or _limpiar_proyecto(reg.get("nombre_proyecto")) or "").strip()
    resolucion = (fd.get("resolucion_serviu")
                  or (reg.get("resolucion_serviu") or {}).get("numero", "")
                  if isinstance(reg.get("resolucion_serviu"), dict)
                  else fd.get("resolucion_serviu") or reg.get("resolucion_serviu") or "")
    con_sub = await _con_subsidio_fd(fd)
    tipo = _tipo_cliente(fd, con_sub)
    inmo, contacto = await _contacto_para(fd, proyecto)
    usada = tipo.startswith("usada")
    if usada:
        v = fd.get("vendedor_usada") or {}
        contacto = ({"contacto": v.get("nombre") or "", "email": v.get("email") or ""}
                    if v.get("email") and v.get("activo", True) is not False else None)
    encargado = (contacto or {}).get("contacto") or ""
    saludo = _saludo_genero(encargado)
    rut = fd.get("rut") or ""
    nombre = fd.get("nombre") or ""
    faltantes = []
    if not rut:
        faltantes.append("RUT del cliente")
    if not usada and not proyecto:
        faltantes.append("Nombre del proyecto")
    # SIN CC: la solicitud sale SOLO al contacto — Victoria/Daniela reciben los
    # documentos por el reenvío automático cuando llega la respuesta (_reenvio_co_rs)
    cc = []
    pie_datos = (f"Cliente: {nombre}\nRUT: {rut or '[RUT del cliente]'}\n")
    doc_sel = doc if doc in ("promesa", "carta_pie") else "promesa"
    alternativas = None
    if tipo == "nueva_con_subsidio":
        docs_hitos = ["carta_oferta", "serviu"]
        asunto = f"Carta Oferta y Resolución SERVIU - {nombre} - {proyecto or '[Nombre del proyecto]'}"
        cuerpo = (f"{saludo} {encargado or '[nombre del destinatario]'},\n\n"
                  f"Solicito por medio de la presente los siguientes documentos:\n\n"
                  f"1. Carta Oferta\n"
                  f"2. Resolución SERVIU\n\n{pie_datos}"
                  f"Proyecto: {proyecto or '[Nombre del proyecto]'}\n")
    elif tipo == "nueva_sin_subsidio":
        docs_hitos = ["carta_oferta"]
        asunto = f"Carta Oferta - {nombre} - {proyecto or '[Nombre del proyecto]'}"
        cuerpo = (f"{saludo} {encargado or '[nombre del destinatario]'},\n\n"
                  f"Solicito por medio de la presente carta oferta de:\n\n{pie_datos}"
                  f"Proyecto: {proyecto or '[Nombre del proyecto]'}\n")
    elif tipo == "usada_con_subsidio":
        docs_hitos = ["promesa", "cert_subsidio"]
        asunto = f"Compromiso de Compraventa y Certificado de Subsidio - {nombre}"
        cuerpo = (f"{saludo} {encargado or '[nombre del vendedor]'},\n\n"
                  f"Solicito por medio de la presente los siguientes documentos:\n\n"
                  f"1. Compromiso de Compraventa firmado\n"
                  f"2. Certificado de Subsidio\n\n"
                  f"Cliente comprador: {nombre}\nRUT: {rut or '[RUT del cliente]'}\n"
                  f"Propiedad: {proyecto or '[Dirección de la propiedad]'}\n")
    else:  # usada_sin_subsidio — el ejecutivo elige
        docs_hitos = [doc_sel]
        alternativas = ["promesa", "carta_pie"]
        asunto = f"{_DOC_LABEL[doc_sel]} - {nombre}"
        cuerpo = (f"{saludo} {encargado or '[nombre del vendedor]'},\n\n"
                  f"Solicito por medio de la presente {'el Compromiso de Compraventa firmado' if doc_sel == 'promesa' else 'la Carta Pie'} de:\n\n"
                  f"Cliente comprador: {nombre}\nRUT: {rut or '[RUT del cliente]'}\n"
                  f"Propiedad: {proyecto or '[Dirección de la propiedad]'}\n")
    cuerpo += "\nAgradeciendo su buena disposición,\nAtentamente,\nCentral Mutuos"
    desde = ""
    for a in mail.ACCOUNTS:
        if a["rol"] == "secundaria":
            desde = a["user"]
    if not desde and mail.ACCOUNTS:
        desde = mail.ACCOUNTS[0]["user"]
    return {"cliente": nombre, "rut": rut, "inmobiliaria": inmo,
            "proyecto": proyecto, "resolucion_serviu": resolucion,
            "usada": usada, "tipo_cliente": tipo, "doc_elegido": doc_sel if alternativas else "",
            "alternativas": alternativas, "requiere_resolucion": False,
            "docs_solicitados": [_DOC_LABEL[h] for h in docs_hitos],
            "tipo": "+".join(docs_hitos),
            "encargado": encargado, "para": (contacto or {}).get("email") or "",
            "configurada": bool((contacto or {}).get("email")), "cc": cc,
            "faltantes": faltantes, "asunto": asunto, "cuerpo": cuerpo, "desde": desde}


@supercarpeta.post("/solicitud-doc/{fid}/enviar")
async def solicitud_doc_enviar(fid: str, payload: dict, request: Request):
    """REGLA ABSOLUTA: solo envía al confirmar el preview — jamás automático ni retroactivo.
    Sin CC (la copia a Gerencia va después, vía reenvío automático). Registro en ADN_CLIENTES_360."""
    user = _exigir_gerencia(request)
    import adn_clientes as _adn
    fd = await db.folders.find_one({"id": fid})
    if not fd:
        raise HTTPException(status_code=404, detail="Carpeta no encontrada")
    para = (payload.get("para") or "").strip()
    con_sub = await _con_subsidio_fd(fd)
    tipo = _tipo_cliente(fd, con_sub)
    usada = tipo.startswith("usada")
    if "@" not in para:
        raise HTTPException(status_code=400,
                            detail="Configure el vendedor del cliente (botón 🏢 Fuentes → Vendedores)" if usada
                            else "Configure el contacto de la inmobiliaria/proyecto (botón 🏢 Fuentes)")
    # VALIDACIÓN BLOQUEANTE: indica exactamente qué campo falta
    rut = (payload.get("rut") or fd.get("rut") or "").strip()
    proyecto = (payload.get("proyecto") or fd.get("proyecto") or "").strip()
    resolucion = str(payload.get("resolucion_serviu") or fd.get("resolucion_serviu") or "").strip()
    faltantes = [] if rut else ["RUT del cliente"]
    if not usada and not proyecto:
        faltantes.append("Nombre del proyecto")
    if faltantes:
        raise HTTPException(status_code=400,
                            detail="ENVÍO BLOQUEADO — falta: " + ", ".join(faltantes))
    doc_sel = payload.get("doc_elegido") if payload.get("doc_elegido") in ("promesa", "carta_pie") else "promesa"
    hitos_marca = {"nueva_con_subsidio": ("carta_oferta", "serviu"),
                   "nueva_sin_subsidio": ("carta_oferta",),
                   "usada_con_subsidio": ("promesa", "cert_subsidio"),
                   "usada_sin_subsidio": (doc_sel,)}[tipo]
    asunto = (payload.get("asunto") or "").strip() or (
        " y ".join(_DOC_LABEL[h] for h in hitos_marca) + f" - {fd.get('nombre')}")
    cuerpo = (payload.get("cuerpo") or "").strip()
    if not cuerpo:
        raise HTTPException(status_code=400, detail="El cuerpo del correo no puede ir vacío")
    # persistir datos completados desde el preview (proyecto / resolución)
    sets_fd = {}
    if proyecto and proyecto != (fd.get("proyecto") or ""):
        sets_fd["proyecto"] = proyecto
    if resolucion and resolucion != (fd.get("resolucion_serviu") or ""):
        sets_fd["resolucion_serviu"] = resolucion
    if sets_fd:
        await db.folders.update_one({"id": fid}, {"$set": {**sets_fd, "updated_at": _now()}})
        await _sync_adn(fid)
    # SIN CC: la solicitud sale SOLO al contacto de la inmobiliaria/vendedor.
    # Victoria y Daniela reciben los documentos completos vía reenvío automático (_reenvio_co_rs).
    cc = []
    html = "<p>" + cuerpo.replace("\n", "<br>") + "</p>"
    html += await _firma_html()
    usuario_envio = user.get("sub") or ""

    async def _envio_bg():
        """ENVÍO EN SEGUNDO PLANO: el throttling SMTP (10s + reintento 60s) jamás
        congela la interfaz ni la respuesta HTTP. El resultado queda en bitácora/alertas."""
        try:
            res = await asyncio.to_thread(
                lambda: mail.send_mail(para, asunto, html, desde="secundaria", cc=cc))
            ok = bool(res.get("success"))
            if ok:
                # APRENDIZAJE AUTOMÁTICO: si el Admin editó el destinatario en la confirmación,
                # queda guardado en el panel de fuentes para ese origen (precargado la próxima vez)
                try:
                    inmo_ap = _inmo_de_folder(fd)
                    if usada:
                        v = fd.get("vendedor_usada") or {}
                        if para.lower() != (v.get("email") or "").lower():
                            v.update({"nombre": (payload.get("encargado") or v.get("nombre") or "").strip(),
                                      "email": para, "activo": True, "actualizado": _now(),
                                      "por": "aprendizaje automático (editado en confirmación)"})
                            await db.folders.update_one({"id": fid}, {"$set": {"vendedor_usada": v}})
                    else:
                        _, cont_prev = await _contacto_para(fd, proyecto)
                        if para.lower() != ((cont_prev or {}).get("email") or "").lower():
                            await db.contactos_carta.update_one(
                                {"inmobiliaria_norm": _norm_inmo(inmo_ap), "proyecto_norm": _norm_inmo(proyecto)},
                                {"$set": {"inmobiliaria": inmo_ap, "inmobiliaria_norm": _norm_inmo(inmo_ap),
                                          "proyecto": proyecto, "proyecto_norm": _norm_inmo(proyecto),
                                          "contacto": (payload.get("encargado") or "").strip(), "email": para,
                                          "activo": True, "actualizado": _now(),
                                          "origen": "aprendizaje automático (editado en confirmación)"},
                                 "$setOnInsert": {"id": str(uuid.uuid4())}}, upsert=True)
                except Exception as ex:
                    logging.warning(f"aprendizaje contacto: {ex}")
                em = fd.get("estados_manuales") or {}
                for h in hitos_marca:
                    if not (em.get(h) or {}).get("estado"):
                        em[h] = {"estado": "Solicitada", "por": usuario_envio,
                                 "en": _now(), "via": "solicitud_email"}
                await db.folders.update_one({"id": fid}, {"$set": {"estados_manuales": em}})
            registro = {"tipo": "+".join(hitos_marca), "tipo_cliente": tipo,
                        "documentos": [_DOC_LABEL[h] for h in hitos_marca], "para": para, "cc": cc,
                        "asunto": asunto, "en": _now(), "por": usuario_envio,
                        "estado": "enviado" if ok else "fallido",
                        "smtp_code": res.get("smtp_code"), "error": res.get("error") or ""}
            await db.folders.update_one({"id": fid}, {
                "$set": {"updated_at": _now()},
                "$push": {"bitacora_solicitudes": registro}})
            # REGISTRO EN ADN_CLIENTES_360: fecha, hora, destinatarios y estado
            if fd.get("rut"):
                await db.adn_clientes_360.update_one(
                    {"rut_norm": _adn._norm_rut(fd["rut"])},
                    {"$push": {"envios_carta_oferta": registro}})
            if not ok:
                await db.alertas.insert_one({"id": str(uuid.uuid4()), "tipo": "solicitud_fallida",
                    "mensaje": f"🔴 Falló el envío de la solicitud de {fd.get('nombre')} a {para}: "
                               f"{res.get('error') or 'error SMTP'} — reintente desde la Supercarpeta",
                    "fecha": _now(), "leida": False})
        except Exception as ex:
            logging.warning(f"solicitud envío bg {fd.get('nombre')}: {ex}")

    asyncio.create_task(_envio_bg())
    return {"ok": True, "estado": "en_envio", "cc": cc,
            "documentos": [_DOC_LABEL[h] for h in hitos_marca]}


# ── COSTOS CBR — extracción desde adjuntos "Simulación" de Mesa ─────────────
_CBR_FILA_RE = re.compile(r"(\bC\.?\s?B\.?\s?R\.?\b|conservador(?:\s+de)?\s+bienes\s+ra[ií]ces|\bconservador\b)", re.I)
_CBR_UF_RE = re.compile(r"(?:UF\s*\$?\s*(\d+(?:[.,]\d+)?)|(\d+(?:[.,]\d+)?)\s*UF\b)", re.I)
_CBR_CLP_RE = re.compile(r"\$\s*(\d{1,3}(?:\.\d{3})+|\d{4,9})|\b(\d{1,3}(?:\.\d{3})+)\b")


def _monto_en_linea(l):
    m_uf = _CBR_UF_RE.search(l)
    if m_uf:
        try:
            return float((m_uf.group(1) or m_uf.group(2)).replace(",", ".")), "UF"
        except ValueError:
            pass
    m_clp = _CBR_CLP_RE.search(l)
    if m_clp:
        try:
            return int((m_clp.group(1) or m_clp.group(2)).replace(".", "")), "CLP"
        except ValueError:
            pass
    m_num = re.search(r"\b(\d{4,9})\b", l)
    if m_num:
        return int(m_num.group(1)), "CLP"
    return None


def _extraer_cbr_texto(texto):
    """REGLA DE HIERRO: solo lee filas reales — jamás inventa valores.
    Si existe la sección 'Gastos Operacionales', busca desde ahí hacia abajo."""
    d = _extraer_gastos_texto(texto)
    return d.get("valor_cbr")


_GASTO_FILAS = [
    ("valor_cbr", _CBR_FILA_RE),
    ("tasacion", re.compile(r"\btasaci[oó]n\b", re.I)),
    ("est_titulos", re.compile(r"estudio\s+de\s+t[ií]tulos|\best\.?\s*t[ií]tulos", re.I)),
]


def _extraer_gastos_texto(texto):
    lineas = [l.strip() for l in (texto or "").splitlines() if l.strip()]
    idx = next((i for i, l in enumerate(lineas) if "gastos operacionales" in l.lower()), None)
    ambito = lineas[idx:] if idx is not None else lineas
    out = {}
    for campo, rx in _GASTO_FILAS:
        for i, l in enumerate(ambito):
            if not rx.search(l):
                continue
            r = _monto_en_linea(l)
            if not r and i + 1 < len(ambito):
                r = _monto_en_linea(ambito[i + 1])
            if r:
                out[campo] = {"valor": r[0], "moneda": r[1], "linea": l[:180]}
                break
    return out


_PROY_LINEA_RE = re.compile(r"^\s*proyecto\s*:?\s+(.{2,90})$", re.I)


def _limpiar_proyecto(raw):
    """'BOETSCH/LAS UVAS Y EL VIENTO' → 'LAS UVAS Y EL VIENTO'. 'CASA USADA/N/D' → ''."""
    v = (raw or "").strip()
    if not v or v.upper() in ("N/D", "ND", "S/I", "POR CONFIRMAR"):
        return ""
    if "/" in v:
        resto = v.split("/", 1)[1].strip()
        return "" if not resto or resto.upper() in ("N/D", "ND", "S/I") else resto
    return v


def _extraer_proyecto_texto(texto):
    for l in (texto or "").splitlines():
        m = _PROY_LINEA_RE.match(l.strip())
        if m:
            return m.group(1).strip()
    return ""


def _extraer_gastos_pdf(pdf_bytes):
    """Abre el adjunto: prioriza la SEGUNDA página (Gastos Operacionales), luego el resto.
    Extrae CBR + Tasación + Estudio de Títulos + Nombre del Proyecto en una sola pasada."""
    import io as _io
    import pdfplumber
    acumulado = {}
    try:
        with pdfplumber.open(_io.BytesIO(pdf_bytes)) as pdf:
            paginas = pdf.pages
            orden = ([1] if len(paginas) > 1 else []) + [i for i in range(len(paginas)) if i != 1]
            for i in orden:
                try:
                    texto = paginas[i].extract_text() or ""
                    d = _extraer_gastos_texto(texto)
                    proy = _extraer_proyecto_texto(texto)
                except Exception:
                    d, proy = {}, ""
                for k, v in d.items():
                    acumulado.setdefault(k, {**v, "pagina": i + 1})
                if proy:
                    acumulado.setdefault("proyecto", {"valor": proy, "pagina": i + 1})
                if len(acumulado) >= len(_GASTO_FILAS) + 1:
                    break
    except Exception:
        pass
    return acumulado


def _extraer_cbr_pdf(pdf_bytes):
    return (_extraer_gastos_pdf(pdf_bytes) or {}).get("valor_cbr")


# COMISIÓN (USO INTERNO — SOLO GERENCIA): % según broker/inmobiliaria
_COMISION_PCT = {"boetsch": 1.0, "boetch": 1.0, "ecomac": 1.0, "poch": 1.0,
                 "comod": 0.8, "usada": 0.5, "word": 0.5, "urbanizate": 0.5}


def _comision_cliente(fd, monto_uf, con_subsidio=None):
    """Regla de cálculo: % sobre el monto del crédito según broker. Sin regla → revisar.
    MAESTRA (regla del dueño): 0,5% sin subsidio (caso Claudia Zurita) / 1% con subsidio.
    ECOMAC (regla del dueño): 0,8% con subsidio / 1% sin subsidio."""
    tipo = (fd.get("tipo_operacion") or "").lower()
    b = mail._sin_acentos((fd.get("inmobiliaria") or fd.get("broker_origen") or "").lower())
    if "maestra" in b:
        if con_subsidio is None:
            return None, "REVISAR CON GERENCIA (Maestra: falta saber si es con o sin subsidio)"
        pct = 1.0 if con_subsidio else 0.5
    elif "ecomac" in b:
        pct = 0.8 if con_subsidio else 1.0
    else:
        pct = 0.5 if ("usad" in tipo or "usada" in b) else next(
            (p for k, p in _COMISION_PCT.items() if k in b), None)
    if pct is None:
        return None, "REVISAR CON GERENCIA"
    com = round(float(monto_uf or 0) * pct / 100, 2) if monto_uf else ""
    return com, f"{pct}%"


async def _ejecutar_cbr():
    import adn_clientes as _adn
    await db.config.update_one({"_key": "cbr_extraccion"}, {"$set": {
        "estado": "en_proceso", "inicio": _now()}}, upsert=True)
    try:
        flota_cfg = await db.config.find_one({"_key": "flota_agosto"}) or {}
        flota = ({n.strip().upper() for n in (flota_cfg.get("nombres") or [])}
                 if flota_cfg.get("activo") else set())
        carpetas = []
        async for fd in db.folders.find({}).sort("nombre", 1):
            if fd.get("oculto_supercarpeta"):
                continue
            if (fd.get("mes_proyeccion") or "2026-08") != "2026-08":
                continue
            nombre_u = (fd.get("nombre") or "").strip().upper()
            if flota and not any(nf in nombre_u or nombre_u in nf for nf in flota):
                continue
            carpetas.append(fd)
        # BÚSQUEDA DIRIGIDA: solo correos de aprobaciones@ que mencionan a cada cliente
        nombres = [" ".join((fd.get("nombre") or "").split()[:2]) for fd in carpetas]
        correos = await asyncio.to_thread(mail.fetch_simulacion_attachments, 60,
                                          "aprobaciones@centralmutuos.cl", nombres)
        adn_map = {}
        async for r in db.adn_clientes_360.find({}, {
                "rut_norm": 1, "financiero": 1, "costo_CBR": 1,
                "costo_tasacion": 1, "costo_estudio_titulos": 1, "comision": 1}):
            adn_map[r.get("rut_norm")] = r
        viejo = await db.config.find_one({"_key": "cbr_extraccion"}) or {}
        old_map = {(r.get("cliente") or "").strip().upper(): r
                   for r in viejo.get("resultados") or []}
        base = await _valores_base()
        resultados = []
        for fd in carpetas:
            nombre_u = (fd.get("nombre") or "").strip().upper()
            toks = [t for t in mail._sin_acentos(nombre_u.lower()).split() if len(t) > 2]
            necesarios = min(2, len(toks)) or 1
            gastos, meta = None, None
            for e in correos:  # ya vienen ordenados del más reciente al más antiguo
                blob = mail._sin_acentos(" ".join(
                    [e.get("subject") or "", e.get("body") or ""]
                    + [p.get("filename") or "" for p in e.get("pdfs") or []]).lower())
                if sum(1 for t in toks if t in blob) < necesarios:
                    continue
                for p in e.get("pdfs") or []:
                    g = await asyncio.to_thread(_extraer_gastos_pdf, p.get("content_bytes") or b"")
                    if g:
                        gastos = g
                        meta = {"fecha_correo": (e.get("date") or "")[:10],
                                "archivo": p.get("filename") or "",
                                "fuente": e.get("cuenta") or "correo"}
                        break
                if gastos:
                    break
            cbr = (gastos or {}).get("valor_cbr")
            reg = adn_map.get(_adn._norm_rut(fd.get("rut"))) if fd.get("rut") else None
            monto_uf = (fd.get("proyeccion_uf")
                        or ((reg or {}).get("financiero") or {}).get("monto_credito_uf") or "")
            subsidio_txt = (fd.get("subsidio_proyeccion")
                            or ("Con Subsidio" if ((reg or {}).get("financiero") or {}).get("con_subsidio")
                                else "Sin Subsidio"))
            comision, pct_txt = _comision_cliente(fd, monto_uf, "con" in subsidio_txt.lower())
            fila = {
                "fid": fd["id"],
                "cliente": fd.get("nombre") or "",
                "rut": fd.get("rut") or "",
                "broker": fd.get("inmobiliaria") or fd.get("broker_origen") or "",
                "proyecto": (fd.get("proyecto")
                             or ((reg or {}).get("propiedad") or {}).get("proyecto") or ""),
                "tipo_propiedad": fd.get("tipo_operacion") or "",
                "subsidio": subsidio_txt,
                "monto_credito": monto_uf,
                "comision": comision if comision is not None else "",
                "pct_aplicado": pct_txt,
                "total_pagado": "",
                "valor_cbr": cbr["valor"] if cbr else "",
                "moneda": cbr["moneda"] if cbr else "",
                "tasacion": "", "tasacion_moneda": "",
                "est_titulos": "", "est_titulos_moneda": "",
                "fecha_correo": (meta or {}).get("fecha_correo", ""),
                "archivo": (meta or {}).get("archivo", ""),
                "linea": (cbr or {}).get("linea", ""),
                "estado": "ENCONTRADO" if cbr else "NO ENCONTRADO"}
            for campo in ("tasacion", "est_titulos"):
                g = (gastos or {}).get(campo)
                if g:
                    fila[campo] = g["valor"]
                    fila[f"{campo}_moneda"] = g["moneda"]
            # respaldo desde la Bóveda ADN si la extracción no trajo el dato
            for campo, mon_key, adn_f in _CBR_CAMPOS:
                if campo == "comision" or fila.get(campo) not in ("", None):
                    continue
                doc = (reg or {}).get(adn_f) or {}
                if doc.get("valor") not in ("", None):
                    fila[campo] = doc["valor"]
                    fila[mon_key] = doc.get("moneda") or "UF"
                    if doc.get("origen") == "manual":
                        fila[f"{campo}_origen"] = "manual"
            # preservar TODOS los valores ingresados manualmente en corridas anteriores
            old = old_map.get(nombre_u) or {}
            for campo in _CBR_NUM + _CBR_TXT:
                if old.get(f"{campo}_origen") == "manual" and old.get(campo) not in ("", None):
                    if campo in ("valor_cbr", "tasacion", "est_titulos") and fila.get(campo) not in ("", None):
                        continue  # dato fresco del correo manda sobre el manual antiguo
                    fila[campo] = old[campo]
                    fila[f"{campo}_origen"] = "manual"
            for _c, mon_key, _f in _CBR_CAMPOS:
                if fila.get(f"{_c}_origen") == "manual" and old.get(mon_key) and not fila.get(mon_key):
                    fila[mon_key] = old[mon_key]
            # VALORES BASE: Tasación 2,5 UF / Est. Títulos 2 UF si el cliente no trae dato
            for campo in ("tasacion", "est_titulos"):
                if fila.get(campo) in ("", None):
                    fila[campo] = base[campo]
                    fila[f"{campo}_moneda"] = "UF"
            resultados.append(fila)
            # ── NOMBRE DEL PROYECTO: simulación → ADN (campo unificado nombre_proyecto) ──
            proy_raw = ((gastos or {}).get("proyecto") or {}).get("valor", "")
            rut_n = _adn._norm_rut(fd["rut"]) if fd.get("rut") else None
            if not proy_raw and reg:
                # unificación de variantes ya existentes en la Bóveda
                proy_raw = ((reg.get("propiedad") or {}).get("proyecto")
                            or reg.get("nombre_proyecto") or reg.get("inmueble") or "")
            if proy_raw:
                limpio = _limpiar_proyecto(proy_raw)
                if rut_n:
                    await db.adn_clientes_360.update_one(
                        {"rut_norm": rut_n}, {"$set": {"nombre_proyecto": proy_raw}})
                if limpio and not (fd.get("proyecto") or "").strip():
                    await db.folders.update_one({"id": fd["id"]}, {"$set": {
                        "proyecto": limpio, "updated_at": _now()}})
                    fd["proyecto"] = limpio
                if limpio and fila.get("proyecto_origen") != "manual":
                    fila["proyecto"] = fila.get("proyecto") or limpio
            # persistencia: carpeta + Bóveda ADN_CLIENTES_360
            sets_f = {}
            for campo, doc_key in (("valor_cbr", "costo_CBR"), ("tasacion", "costo_tasacion"),
                                   ("est_titulos", "costo_estudio_titulos")):
                g = (gastos or {}).get(campo)
                if g:
                    sets_f[doc_key] = {"valor": g["valor"], "moneda": g["moneda"],
                                       "fecha_correo": meta["fecha_correo"], "archivo": meta["archivo"],
                                       "extraido": _now(), "origen": "aprobaciones@centralmutuos.cl"}
            if sets_f:
                await db.folders.update_one({"id": fd["id"]}, {"$set": {
                    **sets_f, "updated_at": _now()}})
                await _sync_adn(fd["id"])
                if fd.get("rut"):
                    await db.adn_clientes_360.update_one(
                        {"rut_norm": _adn._norm_rut(fd["rut"])}, {"$set": sets_f})
        await db.config.update_one({"_key": "cbr_extraccion"}, {"$set": {
            "estado": "completado", "ultima": _now(), "resultados": resultados,
            "remitente": "aprobaciones@centralmutuos.cl",
            "encontrados": sum(1 for r in resultados if r["estado"] == "ENCONTRADO"),
            "total": len(resultados), "correos_revisados": len(correos)}}, upsert=True)
    except Exception as ex:
        logging.warning(f"cbr extraccion: {ex}")
        await db.config.update_one({"_key": "cbr_extraccion"}, {"$set": {
            "estado": f"error: {str(ex)[:150]}", "ultima": _now()}}, upsert=True)


def _es_admin_general(request):
    """ACCESO EXCLUSIVO: solo el Administrador General (Gerardo) ve CBR y comisiones."""
    return ((getattr(request.state, "user", {}) or {}).get("rol") or "") in ("admin", "maestro")


def _exigir_admin_general(request):
    if not _es_admin_general(request):
        raise HTTPException(status_code=403,
            detail="Acceso denegado: módulo de Comisiones y CBR exclusivo del Administrador General")


# campos de gasto: (campo, clave de moneda, campo en ADN/folder)
_CBR_CAMPOS = [("valor_cbr", "moneda", "costo_CBR"),
               ("tasacion", "tasacion_moneda", "costo_tasacion"),
               ("est_titulos", "est_titulos_moneda", "costo_estudio_titulos"),
               ("comision", "comision_moneda", "comision")]
# REGLA ABSOLUTA DE EDITABILIDAD: todos los campos del módulo, sin excepción
_CBR_NUM = ("valor_cbr", "tasacion", "est_titulos", "comision", "monto_credito", "total_pagado")
_CBR_TXT = ("cliente", "rut", "broker", "proyecto", "tipo_propiedad", "subsidio",
            "pct_aplicado", "moneda", "estado")
# campos que viven en la carpeta y se propagan a la Bóveda vía _sync_adn
_CBR_FOLDER_MAP = {"rut": "rut", "broker": "inmobiliaria", "proyecto": "proyecto",
                   "tipo_propiedad": "tipo_operacion", "subsidio": "subsidio_proyeccion",
                   "monto_credito": "proyeccion_uf"}


# ── VALORES BASE OPERACIONALES (Tasación / Estudio de Títulos por defecto) ──
async def _valores_base():
    cfg = await db.config.find_one({"_key": "valores_base"}) or {}
    return {"tasacion": cfg.get("tasacion", 2.5), "est_titulos": cfg.get("est_titulos", 2)}


@supercarpeta.post("/valores-base")
async def valores_base_set(payload: dict, request: Request):
    """Cambia el valor global (clientes nuevos sin dato); los manuales no se tocan."""
    _exigir_admin_general(request)
    sets = {}
    for k in ("tasacion", "est_titulos"):
        if payload.get(k) is not None:
            try:
                sets[k] = round(float(str(payload[k]).replace(",", ".")), 2)
            except ValueError:
                raise HTTPException(status_code=400, detail=f"Valor inválido para {k}")
    if sets:
        await db.config.update_one({"_key": "valores_base"}, {"$set": {
            **sets, "actualizado": _now()}}, upsert=True)
    return {"ok": True, **await _valores_base()}


def _cbr_total_fila(r):
    """Total Pagado por fila: manual manda; si no, CBR + Tasación + Est. Títulos."""
    if r.get("total_pagado_origen") == "manual" and r.get("total_pagado") not in ("", None):
        try:
            return float(r["total_pagado"])
        except (TypeError, ValueError):
            return None
    nums = []
    for k in ("valor_cbr", "tasacion", "est_titulos"):
        v = r.get(k)
        if v in ("", None):
            continue
        try:
            nums.append(float(v))
        except (TypeError, ValueError):
            pass
    return sum(nums) if nums else None


def _cbr_totales(res):
    """REGLA DE CONVERSIÓN: NUNCA se mezclan monedas — UF suma con UF, CLP con CLP."""
    prefijos = {"valor_cbr": "total_cbr", "tasacion": "total_tasacion",
                "est_titulos": "total_titulos", "comision": "total_comision"}
    tot = {f"{p}_{s}": 0.0 for p in prefijos.values() for s in ("uf", "clp")}
    for r in res:
        for campo, mon_key, _ in _CBR_CAMPOS:
            v = r.get(campo)
            if v in ("", None):
                continue
            try:
                v = float(v)
            except (TypeError, ValueError):
                continue
            suf = "clp" if (r.get(mon_key) or "UF").upper() == "CLP" else "uf"
            tot[f"{prefijos[campo]}_{suf}"] += v
    for s in ("uf", "clp"):
        tot[f"gran_total_{s}"] = 0.0
    for r in res:
        v = _cbr_total_fila(r)
        if v is None:
            continue
        suf = "clp" if (r.get("moneda") or "UF").upper() == "CLP" else "uf"
        tot[f"gran_total_{suf}"] += v
    return {k: round(v, 2) for k, v in tot.items()}


@supercarpeta.get("/cbr/estado")
async def cbr_estado(request: Request):
    _exigir_admin_general(request)
    d = (await db.config.find_one({"_key": "cbr_extraccion"}, {"_id": 0})
         or {"estado": "nunca_ejecutado", "resultados": []})
    d.update(_cbr_totales(d.get("resultados") or []))
    d["valores_base"] = await _valores_base()
    return d


@supercarpeta.post("/cbr/manual")
async def cbr_manual(request: Request, payload: dict):
    """REGLA ABSOLUTA DE EDITABILIDAD: todos los campos del módulo son editables por el
    Admin General. Todo cambio se guarda de inmediato en ADN_CLIENTES_360."""
    _exigir_admin_general(request)
    import adn_clientes as _adn
    campo = payload.get("campo")
    if campo not in _CBR_NUM + _CBR_TXT:
        raise HTTPException(status_code=400, detail=f"campo inválido: {campo}")
    bruto = str(payload.get("valor") if payload.get("valor") is not None else "").strip()
    if campo in _CBR_NUM:
        valor = ""
        if bruto:
            try:
                valor = round(float(bruto.replace("%", "").replace(",", ".")), 2)
            except ValueError:
                raise HTTPException(status_code=400, detail="valor numérico inválido")
    else:
        valor = bruto
    cfg = await db.config.find_one({"_key": "cbr_extraccion"}) or {}
    res = cfg.get("resultados") or []
    cliente = (payload.get("cliente") or "").strip().upper()
    fila = next((r for r in res
                 if r.get("fid") == payload.get("fid")
                 or (r.get("cliente") or "").strip().upper() == cliente), None)
    if not fila:
        raise HTTPException(status_code=404, detail="cliente no está en el reporte CBR")
    if campo == "cliente":
        fila.setdefault("cliente_original", fila.get("cliente"))
    fila[campo] = valor
    fila[f"{campo}_origen"] = "manual" if valor != "" else ""
    campos_mon = {c: mk for c, mk, _f in _CBR_CAMPOS}
    if campo in campos_mon and valor != "" and not fila.get(campos_mon[campo]):
        fila[campos_mon[campo]] = "UF"
    # % de comisión editado → recalcular la comisión si no fue sobreescrita manualmente
    if campo == "pct_aplicado" and fila.get("comision_origen") != "manual":
        m_pct = re.search(r"(\d+(?:[.,]\d+)?)", valor or "")
        monto = fila.get("monto_credito")
        if m_pct and monto not in ("", None):
            try:
                fila["comision"] = round(float(monto) * float(m_pct.group(1).replace(",", ".")) / 100, 2)
            except (TypeError, ValueError):
                pass
    await db.config.update_one({"_key": "cbr_extraccion"}, {"$set": {"resultados": res}})
    # ── persistencia inmediata en la Bóveda ADN_CLIENTES_360 ──
    fd = None
    if fila.get("fid"):
        fd = await db.folders.find_one({"id": fila["fid"]})
    if not fd:
        fd = await db.folders.find_one({"nombre": fila.get("cliente_original") or fila.get("cliente")})
    if fd:
        rut_n = _adn._norm_rut(fd["rut"]) if fd.get("rut") else None
        if campo in ("valor_cbr", "tasacion", "est_titulos"):
            adn_field = dict((c, f) for c, _mk, f in _CBR_CAMPOS)[campo]
            doc = {"valor": valor, "moneda": fila.get(campos_mon[campo]) or "UF",
                   "origen": "manual", "actualizado": _now()}
            await db.folders.update_one({"id": fd["id"]}, {"$set": {adn_field: doc, "updated_at": _now()}})
            await _sync_adn(fd["id"])
            if rut_n:
                await db.adn_clientes_360.update_one({"rut_norm": rut_n}, {"$set": {adn_field: doc}})
        elif campo == "comision":
            if rut_n:
                await db.adn_clientes_360.update_one({"rut_norm": rut_n}, {"$set": {"comision": {
                    "valor": valor, "moneda": "UF", "origen": "manual", "actualizado": _now()}}})
        elif campo == "total_pagado":
            if rut_n:
                await db.adn_clientes_360.update_one({"rut_norm": rut_n}, {"$set": {"total_pagado": {
                    "valor": valor, "moneda": fila.get("moneda") or "UF",
                    "origen": "manual", "actualizado": _now()}}})
        elif campo == "cliente":
            if rut_n and valor:
                await db.adn_clientes_360.update_one({"rut_norm": rut_n},
                                                     {"$set": {"identidad.nombre": valor}})
        elif campo in _CBR_FOLDER_MAP:
            await db.folders.update_one({"id": fd["id"]}, {"$set": {
                _CBR_FOLDER_MAP[campo]: valor, "updated_at": _now()}})
            await _sync_adn(fd["id"])
        elif rut_n:
            # campos del reporte (moneda, estado, pct) → registro de overrides en la Bóveda
            await db.adn_clientes_360.update_one({"rut_norm": rut_n}, {"$set": {
                f"cbr_overrides.{campo}": {"valor": valor, "origen": "manual", "actualizado": _now()}}})
    return {"ok": True, **_cbr_totales(res)}


@supercarpeta.post("/cbr/extraer")
async def cbr_extraer(request: Request):
    """Lanza la búsqueda de correos de aprobaciones@ y extrae el CBR de cada cliente."""
    _exigir_admin_general(request)
    from datetime import timedelta
    cfg = await db.config.find_one({"_key": "cbr_extraccion"}) or {}
    if cfg.get("estado") == "en_proceso" and (cfg.get("inicio") or "") > (
            datetime.now(timezone.utc) - timedelta(minutes=10)).isoformat():
        return {"ok": True, "estado": "en_proceso", "nota": "Extracción CBR ya en curso"}
    asyncio.create_task(_ejecutar_cbr())
    return {"ok": True, "lanzado": True, "seguimiento": "GET /api/supercarpeta/cbr/estado"}


@supercarpeta.get("/cbr/excel")
async def cbr_excel(request: Request):
    """ACCESO EXCLUSIVO: el Excel de CBR + comisiones solo para el Administrador General."""
    _exigir_admin_general(request)
    cfg = await db.config.find_one({"_key": "cbr_extraccion"}) or {}
    res = cfg.get("resultados") or []
    if not res:
        raise HTTPException(status_code=404, detail="Sin resultados CBR: ejecute primero la extracción")
    import io as _io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    wb = Workbook()
    ws = wb.active
    ws.title = "Costos CBR"
    ws.append(["N°", "Nombre del cliente", "Broker", "Monto del crédito (UF)",
               "Valor CBR (Inscripción Registro Propiedad + Hipoteca)",
               "Tasación", "Estudio de Títulos", "Total Pagado",
               "Comisión (monto calculado)", "Porcentaje aplicado",
               "Moneda", "Fecha del correo fuente", "Estado CBR"])
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1F2937")
    for num, r in enumerate(res, start=1):
        tp = _cbr_total_fila(r)
        incompleto = (r.get("total_pagado_origen") != "manual"
                      and any(r.get(k) in ("", None) for k in ("valor_cbr", "tasacion", "est_titulos")))
        total_pagado = round(tp, 2) if tp is not None else ""
        ws.append([num, r.get("cliente", ""), r.get("broker", ""), r.get("monto_credito", ""),
                   r.get("valor_cbr", ""), r.get("tasacion", ""), r.get("est_titulos", ""),
                   f"{total_pagado} ⚠ incompleto" if incompleto and total_pagado != "" else total_pagado,
                   r.get("comision", ""), r.get("pct_aplicado", ""),
                   r.get("moneda", ""), r.get("fecha_correo", ""), r.get("estado", "")])
        if incompleto:
            ws.cell(row=ws.max_row, column=8).fill = PatternFill("solid", fgColor="FEF3C7")
        ws.cell(row=ws.max_row, column=13).font = Font(
            bold=True, color="15803D" if r.get("estado") == "ENCONTRADO" else "B91C1C")
    tot = _cbr_totales(res)
    ws.append(["", "TOTAL EN UF", "", "", tot["total_cbr_uf"], tot["total_tasacion_uf"],
               tot["total_titulos_uf"], tot["gran_total_uf"], tot["total_comision_uf"],
               "", "UF", "", ""])
    for c in ws[ws.max_row]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1E3A8A")
    ws.append(["", "TOTAL EN PESOS", "", "", tot["total_cbr_clp"], tot["total_tasacion_clp"],
               tot["total_titulos_clp"], tot["gran_total_clp"], tot["total_comision_clp"],
               "", "CLP", "", ""])
    for c in ws[ws.max_row]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="14532D")
    for col, w in zip("ABCDEFGHIJKLM", (6, 34, 24, 18, 24, 12, 16, 16, 20, 18, 9, 16, 16)):
        ws.column_dimensions[col].width = w
    buf = _io.BytesIO()
    wb.save(buf)
    return Response(content=buf.getvalue(),
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": 'attachment; filename="costos_CBR.xlsx"'})


# ── CUENTA DE BARRIDO (SOLO LECTURA) — categoría propia del Panel ⚙️ ────────
def _dv_rut(cuerpo):
    s, f = 0, 2
    for d in reversed(cuerpo):
        s += int(d) * f
        f = 2 if f == 7 else f + 1
    r = 11 - (s % 11)
    return "0" if r == 11 else "k" if r == 10 else str(r)


async def _cosechar_ruts_faltantes(correos):
    """Regla #66: si un correo trae el nombre de un cliente SIN RUT + exactamente un
    RUT válido (DV verificado), se cosecha y queda ESCRITO en carpeta + Bóveda ADN."""
    sin_rut = await db.folders.find(
        {"oculto_supercarpeta": {"$exists": False},
         "$or": [{"rut": ""}, {"rut": None}, {"rut": {"$exists": False}}]},
        {"id": 1, "nombre": 1}).to_list(200)
    if not sin_rut:
        return 0
    usados = {_rut_limpio(f.get("rut")) async for f in db.folders.find({"rut": {"$nin": ["", None]}}, {"rut": 1})}
    hallados = 0
    for e in correos or []:
        texto = f"{e.get('subject', '')} {e.get('body', '')}".lower()
        ruts = list({_rut_limpio(m) for m in RUT_RE.findall(texto)})
        ruts = [r for r in ruts if len(r) >= 8 and _dv_rut(r[:-1]) == r[-1] and r not in usados]
        if len(ruts) != 1:
            continue
        for fd in list(sin_rut):
            toks = [x for x in (fd.get("nombre") or "").lower().split() if len(x) > 2]
            if len(toks) >= 2 and sum(1 for x in toks if x in texto) >= 2:
                rut_fmt = f"{ruts[0][:-1]}-{ruts[0][-1].upper()}"
                await db.folders.update_one({"id": fd["id"]}, {"$set": {
                    "rut": rut_fmt, "rut_origen": "cuenta_barrido", "updated_at": _now()}})
                await db.alertas.insert_one({"id": str(uuid.uuid4()), "tipo": "cuenta_barrido",
                    "mensaje": f"🧬 RUT cosechado por la Cuenta de Barrido: {fd.get('nombre')} → {rut_fmt}",
                    "fecha": _now(), "leida": False})
                await _sync_adn(fd["id"])
                sin_rut.remove(fd)
                usados.add(ruts[0])
                hallados += 1
                break
    return hallados


async def _ejecutar_barrido_cuenta(dias, origen):
    cfg = await db.config.find_one({"_key": "cuenta_barrido"}) or {}
    rol = cfg.get("rol")
    if not rol:
        return {"error": "sin cuenta designada"}
    await db.config.update_one({"_key": "cuenta_barrido"}, {"$set": {
        "barrido_estado": "en_proceso", "barrido_inicio": _now(), "barrido_origen": origen}}, upsert=True)
    try:
        correos = await asyncio.to_thread(mail.fetch_recent_account, rol, dias)
        res = await _auditar_lote(correos)
        res["ruts_cosechados"] = await _cosechar_ruts_faltantes(correos)
        resumen = {k: v for k, v in res.items() if k != "detalle"}
        await db.config.update_one({"_key": "cuenta_barrido"}, {"$set": {
            "barrido_estado": "completado", "ultima_lectura": _now(),
            "barrido_origen": origen, "ultimo_resultado": resumen}}, upsert=True)
        return resumen
    except Exception as ex:
        logging.warning(f"barrido cuenta: {ex}")
        await db.config.update_one({"_key": "cuenta_barrido"}, {"$set": {
            "barrido_estado": f"error: {str(ex)[:120]}", "ultima_lectura": _now()}}, upsert=True)
        return {"error": str(ex)[:200]}


@supercarpeta.get("/cuenta-barrido")
async def cuenta_barrido_get():
    cfg = await db.config.find_one({"_key": "cuenta_barrido"}, {"_id": 0}) or {}
    disponibles = [{"rol": a["rol"], "correo": a["user"]} for a in mail.ACCOUNTS]
    correo = next((a["user"] for a in mail.ACCOUNTS if a["rol"] == cfg.get("rol")), "")
    return {"configurada": bool(cfg.get("rol")), "rol": cfg.get("rol", ""), "correo": correo,
            "activo": cfg.get("activo", True), "modo": "solo_lectura",
            "ultima_lectura": cfg.get("ultima_lectura", ""),
            "barrido_estado": cfg.get("barrido_estado", ""),
            "ultimo_resultado": cfg.get("ultimo_resultado") or {},
            "cuentas_disponibles": disponibles,
            "nota": "Lectura BODY.PEEK (solo lectura): jamás marca, mueve ni envía correos"}


@supercarpeta.post("/cuenta-barrido")
async def cuenta_barrido_set(payload: dict, request: Request):
    """Designa la casilla existente que actúa como Cuenta de Barrido (Solo Lectura)."""
    user = _exigir_gerencia(request)
    rol = (payload.get("rol") or "").strip().lower()
    if rol and rol not in [a["rol"] for a in mail.ACCOUNTS]:
        raise HTTPException(status_code=400, detail="Esa cuenta no existe en el sistema")
    upd = {"actualizado": _now(), "por": user.get("sub") or "gerencia"}
    if rol:
        upd["rol"] = rol
    if "activo" in payload:
        upd["activo"] = bool(payload.get("activo"))
    await db.config.update_one({"_key": "cuenta_barrido"}, {"$set": upd}, upsert=True)
    correo = next((a["user"] for a in mail.ACCOUNTS if a["rol"] == rol), "")
    if rol:
        await _log_fuente(user, "cuenta_barrido", "todas", "designar", correo,
                          "Cuenta de Barrido (Solo Lectura)")
    cfg = await db.config.find_one({"_key": "cuenta_barrido"}) or {}
    return {"ok": True, "rol": cfg.get("rol", ""), "correo": correo or "",
            "activo": cfg.get("activo", True)}


@supercarpeta.post("/cuenta-barrido/barrer")
async def cuenta_barrido_barrer(request: Request, dias: int = 7):
    """Barrido manual 'Barrer ahora': corre en segundo plano (solo lectura)."""
    _exigir_gerencia(request)
    cfg = await db.config.find_one({"_key": "cuenta_barrido"}) or {}
    if not cfg.get("rol"):
        raise HTTPException(status_code=400, detail="Primero designa la cuenta de barrido en el Panel ⚙️")
    d = min(max(int(dias or 7), 1), 90)
    asyncio.create_task(_ejecutar_barrido_cuenta(d, "manual"))
    return {"ok": True, "lanzado": True, "dias": d,
            "seguimiento": "GET /api/supercarpeta/cuenta-barrido"}


async def _auto_envio_aprobaciones():
    """ENVÍO AUTOMÁTICO EXCLUSIVO PARA APROBACIONES DE MESA: simulación PDF de
    aprobaciones@centralmutuos.cl → extrae CBR (2ª página) → correo ajustado con gastos
    operacionales a gerardo.ext@centralmutuos.cl SIN confirmación. Todo lo demás
    (carta oferta, RS, compromiso, etc.) mantiene el confirm manual obligatorio."""
    import hashlib
    correos = await asyncio.to_thread(mail.fetch_simulacion_attachments, 12,
                                      "aprobaciones@centralmutuos.cl", None)
    vb = (await db.config.find_one({"_key": "valores_base_operacionales"}) or {})
    tas_uf, est_uf = float(vb.get("tasacion_uf") or 2.5), float(vb.get("estudio_titulos_uf") or 2.0)
    for c in (correos or [])[:10]:
        eid = hashlib.sha1(f"{c.get('subject','')}|{c.get('date','')}".encode()).hexdigest()
        if await db.auto_envios_aprobaciones.find_one({"email_id": eid}):
            continue
        pdf = (c.get("pdfs") or [{}])[0]
        gastos = _extraer_gastos_pdf(pdf.get("content_bytes") or b"") or {}
        cbr = gastos.get("valor_cbr") or {}
        registro = {"id": str(uuid.uuid4()), "email_id": eid, "asunto": c.get("subject") or "",
                    "fecha_correo": c.get("date") or "", "archivo": pdf.get("filename") or "",
                    "en": _now()}
        if not cbr.get("valor"):
            # REGLA: sin CBR legible en el PDF → NO se envía; alerta al Admin General
            registro["estado"] = "sin_cbr_alertado"
            await db.auto_envios_aprobaciones.insert_one(registro)
            await db.alertas.insert_one({"id": str(uuid.uuid4()), "tipo": "auto_envio_aprobaciones",
                "mensaje": f"⚠️ Simulación de aprobaciones@ SIN datos CBR legibles en la 2ª página "
                           f"({pdf.get('filename') or 'sin PDF'} — {c.get('subject','')[:70]}). "
                           f"Revisión manual requerida — NO se envió correo automático.",
                "fecha": _now(), "leida": False})
            continue
        moneda = cbr.get("moneda") or "UF"
        val_cbr = cbr.get("valor")
        total_uf = round((val_cbr if moneda == "UF" else 0) + tas_uf + est_uf, 2)
        html = (f"<p>Estimado,</p>"
                f"<p>Simulación de Mesa procesada automáticamente (correo de aprobaciones@centralmutuos.cl):</p>"
                f"<ul>"
                f"<li><b>Asunto original:</b> {c.get('subject') or ''}</li>"
                f"<li><b>Valor CBR (Inscripción Registro Propiedad + Hipoteca):</b> {val_cbr} {moneda}"
                f" (página {cbr.get('pagina') or 2} del PDF)</li>"
                f"<li><b>Tasación (valor base):</b> {tas_uf} UF</li>"
                f"<li><b>Estudio de Títulos (valor base):</b> {est_uf} UF</li>"
                f"<li><b>Total gastos operacionales:</b> {total_uf} UF"
                f"{' (+ CBR en ' + moneda + ')' if moneda != 'UF' else ''}</li>"
                f"</ul><p>Se adjunta el PDF de la simulación procesada.</p>"
                f"<p>Central Mutuos — envío automático (barrido de aprobaciones)</p>")
        adj = [{"filename": pdf.get("filename") or "simulacion.pdf",
                "content_b64": base64.b64encode(pdf["content_bytes"]).decode()}]
        res = await asyncio.to_thread(
            lambda: mail.send_mail("gerardo.ext@centralmutuos.cl",
                                   f"⚙️ Simulación procesada + Gastos Operacionales — {c.get('subject','')[:70]}",
                                   html, attachments=adj, desde="principal"))
        registro["estado"] = "enviado" if res.get("success") else f"error: {res.get('error','')[:80]}"
        registro["cbr"] = cbr
        await db.auto_envios_aprobaciones.insert_one(registro)


async def cuenta_barrido_loop():
    """Cada 20 min barre la cuenta designada (solo lectura) y persiste cada hallazgo
    en la carpeta y la Bóveda ADN (estados, PDFs firmados, RUTs, notaría)."""
    await asyncio.sleep(200)
    while True:
        try:
            cfg = await db.config.find_one({"_key": "cuenta_barrido"}) or {}
            if cfg.get("rol") and cfg.get("activo", True):
                await _ejecutar_barrido_cuenta(2, "automatico")
                try:
                    await _auto_envio_aprobaciones()
                except Exception as ex:
                    logging.warning(f"auto envio aprobaciones: {ex}")
        except Exception as e:
            logging.warning(f"cuenta barrido loop: {e}")
        await asyncio.sleep(1200)


# ── AUDITORÍA DE BÓVEDA (Regla de Hierro): RUT en 4 fuentes + campos de identidad ──
def _rut_en_pdfs(nombre):
    """Busca un RUT válido (DV verificado) en el texto de los PDFs del cliente."""
    from base_historica import validar_rut_chileno as _v
    base = fsvc.folder_dir(nombre or "")
    if not base.exists():
        return "", ""
    from pypdf import PdfReader
    for p in sorted(base.rglob("*.pdf"))[:40]:
        try:
            rd = PdfReader(str(p))
            texto = ""
            for pg in rd.pages[:5]:
                texto += pg.extract_text() or ""
            for m in RUT_RE.findall(texto):
                if _v(m):
                    return m, p.name
        except Exception:
            continue
    return "", ""


async def _buscar_rut_cliente(fd, usados):
    """Orden obligatorio: 1) ficha ADN → 2) EXPEDIENTE_360 → 3) documentos → 4) correos 90d."""
    from base_historica import validar_rut_chileno as _v
    import json as _json
    nombre = (fd.get("nombre") or "").strip()
    toks = [t for t in nombre.lower().split() if len(t) > 2][:2]
    reg = None
    if toks:
        rx = ".*".join(re.escape(t) for t in toks)
        reg = await db.adn_clientes_360.find_one({"identidad.nombre": {"$regex": rx, "$options": "i"}})
    if reg and _v(reg.get("rut")) and _rut_limpio(reg["rut"]) not in usados:
        return reg["rut"], "ficha ADN_CLIENTES_360"
    if reg:
        blob = _json.dumps(reg.get("expediente_360") or {}, ensure_ascii=False, default=str)
        for m in RUT_RE.findall(blob):
            if _v(m) and _rut_limpio(m) not in usados:
                return m, "EXPEDIENTE_360"
    rut_doc, archivo = await asyncio.to_thread(_rut_en_pdfs, nombre)
    if rut_doc and _rut_limpio(rut_doc) not in usados:
        return rut_doc, f"documento ({archivo})"
    try:
        correos = await asyncio.to_thread(mail.fetch_since_text, 90, nombre, 15)
        for e in correos or []:
            texto = f"{e.get('subject', '')} {e.get('body', '')}"
            for m in RUT_RE.findall(texto):
                if _v(m) and _rut_limpio(m) not in usados:
                    return m, f"correo ({(e.get('subject') or '')[:60]})"
    except Exception:
        pass
    return "", ""


async def _auditoria_boveda_run(user_sub):
    """Audita los clientes activos de la flota: RUT (4 fuentes, escritura obligatoria)
    + campos de identidad. Cada hallazgo se ESCRIBE en la Bóveda ADN (Regla #66/#67)."""
    from base_historica import validar_rut_chileno as _v, _fmt_rut
    await db.config.update_one({"_key": "auditoria_boveda"}, {"$set": {
        "estado": "en_proceso", "inicio": _now()}}, upsert=True)
    try:
        flota_cfg = await db.config.find_one({"_key": "flota_agosto"}) or {}
        flota = {n.strip().upper() for n in (flota_cfg.get("nombres") or [])}
        folders = []
        async for fd in db.folders.find({"oculto_supercarpeta": {"$exists": False}}):
            nu = (fd.get("nombre") or "").strip().upper()
            if not flota or any(nf in nu or nu in nf for nf in flota):
                folders.append(fd)
        usados = {_rut_limpio(f.get("rut")) for f in folders if _v(f.get("rut"))}
        rep = {"clientes_auditados": len(folders), "con_rut_boveda": 0, "sin_rut_inicial": 0,
               "ruts_encontrados": [], "ruts_por_confirmar": [],
               "campos_vacios": 0, "campos_poblados": 0, "detalle_campos": [], "fecha": _now()}
        for fd in folders:
            ahora = _now()
            if _v(fd.get("rut")):
                rep["con_rut_boveda"] += 1
                await _sync_adn(fd["id"])
            else:
                rep["sin_rut_inicial"] += 1
                rut, fuente = await _buscar_rut_cliente(fd, usados)
                if rut:
                    rut_f = _fmt_rut(rut)
                    usados.add(_rut_limpio(rut))
                    await db.folders.update_one({"id": fd["id"]}, {"$set": {
                        "rut": rut_f, "rut_origen": f"auditoria_boveda:{fuente}", "updated_at": ahora}})
                    await db.estado_manual_log.insert_one({
                        "id": str(uuid.uuid4()), "folder_id": fd["id"], "cliente": fd.get("nombre") or "",
                        "hito": "identidad_rut", "estado_anterior": str(fd.get("rut") or ""),
                        "estado_nuevo": rut_f, "por": f"auditoria_boveda ({user_sub})",
                        "fecha": ahora, "inmutable": True})
                    await _sync_adn(fd["id"])
                    rep["ruts_encontrados"].append({"cliente": fd.get("nombre"), "rut": rut_f, "fuente": fuente})
                else:
                    await db.alertas.insert_one({"id": str(uuid.uuid4()), "tipo": "auditoria_boveda",
                        "mensaje": f"🔴 RUT Por Confirmar: {fd.get('nombre')} — no está en ficha ADN, "
                                   f"EXPEDIENTE_360, documentos ni correos de 90 días. Requiere ingreso manual.",
                        "fecha": ahora, "leida": False})
                    rep["ruts_por_confirmar"].append(fd.get("nombre"))
            # ── Campos de identidad: si está vacío en carpeta pero existe en ADN, se escribe ──
            rut_actual = fd.get("rut")
            if not _v(rut_actual):
                rut_actual = next((x["rut"] for x in rep["ruts_encontrados"]
                                   if x["cliente"] == fd.get("nombre")), None)
            reg = await db.adn_clientes_360.find_one({"rut_norm": _adn_norm(rut_actual)}) if rut_actual else None
            prop = (reg or {}).get("propiedad") or {}
            exp_p = ((reg or {}).get("expediente_360") or {}).get("propiedad") or {}
            ident = (reg or {}).get("identidad") or {}
            fuentes_campo = {
                "inmobiliaria": prop.get("inmobiliaria") or exp_p.get("inmobiliaria") or "",
                "proyecto": prop.get("proyecto") or exp_p.get("proyecto") or "",
                "ciudad": ident.get("ciudad") or exp_p.get("comuna") or prop.get("comuna") or "",
                "broker_origen": "Mutuaria y Leasing Limitada" if flota else "",
                "notaria": exp_p.get("notaria") or ""}
            upd = {}
            for campo, valor in fuentes_campo.items():
                if fd.get(campo):
                    continue
                if campo == "inmobiliaria" and "usad" in (fd.get("tipo_operacion") or "").lower():
                    continue
                rep["campos_vacios"] += 1
                if valor:
                    upd[campo] = valor
                    rep["campos_poblados"] += 1
                    rep["detalle_campos"].append({"cliente": fd.get("nombre"), "campo": campo, "valor": valor})
            fin_a = (reg or {}).get("financiero") or {}
            if not fd.get("proyeccion_uf"):
                rep["campos_vacios"] += 1
                if fin_a.get("monto_credito_uf"):
                    upd["proyeccion_uf"] = fin_a["monto_credito_uf"]
                    rep["campos_poblados"] += 1
                    rep["detalle_campos"].append({"cliente": fd.get("nombre"), "campo": "monto_uf",
                                                  "valor": fin_a["monto_credito_uf"]})
            if not fd.get("subsidio_proyeccion"):
                rep["campos_vacios"] += 1
                if fin_a.get("con_subsidio") is not None:
                    sub_v = "Con Subsidio" if fin_a["con_subsidio"] else "Sin Subsidio"
                    upd["subsidio_proyeccion"] = sub_v
                    rep["campos_poblados"] += 1
                    rep["detalle_campos"].append({"cliente": fd.get("nombre"), "campo": "subsidio", "valor": sub_v})
            if upd:
                await db.folders.update_one({"id": fd["id"]}, {"$set": {**upd, "updated_at": ahora}})
                await _sync_adn(fd["id"])
        rep["requieren_ingreso_manual"] = len(rep["ruts_por_confirmar"])
        await db.config.update_one({"_key": "auditoria_boveda"}, {"$set": {
            "estado": "completado", "ultima": _now(), "reporte": rep}}, upsert=True)
        return rep
    except Exception as ex:
        logging.warning(f"auditoria boveda: {ex}")
        await db.config.update_one({"_key": "auditoria_boveda"}, {"$set": {
            "estado": f"error: {str(ex)[:120]}", "ultima": _now()}}, upsert=True)
        return {"error": str(ex)[:200]}


def _adn_norm(rut):
    import adn_clientes as _adn
    return _adn._norm_rut(rut)


@supercarpeta.post("/auditoria-boveda")
async def auditoria_boveda_post(request: Request):
    """REGLA DE HIERRO: un cliente sin RUT es falla crítica — auditoría en segundo plano."""
    user = _exigir_gerencia(request)
    asyncio.create_task(_auditoria_boveda_run(user.get("sub") or "gerencia"))
    return {"ok": True, "lanzada": True, "seguimiento": "GET /api/supercarpeta/auditoria-boveda"}


@supercarpeta.get("/auditoria-boveda")
async def auditoria_boveda_get():
    return await db.config.find_one({"_key": "auditoria_boveda"}, {"_id": 0}) or {"estado": "nunca_ejecutada"}


async def migracion_reset_firmas():
    """LIMPIEZA DE ESTADOS FALSOS (una sola vez por BD): resetea TODA firma a Pendiente.
    Solo un correo fuente con evidencia real podrá volver a marcarla (bitácora incluida)."""
    if await db.config.find_one({"_key": "migracion_reset_firmas_v1"}):
        return
    reseteados = 0
    async for fd in db.folders.find({"$or": [
            {"escritura_confirmada_at": {"$exists": True}},
            {"firma_cesion_confirmada_at": {"$exists": True}},
            {"fecha_firma_detectada": {"$exists": True}},
            {"estado_manual.cesion.estado": {"$regex": "firm|confirm", "$options": "i"}},
            {"estado_manual.escritura.estado": {"$regex": "firm|confirm", "$options": "i"}}]}):
        unset = {k: "" for k in ("escritura_confirmada_at", "firma_cesion_confirmada_at",
                                 "fecha_firma_detectada") if fd.get(k)}
        em = fd.get("estado_manual") or {}
        for h in ("cesion", "escritura"):
            est = ((em.get(h) or {}).get("estado") or "")
            if re.search(r"firm|confirm", est, re.I):
                unset[f"estado_manual.{h}"] = ""
                await db.estado_manual_log.insert_one({
                    "id": str(uuid.uuid4()), "folder_id": fd["id"],
                    "cliente": fd.get("nombre") or "", "hito": h,
                    "estado_anterior": est, "estado_nuevo": "Pendiente",
                    "por": "sistema (limpieza de estados falsos de firma)",
                    "fecha": _now(), "inmutable": True})
        if unset:
            await db.folders.update_one({"id": fd["id"]}, {"$unset": unset})
            reseteados += 1
            await db.estado_manual_log.insert_one({
                "id": str(uuid.uuid4()), "folder_id": fd["id"],
                "cliente": fd.get("nombre") or "", "hito": "firma",
                "estado_anterior": "estado de firma sin respaldo de correo fuente",
                "estado_nuevo": "Pendiente",
                "por": "sistema (Regla de Hierro: la firma solo se confirma desde correo)",
                "fecha": _now(), "inmutable": True})
    await db.config.update_one({"_key": "migracion_reset_firmas_v1"}, {"$set": {
        "en": _now(), "carpetas_reseteadas": reseteados}}, upsert=True)
    logging.info(f"🧹 Reset de firmas falsas: {reseteados} carpeta(s) devueltas a Pendiente")


# ── REMITENTES DETECTADOS: revisión y control manual de la captura automática ──
@supercarpeta.get("/remitentes-detectados")
async def remitentes_detectados_get():
    items = []
    async for f in db.folders.find({"fuentes_detectadas": {"$exists": True}},
                                   {"id": 1, "nombre": 1, "fuentes_detectadas": 1}):
        for hito, lst in (f.get("fuentes_detectadas") or {}).items():
            for d in lst or []:
                items.append({**d, "cliente": f.get("nombre"), "folder_id": f["id"], "hito": hito})
    items.sort(key=lambda x: x.get("primera_vez", ""), reverse=True)
    bloqueados = (await db.config.find_one({"_key": "remitentes_bloqueados"}) or {}).get("correos") or []
    registro = [r async for r in db.remitentes_registro.find({}, {"_id": 0}).sort("ultima", -1).limit(50)]
    return {"detectados": items, "bloqueados": bloqueados, "registro": registro,
            "hitos_validos": list(FUENTES_HITOS)}


@supercarpeta.post("/remitentes-detectados/accion")
async def remitentes_accion(payload: dict, request: Request):
    """CONFIRMAR / REUBICAR / ELIMINAR / BLOQUEAR con aprendizaje acumulativo."""
    user = _exigir_gerencia(request)
    fid = payload.get("folder_id") or ""
    hito = (payload.get("hito") or "").strip()
    correo = (payload.get("correo") or "").strip().lower()
    accion = (payload.get("accion") or "").strip().lower()
    destino = (payload.get("hito_destino") or "").strip()
    if accion not in ("confirmar", "reubicar", "eliminar", "bloquear") or not correo:
        raise HTTPException(status_code=400, detail="Acción o correo inválido")
    fd = await db.folders.find_one({"id": fid})
    if not fd:
        raise HTTPException(status_code=404, detail="Carpeta no existe")
    await db.folders.update_one({"id": fid}, {"$pull": {f"fuentes_detectadas.{hito}": {"correo": correo}}})
    if accion == "confirmar":
        await db.folders.update_one({"id": fid}, {"$addToSet": {f"fuentes_doc.{hito}": correo}})
    elif accion == "reubicar":
        if destino not in FUENTES_HITOS:
            raise HTTPException(status_code=400, detail=f"Hito destino inválido ({', '.join(FUENTES_HITOS)})")
        await db.folders.update_one({"id": fid}, {"$addToSet": {f"fuentes_doc.{destino}": correo}})
        reg = await db.remitentes_registro.find_one({"correo": correo}) or {}
        seguidas = (reg.get("reubicaciones") or 0) + 1 if reg.get("ultimo_destino") == destino else 1
        upd = {"ultimo_destino": destino, "reubicaciones": seguidas}
        if seguidas >= 2:
            upd["hito_forzado"] = destino  # aprendizaje permanente tras 2 correcciones iguales
        await db.remitentes_registro.update_one({"correo": correo}, {"$set": upd}, upsert=True)
    elif accion == "bloquear":
        await db.config.update_one({"_key": "remitentes_bloqueados"},
                                   {"$addToSet": {"correos": correo}}, upsert=True)
    await _log_fuente(user, "remitente_detectado", destino or hito or "todas", accion, correo,
                      "captura automática", fd.get("nombre") or "")
    await _sync_adn(fid)
    return {"ok": True, "accion": accion, "correo": correo,
            "aprendizaje": "el sistema ajusta su criterio con cada corrección de Gerencia"}


@supercarpeta.post("/cliente")
async def supercarpeta_cliente_agregar(payload: dict, request: Request):
    """P10 — Gerencia agrega un cliente manualmente: crea la ficha y la suma a la vista.
    Válvula operativa (el flujo normal es la carga automática desde la proyección del broker)."""
    user = _exigir_gerencia(request)
    nombre = (payload.get("nombre") or "").strip().upper()
    if len(nombre) < 5 or len(nombre.split()) < 2:
        raise HTTPException(status_code=400, detail="Nombre inválido (nombre y apellido)")
    if await db.folders.find_one({"nombre": {"$regex": f"^{re.escape(nombre)}$", "$options": "i"},
                                  "oculto_supercarpeta": {"$exists": False}}):
        raise HTTPException(status_code=409, detail="Ese cliente ya existe en la Supercarpeta")
    monto = payload.get("monto_uf")
    try:
        monto = float(str(monto).replace("$", "").replace(" ", "").replace(".", "").replace(",", ".")) if monto else None
    except Exception:
        monto = None
    ahora = _now()
    fd = {"id": str(uuid.uuid4()), "nombre": nombre, "rut": (payload.get("rut") or "").strip(),
          "inmobiliaria": (payload.get("inmobiliaria") or "").strip(),
          "proyecto": (payload.get("proyecto") or "").strip(),
          "ciudad": (payload.get("ciudad") or "").strip(),
          "tipo_operacion": "usada" if "usad" in (payload.get("tipo_propiedad") or "").lower() else
                            ((payload.get("tipo_propiedad") or "").strip().lower() or "nueva"),
          "subsidio_proyeccion": (payload.get("subsidio") or "").strip(),
          "proyeccion_uf": monto,
          "broker_origen": (payload.get("broker") or "Mutuaria y Leasing Limitada").strip(),
          "mes_proyeccion": ((payload.get("mes") or "").strip() or "2026-08"),
          "archivos": [], "created": ahora, "updated_at": ahora,
          "origen_alta": {"manual": True, "por": user.get("sub") or "gerencia", "en": ahora}}
    await db.folders.insert_one(dict(fd))
    cfg = await db.config.find_one({"_key": "flota_agosto"}) or {}
    if cfg.get("activo") and fd["mes_proyeccion"] == "2026-08":
        await db.config.update_one({"_key": "flota_agosto"}, {"$addToSet": {"nombres": nombre}})
    await db.supercarpeta_log.insert_one({
        "id": str(uuid.uuid4()), "accion": "cliente_agregado", "folder_id": fd["id"],
        "cliente": nombre, "por": user.get("sub") or "gerencia", "fecha": ahora, "inmutable": True})
    await _sync_adn(fd["id"])
    return {"ok": True, "id": fd["id"], "cliente": nombre,
            "nota": "Ficha creada y agregada a la vista; queda en bitácora inmutable"}


@supercarpeta.post("/mes-siguiente/{fid}")
async def supercarpeta_mes_siguiente(fid: str, request: Request, payload: dict = None):
    """TRASLADO MENSUAL: mueve al cliente a la proyección del mes siguiente conservando
    TODOS sus datos y estados. NUNCA borra nada; la ficha ADN queda íntegra + bitácora."""
    import adn_clientes as _adn
    user = _exigir_gerencia(request)
    fd = await db.folders.find_one({"id": fid})
    if not fd:
        raise HTTPException(status_code=404, detail="Carpeta no existe")
    actual = fd.get("mes_proyeccion") or "2026-08"
    y, m = int(actual[:4]), int(actual[5:7])
    destino = (((payload or {}).get("mes_destino") or "").strip()
               or (f"{y + 1}-01" if m == 12 else f"{y}-{m + 1:02d}"))
    ahora = _now()
    await db.folders.update_one({"id": fid}, {"$set": {
        "mes_proyeccion": destino, "arrastre_desde": actual, "updated_at": ahora}})
    await db.estado_manual_log.insert_one({
        "id": str(uuid.uuid4()), "folder_id": fid, "cliente": fd.get("nombre") or "",
        "hito": "traslado_mes", "estado_anterior": actual, "estado_nuevo": destino,
        "por": user.get("sub") or "gerencia", "fecha": ahora, "inmutable": True})
    await db.config.update_one({"_key": f"proyeccion_{destino}"},
                               {"$setOnInsert": {"creada": ahora, "meta_uf": 0}}, upsert=True)
    if fd.get("rut"):
        await db.adn_clientes_360.update_one({"rut_norm": _adn._norm_rut(fd["rut"])}, {"$push": {
            "traslados": {"de": actual, "a": destino, "fecha": ahora,
                          "por": user.get("sub") or "gerencia"}}})
    await _sync_adn(fid)
    return {"ok": True, "cliente": fd.get("nombre"), "de": actual, "a": destino,
            "etiqueta": f"Arrastre {actual}",
            "regla": "El traslado nunca borra datos; la ficha ADN_CLIENTES_360 se conserva íntegra"}


@supercarpeta.post("/cliente/{fid}/eliminar")
async def supercarpeta_cliente_eliminar(fid: str, request: Request):
    """P10 — Elimina al cliente de la VISTA Supercarpeta. La ficha ADN_CLIENTES_360 se
    conserva como registro histórico. Bitácora inmutable con fecha, hora y usuario."""
    user = _exigir_gerencia(request)
    fd = await db.folders.find_one({"id": fid})
    if not fd:
        raise HTTPException(status_code=404, detail="Carpeta no existe")
    ahora = _now()
    await db.folders.update_one({"id": fid}, {"$set": {
        "oculto_supercarpeta": {"en": ahora, "por": user.get("sub") or "gerencia"},
        "updated_at": ahora}})
    nombre_u = (fd.get("nombre") or "").strip().upper()
    await db.config.update_one({"_key": "flota_agosto"}, {"$pull": {"nombres": nombre_u}})
    await db.supercarpeta_log.insert_one({
        "id": str(uuid.uuid4()), "accion": "cliente_eliminado_vista", "folder_id": fid,
        "cliente": fd.get("nombre") or "", "por": user.get("sub") or "gerencia",
        "fecha": ahora, "inmutable": True,
        "nota": "Solo se ocultó de la vista; la ficha ADN se conserva como histórico"})
    return {"ok": True, "cliente": fd.get("nombre"),
            "nota": "Eliminado de la Supercarpeta; ficha ADN conservada como registro histórico"}


@supercarpeta.get("/panel/{fid}")
async def supercarpeta_panel(fid: str, hito: str = "tasacion"):
    """SECCIÓN 5 — Panel Lateral: fuentes activas, correos detectados, notas, bitácora."""
    fd = await db.folders.find_one({"id": fid})
    if not fd:
        raise HTTPException(status_code=404, detail="Carpeta no existe")
    nombre = fd.get("nombre") or ""
    fu = await _fuentes_documentos()
    fuentes_hito = _correos_de((fd.get("fuentes_doc") or {}).get(hito)) or fu.get(hito) or []
    correos = await db.hitos_externos.find(
        {"folder_id": fid}, {"_id": 0, "hito": 1, "asunto": 1, "fecha": 1,
                             "fuente": 1, "direccion": 1, "creado": 1}).sort("creado", -1).to_list(15)
    rx = {"$regex": re.escape(nombre[:16]), "$options": "i"}
    envios = await db.correos_smtp_log.find(
        {"subject": rx}, {"_id": 0, "subject": 1, "to": 1, "fecha": 1, "success": 1}).sort("fecha", -1).to_list(8)
    log = await db.estado_manual_log.find(
        {"folder_id": fid, "hito": hito}, {"_id": 0}).sort("fecha", -1).to_list(30)
    notas = ((fd.get("notas_estados") or {}).get(hito)) or []
    bit = None
    if hito in ("tasacion", "estudio"):
        try:
            bit = await _bitacora_hito(fd, hito)
        except Exception:
            bit = None
    reparos = ""
    if hito == "estudio":
        reparos = " | ".join((r.get("texto") or "")[:300]
                             for r in (fd.get("reparos_alertas") or []) if not r.get("resuelto"))[:900]
    return {"cliente": nombre, "hito": hito, "fuentes": fuentes_hito,
            "correos_detectados": correos, "envios": envios,
            "notas": notas, "bitacora_cambios": log, "bitacora_tiempos": bit,
            "detalle_reparos": reparos,
            "estado_manual": (fd.get("estados_manuales") or {}).get(hito) or {}}


async def _hilo_eventos(fd):
    """Construye la línea de tiempo unificada de correos de un cliente (enviados + recibidos)."""
    eventos = []
    # ENVIADOS — solicitudes desde la Supercarpeta (carta oferta / RS / compromiso, etc.)
    for r in (fd.get("bitacora_solicitudes") or []):
        eventos.append({"tipo": "enviado", "en": r.get("en") or "",
                        "asunto": r.get("asunto") or "", "con": r.get("para") or "",
                        "detalle": " + ".join(r.get("documentos") or []) or (r.get("tipo") or ""),
                        "estado": r.get("estado") or "enviado"})
    # ENVIADOS — estudio de título (log del módulo)
    nombre = (fd.get("nombre") or "").strip()
    async for r in db.estudio_titulo_log.find({"nombre": {"$regex": f"^{re.escape(nombre)}$", "$options": "i"}}):
        to = r.get("to") or []
        eventos.append({"tipo": "enviado", "en": r.get("enviado_en") or "",
                        "asunto": f"Solicitud de Antecedentes - Estudio de Título ({r.get('tipo_vivienda') or ''})",
                        "con": ", ".join(to) if isinstance(to, list) else str(to),
                        "detalle": "Estudio de Título", "estado": "enviado"})
    # RECIBIDOS — correos detectados por la malla (hitos externos)
    async for h in db.hitos_externos.find({"folder_id": fd["id"]}):
        eventos.append({"tipo": "recibido", "en": h.get("creado") or h.get("fecha") or "",
                        "asunto": h.get("asunto") or "", "con": h.get("fuente") or h.get("dominio") or "",
                        "detalle": h.get("hito") or "", "estado": "recibido",
                        "adjuntos": h.get("adjuntos") or []})
    eventos.sort(key=lambda e: e.get("en") or "", reverse=True)
    # ADJUNTOS RETROACTIVOS: archivos ya archivados sin vínculo se muestran en el
    # evento MÁS RECIENTE de su tipo (los nuevos correos quedan vinculados 1 a 1)
    pref_map = [("estudio", ["07_estudio_titulo/ESTUDIO_"]), ("tasaci", ["99_otros/TASACION_"]),
                ("set de crédito", ["99_otros/SETCRED_"]), ("promesa", ["99_otros/PROMESA_"]),
                ("carta oferta", ["99_otros/CARTA_OFERTA_"]), ("notaría", ["99_otros/NOTARIA_"]),
                ("vendedor", ["07_estudio_titulo/VENDEDOR_"]),
                ("subsidio", ["99_otros/CERT_SUBSIDIO_"]), ("carta pie", ["99_otros/CARTA_PIE_"])]
    archivos_fd = fd.get("archivos") or []
    ya_vinculados = {a for e in eventos for a in (e.get("adjuntos") or [])}
    for ev in eventos:
        if ev["tipo"] != "recibido" or ev.get("adjuntos"):
            continue
        det = (ev.get("detalle") or "").lower()
        for k, prefs in pref_map:
            if k in det:
                m = [a for a in archivos_fd
                     if any(a.startswith(p) for p in prefs) and a not in ya_vinculados]
                if m:
                    ev["adjuntos"] = m
                    ya_vinculados.update(m)
                break
    return eventos


@supercarpeta.get("/hilo/{fid}")
async def supercarpeta_hilo(fid: str, request: Request):
    """HILO DEL CLIENTE: línea de tiempo unificada de correos enviados y recibidos."""
    _exigir_gerencia(request)
    fd = await db.folders.find_one({"id": fid})
    if not fd:
        raise HTTPException(status_code=404, detail="Carpeta no existe")
    eventos = await _hilo_eventos(fd)
    return {"cliente": (fd.get("nombre") or "").strip(), "eventos": eventos, "total": len(eventos),
            "enviados": sum(1 for e in eventos if e["tipo"] == "enviado"),
            "recibidos": sum(1 for e in eventos if e["tipo"] == "recibido")}


# ─── RESUMEN DEL HILO IA: una línea que dice en qué quedó la conversación ───
def _firma_eventos(eventos):
    base = "|".join(f"{e.get('en')}~{e.get('tipo')}~{e.get('asunto')}" for e in eventos[:15])
    return hashlib.sha256(base.encode()).hexdigest()[:24]


async def _resumen_hilo_generar(fd, eventos):
    """Genera con IA el resumen de una línea del hilo del cliente y lo guarda en la carpeta.
    NORMATIVA FIJA: solo considera eventos de los últimos 90 días."""
    key = os.environ.get("EMERGENT_LLM_KEY", "")
    from datetime import timedelta
    corte = (datetime.now(timezone.utc) - timedelta(days=90)).isoformat()
    eventos = [e for e in eventos if str(e.get("en") or "") >= corte]
    if not key or not eventos:
        return None
    lineas = []
    for e in eventos[:12]:
        rot = f"ENVIADO a {e.get('con') or '—'}" if e.get("tipo") == "enviado" else f"RECIBIDO de {e.get('con') or '—'}"
        lineas.append(f"[{str(e.get('en') or '')[:10]}] {rot}: {e.get('asunto') or e.get('detalle') or ''}"
                      f" ({e.get('detalle') or ''} · {e.get('estado') or ''})")
    hoy = datetime.now(timezone.utc).strftime("%d/%m/%Y")
    try:
        from emergentintegrations.llm.chat import LlmChat, UserMessage
        chat = LlmChat(
            api_key=key, session_id=f"resumen-hilo-{fd['id']}",
            system_message=("Eres el asistente de gestión hipotecaria DashAI. Recibes los correos "
                            "enviados y recibidos de un cliente (del más nuevo al más antiguo). "
                            "Responde SOLO con UNA línea en español (máx 160 caracteres) que resuma "
                            "en qué quedó la conversación: estado actual + quién debe el próximo paso "
                            "+ fecha del último movimiento en formato dd/mm. Ejemplo: 'Esperando "
                            "respuesta de la inmobiliaria por Carta Oferta desde el 12/06'. "
                            "Sin comillas, sin viñetas, sin texto adicional.")
        ).with_model("openai", "gpt-5.4-mini")
        resp = await asyncio.wait_for(
            chat.send_message(UserMessage(text=f"HOY: {hoy}\nCLIENTE: {fd.get('nombre') or ''}\n\n" + "\n".join(lineas))),
            timeout=60)
        texto = (str(resp) or "").strip().strip('"').replace("\n", " ")[:200]
        from constitucion import consultar_cerebro
        await consultar_cerebro(db, "resumen_hilo_ia", texto_ia=texto, modulo="malla_inteligencia.py")
    except Exception as e:
        logging.warning(f"resumen hilo IA ({fd.get('nombre')}): {e}")
        return None
    if not texto:
        return None
    reg = {"texto": texto, "en": _now(), "firma": _firma_eventos(eventos)}
    await db.folders.update_one({"id": fd["id"]}, {"$set": {"resumen_hilo": reg}})
    return reg


@supercarpeta.post("/resumen-hilo/{fid}")
async def supercarpeta_resumen_hilo(fid: str, request: Request):
    """Botón 'Regenerar': fuerza la actualización del resumen IA del hilo del cliente."""
    _exigir_gerencia(request)
    fd = await db.folders.find_one({"id": fid})
    if not fd:
        raise HTTPException(status_code=404, detail="Carpeta no existe")
    eventos = await _hilo_eventos(fd)
    if not eventos:
        return {"ok": False, "texto": "", "nota": "Sin correos registrados: no hay hilo que resumir"}
    reg = await _resumen_hilo_generar(fd, eventos)
    if not reg:
        raise HTTPException(status_code=502,
                            detail="No fue posible generar el resumen (sin correos en los últimos 90 días o error de IA)")
    return {"ok": True, **reg}


async def resumen_hilo_loop():
    """Cada 15 min: si un cliente tiene correos nuevos en su hilo, regenera el resumen IA."""
    await asyncio.sleep(420)
    while True:
        try:
            async for fd in db.folders.find({"oculto_supercarpeta": {"$ne": True}}):
                eventos = await _hilo_eventos(fd)
                if not eventos:
                    continue
                if _firma_eventos(eventos) == (fd.get("resumen_hilo") or {}).get("firma"):
                    continue
                await _resumen_hilo_generar(fd, eventos)
                await asyncio.sleep(3)
        except Exception as e:
            logging.warning(f"resumen hilo loop: {e}")
        await asyncio.sleep(900)


@supercarpeta.post("/nota/{fid}")
async def supercarpeta_nota(fid: str, payload: dict, request: Request):
    """SECCIÓN 5 — Notas manuales por estado, guardadas en la Bóveda ADN."""
    user = getattr(request.state, "user", {}) or {}
    hito = (payload.get("hito") or "").strip().lower()
    texto = (payload.get("texto") or "").strip()
    if hito not in HITOS_VALIDOS or not texto or len(texto) > 600:
        raise HTTPException(status_code=400, detail="Hito o nota inválida")
    fd = await db.folders.find_one({"id": fid})
    if not fd:
        raise HTTPException(status_code=404, detail="Carpeta no existe")
    nota = {"texto": texto, "por": user.get("sub") or "usuario", "en": _now()}
    await db.folders.update_one({"id": fid}, {"$push": {f"notas_estados.{hito}": nota}})
    await _sync_adn(fid)
    return {"ok": True, "nota": nota}


@supercarpeta.get("/archivo/{fid}")
async def supercarpeta_archivo(fid: str, ruta: str):
    """VISUALIZADOR RÁPIDO: sirve el PDF para previsualizar sin entrar a la ficha."""
    fd = await db.folders.find_one({"id": fid})
    if not fd:
        raise HTTPException(status_code=404, detail="Carpeta no existe")
    base = fsvc.folder_dir(fd.get("nombre") or "").resolve()
    p = (base / ruta).resolve()
    if not str(p).startswith(str(base)) or not p.exists() or not p.name.lower().endswith(".pdf"):
        raise HTTPException(status_code=404, detail="Informe no disponible")
    return Response(content=p.read_bytes(), media_type="application/pdf",
                    headers={"Content-Disposition": f'inline; filename="{p.name}"'})
